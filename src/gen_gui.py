"""
gen_gui.py — FPGA 工具箱 GUI (ttkbootstrap 现代化主题)
被 gen_inst.py 的 run_gui() 调用

路径说明:
  - _PROJECT_ROOT = gen_gui.py 所在 src/ 的上级, 即工程根目录
  - 所有对 ip_docs/ _pyserial_lib/ runtime/ assets/ 等的引用都基于此
"""
import tkinter as tk
import os
import re
import csv
import threading
import datetime
import subprocess
import json
import shutil
import sys as _sys
import io as _io

# 工程根目录 (gen_gui.py 在 src/ 子目录, 根目录是它的上级)
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__ or '.')))
from tkinter import filedialog, messagebox

# ── subprocess 编码修复 ──
# Windows 中文系统默认 GBK, gh/git/curl 等输出含 UTF-8 时会触发
# UnicodeDecodeError, 把 subprocess.run 默认编码改为 utf-8 + errors='replace'
# 此外, 对已知会读用户配置/SSH known_hosts 的命令 (git / ssh / gh / glab)
# 自动重试 3 次, 间隔 200ms, 解决内网环境下 WinError 32
# ("另一个程序正在使用此文件") 导致的偶发失败.
_orig_subprocess_run = subprocess.run


def _patched_subprocess_run(*args, **kwargs):
    """默认 text=True 时强制 utf-8 + replace, 避免 GBK 解码崩溃.
    对 git 命令自动注入 GIT_SSH_COMMAND 跳过 host key 验证."""
    if kwargs.get('text') or kwargs.get('universal_newlines'):
        kwargs.setdefault('encoding', 'utf-8')
        kwargs.setdefault('errors', 'replace')

    # 决定是否启用 WinError 32 重试
    cmd = None
    if args and isinstance(args[0], (list, tuple)) and args[0]:
        cmd = str(args[0][0]).lower()
    elif kwargs.get('args') and isinstance(kwargs['args'], (list, tuple)) and kwargs['args']:
        cmd = str(kwargs['args'][0]).lower()
    retryable = cmd in {'git', 'ssh', 'gh', 'glab', 'curl'}

    # ── git/ssh 命令自动跳过 host key 验证(内网 IP/GitLab 常见问题) ──
    # BatchMode=yes 禁止交互输入(避免卡在 passphrase 提示), ConnectTimeout 快速失败
    if cmd in {'git', 'ssh'}:
        env = kwargs.pop('env', None)
        if env is None:
            env = os.environ.copy()
        else:
            env = {**os.environ, **env}
        env.setdefault('GIT_SSH_COMMAND',
            'ssh -o BatchMode=yes -o ConnectTimeout=10'
            ' -o StrictHostKeyChecking=accept-new -o UserKnownHostsFile=NUL')
        kwargs['env'] = env

    if not retryable:
        return _orig_subprocess_run(*args, **kwargs)

    import time as _time
    last_err = None
    for attempt in range(3):
        try:
            return _orig_subprocess_run(*args, **kwargs)
        except (PermissionError, OSError) as e:
            # 只对 WinError 32 (errno 13, "另一个程序正在使用此文件") 重试
            winerr = getattr(e, 'winerror', None)
            errno = getattr(e, 'errno', None)
            is_lock_err = (winerr == 32) or (errno == 13) or (errno == 32)
            if not is_lock_err or attempt == 2:
                last_err = e
                break
            _time.sleep(0.2)
    # 3 次都失败, 抛出原始异常, 让调用方的 except 接住
    if last_err is not None:
        raise last_err
    return _orig_subprocess_run(*args, **kwargs)


subprocess.run = _patched_subprocess_run




# ── 向上查找 .git 目录, 返回仓库根路径; 找不到返回 None ──
def _find_git_root(start):
    d = os.path.abspath(start)
    for _ in range(10):
        if os.path.isdir(os.path.join(d, '.git')):
            return d
        parent = os.path.dirname(d)
        if parent == d:
            break
        d = parent
    return None

# ── git config 安全写入: 解决内网环境下 ~/.gitconfig 被其它工具占用 (WinError 32) ──
def _git_set_identity(user_name, user_email, scope='auto', repo=None):
    """
    安全地写入 Git 身份 (user.name / user.email).
    策略:
      1) scope='auto' 时, 优先写到 repo/.git/config (--local),
         避免抢占 C:\\Users\\<user>\\.gitconfig 文件锁;
      2) repo 为 None 时回退到 --global;
      3) 任何 PermissionError (WinError 32) / OSError / 超时一律吞掉,
         并对 --global 冷却重试 1 次 (其它进程刚好释放锁的情况);
      4) 全失败返回 False, 让上层弹友好提示, 不再裸 traceback.
    返回: True=成功写入至少一处, False=完全失败.
    """
    import time as _time

    scopes = []
    if scope == 'auto':
        if repo:
            scopes.append(('--local', repo))
        scopes.append(('--global', None))
    elif scope == 'local':
        if not repo:
            return False
        scopes.append(('--local', repo))
    else:  # global
        scopes.append(('--global', None))

    # ── 内网环境下 C:\Users\<user>\.gitconfig 经常被 IDE 插件 / 杀软长时间占用,
    #     任何 --global 操作都会撞 WinError 32.  兜底策略: 把 user.name/email
    #     写入一份独立的 include 文件 ~/.gitconfig_fpga_tool, 然后在主
    #     .gitconfig 里加一行 [include] path = "..." 指向它.  写入 include
    #     文件本身不会触发主 .gitconfig 的文件锁 (因为它就是普通文本). ──
    def _try_include_file():
        try:
            inc = os.path.join(os.path.expanduser('~'), '.gitconfig_fpga_tool')
            with open(inc, 'w', encoding='utf-8') as f:
                f.write(f'[user]\n\tname = {user_name}\n\temail = {user_email}\n')
            # 用 git config --global include.path ... 把它注册进去
            r = subprocess.run(
                ['git', 'config', '--global', '--add', 'include.path', inc.replace('\\', '/')],
                capture_output=True, timeout=5)
            return r.returncode == 0
        except (PermissionError, OSError, subprocess.TimeoutExpired):
            return False

    def _try_once(flag, cwd):
        try:
            r1 = subprocess.run(['git', 'config', flag, 'user.name', user_name],
                                capture_output=True, cwd=cwd, timeout=5)
            r2 = subprocess.run(['git', 'config', flag, 'user.email', user_email],
                                capture_output=True, cwd=cwd, timeout=5)
            return r1.returncode == 0 and r2.returncode == 0
        except (PermissionError, OSError, subprocess.TimeoutExpired):
            return None  # None 表示被锁, 与"returncode != 0"区别开

    any_ok = False
    for flag, cwd in scopes:
        result = _try_once(flag, cwd)
        if result is True:
            any_ok = True
            # ── 关键: --local 已经写成功, 就不要再动 --global ──
            # 否则会覆盖用户在主目录 .gitconfig 中已设置的其它身份
            if flag == '--local':
                return True
            continue
        if result is None and flag == '--global':
            # ── 冷却重试: 其它进程 (VSCode 等) 偶尔短暂锁住, 200ms 后通常释放 ──
            _time.sleep(0.2)
            result = _try_once(flag, cwd)
            if result is True:
                any_ok = True
                continue
        # result 是 False/None → 此 scope 失败, 继续下一个 scope (--local 兜底)

    # ── 终极兜底: 上面全部失败, 说明 ~/.gitconfig 被强占, 改写独立 include 文件.
    #     这种方式下 user.name/email 仍然对 git 全局生效, 但不再触碰主 .gitconfig. ──
    if not any_ok:
        if _try_include_file():
            any_ok = True
    return any_ok


def _get_t3_repo_path():
    """
    获取当前 GUI 上下文的"主仓库"路径, 用于 git config --local.
    优先: GUI 上"项目目录"输入框 > 脚本所在目录的上一级 (项目根).
    找不到 .git 时返回 None (调用方应回退到 --global).
    """
    # 1) GUI 上的 t3_dir_var 输入框 (顶栏"项目目录")
    try:
        v = t3_dir_var.get().strip()  # noqa: F821 — 此函数定义在 t3 之前,
                                       # 但 _git_set_identity 调用点都在
                                       # t3_do_login / t3_do_logout 里,
                                       # 那时候 t3_dir_var 一定已存在.
        if v and os.path.isdir(v) and os.path.isdir(os.path.join(v, '.git')):
            return v
    except (NameError, AttributeError):
        pass
    # 2) 脚本所在目录的上一级 (FPGA Toolbox 项目根)
    try:
        proj_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        if os.path.isdir(os.path.join(proj_root, '.git')):
            return proj_root
    except Exception:
        pass
    return None


# ── 永远不尝试 pip install, 直接用标准 ttk ──
_HAS_BOOTSTRAP = False
from tkinter import ttk


def _gen_tcl_wrapper(xdc_text, pins_data):
    """把 XDC 包成 Vivado Tcl 脚本 (source <file>.tcl 直接加载)"""
    header = f'''# ═══════════════════════════════════════════════════════════
# Vivado Pin Constraints (Tcl wrapper)
# Generated by FPGA Toolbox v5.0
# Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
# Pins: {len(pins_data)}
# ═══════════════════════════════════════════════════════════

# 用法: 在 Vivado Tcl Console 中执行
#   source <本文件>.tcl
# 或在 Vivado GUI: Tools → Run Tcl Script

'''
    # 把 XDC 嵌进 readback-writefile 流程
    tcl = []
    tcl.append(header)
    tcl.append('# 临时文件路径 (与本 tcl 同目录)')
    tcl.append('set _xdc_tmp "[file dirname [file normalize [info script]]]/_generated_pins.xdc"')
    tcl.append('')
    tcl.append('# --- 写入 XDC ---')
    # 把 xdc_text 逐行 quote
    tcl.append('set _fp [open $_xdc_tmp w]')
    for line in xdc_text.split('\n'):
        # 转义: \ → \\, " → \", [ → \[
        esc = line.replace('\\', '\\\\').replace('"', '\\"').replace('[', '\\[').replace('$', '\\$')
        tcl.append(f'puts $_fp "{esc}"')
    tcl.append('close $_fp')
    tcl.append('')
    tcl.append('# --- 读取并应用约束 ---')
    tcl.append('if {{[catch {{read_xdc $_xdc_tmp}} err]}} {{')
    tcl.append('    puts "WARNING: read_xdc failed: $err"')
    tcl.append('} else {{')
    tcl.append('    puts "✔ Loaded [llength [get_ports]] ports from $::argv0"')
    tcl.append('}}')
    tcl.append('')
    tcl.append('# --- 清理临时文件 ---')
    tcl.append('file delete -force $_xdc_tmp')
    tcl.append('puts "✔ Pin constraints applied."')
    return '\n'.join(tcl)


def _build_html_report(pins_data, xdc_text, bank_log, src_file):
    """构建完整 HTML 报告 (管脚表 + Bank 检查 + XDC)"""
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    rows_html = []
    for p in pins_data:
        cells = ''.join(
            f'<td>{html_mod.escape(str(p.get(c, "")))}</td>'
            for c in ('port', 'pin', 'dir', 'iostd', 'bank', 'voltage', 'pull', 'note')
        )
        rows_html.append(f'<tr>{cells}</tr>')

    # 统计
    n_total = len(pins_data)
    n_input = sum(1 for p in pins_data if p.get('dir', '').lower() == 'input')
    n_output = sum(1 for p in pins_data if p.get('dir', '').lower() == 'output')
    n_inout = sum(1 for p in pins_data if p.get('dir', '').lower() == 'inout')
    banks = sorted({p.get('bank', '') for p in pins_data if p.get('bank')})

    return f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<title>FPGA 管脚约束报告</title>
<style>
  body {{ font-family: -apple-system, "Microsoft YaHei", sans-serif;
          margin: 0; padding: 24px; background: #f4f6f9; color: #1f2329; }}
  .container {{ max-width: 1100px; margin: 0 auto; background: #fff;
                padding: 32px; border-radius: 8px;
                box-shadow: 0 1px 3px rgba(0,0,0,0.08); }}
  h1 {{ color: #2b6cb0; margin-top: 0; }}
  h2 {{ color: #1f2329; border-bottom: 2px solid #2b6cb0;
        padding-bottom: 6px; margin-top: 32px; }}
  .meta {{ color: #6c757d; font-size: 14px; margin-bottom: 24px; }}
  .stats {{ display: grid; grid-template-columns: repeat(5, 1fr);
            gap: 12px; margin: 20px 0; }}
  .stat {{ background: #f4f6f9; padding: 16px; border-radius: 6px;
           text-align: center; }}
  .stat-num {{ font-size: 24px; font-weight: bold; color: #2b6cb0; }}
  .stat-label {{ color: #6c757d; font-size: 12px; margin-top: 4px; }}
  table {{ width: 100%; border-collapse: collapse; margin: 12px 0;
           font-size: 14px; }}
  th, td {{ padding: 8px 12px; text-align: left;
            border-bottom: 1px solid #e4e7ec; }}
  th {{ background: #f4f6f9; color: #1f2329; font-weight: bold; }}
  tr:hover {{ background: #f8f9fa; }}
  pre {{ background: #1e1e1e; color: #d4d4d4; padding: 16px;
         border-radius: 6px; overflow-x: auto; font-size: 12px;
         line-height: 1.5; }}
  .bank-log {{ background: #f4f6f9; padding: 12px; border-radius: 6px;
               white-space: pre-wrap; font-size: 13px; }}
  .footer {{ color: #6c757d; font-size: 12px; text-align: center;
             margin-top: 32px; padding-top: 16px; border-top: 1px solid #e4e7ec; }}
</style>
</head>
<body>
<div class="container">
  <h1>⚡ FPGA 管脚约束报告</h1>
  <div class="meta">
    生成时间: {now} &nbsp;|&nbsp;
    源文件: {html_mod.escape(src_file or 'N/A')} &nbsp;|&nbsp;
    由 FPGA Toolbox v5.0 生成
  </div>

  <div class="stats">
    <div class="stat"><div class="stat-num">{n_total}</div><div class="stat-label">总引脚</div></div>
    <div class="stat"><div class="stat-num">{n_input}</div><div class="stat-label">Input</div></div>
    <div class="stat"><div class="stat-num">{n_output}</div><div class="stat-label">Output</div></div>
    <div class="stat"><div class="stat-num">{n_inout}</div><div class="stat-label">Inout</div></div>
    <div class="stat"><div class="stat-num">{len(banks)}</div><div class="stat-label">Banks</div></div>
  </div>

  <h2>📋 管脚分配</h2>
  <table>
    <thead>
      <tr>
        <th>Port</th><th>Pin</th><th>Direction</th><th>IO Std</th>
        <th>Bank</th><th>Voltage</th><th>Pull</th><th>Note</th>
      </tr>
    </thead>
    <tbody>
      {''.join(rows_html)}
    </tbody>
  </table>

  <h2>🔍 Bank 电压检查</h2>
  <div class="bank-log">{html_mod.escape(bank_log)}</div>

  <h2>📝 XDC 约束 (源)</h2>
  <pre>{html_mod.escape(xdc_text)}</pre>

  <div class="footer">
    FPGA Toolbox v5.0 · HDL 例化 · 工程压缩 · Git 提交流水线
  </div>
</div>
</body>
</html>
'''


def _build_md_report(pins_data, xdc_text, bank_log, src_file):
    """构建完整 Markdown 报告 (适合 GitLab / 文档)"""
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    n_total = len(pins_data)
    n_input = sum(1 for p in pins_data if p.get('dir', '').lower() == 'input')
    n_output = sum(1 for p in pins_data if p.get('dir', '').lower() == 'output')
    n_inout = sum(1 for p in pins_data if p.get('dir', '').lower() == 'inout')
    banks = sorted({p.get('bank', '') for p in pins_data if p.get('bank')})

    lines = []
    lines.append('# ⚡ FPGA 管脚约束报告')
    lines.append('')
    lines.append(f'- **生成时间**: {now}')
    lines.append(f'- **源文件**: `{src_file or "N/A"}`')
    lines.append(f'- **工具**: FPGA Toolbox v5.0')
    lines.append('')
    lines.append('## 📊 统计')
    lines.append('')
    lines.append('| 指标 | 数量 |')
    lines.append('| --- | ---: |')
    lines.append(f'| 总引脚 | {n_total} |')
    lines.append(f'| Input | {n_input} |')
    lines.append(f'| Output | {n_output} |')
    lines.append(f'| Inout | {n_inout} |')
    lines.append(f'| Bank 数 | {len(banks)} |')
    lines.append('')
    lines.append('## 📋 管脚分配')
    lines.append('')
    lines.append('| Port | Pin | Direction | IO Std | Bank | Voltage | Pull | Note |')
    lines.append('| --- | --- | --- | --- | --- | --- | --- | --- |')
    for p in pins_data:
        cells = [p.get(c, '') for c in
                 ('port', 'pin', 'dir', 'iostd', 'bank', 'voltage', 'pull', 'note')]
        lines.append('| ' + ' | '.join(cells) + ' |')
    lines.append('')
    lines.append('## 🔍 Bank 电压检查')
    lines.append('')
    lines.append('```')
    lines.append(bank_log.rstrip())
    lines.append('```')
    lines.append('')
    lines.append('## 📝 XDC 约束')
    lines.append('')
    lines.append('```tcl')
    lines.append(xdc_text.rstrip())
    lines.append('```')
    lines.append('')
    lines.append('---')
    lines.append('*Generated by FPGA Toolbox v5.0*')
    return '\n'.join(lines)


# import html 模块 (供 _build_html_report 用)
import html as html_mod


def run_gui():
    import sys as _sys
    print(f'[FPGA Toolbox] gen_gui loaded from: {__file__}', file=_sys.stderr, flush=True)
    # ═══ 现代化主题 (ttkbootstrap — flatly: 干净扁平的浅色风) ═══
    if _HAS_BOOTSTRAP:
        import datetime as _dt
        _build_tag = ' [build ' + _dt.datetime.now().strftime('%H:%M:%S') + ']'
        root = ttk.Window(themename='flatly', title='FPGA 工具箱' + _build_tag)
        root.geometry('1280x1280')
        root.minsize(1080, 720)
        s = root.style
    else:
        root = tk.Tk()
        import datetime as _dt
        _build_tag = ' [build ' + _dt.datetime.now().strftime('%H:%M:%S') + ']'
        root.title('FPGA 工具箱' + _build_tag)
        root.geometry('1280x1280')
        root.minsize(1080, 720)
        s = ttk.Style()
        s.theme_use('clam')

    # ═══ 精致浅色配色 (用于 tk.Text / tk.Label / tk.Listbox 等非 ttk 控件) ═══
    # 高级灰阶 + 沉稳强调色, 替代原来过饱和的 Bootstrap 默认色
    C = {
        'bg':   '#f4f6f9', 'card': '#ffffff', 'bd': '#e4e7ec',
        'fg':   '#1f2329', 'sub':  '#8a929e',
        'blue': '#2b6cb0', 'green':'#2f855a', 'red':  '#c53030',
        'yellow':'#b7791f','purple':'#6b46c1',
        'btn':  '#eef1f5', 'btn_h':'#e0e4ea', 'ebg': '#fcfcfd',
    }
    F = 'Microsoft YaHei'
    M = 'Cascadia Code'

    # ═══ 样式精修 (ttkbootstrap) ═══
    if _HAS_BOOTSTRAP:
        # 按钮统一圆角 padding, 更舒展精致
        s.configure('TButton', padding=(14, 7), font=(F, 10))
        s.configure('Accent.TButton', font=(F, 10, 'bold'), padding=(16, 8))
        s.configure('Success.TButton', font=(F, 10, 'bold'), padding=(16, 8))
        s.configure('Info.TButton', font=(F, 10), padding=(14, 7))
        s.configure('Danger.TButton', font=(F, 10, 'bold'), padding=(14, 7))
        s.configure('Warning.TButton', font=(F, 10), padding=(14, 7))
        s.configure('Normal.TButton', font=(F, 10), padding=(14, 7))
        s.configure('Small.TButton', font=(F, 11, 'bold'), padding=(8, 2))

        # 卡片式 LabelFrame: 蓝色加粗标题, 提升信息层次
        s.configure('TLabelframe.Label', font=(F, 10, 'bold'),
                    foreground=C['blue'])

        # Notebook Tab: 精修 — padding + 选中态蓝色下划线感
        # (ttkbootstrap flatly 主题已较精致, 此处微调)
        # 缩小 padding/font 确保 11 个 Tab 在最小窗口(880px)内不截断
        s.configure('TNotebook.Tab', padding=(10, 6), font=(F, 9))
        s.configure('TNotebook', tabmargins=(3, 2, 0, 0))

        # Treeview: 行高 + 表头加粗
        s.configure('Treeview', rowheight=28, font=(F, 9))
        s.configure('Treeview.Heading', font=(F, 9, 'bold'),
                    padding=(6, 5))

        # Entry / Combobox 内边距, 输入区更通透
        s.configure('TEntry', padding=(7, 5))
        s.configure('Big.TEntry', padding=(8, 7), font=(F, 12, 'bold'))
        s.configure('TCombobox', padding=(7, 5))
        s.configure('TCheckbutton', font=(F, 10))

        # 滚动条更纤细 (ttkbootstrap 部分主题不允许, 用 try 保护)
        try:
            s.configure('Vertical.TScrollbar', arrowsize=14)
            s.configure('Horizontal.TScrollbar', arrowsize=14)
        except tk.TclError:
            pass
    else:
        # 标准 ttk (clam) 回退 — 手工着色让无 ttkbootstrap 时也精致
        s.configure('TFrame', background=C['bg'])
        s.configure('TLabel', background=C['bg'], font=(F, 10))
        s.configure('TButton', padding=(12, 6), font=(F, 10))
        s.configure('Accent.TButton', font=(F, 10, 'bold'),
                    background=C['blue'], foreground='#ffffff')
        s.map('Accent.TButton', background=[('active', '#22589c')])
        s.configure('Success.TButton', font=(F, 10, 'bold'),
                    background=C['green'], foreground='#ffffff')
        s.map('Success.TButton', background=[('active', '#276748')])
        s.configure('Normal.TButton', font=(F, 10),
                    background=C['card'], foreground=C['fg'])
        s.map('Normal.TButton', background=[('active', C['btn_h'])])
        s.configure('Small.TButton', font=(F, 11, 'bold'),
                    background=C['btn'], foreground=C['fg'])
        s.map('Small.TButton', background=[('active', C['btn_h'])])
        s.configure('TLabelframe', background=C['card'],
                    bordercolor=C['bd'], borderwidth=1, relief='solid')
        s.configure('TLabelframe.Label', background=C['card'],
                    foreground=C['blue'], font=(F, 10, 'bold'))
        s.configure('Card.TLabelframe', background=C['card'])
        s.configure('Card.TLabelframe.Label', background=C['card'],
                    foreground=C['blue'], font=(F, 10, 'bold'))
        s.configure('TNotebook', background=C['bg'], tabmargins=(3, 2, 0, 0))
        s.configure('TNotebook.Tab', padding=(8, 4), font=(F, 9),
                    background=C['btn'])
        s.map('TNotebook.Tab',
              background=[('selected', C['card'])],
              foreground=[('selected', C['blue'])])
        s.configure('Treeview', rowheight=26, font=(F, 9),
                    background=C['ebg'], fieldbackground=C['ebg'])
        s.configure('Treeview.Heading', font=(F, 9, 'bold'))
        s.configure('TEntry', padding=(6, 4))
        s.configure('Big.TEntry', padding=(8, 7), font=(F, 12, 'bold'))
        s.configure('TCombobox', padding=(6, 4))

    # ═══ 全局 Header 栏 (白卡片 + 立体感 + 状态徽章) ═══
    header_bg = tk.Frame(root, bg=C['bg'])
    header_bg.pack(side='top', fill='x', padx=14, pady=(14, 0))

    # 白色卡片 + 投影感 (用双层 Frame 模拟)
    header_shadow = tk.Frame(header_bg, bg='#d4dae3', height=2)
    header_shadow.pack(side='bottom', fill='x', padx=2)

    header = tk.Frame(header_bg, bg=C['card'],
                      highlightbackground=C['bd'],
                      highlightthickness=1, bd=0)
    header.pack(side='top', fill='x')

    # 左侧: 标题区
    title_box = tk.Frame(header, bg=C['card'])
    title_box.pack(side='left', anchor='w', padx=18, pady=14)

    # Logo (皮卡丘图标)
    _logo_path = os.path.join(_PROJECT_ROOT, 'assets', 'pikachu.png')
    _logo_img_ref = [None]  # 防止被 GC
    try:
        from PIL import Image, ImageTk
        _img = Image.open(_logo_path)
        _img = _img.resize((64, 64), Image.LANCZOS)
        _logo_img_ref[0] = ImageTk.PhotoImage(_img)
        logo = tk.Label(title_box, image=_logo_img_ref[0], bg=C['card'])
        logo.pack(side='left', padx=(0, 14))
    except Exception:
        # PIL 缺失或图片读取失败时, 回退到黄色 emoji 占位
        logo = tk.Frame(title_box, bg='#f6c84a', width=64, height=64)
        logo.pack(side='left', padx=(0, 14))
        logo.pack_propagate(False)
        tk.Label(logo, text='⚡', bg='#f6c84a', fg='#ffffff',
                 font=('Segoe UI Emoji', 32, 'bold')).pack(expand=True)

    # 标题 + 副标题
    title_text = tk.Frame(title_box, bg=C['card'])
    title_text.pack(side='left')
    tk.Label(title_text, text='FPGA 工具箱', bg=C['card'],
             fg=C['fg'], font=(F, 15, 'bold')).pack(anchor='w')
    tk.Label(title_text, text='HDL 例化 · 工程压缩 · Git 提交流水线',
             bg=C['card'], fg=C['sub'],
             font=(F, 9)).pack(anchor='w')

    # 右侧: 状态徽章 + 版本
    right_box = tk.Frame(header, bg=C['card'])
    right_box.pack(side='right', anchor='e', padx=18, pady=14)

    # 版本标签
    tk.Label(right_box, text='v5.0', bg=C['card'], fg=C['sub'],
             font=(F, 9)).pack(side='right', padx=(10, 0))

    # 绿色徽章
    badge_frame = tk.Frame(right_box, bg='#e6f4ec',
                           highlightbackground='#b8e0c7',
                           highlightthickness=1, bd=0)
    badge_frame.pack(side='right', padx=2)
    # 绿色圆点 (用 Label 模拟)
    dot = tk.Canvas(badge_frame, width=10, height=10, bg='#e6f4ec',
                    highlightthickness=0, bd=0)
    dot.create_oval(2, 2, 8, 8, fill=C['green'], outline='')
    dot.pack(side='left', padx=(8, 4), pady=6)
    tk.Label(badge_frame, text='就绪', bg='#e6f4ec', fg=C['green'],
             font=(F, 9, 'bold')).pack(side='left', padx=(0, 10), pady=6)

    # ═══ Notebook 容器 (单独白卡片包裹, 让 Tab 与内容成为一体) ═══
    body_bg = tk.Frame(root, bg=C['bg'])
    body_bg.pack(side='top', fill='both', expand=True, padx=14, pady=10)
    body_bg.pack_propagate(False)  # 锁定, 只跟随 root 缩放

    # 白色卡片包裹
    body_card = tk.Frame(body_bg, bg=C['card'],
                         highlightbackground=C['bd'],
                         highlightthickness=1, bd=0)
    body_card.pack(fill='both', expand=True)
    body_card.pack_propagate(False)  # 锁定, 只跟随 body_bg 缩放

    nb = ttk.Notebook(body_card)
    nb.pack(fill='both', expand=True, padx=0, pady=0)

    # ── 自适应: 所有 Tab Frame 禁止子控件撑大父级, 让内容跟随窗口缩放 ──
    # 类级别 monkey-patch, 覆盖主 Notebook 及所有子 Notebook (t1_nb / calc_nb / aux_nb / a1_nb 等)
    _orig_cls_add = ttk.Notebook.add
    def _cls_add_with_resize(self, child, **kw):
        child.grid_propagate(False)
        child.pack_propagate(False)
        return _orig_cls_add(self, child, **kw)
    ttk.Notebook.add = _cls_add_with_resize

    # ═══ 工具函数 ═══
    def _dcenter(dlg):
        """把 Toplevel 居中到 root 窗口中央 (只设位置, 尺寸由 pack 决定)"""
        try:
            dlg.update_idletasks()
            root.update_idletasks()
            root.update()
            x = root.winfo_rootx() + max(0, (root.winfo_width()  - dlg.winfo_reqwidth())  // 2)
            y = root.winfo_rooty() + max(0, (root.winfo_height() - dlg.winfo_reqheight()) // 2)
            dlg.geometry(f'+{x}+{y}')
        except Exception:
            pass

    def _log_widget(parent, h=8):
        # 容器: 浅色边框 + 圆角感
        f = tk.Frame(parent, bg=C['card'],
                     highlightbackground=C['bd'],
                     highlightthickness=1, bd=0)
        f.pack_propagate(False)  # 禁止子控件撑大, 跟随父级缩放
        inner = ttk.Frame(f, style='TFrame')
        inner.pack(fill='both', expand=True, padx=1, pady=1)

        t = tk.Text(inner, font=(M, 9), bg=C['ebg'], fg=C['fg'],
                    insertbackground=C['fg'], relief='flat',
                    padx=12, pady=10, wrap='word', height=h,
                    selectbackground='#cfe2f3',
                    selectforeground=C['fg'])
        t.pack(fill='both', expand=True)
        # 阻止编辑但允许 Ctrl+C / Ctrl+A / 鼠标选中
        def _sel_all(e):
            t.tag_add('sel', '1.0', 'end')
            return 'break'
        def _do_copy(e):
            if not t.tag_ranges('sel'):
                t.tag_add('sel', '1.0', 'end')
            try:
                t.clipboard_clear()
                t.clipboard_append(t.get('sel.first', 'sel.last'))
            except Exception:
                pass
            return 'break'
        t.bind('<Control-a>', _sel_all)
        t.bind('<Control-c>', _do_copy)
        t.bind('<Control-C>', _do_copy)
        t.bind('<Control-Insert>', _do_copy)
        def _block_edit(e):
            if e.state & 4:
                if e.keysym in ('c','C','a','A','Insert'):
                    return None
            if e.keysym in ('Up','Down','Left','Right','Prior','Next',
                            'Home','End','Shift_L','Shift_R','Control_L',
                            'Control_R','Tab','Escape'):
                return None
            return 'break'
        t.bind('<Key>', _block_edit)
        for ev in ('<<Paste>>', '<<Cut>>', '<<Clear>>'):
            t.bind(ev, lambda e: 'break')
        sb = tk.Scrollbar(inner, orient='vertical', bg=C['card'],
                          troughcolor=C['bg'], relief='flat',
                          width=10)
        sb.pack(side='right', fill='y')
        t.config(yscrollcommand=sb.set)
        sb.config(command=t.yview)
        return f, t

    def _log(txt, msg, color=None):
        if color is None:
            color = C['fg']
        tag = 'c_' + color.lstrip('#')  # 根据颜色创建唯一tag
        txt.tag_configure(tag, foreground=color)
        txt.insert('end', msg + '\n', (tag,))
        txt.see('end')

    def _code_panel(parent):
        """代码预览面板 (Text + 滚动条 + 语法高亮, 圆角卡片感)"""
        f = tk.Frame(parent, bg=C['card'],
                     highlightbackground=C['bd'],
                     highlightthickness=1, bd=0)
        f.grid_propagate(False)  # 禁止子控件撑大, 跟随父级缩放
        f.grid_rowconfigure(0, weight=1)
        f.grid_columnconfigure(0, weight=1)
        t = tk.Text(f, font=(M, 10), bg=C['ebg'], fg=C['fg'],
                    insertbackground=C['fg'], relief='flat',
                    padx=16, pady=14, wrap='none',
                    selectbackground='#cfe2f3',
                    selectforeground=C['fg'])
        # 阻止编辑但允许 Ctrl+C / Ctrl+A / 鼠标选中
        def _sel_all(e):
            t.tag_add('sel', '1.0', 'end')
            return 'break'
        def _do_copy(e):
            if not t.tag_ranges('sel'):
                t.tag_add('sel', '1.0', 'end')
            try:
                t.clipboard_clear()
                t.clipboard_append(t.get('sel.first', 'sel.last'))
            except Exception:
                pass
            return 'break'
        t.bind('<Control-a>', _sel_all)
        t.bind('<Control-c>', _do_copy)
        t.bind('<Control-C>', _do_copy)
        t.bind('<Control-Insert>', _do_copy)
        def _block_edit(e):
            if e.state & 4 and e.keysym in ('c','C','a','A','Insert'):
                return None
            if e.keysym in ('Up','Down','Left','Right','Prior','Next',
                            'Home','End','Shift_L','Shift_R','Control_L',
                            'Control_R','Tab','Escape'):
                return None
            return 'break'
        t.bind('<Key>', _block_edit)
        for ev in ('<<Paste>>', '<<Cut>>', '<<Clear>>'):
            t.bind(ev, lambda e: 'break')
        t.grid(row=0, column=0, sticky='nsew', padx=1, pady=(1, 0))
        sy = tk.Scrollbar(f, orient='vertical', bg=C['card'],
                          troughcolor=C['bg'], relief='flat', width=10)
        sy.grid(row=0, column=1, sticky='ns', pady=(1, 0))
        t.config(yscrollcommand=sy.set)
        sy.config(command=t.yview)
        sx = tk.Scrollbar(f, orient='horizontal', bg=C['card'],
                          troughcolor=C['bg'], relief='flat', width=10)
        sx.grid(row=1, column=0, sticky='ew')
        t.config(xscrollcommand=sx.set)
        sx.config(command=t.xview)
        t.tag_configure('comment', foreground='#5a8a6a', font=(M, 10, 'italic'))
        t.tag_configure('keyword', foreground=C['blue'], font=(M, 10, 'bold'))
        t.tag_configure('string', foreground=C['purple'])
        t.tag_configure('number', foreground=C['yellow'])
        return f, t

    def _highlight(tw, code):
        for m in re.finditer(r'//.*$|--.*$', code, re.MULTILINE):
            try:
                tw.tag_add('comment',
                           f'1.0+{m.start()}c', f'1.0+{m.end()}c')
            except tk.TclError:
                pass
        for m in re.finditer(
                r'\b(module|endmodule|input|output|inout|wire|reg|'
                r'entity|end|component|port|generic|signal|begin|'
                r'architecture|std_logic|std_logic_vector|'
                r'parameter|localparam|assign|always|initial|'
                r'if|else|case|endcase|for|while|function|task)\b',
                code, re.IGNORECASE):
            try:
                tw.tag_add('keyword',
                           f'1.0+{m.start()}c', f'1.0+{m.end()}c')
            except tk.TclError:
                pass
        # 数字高亮
        for m in re.finditer(r'\b\d+\'?[bdh][a-fA-F0-9]+\b|\b\d+\b', code):
            try:
                tw.tag_add('number',
                           f'1.0+{m.start()}c', f'1.0+{m.end()}c')
            except tk.TclError:
                pass

    def _set_code(tw, code):
        tw.delete('1.0', 'end')
        tw.insert('1.0', code)
        _highlight(tw, code)

    # ══════════════════════════════════════
    # TAB 1 — 例化模板
    # ══════════════════════════════════════
    from gen_inst import parse_file, generate_templates

    t1 = ttk.Frame(nb, style='TFrame')
    nb.add(t1, text='⚡ 例化模板')
    t1.grid_rowconfigure(2, weight=1)
    t1.grid_columnconfigure(0, weight=1)

    t1_mod = [None]
    t1_fp = [None]
    t1_v = [None]
    t1_h = [None]

    # 源文件卡片
    fc1 = ttk.LabelFrame(t1, text=' 源文件 ', )
    fc1.grid(row=0, column=0, sticky='ew', padx=12, pady=(10, 4))
    fc1.grid_columnconfigure(0, weight=1)

    t1_info = ttk.Label(
        fc1,
        text='\U0001f4c4 点击按钮或将 .v/.vhd 拖入窗口',
        foreground=C['sub'], font=(F, 10))
    t1_info.grid(row=0, column=0, sticky='w', padx=14, pady=(10, 4))

    br1 = ttk.Frame(fc1, style='TFrame')
    br1.grid(row=1, column=0, sticky='ew', padx=14, pady=(4, 12))

    def t1_load(fp):
        try:
            t1_fp[0] = fp
            t1_mod[0] = parse_file(fp)
            bn = os.path.basename(fp)
            m = t1_mod[0]
            t1_info.config(
                text=(f'\u2714 {bn}  |  {m.name} ({m.lang.upper()})  |  '
                      f'端口:{len(m.ports)}  参数:{len(m.params)}'),
                foreground=C['green'])
            t1_refresh()
        except Exception as e:
            messagebox.showerror('解析错误', str(e))

    def t1_choose():
        fp = filedialog.askopenfilename(
            title='选择硬件描述文件',
            filetypes=[('HDL 文件', '*.v *.sv *.vhd *.vhdl'),
                       ('Verilog', '*.v *.sv'),
                       ('VHDL', '*.vhd *.vhdl')])
        if fp:
            t1_load(fp)

    ttk.Button(br1, text='\U0001f4c2 选择文件',
               command=t1_choose,
               style='Accent.TButton').grid(row=0, column=0, sticky='w')

    # 设置卡片
    sc1 = ttk.LabelFrame(t1, text=' 设置 ', )
    sc1.grid(row=1, column=0, sticky='ew', padx=12, pady=(2, 4))

    si1 = ttk.Frame(sc1, style='TFrame')
    si1.grid(row=0, column=0, sticky='w', padx=14, pady=10)

    ttk.Label(si1, text='例化个数:', font=(F, 10)).grid(
        row=0, column=0, sticky='w', padx=(0, 8))

    t1_cnt = tk.StringVar(value='1')
    ttk.Entry(si1, textvariable=t1_cnt, width=6,
              style='Big.TEntry', justify='center').grid(
        row=0, column=1, padx=(0, 4))

    def _cnt_chg(d):
        try:
            v = int(t1_cnt.get() or '1') + d
            if v >= 1:
                t1_cnt.set(str(v))
                t1_refresh()
        except ValueError:
            pass

    ttk.Button(si1, text='+', command=lambda: _cnt_chg(1),
               style='Small.TButton').grid(row=0, column=2, padx=1)
    ttk.Button(si1, text='\u2212', command=lambda: _cnt_chg(-1),
               style='Small.TButton').grid(row=0, column=3, padx=(1, 20))

    ttk.Label(si1, text='F5 刷新  Ctrl+C 复制  Ctrl+S 保存',
              foreground=C['bd'], font=(F, 9)).grid(
        row=0, column=4, padx=(20, 0))

    # 代码预览卡片
    cc1 = ttk.LabelFrame(t1, text=' 生成代码 ', )
    cc1.grid(row=2, column=0, sticky='nsew', padx=12, pady=(2, 8))
    cc1.grid_rowconfigure(0, weight=1)
    cc1.grid_columnconfigure(0, weight=1)

    t1_nb = ttk.Notebook(cc1)
    t1_nb.grid(row=0, column=0, sticky='nsew', padx=4, pady=(2, 6))

    t1_vf, t1_vt = _code_panel(t1_nb)
    t1_nb.add(t1_vf, text='  Verilog  ')
    t1_hf, t1_ht = _code_panel(t1_nb)
    t1_nb.add(t1_hf, text='  VHDL  ')

    def t1_refresh(*_):
        if not t1_mod[0]:
            return
        try:
            n = int(t1_cnt.get() or '1')
        except ValueError:
            n = 1
        if n < 1:
            n = 1
            t1_cnt.set('1')
        t1_v[0], t1_h[0] = generate_templates(t1_mod[0], n)
        _set_code(t1_vt, t1_v[0])
        _set_code(t1_ht, t1_h[0])

    # 底部按钮
    bb1 = ttk.Frame(t1, style='TFrame')
    bb1.grid(row=3, column=0, sticky='nsew', padx=12, pady=(0, 10))
    bb1.grid_columnconfigure(100, weight=1)

    t1_copy_btn = ttk.Button(bb1, text='\U0001f4cb 复制',
                             style='Normal.TButton')
    t1_copy_btn.grid(row=0, column=0, padx=2)

    def t1_copy():
        code = (t1_v[0]
                if t1_nb.index(t1_nb.select()) == 0
                else t1_h[0])
        root.clipboard_clear()
        root.clipboard_append(code)
        root.update()
        t1_copy_btn.config(text='\u2714 已复制!')
        root.after(1500,
                   lambda: t1_copy_btn.config(text='\U0001f4cb 复制'))

    t1_copy_btn.config(command=t1_copy)

    def t1_save(ext):
        if not t1_mod[0]:
            messagebox.showinfo('提示', '请先加载文件')
            return
        code = t1_v[0] if ext == '.v' else t1_h[0]
        d = (os.path.dirname(t1_fp[0])
             if t1_fp[0] else os.path.expanduser('~'))
        bn = (os.path.splitext(os.path.basename(t1_fp[0]))[0]
              if t1_fp[0] else t1_mod[0].name)
        fp = filedialog.asksaveasfilename(
            initialdir=d,
            initialfile=f'{bn}_inst{ext}',
            title=f'保存 {ext} 例化文件',
            filetypes=([('Verilog', '*.v')]
                       if ext == '.v' else [('VHDL', '*.vhd')]))
        if fp:
            with open(fp, 'w', encoding='utf-8') as f:
                f.write(code)
            messagebox.showinfo('保存成功', f'已保存到:\n{fp}')

    def t1_save_both():
        if not t1_mod[0]:
            messagebox.showinfo('提示', '请先加载文件')
            return
        d = (os.path.dirname(t1_fp[0])
             if t1_fp[0] else os.path.expanduser('~'))
        bn = (os.path.splitext(os.path.basename(t1_fp[0]))[0]
              if t1_fp[0] else t1_mod[0].name)
        dp = filedialog.askdirectory(initialdir=d, title='选择保存目录')
        if dp:
            vp = os.path.join(dp, f'{bn}_inst.v')
            hp = os.path.join(dp, f'{bn}_inst.vhd')
            with open(vp, 'w', encoding='utf-8') as f:
                f.write(t1_v[0])
            with open(hp, 'w', encoding='utf-8') as f:
                f.write(t1_h[0])
            messagebox.showinfo('保存成功', f'{vp}\n{hp}')

    ttk.Button(bb1, text=' 保存 .v ',
               command=lambda: t1_save('.v'),
               style='Normal.TButton').grid(row=0, column=1, padx=2)
    ttk.Button(bb1, text=' 保存 .vhd ',
               command=lambda: t1_save('.vhd'),
               style='Normal.TButton').grid(row=0, column=2, padx=2)
    ttk.Button(bb1, text='\u2605 同时保存',
               command=t1_save_both,
               style='Success.TButton').grid(row=0, column=3, padx=2)

    # ══════════════════════════════════════
    # TAB 2 — 工程压缩
    # ══════════════════════════════════════
    from gen_inst import (compress_project, archive_project,
                          _generate_fpga_gitignore, _fmt_size,
                          auto_export_vivado, _load_config, _save_config)

    t2 = ttk.Frame(nb, style='TFrame')
    nb.add(t2, text='📦 工程压缩')
    t2.grid_rowconfigure(3, weight=1)
    t2.grid_columnconfigure(0, weight=1)
    t2.grid_rowconfigure(0, weight=0)

    # ============================================
    # Vivado 路径 — 来自 ⚙设置 Tab 全局配置
    # ============================================
    from src.app_config import get_vivado_bin_dirs as _t2_viv_dirs

    vc2 = ttk.LabelFrame(t2, text=' Vivado 路径 (来自 ⚙设置 Tab) ', )
    vc2.grid(row=0, column=0, sticky='ew', padx=12, pady=(10, 4))
    vc2.grid_columnconfigure(0, weight=1)

    t2_viv_path_var = tk.StringVar()
    t2_viv_label = ttk.Label(vc2, textvariable=t2_viv_path_var,
                             font=(F, 9), foreground=C['sub'])
    t2_viv_label.grid(row=0, column=0, sticky='w', padx=14, pady=(8, 4))
    _t2_viv_btn = ttk.Button(vc2, text='⚙ 打开设置', command=lambda: nb.select(t16),
                              style='Small.TButton')

    # 从 Vivado 路径提取版本号 (延迟到 t2_ver_combo 创建后调用)
    def _t2_extract_versions():
        vers = []
        for d in _t2_viv_dirs():
            d = d.replace('\\', '/')
            parts = d.split('/')
            for p in parts:
                if re.match(r'^\d{4}\.\d$', p): vers.append(p); break
        return vers

    def t2_refresh_viv_path():
        dirs = _t2_viv_dirs()
        if dirs:
            t2_viv_path_var.set(f'✔ 已配置 {len(dirs)} 个 Vivado 路径')
            t2_viv_label.config(foreground=C['green'])
            _t2_viv_btn.grid_remove()
        else:
            t2_viv_path_var.set('未配置 Vivado 路径 — 请到 ⚙设置 Tab 添加')
            t2_viv_label.config(foreground=C['red'])
            _t2_viv_btn.grid(row=0, column=1, padx=(0, 14), pady=(8, 4))
        if t2_ver_combo is not None:
            t2_ver_combo['values'] = _t2_extract_versions()

    def _t2_find_exe(d):
        for exe in ['vitis.exe','vitis.bat','vivado.exe','vivado.bat','vivado','vitis']:
            fp = os.path.join(d, exe)
            if os.path.isfile(fp): return fp
        return None

    def t2_get_selected_vivado_bin():
        """从全局设置读取 Vivado bin 路径: 按版本号匹配"""
        dirs = _t2_viv_dirs()
        if not dirs:
            return None
        ver = t2_ver_var.get().strip()
        # 按版本匹配
        if ver and dirs:
            for d in dirs:
                if ver in d.replace('\\', '/'):
                    exe = _t2_find_exe(d)
                    if exe: return exe
        # 回退第一个
        if dirs:
            exe = _t2_find_exe(dirs[0])
            if exe: return exe
        return None

    # 提前定义，避免闭包引用未赋值变量
    t2_src = tk.StringVar()
    t2_ver_var = tk.StringVar()
    t2_ver_combo = None  # 稍后替换为实际 Combobox

    fc2 = ttk.LabelFrame(t2, text=' 路径设置 ', )
    fc2.grid(row=1, column=0, sticky='ew', padx=12, pady=(6, 4))
    fc2.grid_columnconfigure(1, weight=1)

    ttk.Label(fc2, text='工程路径:', font=(F, 10)).grid(
        row=0, column=0, sticky='w', padx=14, pady=(10, 2))
    ttk.Entry(fc2, textvariable=t2_src, font=(F, 10)).grid(
        row=0, column=1, sticky='ew', padx=(4, 4), pady=(10, 2))

    def t2_br_src():
        d = filedialog.askdirectory(title='选择 FPGA 工程根目录')
        if d:
            t2_src.set(d)

    ttk.Button(fc2, text='浏览', command=t2_br_src,
               style='Normal.TButton').grid(
        row=0, column=2, padx=(0, 14), pady=(10, 2))
    ttk.Label(fc2, text='  ℹ 含 .xpr 的 Vivado 工程根目录',
              foreground=C['sub'], font=(F, 8)).grid(
        row=1, column=1, columnspan=2, sticky='w', padx=(4, 14), pady=(0, 4))

    # Vivado 版本 (从设置的路径自动提取)
    def _t2_extract_versions():
        vers = []
        for d in _t2_viv_dirs():
            d = d.replace('\\', '/')
            parts = d.split('/')
            for p in parts:
                if re.match(r'^\d{4}\.\d$', p): vers.append(p); break
        return vers
    ttk.Label(fc2, text='Vivado 版本:', font=(F, 10)).grid(
        row=2, column=0, sticky='w', padx=14, pady=(6, 2))
    t2_ver_combo = ttk.Combobox(fc2, textvariable=t2_ver_var, font=(F, 10),
                                 width=10, values=_t2_extract_versions())
    t2_ver_combo.grid(row=2, column=1, sticky='w', padx=(4, 4), pady=(6, 2))
    t2_refresh_viv_path()  # Combobox 创建后调用

    # t2_ver_combo 创建后才能加载和监听
    # 从旧配置恢复版本号 (仅版本号, 路径从设置Tab读)
    def _t2_load_version():
        cfg = _load_config()
        last_ver = cfg.get('last_vivado_ver', '')
        if last_ver:
            t2_ver_var.set(last_ver)
        lcp = cfg.get('last_compress_path', '')
        if lcp and os.path.isdir(lcp):
            t2_src.set(lcp)
    def _t2_save_version(*args):
        cfg = _load_config()
        cfg['last_vivado_ver'] = t2_ver_var.get().strip()
        if t2_src.get().strip():
            cfg['last_compress_path'] = t2_src.get().strip()
        _save_config(cfg)
    _t2_load_version()
    t2_ver_var.trace_add('write', lambda *_: _t2_save_version())
    ttk.Label(fc2, text='留空=不自动导出  参考: 2023.2',
              foreground=C['sub'], font=(F, 8)).grid(
        row=2, column=2, sticky='w', padx=(0, 14), pady=(6, 2))

    ttk.Label(fc2,
              text='输出路径 (留空=原地压缩, 填写=归档导出):',
              font=(F, 10), foreground=C['sub']).grid(
        row=3, column=0, sticky='w', padx=14, pady=(2, 2))

    t2_dst = tk.StringVar()
    ttk.Entry(fc2, textvariable=t2_dst, font=(F, 10)).grid(
        row=3, column=1, sticky='nsew', padx=(4, 4), pady=(2, 10))

    def t2_br_dst():
        d = filedialog.askdirectory(title='选择输出目录')
        if d:
            t2_dst.set(d)

    ttk.Button(fc2, text='浏览', command=t2_br_dst,
               style='Normal.TButton').grid(
        row=3, column=2, padx=(0, 14), pady=(2, 10))

    # 自动导出始终开启，不显示在 UI 上
    t2_auto_export = tk.BooleanVar(value=True)

    t2_lf, t2_log = _log_widget(t2, 10)
    t2_lf.grid(row=3, column=0, sticky='nsew', padx=12, pady=(2, 4))

    def t2_run():
        p = t2_src.get().strip()
        d = t2_dst.get().strip()
        if not p or not os.path.isdir(p):
            messagebox.showerror('错误', '请选择有效的工程目录')
            return
        if d and os.path.abspath(p) == os.path.abspath(d):
            messagebox.showerror('错误', '输出路径不能与源路径相同')
            return

        ae = t2_auto_export.get()
        vb = t2_get_selected_vivado_bin()
        ver = t2_ver_var.get().strip()  # vivado 版本 (如 2023.2)
        _t2_save_version()  # persist compress path

        # 检查 Vivado 路径是否已配置
        if not vb or not _t2_viv_dirs():
            messagebox.showerror(
                '未配置 Vivado',
                '请先到 ⚙设置 Tab 添加 Vivado bin 路径,\n'
                '然后返回本页执行压缩。\n\n'
                '没有 Vivado 无法自动导出 BD Tcl / HDF 等文件。')
            return
        if not vb:
            messagebox.showerror(
                '未选择版本',
                '请在 "版本标识" 下拉框中选择本次使用的 Vivado 版本。')
            return

        archive = bool(d)

        # dry-run 扫描
        old = _sys.stdout
        _sys.stdout = _io.StringIO()
        try:
            if archive:
                cnt, sz = archive_project(p, d, True, auto_export=ae, vivado_bin=vb, vivado_ver=ver)
            else:
                cnt, sz = compress_project(p, dry_run=True, auto_export=ae, vivado_bin=vb, vivado_ver=ver)
            out = _sys.stdout.getvalue()
        except Exception as e:
            out = str(e)
            cnt, sz = 0, 0
        finally:
            _sys.stdout = old

        if cnt == 0:
            _log(t2_log, out + '\n无需操作')
            return

        if archive:
            msg = (f'将复制 {cnt} 个文件到:\n{d}\n'
                   f'总计 {_fmt_size(sz)}\n\n源文件不会改动。确认？')
        else:
            msg = f'将删除 {cnt} 项, 释放 {_fmt_size(sz)}\n\n确认执行？'

        if not messagebox.askyesno('确认操作', msg):
            return

        _log(t2_log, out + '\n执行中...')

        def _bg():
            try:
                old = _sys.stdout
                _sys.stdout = _io.StringIO()
                if archive:
                    c2, s2 = archive_project(p, d, False, auto_export=ae, vivado_bin=vb, vivado_ver=ver)
                else:
                    c2, s2 = compress_project(p, False, True, auto_export=ae, vivado_bin=vb, vivado_ver=ver)
                out2 = _sys.stdout.getvalue()
                _sys.stdout = old
                root.after(0, lambda: _log(
                    t2_log,
                    out2 + f'\n\u2714 完成! {c2} 项, {_fmt_size(s2)}',
                    C['green']))
            except Exception as e:
                root.after(0, lambda: _log(t2_log, str(e), C['red']))

        threading.Thread(target=_bg, daemon=True).start()

    ttk.Button(t2, text='▶ 执行', command=t2_run,
               style='Accent.TButton').grid(
        row=4, column=0, sticky='w', padx=12, pady=(2, 8))

    # ── 整理为 Git 工程 (新建文件夹, 审核 + 用户勾选 + 仅补全勾选项) ──
    def t2_organize_to_git():
        p = t2_src.get().strip()
        if not p or not os.path.isdir(p):
            messagebox.showerror('错误', '请选择有效的工程目录')
            return
        vb = t2_get_selected_vivado_bin()
        ver = t2_ver_var.get().strip()

        from gen_inst import audit_project, organize_project_to_git

        # 1) 审核 (只读, 不修改工程)
        try:
            items, summary = audit_project(
                p, vivado_bin=vb, vivado_ver_str=ver)
        except Exception as e:
            messagebox.showerror('审核失败', str(e))
            return

        if not items:
            messagebox.showwarning('无项目', '该目录没有任何需要补全的项')
            return

        # 2) 弹窗: 显示清单 + 复选框
        dlg = tk.Toplevel(root)
        dlg.title('审核清单 - 选择要补全的项目')
        dlg.geometry('900x560')
        dlg.transient(root)
        dlg.configure(bg=C['card'])

        # 顶部: 工程概要
        info = (
            f"工程: {summary['proj_name']}    "
            f".xpr: {'有' if summary['has_xpr'] else '无'}    "
            f".bd: {summary['bd_count']}    "
            f".xdc: {summary['xdc_count']}    "
            f"app: {summary['app_count']}    "
            f"IP: {summary['ip_count']}\n"
            f"源: {summary['root']}    "
            f"Vivado: {summary['vivado_ver'] or '(未指定)'}    "
            f"Vivado bin: {summary['vivado_bin'] or '(未指定)'}"
        )
        tk.Label(dlg, text=info, bg=C['card'], fg=C['sub'],
                 font=(F, 9), justify='left',
                 padx=14, pady=(12, 4), anchor='w').pack(fill='x')

        tk.Label(dlg,
                 text='  复选框打勾 → 补全该项;   不勾 → 跳过该项   '
                      '(依据 FPGA_GIT_GUIDE.md 3.1/3.2/4.2/5.1/6/7.1 节)',
                 bg=C['card'], fg=C['sub'],
                 font=(F, 9), padx=14, pady=(0, 4), anchor='w').pack(fill='x')

        # 工具按钮
        tb = ttk.Frame(dlg, style='TFrame')
        tb.pack(fill='x', padx=14, pady=(0, 4))

        def _check_all():
            for it in item_state:
                it['check'].set(True)
        def _uncheck_all():
            for it in item_state:
                it['check'].set(False)
        def _restore_default():
            for it in item_state:
                it['check'].set(it['item']['default'])

        ttk.Button(tb, text='☑ 全选', command=_check_all,
                   style='Normal.TButton').pack(side='left', padx=(0, 4))
        ttk.Button(tb, text='☐ 全不选', command=_uncheck_all,
                   style='Normal.TButton').pack(side='left', padx=4)
        ttk.Button(tb, text='↺ 恢复默认', command=_restore_default,
                   style='Normal.TButton').pack(side='left', padx=4)

        # Treeview
        cols = ('check', 'group', 'name', 'status', 'desc', 'note')
        tree_frame = ttk.Frame(dlg)
        tree_frame.pack(fill='both', expand=True, padx=14, pady=(0, 4))

        tree = ttk.Treeview(tree_frame, columns=cols, show='headings',
                            selectmode='browse', height=14)
        tree.heading('check', text='☑')
        tree.heading('group', text='类别')
        tree.heading('name', text='名称 (相对路径)')
        tree.heading('status', text='状态')
        tree.heading('desc', text='规则说明')
        tree.heading('note', text='备注')

        tree.column('check', width=42, anchor='center', stretch=False)
        tree.column('group', width=80, anchor='center', stretch=False)
        tree.column('name', width=220, anchor='w', stretch=False)
        tree.column('status', width=60, anchor='center', stretch=False)
        tree.column('desc', width=300, anchor='w', stretch=True)
        tree.column('note', width=160, anchor='w', stretch=False)

        # 交替行底色
        tree.tag_configure('exists',  foreground=C['sub'])
        tree.tag_configure('missing', foreground=C['fg'])
        tree.tag_configure('nonbd',   foreground=C['sub'])

        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')

        # 保存 item + 复选状态 (iid -> dict)
        item_state = []  # [{iid, item, check:BooleanVar}]

        def _refresh_row(iid, checked):
            """更新 #0 (☐/☑) 和 check 列"""
            tree.set(iid, 'check', '☑' if checked else '☐')

        for it in items:
            tag = 'exists' if it['exists'] else 'missing'
            if it['reason_skip']:
                tag = 'nonbd'
            iid = tree.insert('', 'end',
                              values=('☑' if it['default'] else '☐',
                                      it['group'], it['name'],
                                      it['status'], it['desc'],
                                      it['note']),
                              tags=(tag,))
            cv = tk.BooleanVar(value=it['default'])
            cv.trace_add('write',
                         lambda *_a, iid=iid, c=cv: _refresh_row(iid, c.get()))
            item_state.append({'iid': iid, 'item': it, 'check': cv})

        # 单击 "☑" 列切换
        def _on_click(event):
            region = tree.identify('region', event.x, event.y)
            if region != 'cell':
                return
            col = tree.identify_column(event.x)
            if col != '#1':  # 只在 check 列切换
                return
            iid = tree.identify_row(event.y)
            if not iid:
                return
            for s in item_state:
                if s['iid'] == iid:
                    s['check'].set(not s['check'].get())
                    break

        tree.bind('<Button-1>', _on_click)

        # 双击整行 = 切换
        def _on_dblclick(event):
            iid = tree.identify_row(event.y)
            if not iid:
                return
            for s in item_state:
                if s['iid'] == iid:
                    s['check'].set(not s['check'].get())
                    break
        tree.bind('<Double-Button-1>', _on_dblclick)

        # 底部: 状态/按钮
        bot = ttk.Frame(dlg, style='TFrame')
        bot.pack(fill='x', padx=14, pady=(4, 12))

        summary_var = tk.StringVar(value='')
        def _update_summary():
            n_total = len(item_state)
            n_checked = sum(1 for s in item_state if s['check'].get())
            n_missing = sum(1 for s in item_state
                            if not s['item']['exists'])
            n_skip = n_missing - n_checked
            summary_var.set(
                f'共 {n_total} 项, 已勾选 {n_checked} 项, '
                f'将跳过 {n_skip} 项缺失项')
        for s in item_state:
            s['check'].trace_add('write', lambda *_a: _update_summary())
        _update_summary()
        ttk.Label(bot, textvariable=summary_var,
                  font=(F, 9), foreground=C['sub']).pack(side='left')

        def _on_cancel():
            dlg.destroy()

        def _on_execute():
            # 收集用户选择的 keys
            checked_keys = {s['item']['key'] for s in item_state
                            if s['check'].get()}
            all_keys = {s['item']['key'] for s in item_state}
            skip_keys = all_keys - checked_keys

            n_skip = len(skip_keys)
            if n_skip > 0:
                skip_names = [s['item']['name'] for s in item_state
                              if s['item']['key'] in skip_keys]
                if not messagebox.askyesno(
                        '确认',
                        f'将跳过 {n_skip} 项 (用户未勾选):\n  '
                        + '\n  '.join(skip_names[:6])
                        + ('\n  ...' if len(skip_names) > 6 else '')
                        + '\n\n确认执行?',
                        parent=dlg):
                    return

            dlg.destroy()

            # 3) 后台执行: 复制 + 重组 + 仅补全勾选项
            prog = tk.Toplevel(root)
            prog.title('整理为 Git 工程')
            prog.geometry('720x420')
            prog.transient(root)
            prog.configure(bg=C['card'])
            tk.Label(prog, text='正在整理工程 (不破坏原工程)...',
                     bg=C['card'], fg=C['blue'],
                     font=(F, 11, 'bold'),
                     padx=20, pady=(14, 6)).pack(anchor='w')
            log_txt = tk.Text(prog, font=(M, 9), height=18, width=88,
                              bg=C['card'], fg=C['fg'], relief='flat',
                              padx=10, pady=6, wrap='word')
            log_txt.pack(fill='both', expand=True, padx=14, pady=(0, 8))

            def _bg():
                try:
                    new_dir, log_str = organize_project_to_git(
                        p, vivado_bin=vb, vivado_ver=ver,
                        skip_keys=skip_keys)
                    root.after(0, lambda: _show_result(new_dir, log_str))
                except Exception as e:
                    root.after(0, lambda: _show_result(None, '错误: ' + str(e)))

            def _show_result(new_dir, log_str):
                log_txt.insert('end', log_str)
                log_txt.see('end')
                btn_frame = ttk.Frame(prog, style='TFrame')
                btn_frame.pack(fill='x', padx=14, pady=(0, 10))
                if new_dir:
                    def _open():
                        try:
                            if _sys.platform == 'win32':
                                os.startfile(new_dir)
                            elif _sys.platform == 'darwin':
                                subprocess.run(['open', new_dir])
                            else:
                                subprocess.run(['xdg-open', new_dir])
                        except Exception:
                            pass
                    ttk.Button(btn_frame, text='📂 打开新目录', command=_open,
                               style='Accent.TButton').pack(side='left', padx=(0, 6))
                ttk.Button(btn_frame, text='完成', command=prog.destroy,
                           style='Normal.TButton').pack(side='right')

            threading.Thread(target=_bg, daemon=True).start()

        ttk.Button(bot, text='取消', command=_on_cancel,
                   style='Normal.TButton').pack(side='right', padx=(6, 0))
        ttk.Button(bot, text='执行选中项', command=_on_execute,
                   style='Accent.TButton').pack(side='right')

        # 居中显示
        dlg.update_idletasks()
        x = root.winfo_x() + (root.winfo_width() - dlg.winfo_width()) // 2
        y = root.winfo_y() + (root.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f'+{max(x, 0)}+{max(y, 0)}')

    ttk.Button(t2, text='📁 整理为 Git 工程', command=t2_organize_to_git,
               style='Normal.TButton').grid(
        row=4, column=0, sticky='w', padx=(110, 0), pady=(2, 8))

    # ══════════════════════════════════════
    # TAB 3 — Git 提交
    # ══════════════════════════════════════
    t3 = ttk.Frame(nb, style='TFrame')
    nb.add(t3, text='🔀 Git提交')
    t3.grid_rowconfigure(5, weight=1)
    t3.grid_columnconfigure(0, weight=1)

    # 提前声明，避免闭包引用未定义变量
    t3_log = None
    t3_lf = None
    t3_log_obj = None

    # ── Git 账户状态栏 ──
    ac3 = ttk.LabelFrame(t3, text=' Git 账户 ', )
    ac3.grid(row=0, column=0, sticky='ew', padx=12, pady=(10, 2))
    ac3.grid_columnconfigure(2, weight=1)

    t3_user_var = tk.StringVar(value='检测中...')

    def t3_check_git_user():
        """读取全局 git 配置, 更新状态显示"""
        gu = ge = ''
        try:
            r = subprocess.run(['git', 'config', '--global', 'user.name'],
                              capture_output=True, text=True)
            gu = r.stdout.strip()
        except Exception: pass
        try:
            r = subprocess.run(['git', 'config', '--global', 'user.email'],
                              capture_output=True, text=True)
            ge = r.stdout.strip()
        except Exception: pass
        if gu:
            t3_user_var.set(f'已登录: {gu} <{ge or "无邮箱"}>')
            t3_login_btn.configure(text='切换')
            t3_logout_btn.grid()
        else:
            t3_user_var.set('未登录 — 请先登录才能提交')
            t3_login_btn.configure(text='登录')
            t3_logout_btn.grid_remove()

    def _t3_choose_local_repo():
        """从本地已克隆的 git 仓库中选择 (备选方案)"""
        d = filedialog.askdirectory(title='选择本地已克隆的 git 仓库')
        if not d:
            return
        # 提取远程地址
        try:
            r = subprocess.run(
                ['git', '-C', d, 'remote', 'get-url', 'origin'],
                capture_output=True, text=True, timeout=5)
            if r.returncode == 0 and r.stdout.strip():
                t3_remote.set(r.stdout.strip())
                _log(t3_log, f'✔ 已选择本地仓库: {d}\n  远程: {r.stdout.strip()}',
                     C['green'])
            else:
                t3_path.set(d)
                _log(t3_log, f'✔ 本地仓库: {d} (无 origin 远程)',
                     C['yellow'])
        except Exception as e:
            _log(t3_log, f'读取远程地址失败: {e}', C['red'])

    def t3_manual_input_repo():
        """手动输入仓库地址的快捷助手"""
        dlg = tk.Toplevel(root)
        dlg.title('手动输入仓库地址')
        dlg.resizable(False, False)
        dlg.transient(root)
        dlg.grab_set()
        dlg.configure(bg=C['card'])

        # 注意: tk.Label 构造时不能用 padx/pady tuple (会触发 bad screen distance)
        # 只能用 .pack()/.grid() 的 padx/pady tuple
        tk.Label(dlg, text='仓库地址', bg=C['card'],
                 fg=C['fg'], font=(F, 10, 'bold')
                 ).pack(anchor='w', padx=20, pady=(14, 4))

        tk.Label(dlg, text='支持任意 Git 远程地址:\n'
                            '  git@github.com:owner/repo.git\n'
                            '  git@gitlab.internal:group/repo.git\n'
                            '  https://github.com/owner/repo.git\n'
                            '  http://192.168.1.100/gitlab/ns/repo.git',
                 bg=C['card'], fg=C['sub'],
                 font=(F, 9), justify='left'
                 ).pack(anchor='w', padx=20, pady=(0, 8))

        v = tk.StringVar(value=t3_remote.get() or
                         'git@github.com:')
        entry = ttk.Entry(dlg, textvariable=v, font=(F, 10), width=44)
        entry.pack(padx=20, pady=(0, 8))
        entry.focus_set()
        entry.select_range(0, 'end')

        def _ok():
            url = v.get().strip()
            if not url or '@' not in url and '://' not in url and '.git' not in url:
                messagebox.showwarning('提示', '请输入有效的 Git 仓库地址\n(SSH 或 HTTPS 格式)', parent=dlg)
                return
            t3_remote.set(url)
            dlg.destroy()
            t3_query_branches()

        bf = ttk.Frame(dlg)
        bf.pack(padx=20, pady=(4, 14))
        ttk.Button(bf, text='确定', command=_ok,
                   style='Accent.TButton').pack(side='left', padx=(0, 6))
        ttk.Button(bf, text='取消', command=dlg.destroy,
                   style='Normal.TButton').pack(side='left')

        dlg.bind('<Return>', lambda e: _ok())
        dlg.bind('<Escape>', lambda e: dlg.destroy())

        dlg.update_idletasks()
        x = root.winfo_x() + (root.winfo_width() - dlg.winfo_width()) // 2
        y = root.winfo_y() + (root.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f'+{x}+{y}')

    def t3_do_login():
        """弹出登录框, 录入用户名/邮箱并保存到全局 git config"""
        dlg = tk.Toplevel(root)
        dlg.title('Git 登录')
        dlg.resizable(False, False)
        dlg.transient(root)
        dlg.grab_set()
        dlg.configure(bg=C['card'])

        tk.Label(dlg, text='Git 提交身份', bg=C['card'],
                 fg=C['fg'], font=(F, 10, 'bold')).pack(padx=24, pady=(14, 8))
        tk.Label(dlg, text='用户名:', bg=C['card'], fg=C['sub'],
                 font=(F, 9)).pack(padx=24, anchor='w')
        vu = tk.StringVar()
        vu_entry = ttk.Entry(dlg, textvariable=vu, font=(F, 11), width=32)
        vu_entry.pack(padx=24, pady=(0, 8))
        tk.Label(dlg, text='邮箱:', bg=C['card'], fg=C['sub'],
                 font=(F, 9)).pack(padx=24, anchor='w')
        ve = tk.StringVar()
        ve_entry = ttk.Entry(dlg, textvariable=ve, font=(F, 11), width=32)
        ve_entry.pack(padx=24, pady=(0, 10))

        def _do_save():
            u = vu.get().strip()
            e = ve.get().strip()
            if not u:
                messagebox.showwarning('提示', '用户名不能为空', parent=dlg)
                return
            if not e:
                e = u + '@users.noreply.github.com'
            # ── 改用 _git_set_identity: 优先写到当前仓库的 .git/config (--local),
            #  避免与其它工具 (VSCode / SourceTree / IDE 插件) 抢占
            #  C:\Users\<user>\.gitconfig 文件锁 (WinError 32).
            #  当前仓库为 None 时 (无 repo 上下文) 才回退到 --global. ──
            repo = _get_t3_repo_path()  # 可能为 None
            try:
                ok_global = _git_set_identity(u, e, scope='auto', repo=repo)
            except PermissionError as ex:
                # ── 兜底: 极端情况下 _patched_subprocess_run 重试 3 次仍失败
                #  (比如杀软正在扫描 git.exe 本身), 把 PermissionError 转成
                #  友好提示, 不再让 traceback 弹到 cmd 黑色窗口. ──
                if getattr(ex, 'winerror', None) == 32 or 'WinError 32' in str(ex):
                    messagebox.showwarning(
                        'Git 登录',
                        'git.exe 正在被其它程序占用 (常见原因:\n'
                        '  • 杀软 (Windows Defender / 360 / 火绒) 正在扫描 git.exe\n'
                        '  • 另一个 git 进程正在运行\n'
                        '  • 资源管理器正在预览 .git 目录\n\n'
                        '请等待 5-10 秒后重试, 或临时关闭杀软实时监控后重试。',
                        parent=dlg)
                else:
                    messagebox.showerror('Git 登录', f'保存失败: {ex}', parent=dlg)
                return
            except Exception as ex:
                messagebox.showerror('Git 登录', f'保存失败: {ex}', parent=dlg)
                return
            if not ok_global:
                # 实在写不进 (文件被强占) — 弹出友好提示, 不再裸 traceback
                messagebox.showwarning(
                    'Git 登录',
                    '无法保存 Git 身份 — 配置文件正被其它程序占用。\n\n'
                    '请关闭 VSCode / SourceTree / GitHub Desktop 等正在使用\n'
                    'git 的工具, 然后重试。\n\n'
                    '(本工具下次启动会再次尝试。)',
                    parent=dlg)
            dlg.destroy()
            t3_check_git_user()

        bf = ttk.Frame(dlg)
        bf.pack(padx=24, pady=(4, 14))
        ttk.Button(bf, text='保存并登录', command=_do_save,
                   style='Accent.TButton').pack(side='left', padx=(0, 6))
        ttk.Button(bf, text='取消', command=dlg.destroy,
                   style='Normal.TButton').pack(side='left')

        dlg.update_idletasks()
        x = root.winfo_x() + (root.winfo_width()-dlg.winfo_width())//2
        y = root.winfo_y() + (root.winfo_height()-dlg.winfo_height())//2
        dlg.geometry(f'+{x}+{y}')
        vu_entry.focus_set()
        dlg.bind('<Return>', lambda e: _do_save())
        dlg.bind('<Escape>', lambda e: dlg.destroy())

    def t3_do_logout():
        if messagebox.askyesno('确认注销', '将清除全局 Git 用户名和邮箱, 确认?'):
            # ── 注销时同时尝试 --local + --global, 谁存在就清谁 ──
            # ── 同时吞掉 WinError 32 (文件被占), 不让用户看到 traceback ──
            repo = _get_t3_repo_path()
            for scope_args in (['--local'], ['--global']):
                if repo is None and scope_args == ['--local']:
                    continue
                cwd = repo if scope_args == ['--local'] else None
                try:
                    subprocess.run(['git', 'config', *scope_args, '--unset', 'user.name'],
                                   capture_output=True, cwd=cwd, timeout=5)
                    subprocess.run(['git', 'config', *scope_args, '--unset', 'user.email'],
                                   capture_output=True, cwd=cwd, timeout=5)
                except (PermissionError, OSError, subprocess.TimeoutExpired):
                    continue
            t3_check_git_user()

    ttk.Label(ac3, textvariable=t3_user_var, font=(F, 10),
              foreground=C['green']).grid(row=0, column=0, sticky='w',
              padx=14, pady=(6, 6))
    ttk.Label(ac3, text='', font=(F, 10)).grid(row=0, column=1)  # spacer
    t3_login_btn = ttk.Button(ac3, text='登录', command=t3_do_login,
                               style='Normal.TButton')
    t3_login_btn.grid(row=0, column=2, sticky='e', padx=(4, 2), pady=(6, 6))
    t3_logout_btn = ttk.Button(ac3, text='注销', command=t3_do_logout,
                                style='Normal.TButton')
    t3_logout_btn.grid(row=0, column=3, sticky='e', padx=(0, 14), pady=(6, 6))

    t3_check_git_user()

    fc3 = ttk.LabelFrame(t3, text=' 提交设置 ', )
    fc3.grid(row=1, column=0, sticky='ew', padx=12, pady=(4, 4))
    fc3.grid_columnconfigure(1, weight=1)  # Entry 区域
    # 按钮列 (2/3/4) 宽度固定, 最后一列右贴边

    # 本地路径
    ttk.Label(fc3, text='本地工程:', font=(F, 10)).grid(
        row=0, column=0, sticky='w', padx=14, pady=(8, 2))
    t3_path = tk.StringVar()
    ttk.Entry(fc3, textvariable=t3_path, font=(F, 10)).grid(
        row=0, column=1, sticky='ew', padx=(4, 4), pady=(8, 2))

    def t3_br():
        d = filedialog.askdirectory(title='选择本地工程目录')
        if d:
            t3_path.set(d)

    ttk.Button(fc3, text='浏览', command=t3_br,
               style='Normal.TButton').grid(
        row=0, column=2, padx=(0, 4), pady=(8, 2), sticky='e')

    # 远程仓库
    ttk.Label(fc3, text='远程仓库:', font=(F, 10)).grid(
        row=1, column=0, sticky='w', padx=14, pady=2)
    t3_remote = tk.StringVar()
    ttk.Entry(fc3, textvariable=t3_remote, font=(F, 10)).grid(
        row=1, column=1, sticky='ew', padx=(4, 4), pady=2)
    ttk.Button(fc3, text='\U0001f50d 查询分支', command=None,  # 函数稍后赋值
               style='Normal.TButton').grid(
        row=1, column=2, padx=(0, 4), pady=2, sticky='e')
    ttk.Button(fc3, text='\U0001f310 获取仓库',
               command=lambda: t3_fetch_my_repos(),
               style='Normal.TButton').grid(
        row=1, column=3, padx=(0, 4), pady=2, sticky='e')
    ttk.Button(fc3, text='✏ 手动输入',
               command=t3_manual_input_repo,
               style='Normal.TButton').grid(
        row=1, column=4, padx=(0, 4), pady=2, sticky='e')

    # "获取仓库" 函数 (需在按钮引用前定义)
    # ── 通用 CLI 检测 ──
    def _check_cli(cmd):
        """检查 cmd 是否可执行, 返回 (bool, version_string)"""
        try:
            r = subprocess.run([cmd, '--version'],
                               capture_output=True, text=True, timeout=5)
            if r.returncode == 0:
                return True, r.stdout.strip().split('\n')[0]
        except (FileNotFoundError, Exception):
            pass
        return False, ''

    # ── 获取 GitLab 仓库 (通过 API) ──
    def _fetch_gitlab_repos(api_url, token=None):
        """通过 GitLab REST API 获取仓库列表
        返回 (repos, err_msg):
          - repos: 成功时为 list, 失败时为 []
          - err_msg: 失败时的可读错误 (空字符串=成功)
        """
        url = f'{api_url.rstrip("/")}/api/v4/projects?membership=true&per_page=50&order_by=last_activity_at'
        try:
            cmd = ['curl', '-sS', '--max-time', '15']
            if token:
                cmd += ['-H', f'PRIVATE-TOKEN: {token}']
            cmd.append(url)
            r = subprocess.run(cmd, capture_output=True, text=True, timeout=20)
            if r.returncode != 0:
                # curl 自身失败 (网络/DNS/SSL)
                curl_err = (r.stderr or '').strip() or f'curl 返回码 {r.returncode}'
                return [], f'curl 调用失败: {curl_err}'
            if not r.stdout.strip():
                return [], '服务器返回空内容'
            try:
                data = json.loads(r.stdout)
            except json.JSONDecodeError as e:
                return [], f'返回内容非 JSON: {e}\n前 200 字符: {r.stdout[:200]}'
            if isinstance(data, dict):
                err_msg = data.get('message') or data.get('error') or json.dumps(data, ensure_ascii=False)[:200]
                return [], f'API 错误: {err_msg}'
            if not isinstance(data, list):
                return [], f'返回类型异常: {type(data).__name__}'
            return data, ''
        except subprocess.TimeoutExpired:
            return [], '请求超时 (15 秒未响应)'
        except FileNotFoundError:
            return [], '未找到 curl, 请先安装 Git for Windows 或把 curl 加入 PATH'
        except Exception as e:
            return [], f'未知异常: {type(e).__name__}: {e}'

    def t3_fetch_my_repos():
        """通用仓库获取: 支持 GitHub (gh)/GitLab (glab/API)/手动/本地 — 不弹打扰窗"""
        # 检查可用 CLI
        gh_ok, gh_ver = _check_cli('gh')
        glab_ok, glab_ver = _check_cli('glab')

        # === 构建选择弹窗 ===
        dlg = tk.Toplevel(root)
        dlg.title('获取远程仓库')
        dlg.resizable(False, False)
        dlg.transient(root)
        dlg.configure(bg='white')
        # 关键: 先给 dlg 一个固定尺寸, 否则子控件算不出 reqsize
        dlg.geometry('560x460+200+100')

        # 用一个 Frame 占满 dlg, 强制白色背景, 防止系统主题污染
        root_frame = tk.Frame(dlg, bg='white', bd=0, highlightthickness=0)
        root_frame.pack(fill='both', expand=True)

        # 标题栏
        header = tk.Frame(root_frame, bg='#2b6cb0', height=44)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text='🔗  获取远程仓库', bg='#2b6cb0', fg='white',
                 font=(F, 11, 'bold'),
                 padx=20, pady=12).pack(anchor='w', side='left')

        # 主内容区
        body = tk.Frame(root_frame, bg='white')
        body.pack(fill='both', expand=True, padx=20, pady=14)

        tk.Label(body, text='选择获取方式:', bg='white', fg='#1f2329',
                 font=(F, 10, 'bold'),
                 anchor='w').grid(row=0, column=0, columnspan=2, sticky='w',
                                  pady=(0, 10))

        def _add_opt(row, text, desc, cmd, color='#1f2329'):
            b = tk.Button(body, text=text, command=cmd,
                          bg='#eef1f5', fg='#1f2329', font=(F, 9),
                          activebackground='#e0e4ea', activeforeground='#1f2329',
                          relief='flat', bd=0, padx=14, pady=6, cursor='hand2',
                          anchor='w')
            b.grid(row=row, column=0, sticky='w', pady=3, padx=(0, 10))
            tk.Label(body, text=desc, bg='white', fg=color,
                     font=(F, 9)).grid(row=row, column=1, sticky='w', pady=3)

        row = 1
        if gh_ok:
            _add_opt(row, '📦 GitHub (gh CLI)', f'已安装 · {gh_ver}',
                     lambda: (dlg.destroy(), _t3_fetch_github_repos()),
                     '#2f855a')
        else:
            _add_opt(row, '📦 GitHub (gh CLI)', '未安装 — 点击查看安装指引',
                     lambda: (dlg.destroy(), _t3_show_gh_install()))
        row += 1

        if glab_ok:
            _add_opt(row, '🦊 GitLab (glab CLI)', f'已安装 · {glab_ver}',
                     lambda: (dlg.destroy(), _t3_fetch_gitlab_repos()),
                     '#2f855a')
        else:
            _add_opt(row, '🦊 GitLab (glab CLI)', '未安装 — 点击查看安装指引',
                     lambda: (dlg.destroy(), _t3_show_glab_install()))
        row += 1

        _add_opt(row, '🌐 GitLab API 获取', '适用于内网 GitLab (需 Access Token)',
                 lambda: (dlg.destroy(), _t3_fetch_gitlab_api()))
        row += 1

        # 分组标题
        sep = tk.Label(body, text='💡  无需任何 CLI 的方式:', bg='white', fg='#8a929e',
                       font=(F, 9, 'bold'), anchor='w')
        sep.grid(row=row, column=0, columnspan=2, sticky='w', pady=(14, 6))
        row += 1

        _add_opt(row, '✏  手动输入仓库地址', '支持 SSH / HTTPS 格式',
                 lambda: (dlg.destroy(), t3_manual_input_repo()))
        row += 1
        _add_opt(row, '📂 从本地已克隆仓库选择', '自动提取 origin 远程地址',
                 lambda: (dlg.destroy(), _t3_choose_local_repo()))

        # 底部
        tk.Button(root_frame, text='关闭', command=dlg.destroy,
                  bg='#eef1f5', fg='#1f2329', font=(F, 10),
                  activebackground='#e0e4ea', activeforeground='#1f2329',
                  relief='flat', bd=0, padx=24, pady=8, cursor='hand2'
                  ).pack(side='bottom', pady=(6, 12))

        # 居中到 root
        dlg.update_idletasks()
        try:
            rx, ry = root.winfo_rootx(), root.winfo_rooty()
            rw, rh = root.winfo_width(), root.winfo_height()
            dw, dh = 560, 460
            dlg.geometry(f'+{rx + max(0, (rw - dw) // 2)}+{ry + max(0, (rh - dh) // 2)}')
        except Exception:
            pass
        dlg.update()

    def _t3_show_gh_install():
        """⚠ 精简的 gh 安装提示"""
        msg = ('未安装 gh (GitHub CLI)。\n\n'
               '安装命令:\n'
               '  winget install --id GitHub.cli\n'
               '  (或: scoop install gh)\n\n'
               '安装后需登录:\n'
               '  gh auth login\n\n'
               '💡  目前可用替代方式:\n'
               '  · 手动输入仓库地址\n'
               '  · 从本地已克隆仓库选择')
        messagebox.showinfo('gh CLI 安装指引', msg)

    def _t3_show_glab_install():
        """⚠ 精简的 glab 安装提示"""
        msg = ('未安装 glab (GitLab CLI)。\n\n'
               '安装命令:\n'
               '  winget install --id GitLab.GitLabCLI\n'
               '  (或: scoop install glab)\n\n'
               '安装后需配置:\n'
               '  glab auth login --hostname gitlab.example.com\n\n'
               '💡  目前可用替代方式:\n'
               '  · GitLab API 获取 (需 Token)\n'
               '  · 手动输入仓库地址\n'
               '  · 从本地已克隆仓库选择')
        messagebox.showinfo('glab CLI 安装指引', msg)

    def _t3_fetch_github_repos():
        """通过 gh CLI 获取 GitHub 仓库 (同时返回 SSH 和 HTTPS 地址)"""
        try:
            r = subprocess.run(
                ['gh', 'repo', 'list', '--json', 'nameWithOwner,sshUrl,url',
                 '--limit', '50', '--no-archived'],
                capture_output=True, text=True, timeout=15,
                env={**os.environ, 'GIT_TERMINAL_PROMPT': '0'})
        except FileNotFoundError:
            messagebox.showerror('错误', '未找到 gh CLI')
            return
        except Exception as e:
            messagebox.showerror('错误', f'gh 调用失败:\n{e}')
            return
        if r.returncode != 0:
            messagebox.showerror('gh 错误', (r.stderr or r.stdout or '未知')[:300])
            return
        try:
            data = json.loads(r.stdout)
        except json.JSONDecodeError:
            messagebox.showerror('错误', 'gh 输出非 JSON 格式')
            return
        if not data:
            messagebox.showinfo('提示', '未找到 GitHub 仓库')
            return
        # 每个仓库生成两条: SSH 和 HTTPS, 用户根据网络环境选择
        repos = []
        for r in data:
            name = r.get('nameWithOwner', '')
            ssh = r.get('sshUrl', '')
            https = r.get('url', '')
            if ssh:
                repos.append({'url': ssh, 'name': f'{name}  [SSH]'})
            if https:
                repos.append({'url': https, 'name': f'{name}  [HTTPS]'})
        _show_repo_list_dlg('GitHub', repos,
                            lambda r: r['url'],
                            lambda r: r['name'])

    def _t3_fetch_gitlab_repos():
        """通过 glab CLI 获取 GitLab 仓库 (同时返回 SSH 和 HTTP 地址)"""
        try:
            r = subprocess.run(
                ['glab', 'repo', 'list', '--all', '--per-page', '50',
                 '--output', 'json'],
                capture_output=True, text=True, timeout=15,
                env={**os.environ, 'GIT_TERMINAL_PROMPT': '0'})
        except FileNotFoundError:
            messagebox.showerror('错误', '未找到 glab CLI')
            return
        except Exception as e:
            messagebox.showerror('错误', f'glab 调用失败:\n{e}')
            return
        if r.returncode != 0:
            # 尝试备用格式 (纯文本)
            try:
                r = subprocess.run(
                    ['glab', 'repo', 'list', '--all', '--per-page', '50'],
                    capture_output=True, text=True, timeout=15,
                    env={**os.environ, 'GIT_TERMINAL_PROMPT': '0'})
                if r.returncode != 0:
                    messagebox.showerror('glab 错误',
                                         (r.stderr or '未知错误')[:300])
                    return
                # 解析文本格式: url\tpath
                repos = []
                for line in r.stdout.strip().split('\n'):
                    parts = line.split('\t')
                    if len(parts) >= 2:
                        repos.append({'url': parts[0], 'name': parts[1]})
            except Exception as e:
                messagebox.showerror('glab 错误', str(e)[:300])
                return
        else:
            try:
                data = json.loads(r.stdout)
            except json.JSONDecodeError:
                messagebox.showerror('错误', 'glab 输出非 JSON 格式')
                return
            # 每个仓库生成两条: SSH 和 HTTP
            repos = []
            for r in data:
                name = (r.get('nameWithOwner')
                        or r.get('path_with_namespace')
                        or r.get('name', ''))
                ssh = r.get('sshUrl') or r.get('ssh_url_to_repo', '')
                http = r.get('httpUrl') or r.get('http_url_to_repo', '')
                if ssh:
                    repos.append({'url': ssh, 'name': f'{name}  [SSH]'})
                if http:
                    repos.append({'url': http, 'name': f'{name}  [HTTP]'})
        if not repos:
            messagebox.showinfo('提示', '未找到 GitLab 仓库')
            return
        _show_repo_list_dlg('GitLab', repos,
                            lambda r: r['url'],
                            lambda r: r['name'])

    def _t3_fetch_gitlab_api():
        """通过 GitLab REST API 获取仓库, 适用于内网 GitLab
        单一登录态: 1 个 token 槽位, 登录/登出按钮, 自动记忆"""
        dlg = tk.Toplevel(root)
        dlg.title('GitLab API 配置')
        dlg.resizable(False, False)
        dlg.transient(root)
        dlg.configure(bg='white')
        dlg.geometry('520x420+250+150')

        root_frame = tk.Frame(dlg, bg='white', bd=0, highlightthickness=0)
        root_frame.pack(fill='both', expand=True)
        root_frame.columnconfigure(0, weight=1)
        root_frame.columnconfigure(1, weight=1)

        # ── 标题 + 当前登录态 ──
        tk.Label(root_frame, text='🛰  GitLab API 配置', bg='white', fg='#1f2329',
                 font=(F, 11, 'bold'), anchor='w'
                 ).grid(row=0, column=0, sticky='w', padx=20, pady=(14, 4))

        # 右上角: 登录状态徽章
        gl_cfg = _load_config()
        gl_default_url = gl_cfg.get('gitlab_url') or 'https://gitlab.com'
        gl_stored_token = gl_cfg.get('gitlab_token', '')

        def _update_status():
            """根据当前状态刷新徽章颜色"""
            tok_now = token_var.get().strip() or gl_stored_token
            if tok_now:
                status_lbl.config(text='● 已登录', fg='#38a169')
            else:
                status_lbl.config(text='○ 未登录', fg='#a0aec0')

        status_lbl = tk.Label(root_frame, text='', bg='white',
                              font=(F, 9, 'bold'), anchor='e')
        status_lbl.grid(row=0, column=1, sticky='e', padx=(0, 20), pady=(14, 4))

        tip = ('输入内网 GitLab 地址 + Access Token, 点 "登录" 保存并使用。\n'
               'Token 获取: GitLab → Settings → Access Tokens\n'
               '勾选 read_api + read_repository 权限即可。')
        tk.Label(root_frame, text=tip, bg='white', fg='#8a929e',
                 font=(F, 9), justify='left', wraplength=460, anchor='w'
                 ).grid(row=1, column=0, columnspan=2, sticky='ew',
                        padx=20, pady=(0, 10))

        # ── GitLab 地址 ──
        tk.Label(root_frame, text='GitLab 地址:', bg='white', fg='#1f2329',
                 font=(F, 10), anchor='w'
                 ).grid(row=2, column=0, columnspan=2, sticky='w',
                        padx=20, pady=(4, 0))
        url_var = tk.StringVar(value=gl_default_url)
        url_entry = tk.Entry(root_frame, textvariable=url_var, font=(F, 10),
                             relief='solid', bd=1, bg='white', fg='#1f2329',
                             highlightthickness=1, highlightbackground='#cccccc',
                             highlightcolor='#2b6cb0')
        url_entry.grid(row=3, column=0, columnspan=2, sticky='ew',
                       padx=20, pady=(2, 8))

        # ── Access Token ──
        tk.Label(root_frame, text='Access Token:', bg='white', fg='#1f2329',
                 font=(F, 10), anchor='w'
                 ).grid(row=4, column=0, columnspan=2, sticky='w',
                        padx=20, pady=(4, 0))
        token_var = tk.StringVar(value=gl_stored_token)
        token_entry = tk.Entry(root_frame, textvariable=token_var, font=(F, 10),
                               show='*', relief='solid', bd=1, bg='white',
                               fg='#1f2329',
                               highlightthickness=1, highlightbackground='#cccccc',
                               highlightcolor='#2b6cb0')
        token_entry.grid(row=5, column=0, columnspan=2, sticky='ew',
                         padx=20, pady=(2, 6))
        token_var.trace_add('write', lambda *_: _update_status())
        _update_status()

        # 显示切换 checkbox
        show_var = tk.BooleanVar(value=False)
        tk.Checkbutton(root_frame, text='显示明文 Token',
                       variable=show_var, bg='white', font=(F, 9),
                       command=lambda: token_entry.config(show='' if show_var.get() else '*')
                       ).grid(row=6, column=0, columnspan=2, sticky='w',
                              padx=20, pady=(0, 4))

        # ── 登录 / 登出 按钮 ──
        login_frame = tk.Frame(root_frame, bg='white')
        login_frame.grid(row=7, column=0, columnspan=2, sticky='ew', pady=(6, 4))
        login_frame.columnconfigure(0, weight=0)
        login_frame.columnconfigure(1, weight=0)
        login_frame.columnconfigure(2, weight=1)

        def _do_login():
            """登录: 保存 token + url 到配置文件"""
            url = url_var.get().strip()
            tok = token_var.get().strip()
            if not url:
                messagebox.showwarning('提示', '请输入 GitLab 地址', parent=dlg)
                return
            if not tok:
                messagebox.showwarning('提示', '请输入 Access Token', parent=dlg)
                return
            cfg = _load_config()
            cfg['gitlab_url'] = url
            cfg['gitlab_token'] = tok
            _save_config(cfg)
            gl_stored_token = tok  # 局部缓存
            _update_status()
            messagebox.showinfo('登录成功', f'已登录 GitLab:\n{url}', parent=dlg)

        def _do_logout():
            """登出: 清空配置文件中的 token (其他人可重新登录)"""
            nonlocal gl_stored_token
            if not gl_stored_token and not token_var.get().strip():
                messagebox.showinfo('提示', '当前未登录任何 GitLab 账号', parent=dlg)
                return
            if not messagebox.askyesno('确认登出',
                                       '登出后 token 会被清除, 下次使用需重新输入。\n'
                                       '确认登出?', parent=dlg):
                return
            cfg = _load_config()
            cfg.pop('gitlab_token', None)
            _save_config(cfg)
            gl_stored_token = ''
            token_var.set('')
            _update_status()
            messagebox.showinfo('已登出', 'GitLab 账号已登出', parent=dlg)

        tk.Button(login_frame, text='🔑  登录', command=_do_login,
                  bg='#38a169', fg='white', font=(F, 10, 'bold'),
                  activebackground='#2f855a', activeforeground='white',
                  relief='flat', bd=0, padx=20, pady=6, cursor='hand2'
                  ).grid(row=0, column=0, padx=(20, 6))
        tk.Button(login_frame, text='🚪  登出', command=_do_logout,
                  bg='#ed8936', fg='white', font=(F, 10),
                  activebackground='#dd6b20', activeforeground='white',
                  relief='flat', bd=0, padx=20, pady=6, cursor='hand2'
                  ).grid(row=0, column=1, padx=6)
        tk.Label(login_frame, text='(登录后 token 自动保存, 关闭窗口不影响)',
                 bg='white', fg='#8a929e', font=(F, 8)
                 ).grid(row=0, column=2, padx=(12, 0), sticky='w')

        # ── 底部: 查询仓库 / 取消 ──
        btn_frame = tk.Frame(root_frame, bg='white')
        btn_frame.grid(row=8, column=0, columnspan=2, sticky='ew', pady=(8, 14))
        btn_frame.columnconfigure(0, weight=0)
        btn_frame.columnconfigure(1, weight=0)

        def _do_fetch():
            url = url_var.get().strip()
            # 优先用输入框临时值, 否则用存储的
            token = token_var.get().strip() or gl_stored_token
            if not url:
                messagebox.showwarning('提示', '请输入 GitLab 地址', parent=dlg)
                return
            if not token:
                if not messagebox.askyesno('无 Token',
                                           '未输入 Token, 将尝试匿名访问 (可能失败)。\n'
                                           '继续?', parent=dlg):
                    return
            dlg.destroy()
            repos, err = _fetch_gitlab_repos(url, token)
            if not repos:
                # 把具体错误原因告诉用户, 不再用空泛的"网络/curl"清单
                msg = '未能获取仓库列表。\n\n'
                if err:
                    msg += f'原因: {err}\n\n'
                msg += ('可能原因:\n'
                        '  · GitLab 地址是否正确 (含子路径如 /gitlab)\n'
                        '  · Token 是否有效 + 已勾选 read_api 权限\n'
                        '  · 网络是否可达 (能否浏览器打开此 URL)\n'
                        '  · 是否安装了 curl (Git for Windows 自带)\n'
                        '  · 内网是否需要代理 (HTTP_PROXY 环境变量)')
                messagebox.showerror('获取失败', msg)
                return
            _show_repo_list_dlg('GitLab API', repos,
                                lambda r: r.get('ssh_url_to_repo',
                                         r.get('http_url_to_repo', '')),
                                lambda r: r.get('path_with_namespace',
                                         r.get('name', '')))

        tk.Button(btn_frame, text='查询仓库', command=_do_fetch,
                  bg='#2b6cb0', fg='white', font=(F, 10, 'bold'),
                  activebackground='#1f5380', activeforeground='white',
                  relief='flat', bd=0, padx=24, pady=8, cursor='hand2'
                  ).grid(row=0, column=0, sticky='w', padx=(20, 6))
        tk.Button(btn_frame, text='取消', command=dlg.destroy,
                  bg='#eef1f5', fg='#1f2329', font=(F, 10),
                  activebackground='#e0e4ea', activeforeground='#1f2329',
                  relief='flat', bd=0, padx=24, pady=8, cursor='hand2'
                  ).grid(row=0, column=1, sticky='w', padx=(6, 0))

        url_entry.focus_set()

        # 居中到 root (固定 520x420)
        dlg.update_idletasks()
        try:
            rx, ry = root.winfo_rootx(), root.winfo_rooty()
            rw, rh = root.winfo_width(), root.winfo_height()
            dlg.geometry(f'+{rx + max(0, (rw - 520) // 2)}+{ry + max(0, (rh - 420) // 2)}')
        except Exception:
            pass
        dlg.update()

    def _show_repo_list_dlg(title_prefix, repos, url_getter, name_getter):
        """通用仓库列表选择弹窗 (grid 布局, 搜索/列表/按钮互不遮挡)"""
        dlg = tk.Toplevel(root)
        dlg.title(f'选择 {title_prefix} 仓库')
        dlg.resizable(True, True)
        dlg.transient(root)
        dlg.grab_set()
        dlg.configure(bg=C['card'])
        dlg.geometry('560x480')

        # grid 布局: row0 标题, row1 搜索, row2 列表(weight=1), row3 按钮
        dlg.columnconfigure(0, weight=1)
        dlg.rowconfigure(2, weight=1)

        # 标题
        tk.Label(dlg, text=f'{title_prefix}: 共 {len(repos)} 个仓库',
                 bg=C['card'], fg=C['fg'], font=(F, 10, 'bold')
                 ).grid(row=0, column=0, sticky='ew', padx=14, pady=(12, 4))

        # 搜索框
        sf = ttk.Frame(dlg)
        sf.grid(row=1, column=0, sticky='ew', padx=14, pady=(0, 4))
        sf.columnconfigure(0, weight=1)
        sv = tk.StringVar()
        ttk.Entry(sf, textvariable=sv, font=(F, 10)).grid(
            row=0, column=0, sticky='ew', padx=(0, 6))
        ttk.Label(sf, text='搜索过滤', foreground=C['sub'],
                  font=(F, 8)).grid(row=0, column=1)

        # 列表 (Listbox + Scrollbar)
        lf = ttk.Frame(dlg)
        lf.grid(row=2, column=0, sticky='nsew', padx=14, pady=(0, 4))
        lf.columnconfigure(0, weight=1)
        lf.rowconfigure(0, weight=1)
        lb = tk.Listbox(lf, font=(F, 10), bg=C['ebg'], fg=C['fg'],
                        relief='flat', selectbackground='#cfe2f3',
                        selectforeground=C['blue'],
                        highlightthickness=0, bd=0,
                        activestyle='dotbox')
        lb.grid(row=0, column=0, sticky='nsew')
        sb = ttk.Scrollbar(lf, orient='vertical', command=lb.yview)
        sb.grid(row=0, column=1, sticky='ns')
        lb.configure(yscrollcommand=sb.set)

        # 填充列表
        repo_urls = []
        repo_names = []
        for r in repos:
            name = name_getter(r)
            url = url_getter(r)
            repo_names.append(name)
            repo_urls.append(url)
            lb.insert('end', name if name else '(无名称)')
        print(f'[_show_repo_list_dlg] {title_prefix} 填充 {len(repo_names)} 项, '
              f'前 3 项示例: {repo_names[:3]}', file=_sys.stderr)

        # 搜索过滤
        def on_search(*_):
            q = sv.get().strip().lower()
            lb.delete(0, 'end')
            for name in repo_names:
                if q in name.lower():
                    lb.insert('end', name if name else '(无名称)')
        sv.trace_add('write', on_search)

        def on_select(e=None):
            sel = lb.curselection()
            if not sel:
                return
            idx = sel[0]
            selected_name = lb.get(idx)
            try:
                orig_idx = repo_names.index(selected_name)
            except ValueError:
                orig_idx = idx
            if orig_idx < len(repo_urls):
                t3_remote.set(repo_urls[orig_idx])
                if not t3_branch.get().strip():
                    t3_branch.set('feature/update')
            dlg.destroy()
            t3_query_branches()

        lb.bind('<Double-Button-1>', on_select)
        lb.bind('<Return>', on_select)

        # 底部按钮
        bf = ttk.Frame(dlg)
        bf.grid(row=3, column=0, sticky='ew', padx=14, pady=(4, 10))
        ttk.Button(bf, text='选择', command=on_select,
                   style='Accent.TButton').pack(side='left', padx=(0, 6))
        ttk.Button(bf, text='取消', command=dlg.destroy,
                   style='Normal.TButton').pack(side='left')

        _dcenter(dlg)

    # 分支查询结果显示
    t3_branch_info = tk.StringVar(value='输入仓库链接或点 "获取仓库" 自动填充')
    ttk.Label(fc3, textvariable=t3_branch_info, font=(F, 9),
              foreground=C['blue'], anchor='w').grid(
        row=2, column=1, columnspan=2, sticky='nsew', padx=(4, 14), pady=(0, 6))

    # 日期 + 改动
    ttk.Label(fc3, text='日期:', font=(F, 10)).grid(
        row=3, column=0, sticky='w', padx=14, pady=2)
    t3_date = tk.StringVar(value=datetime.date.today().isoformat())
    ttk.Entry(fc3, textvariable=t3_date, font=(F, 10), width=14).grid(
        row=3, column=1, sticky='w', padx=(4, 4), pady=2)

    ttk.Label(fc3, text='改动说明:', font=(F, 10)).grid(
        row=4, column=0, sticky='w', padx=14, pady=(2, 2))
    t3_msg = tk.StringVar()
    ttk.Entry(fc3, textvariable=t3_msg, font=(F, 10)).grid(
        row=4, column=1, columnspan=2, sticky='nsew',
        padx=(4, 14), pady=(2, 2))

    ttk.Label(fc3, text='分支名:', font=(F, 10)).grid(
        row=5, column=0, sticky='w', padx=14, pady=(2, 8))
    t3_branch = tk.StringVar()
    ttk.Entry(fc3, textvariable=t3_branch, font=(F, 10), width=16).grid(
        row=5, column=1, sticky='w', padx=(4, 4), pady=(2, 8))
    ttk.Label(fc3, text='留空=当前分支; 填名称=新建/切换到该分支',
              foreground=C['sub'], font=(F, 8)).grid(
        row=5, column=2, sticky='w', padx=(0, 14), pady=(2, 8))

    t3_gen_gi = tk.BooleanVar(value=True)
    ttk.Checkbutton(fc3, text='自动生成 .gitignore (文件全被过滤时可取消)',
                    variable=t3_gen_gi).grid(
        row=6, column=0, columnspan=3, sticky='w', padx=14, pady=(2, 8))

    # 按钮
    b3 = ttk.Frame(t3, style='TFrame')
    b3.grid(row=2, column=0, sticky='nsew', padx=12, pady=(2, 4))

    t3_submit_btn = ttk.Button(b3, text='\u25b6 提交到当前分支',
                               style='Accent.TButton')
    t3_submit_btn.grid(row=0, column=0, padx=(0, 6))
    t3_merge_btn = ttk.Button(b3, text='\u21c4 合并到 main',
                              style='Success.TButton')
    t3_merge_btn.grid(row=0, column=1, padx=(0, 6))
    t3_clone_btn = ttk.Button(b3, text='\u2b07 下载仓库',
                              style='Normal.TButton')
    t3_clone_btn.grid(row=0, column=2, padx=(0, 6))
    t3_refresh_btn = ttk.Button(b3, text='\u21bb 刷新历史',
                                style='Normal.TButton')
    t3_refresh_btn.grid(row=0, column=3)
    t3_delete_btn = ttk.Button(b3, text='\U0001f5d1 删除仓库文件',
                                style='Danger.TButton')
    t3_delete_btn.grid(row=0, column=4, padx=(6, 0))

    def t3_delete_repo_files():
        """删除远程仓库文件: 临时克隆→删→推→销毁, 不碰本地任何文件"""
        remote_url = t3_remote.get().strip()
        if not remote_url:
            messagebox.showerror('错误', '请先填写远程仓库地址 (SSH 或 HTTP)')
            return

        # 从远程获取分支列表
        _log(t3_log, '')
        _log(t3_log, f'🔍 查询远程仓库...', C['blue'])

        def _bg_list():
            def l(msg, color=None):
                root.after(0, lambda: _log(t3_log, msg, color or C['fg']))

            try:
                rv = subprocess.run(['git', 'ls-remote', '--heads', remote_url],
                                    capture_output=True, text=True, timeout=15)
            except subprocess.TimeoutExpired:
                l('查询超时', C['red'])
                return
            except Exception as e:
                l(f'查询失败: {e}', C['red'])
                return
            if rv.returncode != 0:
                l((rv.stderr or '未知错误').strip()[:300], C['red'])
                return

            branches = []
            for line in rv.stdout.strip().split('\n'):
                if line and 'refs/heads/' in line:
                    b = line.split('refs/heads/')[-1].strip()
                    if b:
                        branches.append(b)
            if not branches:
                l('未找到任何分支', C['yellow'])
                return

            l(f'共 {len(branches)} 个远程分支', C['sub'])

            # 直接用当前分支名
            cur_branch = (t3_branch.get().strip()
                          or t3_branch_info.get().split('|')[0].replace('\u2714', '').strip()
                          or 'main')
            l(f'分支: {cur_branch}', C['sub'])
            root.after(0, lambda: _do_clone_and_show(cur_branch))

            def _do_clone_and_show(cur_branch):
                import tempfile
                tmp = tempfile.mkdtemp(prefix='git_del_')
                l(f'⏳ 克隆 {cur_branch} ...', C['sub'])

                try:
                    rv = subprocess.run(
                        ['git', 'clone', '--depth=1', '--single-branch', '-b', cur_branch,
                         remote_url, tmp],
                        capture_output=True, text=True, timeout=120,
                        env={**os.environ, 'GIT_TERMINAL_PROMPT': '0',
                             'GCM_INTERACTIVE': 'Never'})
                except subprocess.TimeoutExpired:
                    l('克隆超时 (120s)', C['red'])
                    shutil.rmtree(tmp, ignore_errors=True)
                    return
                if rv.returncode != 0:
                    err = (rv.stderr or rv.stdout).strip()[:400]
                    l(f'克隆失败: {err}', C['red'])
                    if 'authentication' in err.lower() or 'Permission denied' in err:
                        l('提示: 内网 GitLab 可能只支持 HTTP, 非 SSH', C['yellow'])
                    shutil.rmtree(tmp, ignore_errors=True)
                    return

                l(f'✔ {cur_branch} 克隆完成', C['green'])

                # 获取文件列表
                rv2 = subprocess.run(['git', '-C', tmp, 'ls-files', '-z'],
                                    capture_output=True, timeout=15)
                enc = _sys.getfilesystemencoding()
                files = sorted([f for f in rv2.stdout.decode(enc, errors='replace').split('\0') if f.strip()])

                if not files:
                    l(f'{cur_branch} 分支无文件', C['yellow'])
                    shutil.rmtree(tmp, ignore_errors=True)
                    return

                l(f'文件: {len(files)} 个', C['sub'])

                # ── 弹窗 ──
                def _show_dlg():
                    dlg = tk.Toplevel(root)
                    dlg.title(f'删除远程文件 — {cur_branch}')
                    dlg.resizable(True, True)
                    dlg.transient(root)
                    dlg.grab_set()
                    dlg.configure(bg=C['card'])
                    dlg.geometry('700x550')
                    dlg.minsize(500, 350)
    
                    tk.Label(dlg, text=f'远程: {remote_url}', bg=C['card'],
                             fg=C['sub'], font=(F, 9)).pack(anchor='w', padx=16, pady=(12, 0))
                    tk.Label(dlg, text=f'分支: {cur_branch}  |  文件: {len(files)} 个  |  ⚠ 不影响本地文件',
                             bg=C['card'], fg=C['sub'], font=(F, 9)).pack(anchor='w', padx=16, pady=(2, 8))
    
                    sf = ttk.Frame(dlg, style='TFrame')
                    sf.pack(fill='x', padx=16, pady=(0, 4))
                    tk.Label(sf, text='过滤:', bg=C['card'], fg=C['fg'], font=(F, 9)).pack(side='left', padx=(0, 6))
                    filter_var = tk.StringVar()
                    fe = ttk.Entry(sf, textvariable=filter_var, font=(F, 9))
                    fe.pack(side='left', fill='x', expand=True)
    
                    lf = ttk.Frame(dlg, style='TFrame')
                    lf.pack(fill='both', expand=True, padx=16, pady=(4, 8))
                    lb = tk.Listbox(lf, selectmode='extended', font=(F, 9),
                                    bg='#ffffff', fg=C['fg'], activestyle='none',
                                    selectbackground='#2b6cb0', selectforeground='#ffffff',
                                    exportselection=False)
                    sb = ttk.Scrollbar(lf, orient='vertical', command=lb.yview)
                    lb.configure(yscrollcommand=sb.set)
                    lb.pack(side='left', fill='both', expand=True)
                    sb.pack(side='right', fill='y')
    
                    def _refresh_list(*_):
                        filt = filter_var.get().lower()
                        lb.delete(0, 'end')
                        for f in files:
                            if not filt or filt in f.lower():
                                lb.insert('end', f)
                    _refresh_list()
                    filter_var.trace_add('write', _refresh_list)
    
                    bf = ttk.Frame(dlg, style='TFrame')
                    bf.pack(fill='x', padx=16, pady=(0, 12))
                    ttk.Button(bf, text='全选', command=lambda: lb.selection_set(0, lb.size()-1),
                               style='Normal.TButton').pack(side='left', padx=(0, 6))
                    ttk.Button(bf, text='取消全选', command=lambda: lb.selection_clear(0, lb.size()-1),
                               style='Normal.TButton').pack(side='left')
    
                    def _do_delete():
                        indices = lb.curselection()
                        if not indices:
                            messagebox.showwarning('提示', '请至少选择一个文件', parent=dlg)
                            return
                        sel_files = [lb.get(i) for i in indices]
                        if not messagebox.askyesno('确认删除',
                                                   f'从远程仓库删除 {len(sel_files)} 个文件\n'
                                                   f'(不影响本地文件)\n\n'
                                                   f'分支: {cur_branch}\n远程: {remote_url}\n\n' +
                                                   '\n'.join(sel_files[:15]) +
                                                   (f'\n... 共 {len(sel_files)} 个' if len(sel_files) > 15 else ''),
                                                   parent=dlg):
                            return
                        dlg.destroy()
    
                        _log(t3_log, '')
                        _log(t3_log, f'🗑 删除远程 {len(sel_files)} 个文件... {cur_branch}', C['blue'])
    
                        def _bg_del():
                            def ll(msg, color=None):
                                root.after(0, lambda: _log(t3_log, msg, color or C['fg']))
    
                            # git rm + commit + push 全部在临时目录
                            enc = _sys.getfilesystemencoding()
                            rv = subprocess.run(['git', '-C', tmp, 'rm'] + sel_files,
                                                capture_output=True, text=True, timeout=30,
                                                encoding=enc, errors='replace')
                            if rv.returncode != 0:
                                ll(f'✘ git rm 失败: {(rv.stderr or rv.stdout).strip()[:300]}', C['red'])
                                shutil.rmtree(tmp, ignore_errors=True)
                                return
    
                            rv = subprocess.run(['git', '-C', tmp, 'commit', '-m',
                                                 f'delete(remote): remove {len(sel_files)} file(s)'],
                                                capture_output=True, text=True, timeout=30)
                            if rv.returncode != 0:
                                err = (rv.stderr or rv.stdout).strip()[:300]
                                if 'nothing to commit' in err.lower():
                                    ll('没有需要提交的变更', C['yellow'])
                                else:
                                    ll(f'✘ git commit 失败: {err}', C['red'])
                                shutil.rmtree(tmp, ignore_errors=True)
                                return
                            ll(rv.stdout.strip(), C['green'])
    
                            ll(f'↑ 推送 {cur_branch} → origin ...', C['blue'])
                            try:
                                rv = subprocess.run(['git', '-C', tmp, 'push', '--force', 'origin', cur_branch],
                                                    capture_output=True, text=True, timeout=60,
                                                    env={**os.environ, 'GIT_TERMINAL_PROMPT': '0',
                                                         'GCM_INTERACTIVE': 'Never'})
                            except subprocess.TimeoutExpired:
                                ll('推送超时 (60s)', C['red'])
                                shutil.rmtree(tmp, ignore_errors=True)
                                return
                            if rv.returncode == 0:
                                ll(rv.stdout.strip(), C['green'])
                                if rv.stderr:
                                    ll(rv.stderr.strip(), C['yellow'])
                                ll(f'✔ 远程已删除 {len(sel_files)} 个文件', C['green'])
                                # 同步本地仓库引用, 避免后续提交出现 fetch first
                                local_repo = t3_path.get().strip()
                                if local_repo and os.path.isdir(os.path.join(local_repo, '.git')):
                                    subprocess.run(['git', '-C', local_repo, 'fetch', 'origin', cur_branch],
                                                   capture_output=True, timeout=15)
                                    ll('  已同步本地引用', C['sub'])
                            else:
                                ll((rv.stderr or rv.stdout).strip()[:500], C['red'])
                                if 'rejected' in (rv.stderr or '').lower():
                                    ll('提示: 推送被拒绝，可能需要先 pull', C['yellow'])
    
                            # 清理临时目录
                            shutil.rmtree(tmp, ignore_errors=True)
                            root.after(100, t3_history)
    
                        threading.Thread(target=_bg_del, daemon=True).start()
    
                    ttk.Button(bf, text='🗑 确认删除', command=_do_delete,
                               style='Danger.TButton').pack(side='right')
                    ttk.Button(bf, text='取消', command=lambda: (dlg.destroy(), shutil.rmtree(tmp, ignore_errors=True)),
                               style='Normal.TButton').pack(side='right', padx=(0, 6))
                    fe.focus_set()
    
                root.after(0, _show_dlg)

        threading.Thread(target=_bg_list, daemon=True).start()

    t3_delete_btn.config(command=t3_delete_repo_files)

    # ── 删除远程分支按钮 ──
    t3_del_branch_btn = ttk.Button(b3, text='\U0001f5d1 删除分支',
                                    style='Danger.TButton')
    t3_del_branch_btn.grid(row=0, column=5, padx=(6, 0))

    def t3_delete_branch():
        remote_url = t3_remote.get().strip()
        cur_branch = (t3_branch.get().strip()
                      or t3_branch_info.get().split('|')[0].replace('\u2714', '').strip()
                      or '')
        if not remote_url:
            messagebox.showerror('错误', '请先填写远程仓库地址')
            return
        if not cur_branch:
            cur_branch = tk.simpledialog.askstring('分支名', '请输入要删除的远程分支名:', parent=root)
            if not cur_branch:
                return
        if not messagebox.askyesno('确认删除分支',
                                   f'确定删除远程分支 {cur_branch} ?\n\n'
                                   f'远程: {remote_url}\n此操作不可恢复!',
                                   parent=root):
            return

        _log(t3_log, '')
        _log(t3_log, f'\U0001f5d1 删除远程分支 {cur_branch} ...', C['blue'])

        def _bg():
            def l(msg, color=None):
                root.after(0, lambda: _log(t3_log, msg, color or C['fg']))
            try:
                rv = subprocess.run(['git', 'push', 'origin', '--delete', cur_branch],
                                    capture_output=True, text=True, timeout=30,
                                    env={**os.environ, 'GIT_TERMINAL_PROMPT': '0', 'GCM_INTERACTIVE': 'Never'})
            except subprocess.TimeoutExpired:
                l('推送超时 (30s)', C['red']); return
            if rv.returncode == 0:
                l(rv.stdout.strip(), C['green'])
                if rv.stderr: l(rv.stderr.strip(), C['yellow'])
                l(f'\u2714 远程分支 {cur_branch} 已删除', C['green'])
                # 同步本地仓库引用
                local_repo = t3_path.get().strip()
                if local_repo and os.path.isdir(os.path.join(local_repo, '.git')):
                    subprocess.run(['git', '-C', local_repo, 'fetch', 'origin', '--prune'],
                                   capture_output=True, timeout=15)
                root.after(500, t3_query_branches)
            else:
                l((rv.stderr or rv.stdout).strip()[:400], C['red'])
            root.after(100, t3_history)
        threading.Thread(target=_bg, daemon=True).start()

    t3_del_branch_btn.config(command=t3_delete_branch)

    def t3_query_branches():
        """查询远程仓库的分支列表 (git ls-remote)"""
        url = t3_remote.get().strip()
        if not url:
            t3_branch_info.set('请先输入远程仓库地址')
            return

        _log(t3_log, f'\U0001f50d 正在查询 {url} ...', C['blue'])
        t3_branch_info.set('查询中...')

        def _bg():
            try:
                r = subprocess.run(
                    ['git', 'ls-remote', '--heads', url],
                    capture_output=True, text=True, timeout=30,
                    env={**os.environ, 'GIT_TERMINAL_PROMPT': '0',
                         'GCM_INTERACTIVE': 'Never'})

                if r.returncode != 0:
                    err = (r.stderr or '查询失败').strip()[:200]
                    root.after(0, lambda: _log(t3_log, f'✘ {err}', C['red']))
                    root.after(0, lambda: t3_branch_info.set(f'查询失败: {err[:60]}'))
                    return

                lines = [l for l in r.stdout.strip().split('\n') if l.strip()]
                branches = []
                for line in lines:
                    parts = line.split('\t')
                    if len(parts) >= 2:
                        ref = parts[-1]
                        if ref.startswith('refs/heads/'):
                            branches.append(ref[11:])

                if not branches:
                    root.after(0, lambda: _log(t3_log, '未找到任何分支', C['yellow']))
                    root.after(0, lambda: t3_branch_info.set('此仓库没有分支'))
                    return

                # 识别主分支
                main_branch = None
                for b in branches:
                    if b.lower() in ('main', 'master'):
                        main_branch = b
                        break
                if not main_branch:
                    main_branch = branches[0]  # 第一个作为默认主分支

                # 更新摘要标签
                cnt = len(branches)
                info = f'✔ 共 {cnt} 个分支 | 主分支: {main_branch}'
                root.after(0, lambda: t3_branch_info.set(info))

                # 日志输出详细列表
                root.after(0, lambda: _log(
                    t3_log,
                    f'✔ 远程仓库包含 {cnt} 个分支 (主分支: {main_branch}):',
                    C['green']))
                for b in branches:
                    marker = '★' if b == main_branch else '  '
                    clr = C['blue'] if b == main_branch else C['fg']
                    # 闭包陷阱修复
                    root.after(0, lambda b=b, marker=marker, clr=clr:
                        _log(t3_log, f'{marker} {b}', clr))

            except subprocess.TimeoutExpired:
                root.after(0, lambda: _log(t3_log, '查询超时 (30s)', C['red']))
                root.after(0, lambda: t3_branch_info.set('查询超时'))
            except Exception as e:
                # ── 把 WinError 32 / PermissionError 翻译成友好提示 ──
                # 内网常因其它 SSH 客户端 (VSCode Remote / SourceTree /
                # Termius 等) 占用 ~/.ssh/config 或 known_hosts 而短暂失败,
                # _patched_subprocess_run 已重试 3 次仍失败, 说明占用较久.
                err_str = str(e)
                winerr = getattr(e, 'winerror', None)
                is_lock = (winerr == 32) or ('WinError 32' in err_str)
                if is_lock:
                    msg = ('SSH 配置文件正被其它程序占用 (WinError 32)。\n'
                           '请临时关闭以下任一工具, 然后重新点 "查询分支":\n'
                           '  • VSCode (Remote-SSH 扩展)\n'
                           '  • SourceTree / GitHub Desktop / Fork\n'
                           '  • Termius / Xshell / WinSCP / MobaXterm\n'
                           '  • 任何正在用此 SSH 密钥的 ssh.exe 进程\n'
                           '  (占用文件:  C:\\Users\\<你>\\.ssh\\config 或 known_hosts)')
                else:
                    msg = f'查询失败: {err_str[:200]}'
                root.after(0, lambda: _log(t3_log, msg, C['red']))
                root.after(0, lambda: t3_branch_info.set('查询失败 (见日志)'))

        threading.Thread(target=_bg, daemon=True).start()

    # 将查询按钮的回调绑定到已定义的函数
    for child in fc3.grid_slaves(row=1, column=2):
        if isinstance(child, ttk.Button):
            child.config(command=t3_query_branches)

    def t3_submit(merge_to_main=False):
        p = t3_path.get().strip()
        r = t3_remote.get().strip()
        m = t3_msg.get().strip()
        d = t3_date.get().strip()

        if not p or not os.path.isdir(p):
            messagebox.showerror('错误', '请选择有效的工程目录')
            return
        if not m:
            messagebox.showerror('错误', '请填写改动说明')
            return

        _log(t3_log, '', '')
        _log(t3_log, '\u25b6 开始提交...', C['blue'])
        cm = f'[{d}] {m}'

        # 收集 git 用户信息（顶部 Git 账户栏已处理登录）
        git_user = ''
        git_email = ''
        try:
            gu = subprocess.run(['git', 'config', '--global', 'user.name'],
                                capture_output=True, text=True)
            git_user = gu.stdout.strip()
        except Exception:
            pass
        try:
            ge = subprocess.run(['git', 'config', '--global', 'user.email'],
                                capture_output=True, text=True)
            git_email = ge.stdout.strip()
        except Exception:
            pass

        if not git_user:
            root.after(0, lambda: _log(
                t3_log,
                '✘ 未登录 Git 账户! 请点击顶部 "登录" 按钮设置用户名和邮箱',
                C['red']))
            root.after(0, t3_history)
            return

        def _bg():
            def l(msg, color=None):
                if color is None:
                    color = C['fg']
                root.after(0, lambda: _log(t3_log, msg, color))

            # 1. 检查 git
            try:
                subprocess.run(['git', '--version'],
                               capture_output=True, check=True)
            except Exception:
                l('\u2718 未找到 Git，请先安装 Git', C['red'])
                root.after(0, t3_history)
                return

            # 2. .gitignore (可选)
            if t3_gen_gi.get():
                gi = os.path.join(p, '.gitignore')
                if not os.path.exists(gi):
                    with open(gi, 'w', encoding='utf-8') as f:
                        f.write(_generate_fpga_gitignore())
                    l('\u2714 已自动创建 .gitignore', C['green'])
            else:
                l('已跳过 .gitignore 创建', C['sub'])

            # 3. git init
            dg = os.path.join(p, '.git')
            if not os.path.exists(dg):
                rv = subprocess.run(['git', 'init'],
                                    capture_output=True, text=True, cwd=p)
                l(rv.stdout.strip())
                if rv.stderr:
                    l(rv.stderr.strip(), C['yellow'])

            # 3.5 切换到指定分支
            target_branch = t3_branch.get().strip()
            if target_branch:
                rv = subprocess.run(['git', 'checkout', '-B', target_branch],
                                    capture_output=True, text=True, cwd=p)
                if rv.returncode == 0:
                    l(f'\u21c4 已切换到分支: {target_branch}', C['green'])
                else:
                    l((rv.stderr or rv.stdout).strip()[:120], C['yellow'])

                        # 3.6 提交用户信息
            l(f'提交用户: {git_user} <{git_email}>', C['sub'])

            # 3.7 自动创建/更新 说明.txt (含上传时间 + 改动说明)
            try:
                notes_path = os.path.join(p, '说明.txt')
                now_ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                note_lines = [
                    '═══════════════════════════════════════════════════════════════',
                    f'  FPGA 工程上传说明',
                    '═══════════════════════════════════════════════════════════════',
                    '',
                    f'  上传时间:  {now_ts}',
                    f'  提交用户:  {git_user} <{git_email}>',
                    f'  改动说明:  {m}',
                    f'  日期标签:  [{d}]',
                ]
                if t3_branch.get().strip():
                    note_lines.append(f'  分支:      {t3_branch.get().strip()}')
                note_lines += [
                    '',
                    '───────────────────────────────────────────────────────────────',
                    '  历史记录 (从下到上 = 从旧到新):',
                    '───────────────────────────────────────────────────────────────',
                    '',
                ]
                # 若已存在说明.txt, 把旧内容追加到下面 (历史)
                old_content = ''
                if os.path.isfile(notes_path):
                    try:
                        with open(notes_path, 'r', encoding='utf-8') as fh:
                            old_content = fh.read()
                    except Exception:
                        pass
                full = '\n'.join(note_lines) + '\n' + old_content
                with open(notes_path, 'w', encoding='utf-8') as fh:
                    fh.write(full)
                l(f'✔ 已生成说明.txt (含本次上传信息)', C['green'])
            except Exception as e:
                l(f'· 说明.txt 生成失败 (非致命): {e}', C['yellow'])

            # 4. git add
            rv = subprocess.run(['git', 'add', '.'],
                                capture_output=True, text=True, cwd=p)
            if rv.stdout:
                l(rv.stdout.strip())
            if rv.stderr:
                l(rv.stderr.strip(), C['yellow'])

            # 4.5 检查暂存情况
            rv2 = subprocess.run(['git', 'diff', '--cached', '--name-only'],
                                 capture_output=True, text=True, cwd=p)
            staged = [x for x in rv2.stdout.strip().split('\n') if x]
            if staged:
                l(f'\u2714 已暂存 {len(staged)} 个文件:', C['green'])
                for f in staged[:20]:
                    l(f'  + {f}', C['green'])
                if len(staged) > 20:
                    l(f'    ... 共 {len(staged)} 个', C['sub'])
            else:
                l('\u26a0 没有文件被暂存!', C['red'])
                l('  .gitignore 可能过滤了所有文件', C['yellow'])
                l('  请检查工程目录或删除 .gitignore 重试', C['sub'])
                # 显示 git status
                rv3 = subprocess.run(['git', 'status', '--short'],
                                     capture_output=True, text=True, cwd=p)
                if rv3.stdout.strip():
                    l('  git status:', C['sub'])
                    for line in rv3.stdout.strip().split('\n')[:10]:
                        l(f'    {line}', C['sub'])

            # 5. git commit (使用全局配置)
            commit_cmd = ['git', 'commit', '-m', cm]
            rv = subprocess.run(commit_cmd, capture_output=True, text=True, cwd=p)
            if rv.returncode == 0:
                l(rv.stdout.strip(), C['green'])
            else:
                # 可能是 nothing to commit 或其它错误
                out = rv.stdout.strip() or rv.stderr.strip()
                l(out, C['yellow'])
                if 'nothing to commit' in out.lower():
                    l('没有需要提交的变更', C['sub'])
                elif 'not a git repository' in out.lower():
                    return  # 致命错误
                else:
                    # 非致命，继续尝试 push
                    pass

            # 6. 获取当前分支名
            target_branch = t3_branch.get().strip()
            if target_branch:
                cur_branch = target_branch
            else:
                cur_branch = 'master'
                rv = subprocess.run(['git', 'branch', '--show-current'],
                                    capture_output=True, text=True, cwd=p)
                if rv.stdout.strip():
                    cur_branch = rv.stdout.strip()

            # 7. git remote + push
            if r:
                rv = subprocess.run(['git', 'remote', '-v'],
                                    capture_output=True, text=True, cwd=p)
                if 'origin' in rv.stdout:
                    l('origin 已存在', C['yellow'])
                else:
                    r2 = subprocess.run(
                        ['git', 'remote', 'add', 'origin', r],
                        capture_output=True, text=True, cwd=p)
                    l(f'\u2714 remote add: {r}', C['green'])
                    if r2.stderr:
                        l(r2.stderr.strip(), C['yellow'])

                # 决定推送目标分支
                if merge_to_main:
                    l(f'⇄ 合并 {cur_branch} → main ...', C['blue'])

                    # 1) 确保 main 分支存在
                    rv_m = subprocess.run(
                        ['git', 'show-ref', '--verify', '--quiet',
                         'refs/heads/main'],
                        capture_output=True, cwd=p)
                    if rv_m.returncode != 0:
                        subprocess.run(['git', 'branch', 'main', 'HEAD'],
                                       capture_output=True, cwd=p)
                        l('  已创建本地 main 分支', C['sub'])

                    # 2) 切到 main，合并功能分支
                    subprocess.run(['git', 'checkout', 'main'],
                                   capture_output=True, cwd=p)
                    mr_rv = subprocess.run(
                        ['git', 'merge', cur_branch, '--allow-unrelated-histories',
                         '--no-edit'],
                        capture_output=True, text=True, cwd=p,
                        env={**os.environ, 'GIT_TERMINAL_PROMPT': '0'})
                    if mr_rv.returncode == 0:
                        l(f'  已合并 {cur_branch} → main', C['green'])
                    elif mr_rv.stderr:
                        l(f'  合并警告: {(mr_rv.stderr or mr_rv.stdout).strip()[:120]}',
                          C['yellow'])

                    push_branch = 'main'
                else:
                    push_branch = cur_branch

                l(f'↑ 推送 {push_branch} → origin ...', C['blue'])
                push_ok = False
                try:
                    push_cmd = ['git', 'push', '-u', 'origin', push_branch]
                    if merge_to_main:
                        push_cmd.insert(2, '--force-with-lease')
                    rv = subprocess.run(
                        push_cmd,
                        capture_output=True, text=True, cwd=p, timeout=90,
                        env={**os.environ, 'GIT_TERMINAL_PROMPT': '0',
                             'GCM_INTERACTIVE': 'Never'})
                except subprocess.TimeoutExpired:
                    l('推送超时 (90s) — 认证卡住', C['red'])
                    l('建议: 确认 SSH 密钥已添加到 GitHub/GitLab', C['yellow'])
                    return
                if rv.returncode == 0:
                    l(rv.stdout.strip(), C['green'])
                    if rv.stderr:
                        l(rv.stderr.strip(), C['yellow'])
                    push_ok = True
                else:
                    err = (rv.stderr or rv.stdout).strip()
                    l(err, C['red'])
                    if 'authentication' in err.lower() or 'credentials' in err.lower():
                        l('提示: 需要 GitHub 认证。请使用 Personal Access Token', C['yellow'])
                        l('  Settings → Developer settings → Tokens (classic)', C['sub'])
                    elif 'remote' in err.lower() and 'not found' in err.lower():
                        l('提示: 远程仓库不存在或无权限访问', C['yellow'])
                    elif 'rejected' in err.lower():
                        l('提示: 推送被拒绝，可能需要先 git pull', C['yellow'])

                # 3) 推送成功后删除源分支 (本地 + 远程)
                if merge_to_main and push_ok and cur_branch != 'main':
                    l(f'\U0001f5d1 删除分支 {cur_branch} ...', C['sub'])
                    # 本地删除
                    dl = subprocess.run(
                        ['git', 'branch', '-d', cur_branch],
                        capture_output=True, text=True, cwd=p)
                    if dl.returncode == 0:
                        l(f'  本地 {cur_branch} 已删除', C['sub'])
                    elif dl.stderr:
                        l(f'  (本地) {(dl.stderr or dl.stdout).strip()[:120]}',
                          C['yellow'])
                    # 远程删除
                    dr = subprocess.run(
                        ['git', 'push', 'origin', '--delete', cur_branch],
                        capture_output=True, text=True, cwd=p, timeout=20,
                        env={**os.environ, 'GIT_TERMINAL_PROMPT': '0',
                             'GCM_INTERACTIVE': 'Never'})
                    if dr.returncode == 0:
                        l(f'  远程 {cur_branch} 已删除', C['green'])
                    elif dr.stderr:
                        l(f'  (远程) {(dr.stderr or dr.stdout).strip()[:120]}',
                          C['yellow'])

                # 最终确认停留在 main
                if merge_to_main:
                    subprocess.run(['git', 'checkout', 'main'],
                                   capture_output=True, cwd=p)

            l('✔ 提交完成', C['green'])
            root.after(100, t3_history)

        threading.Thread(target=_bg, daemon=True).start()

    t3_submit_btn.config(command=lambda: t3_submit(merge_to_main=False))
    t3_merge_btn.config(command=lambda: t3_submit(merge_to_main=True))

    # ── 下载仓库 (git clone 到本地工程路径) ──
    def t3_clone_repo():
        """把选好的仓库+分支下载到 t3_path 指定的本地目录"""
        url = t3_remote.get().strip()
        local = t3_path.get().strip()
        branch = t3_branch.get().strip()
        if not url:
            messagebox.showerror('错误', '请先填写远程仓库地址')
            return
        if not local:
            messagebox.showerror('错误', '请先填写本地工程路径 (作为下载目标)')
            return
        # local 是"父目录", 仓库会克隆到 <local>/<repo_name>/
        if not os.path.isdir(local):
            try:
                os.makedirs(local, exist_ok=True)
            except Exception as e:
                messagebox.showerror('错误', f'无法创建本地目录: {e}')
                return
        _log(t3_log, '')
        _log(t3_log, f'⬇ 开始下载仓库...', C['blue'])
        _log(t3_log, f'   URL:   {url}')
        _log(t3_log, f'   本地:  {local}')
        if branch:
            _log(t3_log, f'   分支:  {branch}')

        def _bg():
            def l(msg, color=None):
                if color is None:
                    color = C['fg']
                root.after(0, lambda: _log(t3_log, msg, color))

            try:
                cmd = ['git', 'clone', url]
                if branch:
                    cmd += ['-b', branch]
                # 推到 local/<repo_name>/
                # 先解析仓库名
                import re as _re
                m = _re.search(r'/([^/]+?)(?:\.git)?/?$', url)
                repo_name = m.group(1) if m else 'repo'
                dst = os.path.join(local, repo_name)
                if os.path.exists(dst):
                    l(f'⚠ 目标已存在: {dst}', C['yellow'])
                    l('  请先删除或重命名后再下载', C['sub'])
                    root.after(100, t3_history)
                    return

                # ── 预检: 目标父目录里的压缩/工程制品可能会被 WinRAR / 资源管理器
                #  预览长期占用, git clone 时容易撞 WinError 32.
                #  仅警告, 不阻塞 (用户可能故意放在这里) ──
                _WARN_EXTS = ('.rar', '.7z', '.zip', '.tar', '.gz', '.tgz')
                _found_lockable = []
                try:
                    for _fn in os.listdir(local):
                        if _fn.lower().endswith(_WARN_EXTS):
                            _found_lockable.append(_fn)
                except OSError:
                    pass
                if _found_lockable:
                    l(f'⚠ 目标目录里有 {len(_found_lockable)} 个压缩文件, '
                      f'可能与 WinRAR 冲突:', C['yellow'])
                    for _fn in _found_lockable[:5]:
                        l(f'  · {_fn}', C['sub'])
                    if len(_found_lockable) > 5:
                        l(f'  · ... 共 {len(_found_lockable)} 个', C['sub'])
                    l('  建议: 关闭 WinRAR / 资源管理器对此目录的预览后重试',
                      C['yellow'])

                cmd.append(dst)
                l(f'  $ git clone {url}' + (f' -b {branch}' if branch else ''), C['sub'])
                rv = subprocess.run(
                    cmd, capture_output=True, text=True, timeout=300,
                    env={**os.environ, 'GIT_TERMINAL_PROMPT': '0',
                         'GCM_INTERACTIVE': 'Never'})
                if rv.returncode == 0:
                    l(rv.stdout.strip(), C['green'])
                    if rv.stderr:
                        l(rv.stderr.strip(), C['yellow'])
                    l(f'✔ 下载完成: {dst}', C['green'])
                    # 自动填充本地路径
                    root.after(0, lambda: t3_path.set(dst))
                else:
                    err = (rv.stderr or rv.stdout).strip()
                    l(f'✘ 克隆失败: {err[:300]}', C['red'])
                    if 'authentication' in err.lower() or 'credentials' in err.lower():
                        l('提示: 需要认证. 使用 SSH 地址 git@github.com:user/repo.git',
                          C['yellow'])
                    # ── WinError 32 归因: 多由 WinRAR / 资源管理器预览占用 .rar 等压缩文件 ──
                    elif ('WinError 32' in err
                          or '另一个程序正在使用' in err
                          or 'unable to access' in err.lower()
                          or 'could not create directory' in err.lower()):
                        l('提示: 失败原因可能与 WinRAR / 资源管理器预览冲突。',
                          C['yellow'])
                        l('  解决方法: 关闭 WinRAR, 然后在资源管理器里', C['sub'])
                        l(f'  右键目录 "{local}" → 属性 → 自定义 → 优化"常规项",', C['sub'])
                        l('  取消"始终显示图标,从不显示缩略图"以外的预览选项。', C['sub'])
                        l('  或者把"工程路径"改到一个没有 .rar/.7z 文件的目录。', C['sub'])
            except subprocess.TimeoutExpired:
                l('克隆超时 (5分钟)', C['red'])
            except Exception as e:
                l(f'错误: {e}', C['red'])
            root.after(100, t3_history)

        threading.Thread(target=_bg, daemon=True).start()

    t3_clone_btn.config(command=t3_clone_repo)

    def t3_history():
        t3_tree.delete(*t3_tree.get_children())
        p = t3_path.get().strip()
        if not p or not os.path.exists(os.path.join(p, '.git')):
            return
        try:
            r = subprocess.run(
                ['git', '-C', p, 'log', '--oneline',
                 '--format=%ad|%h|%s', '--date=short', '-20'],
                capture_output=True, text=True)
            for ln in r.stdout.strip().split('\n'):
                parts = ln.split('|', 2)
                if len(parts) == 3:
                    t3_tree.insert('', 'end', values=parts)
        except Exception:
            pass

    t3_refresh_btn.config(command=t3_history)

    # 日志 + 历史
    pan = ttk.Panedwindow(t3, orient='vertical')
    pan.grid(row=5, column=0, sticky='nsew', padx=12, pady=(2, 8))

    t3_lf, t3_log = _log_widget(pan, 5)
    pan.add(t3_lf, weight=1)

    hf = ttk.LabelFrame(pan, text=' 提交历史 ',
                         )
    pan.add(hf, weight=2)

    t3_tree = ttk.Treeview(hf,
                           columns=('date', 'hash', 'message'),
                           show='headings', height=5)
    t3_tree.heading('date', text='日期')
    t3_tree.heading('hash', text='哈希')
    t3_tree.heading('message', text='提交信息')
    t3_tree.column('date', width=90)
    t3_tree.column('hash', width=100)
    t3_tree.column('message', width=400)
    t3_tree.pack(side='left', fill='both', expand=True)

    tsb = tk.Scrollbar(hf, orient='vertical', bg=C['card'],
                       troughcolor=C['bg'], relief='flat')
    tsb.pack(side='right', fill='y')
    t3_tree.config(yscrollcommand=tsb.set)
    tsb.config(command=t3_tree.yview)

    # ══════════════════════════════════════
    # TAB 4 — 国产化导出
    # ══════════════════════════════════════
    from src.app_config import get_vivado_bin_dirs as _t4_viv_dirs

    t4 = ttk.Frame(nb, style='TFrame')
    nb.add(t4, text='🏭 国产化')
    t4.grid_rowconfigure(4, weight=1)
    t4.grid_columnconfigure(0, weight=1)

    # Vivado 路径 — 来自 ⚙设置 Tab (对齐 Tab2 样式)
    _t4_viv_bar = ttk.LabelFrame(t4, text=' Vivado 路径 (来自 ⚙设置 Tab) ', )
    _t4_viv_bar.grid(row=0, column=0, sticky='ew', padx=12, pady=(8, 2))
    _t4_viv_bar.grid_columnconfigure(0, weight=1)
    _t4_viv_var = tk.StringVar()
    _t4_viv_label = ttk.Label(_t4_viv_bar, textvariable=_t4_viv_var,
                               font=(F, 9), foreground=C['sub'])
    _t4_viv_label.grid(row=0, column=0, sticky='w', padx=14, pady=(8, 4))
    _t4_viv_btn = ttk.Button(_t4_viv_bar, text='⚙ 打开设置', command=lambda: nb.select(t16),
                              style='Small.TButton')

    # 从 Vivado 路径提取版本号 (延迟到 t4_ver_combo 创建后)
    def _t4_extract_versions():
        vers = []
        for d in _t4_viv_dirs():
            d = d.replace('\\', '/')
            parts = d.split('/')
            for p in parts:
                if re.match(r'^\d{4}\.\d$', p): vers.append(p); break
        return vers

    def _t4_refresh_viv():
        dirs = _t4_viv_dirs()
        if dirs:
            _t4_viv_var.set(f'✔ 已配置 {len(dirs)} 个 Vivado 路径')
            _t4_viv_label.config(foreground=C['green'])
            _t4_viv_btn.grid_remove()
        else:
            _t4_viv_var.set('✘ 未配置 Vivado 路径 — 请到 ⚙设置 Tab 添加')
            _t4_viv_label.config(foreground=C['red'])
            _t4_viv_btn.grid(row=0, column=1, padx=(0, 14), pady=(8, 4))
        if t4_ver_combo is not None:
            vers = _t4_extract_versions()
            t4_ver_combo['values'] = vers
            if vers and not t4_ver.get():
                t4_ver.set(vers[0])

    fc4 = ttk.LabelFrame(t4, text=' 导出设置 ', )
    fc4.grid(row=1, column=0, sticky='ew', padx=12, pady=(4, 4))
    fc4.grid_columnconfigure(1, weight=1)

    ttk.Label(fc4, text='工程路径:', font=(F, 10)).grid(
        row=0, column=0, sticky='w', padx=14, pady=(8, 2))
    t4_src = tk.StringVar()
    ttk.Entry(fc4, textvariable=t4_src, font=(F, 10)).grid(
        row=0, column=1, sticky='ew', padx=(4, 4), pady=(8, 2))
    ttk.Button(fc4, text='浏览',
               command=lambda: (d := filedialog.askdirectory(title='选择 FPGA 工程根目录')) and t4_src.set(d),
               style='Normal.TButton').grid(row=0, column=2, padx=(0, 14), pady=(8, 2))

    ttk.Label(fc4, text='导出路径:', font=(F, 10)).grid(
        row=1, column=0, sticky='w', padx=14, pady=(2, 2))
    t4_dst = tk.StringVar()
    ttk.Entry(fc4, textvariable=t4_dst, font=(F, 10)).grid(
        row=1, column=1, sticky='ew', padx=(4, 4), pady=(2, 2))
    ttk.Button(fc4, text='浏览',
               command=lambda: (d := filedialog.askdirectory(title='选择导出输出目录')) and t4_dst.set(d),
               style='Normal.TButton').grid(row=1, column=2, padx=(0, 14), pady=(2, 2))

    ttk.Label(fc4, text='国产平台:', font=(F, 10)).grid(
        row=2, column=0, sticky='w', padx=14, pady=(2, 2))
    t4_plat = tk.StringVar(value='复旦微')
    t4_plat_combo = ttk.Combobox(fc4, textvariable=t4_plat,
                                 values=['复旦微', '国微'], font=(F, 10),
                                 width=10, state='readonly')
    t4_plat_combo.grid(row=2, column=1, sticky='w', padx=(4, 4), pady=(2, 2))

    ttk.Label(fc4, text='Vivado 版本:', font=(F, 10)).grid(
        row=3, column=0, sticky='w', padx=14, pady=(2, 8))
    t4_ver = tk.StringVar()
    t4_ver_combo = ttk.Combobox(fc4, textvariable=t4_ver, font=(F, 10),
                                 width=10, values=[])
    t4_ver_combo.grid(row=3, column=1, sticky='w', padx=(4, 4), pady=(2, 8))
    _t4_refresh_viv()  # Combobox 创建后调用

    t4_btn = ttk.Button(t4, text='\u25b6 执行导出', style='Accent.TButton')
    t4_btn.grid(row=2, column=0, sticky='w', padx=12, pady=(2, 4))

    t4_lf, t4_log = _log_widget(t4, 12)
    t4_lf.grid(row=4, column=0, sticky='nsew', padx=12, pady=(2, 8))

    def t4_execute():
        p = t4_src.get().strip()
        if not p or not os.path.isdir(p):
            messagebox.showerror('错误', '请选择有效的工程目录')
            return

        # 递归查找 .xpr 文件
        xpr = _find_xpr_recursive(p)
        if not xpr:
            _log(t4_log, '\u2718 未找到 .xpr 工程文件 (已递归查找)', C['red'])
            return
        proj_name = os.path.splitext(os.path.basename(xpr))[0]

        out_dir = t4_dst.get().strip() or os.path.join(p, 'output_soft')
        os.makedirs(out_dir, exist_ok=True)

        _log(t4_log, f'工程: {proj_name}', C['fg'])
        _log(t4_log, f'平台: {t4_plat.get()}', C['blue'])
        _log(t4_log, f'输出: {out_dir}', C['sub'])
        _log(t4_log, '', '')

        def _bg():
            def l(msg, color=None):
                if color is None:
                    color = C['fg']
                root.after(0, lambda: _log(t4_log, msg, color))

            # 辅助: 在 bin 目录找 vivado/vitis 可执行文件
            def _find_vivado(d):
                # 只找 vivado (支持 -mode batch), vitis 不支持批量模式
                for exe in ['vivado.exe','vivado.bat','vivado'] if _sys.platform=='win32' else ['vivado']:
                    fp = os.path.join(d, exe)
                    if os.path.isfile(fp): return fp
                return None

            # 优先按版本匹配合适的 Vivado, 回退第一个 / PATH
            _vdirs = _t4_viv_dirs()
            vivado_bin = None
            ver = t4_ver.get().strip()
            if ver and _vdirs:
                for d in _vdirs:
                    if ver in d.replace('\\', '/'):
                        vivado_bin = _find_vivado(d)
                        break
            if not vivado_bin and _vdirs:
                vivado_bin = _find_vivado(_vdirs[0])
            if not vivado_bin:
                vivado_bin = 'vivado'

            # 0. 检测器件 + BD (从 .xpr XML 解析 + 文件扫描)
            import xml.etree.ElementTree as ET
            is_zynq = False
            bd_files = []
            try:
                tree = ET.parse(xpr)
                for el in tree.iter():
                    if el.attrib.get('Name') == 'Part':
                        part = el.attrib.get('Val', '')
                        is_zynq = any(k in part.lower() for k in ('xcz', 'xc7z', 'zynq'))
                        l(f'  [检测] 器件: {part}', C['blue'] if is_zynq else C['sub'])
                        if is_zynq: l('  [检测] Zynq 器件', C['blue'])
                        break
                # 扫描 BD 文件
                for dirpath, _, fnames in os.walk(p):
                    for f in fnames:
                        if f.endswith('.bd'):
                            bd_files.append(os.path.join(dirpath, f))
                if bd_files:
                    l(f'  [检测] BD: {len(bd_files)} 个 ({os.path.basename(bd_files[0])})', C['blue'])
                else:
                    l('  [检测] 无 Block Design', C['sub'])
            except Exception as e:
                l(f'  [检测失败] {str(e)[:80]}', C['red'])

            # 1. 导出 BD 文件 (.bd) — 保留原始文件名
            if not bd_files:
                l('  [跳过] 无 BD', C['sub'])
            else:
                bd_name = os.path.basename(bd_files[0])
                bd_out = os.path.join(out_dir, bd_name)
                shutil.copy2(bd_files[0], bd_out)
                l(f'  \u2714 BD: {bd_name} ({_fmt_size(os.path.getsize(bd_out))})', C['green'])

            # 2. 导出 HDF/XSA (仅 BD 工程需要 — 先搜 .sdk 目录)
            bd_base = os.path.splitext(os.path.basename(bd_files[0]))[0] if bd_files else 'design_1'
            xsa = os.path.join(out_dir, f'{bd_base}_wrapper.xsa')
            hdf = os.path.join(out_dir, f'{bd_base}_wrapper.hdf')
            if not bd_files:
                l('  [跳过] 无 BD, 不导出 HDF/XSA', C['sub'])
                hdf = None; xsa = None
            else:
                # 先在 .sdk 目录搜已有 HDF/XSA
                sdk_dir = os.path.join(p, f'{proj_name}.sdk')
                found_hdf = None
                if os.path.isdir(sdk_dir):
                    for root_d, _, files in os.walk(sdk_dir):
                        for f in files:
                            if f.lower().endswith(('.hdf', '.xsa')):
                                found_hdf = os.path.join(root_d, f); break
                        if found_hdf: break
                if found_hdf:
                    shutil.copy2(found_hdf, os.path.join(out_dir, os.path.basename(found_hdf)))
                    l(f'  \u2714 HDF/XSA: {os.path.basename(found_hdf)} ({_fmt_size(os.path.getsize(found_hdf))})', C['green'])
                    hdf = None; xsa = None  # 已复制, 不再生成
                else:
                    l('  [创建] 正在导出 HDF/XSA ...', C['blue'])
                    # 版本判断: ≤2019 → write_hwdef(.hdf), ≥2020 → XSA+HDF
                    viv_ver = ver or ''
                    try: viv_major = int(re.match(r'(\d{4})', viv_ver).group(1))
                    except: viv_major = 0

                    # ╺╺╺ 备份工程到临时目录, 避免 Vivado 污染源工程 ╺╺╺
                    import shutil as _shutil
                    tmp_proj = os.path.join(out_dir, '__tmp_proj')
                    if os.path.isdir(tmp_proj):
                        _shutil.rmtree(tmp_proj, ignore_errors=True)
                    l('  [备份] 复制工程到临时目录 ...', C['sub'])
                    try:
                        _shutil.copytree(p, tmp_proj)
                    except Exception as e:
                        l(f'  [备份失败] {e}', C['red']); tmp_proj = p  # 回退直接用原目录

                    # 重新定位 xpr/bd 到备份目录
                    proj_base = os.path.basename(p.rstrip(os.sep))
                    tmp_xpr = os.path.join(tmp_proj, os.path.relpath(xpr, p)).replace('\\', '/')
                    bd_path = bd_files[0].replace('\\', '/')
                    tmp_bd = os.path.join(tmp_proj, os.path.relpath(bd_path, p)).replace('\\', '/')

                    tmp_tcl = os.path.join(out_dir, '_t4_hdf.tcl')
                    if viv_major >= 2020:
                        # 新版: open_project + upgrade_ip + generate_target + XSA + HDF
                        with open(tmp_tcl, 'w', encoding='utf-8') as f:
                            f.write(f'open_project {{{tmp_xpr}}}\n'
                                    f'open_bd_design {{{tmp_bd}}}\n'
                                    f'upgrade_ip [get_ips]\n'
                                    f'generate_target all [get_files {os.path.basename(tmp_bd)}]\n'
                                    f'write_hw_platform -fixed -force -file {{{xsa}}}\n'
                                    f'write_hwdef -force -file {{{hdf}}}\n'
                                    f'close_project\n')
                    else:
                        # 旧版(≤2019): 清理 xpr 中不兼容选项 → open_project → write_hwdef → .hdf
                        import re as _re
                        with open(tmp_xpr, 'r', encoding='utf-8') as f:
                            content = f.read()
                        content = _re.sub(r'<Option Name="Simulator\w+".*?/>\n?', '', content)
                        content = _re.sub(r'<Option Name="DAElab\w*Attr".*?/>\n?', '', content)
                        with open(tmp_xpr, 'w', encoding='utf-8') as f:
                            f.write(content)
                        with open(tmp_tcl, 'w', encoding='utf-8') as f:
                            f.write(f'open_project {{{tmp_xpr}}}\n'
                                    f'open_bd_design {{{tmp_bd}}}\n'
                                    f'write_hwdef -force -file {{{hdf}}}\n'
                                    f'close_project\n')
                    r = subprocess.run([vivado_bin, '-mode', 'batch', '-source', tmp_tcl, '-notrace'],
                                       capture_output=True, text=True, timeout=300, cwd=tmp_proj)
                    for line in (r.stdout + r.stderr).split('\n'):
                        if 'failed' in line.lower() or 'error' in line.lower():
                            l(f'  Vivado: {line.strip()[:120]}', C['red'])

                    # ╺╺╺ 清理临时文件 ╺╺╺
                    try: os.remove(tmp_tcl)
                    except: pass
                    if tmp_proj != p:
                        _shutil.rmtree(tmp_proj, ignore_errors=True)

                    if os.path.exists(xsa) and os.path.getsize(xsa) > 0:
                        l(f'  \u2714 XSA: {_fmt_size(os.path.getsize(xsa))}', C['green'])
                    elif os.path.exists(hdf) and os.path.getsize(hdf) > 0:
                        l(f'  \u2714 HDF: {_fmt_size(os.path.getsize(hdf))}', C['green'])
                    else:
                        l('  \u26a0 HDF/XSA 未生成', C['yellow'])

            # 3. BIT — 保留原始文件名
            found_bit = False
            runs_dir = os.path.join(p, f'{proj_name}.runs')
            if os.path.isdir(runs_dir):
                for root_d, _, files in os.walk(runs_dir):
                    for f in files:
                        if f.endswith('.bit'):
                            bit = os.path.join(out_dir, f)
                            shutil.copy2(os.path.join(root_d, f), bit)
                            l(f'  \u2714 BIT: {f} ({_fmt_size(os.path.getsize(bit))})', C['green'])
                            found_bit = True; break
                    if found_bit: break
            if not found_bit:
                l('  \u26a0 BIT 未找到 (需先完成实现)', C['yellow'])

            # 4. Zynq: 导出 PS IP XCI — 保留原始文件名
            if is_zynq:
                ps_xci = None
                for dirpath, _, fnames in os.walk(p):
                    for f in fnames:
                        if f.endswith('.xci') and 'processing_system' in f.lower():
                            ps_xci = os.path.join(dirpath, f); break
                    if ps_xci: break
                if ps_xci:
                    xci_name = os.path.basename(ps_xci)
                    shutil.copy2(ps_xci, os.path.join(out_dir, xci_name))
                    l(f'  \u2714 Zynq XCI: {xci_name} ({_fmt_size(os.path.getsize(os.path.join(out_dir, xci_name)))})', C['green'])
                else:
                    l('  \u26a0 Zynq XCI 未找到', C['yellow'])

            l('')
            l(f'导出完成 → {out_dir}', C['green'])
            exported = [f for f in os.listdir(out_dir) if not f.startswith('_')]
            l(f'文件 ({len(exported)}): ' + ', '.join(exported), C['sub'])

        threading.Thread(target=_bg, daemon=True).start()

    t4_btn.config(command=t4_execute)

    # ══════════════════════════════════════
    # TAB 5 — 代码规范整理
    # ══════════════════════════════════════
    from gen_inst import format_fpga_code

    t5 = ttk.Frame(nb, style='TFrame')
    nb.add(t5, text='🧹 代码整理')
    # ── 布局: row 0 (fc5) + row 3 (日志) 都设 weight=1, 平分 t5 高度, 保证日志框不被挤掉
    t5.grid_columnconfigure(0, weight=1)
    t5.grid_rowconfigure(0, weight=1)   # 整理设置框占 1 份
    t5.grid_rowconfigure(3, weight=40)  # 日志框占 40 份 (1:40 比例)

    fc5 = ttk.LabelFrame(t5, text=' 整理设置 ', )
    fc5.grid(row=0, column=0, sticky='nsew', padx=12, pady=(10, 4))
    fc5.grid_columnconfigure(1, weight=1)
    fc5.grid_rowconfigure(1, weight=1)   # fc5 内 row=1 (路径列表) 撑开吃空白

    t5_src = tk.StringVar()
    ttk.Entry(fc5, textvariable=t5_src, font=(F, 10)).grid(
        row=0, column=1, sticky='ew', padx=(4, 4), pady=(8, 2))
    ttk.Button(fc5, text='浏览', style='Normal.TButton',
               command=lambda: (d := filedialog.askdirectory(title='选择目录')) and t5_add_path(d)
               ).grid(row=0, column=2, padx=(0, 4), pady=(8, 2))
    ttk.Button(fc5, text='➕ 添加文件', style='Normal.TButton',
               command=lambda: t5_add_files()
               ).grid(row=0, column=3, padx=(0, 14), pady=(8, 2))

    # ── 路径列表 ──
    t5_paths_frame = ttk.Frame(fc5, style='TFrame')
    t5_paths_frame.grid(row=1, column=0, columnspan=4, sticky='nsew',
                        padx=14, pady=(2, 4))
    t5_paths_frame.grid_columnconfigure(0, weight=1)

    t5_paths_list = tk.Listbox(t5_paths_frame, font=(F, 9), height=8,
                               relief='flat', borderwidth=1,
                               bg=C['bg'], fg=C['fg'],
                               selectbackground=C['blue'], selectforeground='white',
                               activestyle='none')
    t5_paths_list.pack(side='left', fill='both', expand=True)
    t5_paths_sb = tk.Scrollbar(t5_paths_frame, orient='vertical',
                                bg='#8a96a3', troughcolor=C['bg'],
                                activebackground='#5a6573',
                                highlightthickness=0, relief='flat',
                                width=16, borderwidth=1)
    t5_paths_sb.pack(side='right', fill='y')
    t5_paths_list.config(yscrollcommand=t5_paths_sb.set)
    t5_paths_sb.config(command=t5_paths_list.yview)

    t5_paths_btn_frame = ttk.Frame(fc5, style='TFrame')
    t5_paths_btn_frame.grid(row=1, column=4, sticky='ns', padx=(4, 14), pady=(2, 4))

    def t5_add_path(d):
        if not d or not os.path.isdir(d):
            return
        d = os.path.abspath(d)
        # 去重
        for i in range(t5_paths_list.size()):
            if t5_paths_list.get(i) == d:
                return
        t5_paths_list.insert('end', d)

    def t5_add_files():
        files = filedialog.askopenfilenames(
            title='选择 HDL 源文件',
            filetypes=[('HDL files', '*.v *.sv *.vhd *.vhdl'),
                       ('All files', '*.*')])
        for f in files:
            f = os.path.abspath(f)
            # 去重
            for i in range(t5_paths_list.size()):
                if t5_paths_list.get(i) == f:
                    break
            else:
                t5_paths_list.insert('end', f)

    def t5_del_path():
        sel = t5_paths_list.curselection()
        if sel:
            for i in reversed(sel):
                t5_paths_list.delete(i)

    ttk.Button(t5_paths_btn_frame, text='删除', command=t5_del_path,
               style='Normal.TButton').pack(pady=(0, 4))

    ttk.Label(fc5, text='可混合添加目录(递归扫 .v/.sv/.vhd) 和单独文件  添加注释头 / 统一缩进 / 对齐声明',
              foreground=C['sub'], font=(F, 8)).grid(
        row=2, column=0, columnspan=5, sticky='w', padx=14, pady=(0, 8))

    t5_btn_row = ttk.Frame(t5)
    t5_btn_row.grid(row=1, column=0, sticky='w', padx=12, pady=(2, 4))
    t5_btn = ttk.Button(t5_btn_row, text='▶ 执行整理', style='Accent.TButton')
    t5_btn.pack(side='left', padx=(0, 6))

    t5_lf, t5_log = _log_widget(t5, 6)
    t5_lf.grid(row=3, column=0, sticky='nsew', padx=12, pady=(2, 8))

    def _t5_collect_hdl_files():
        """收集列表中所有目录和文件的 HDL 文件, 返回 (file_paths, base_dir)"""
        results = []
        seen = set()
        default_base = None

        for i in range(t5_paths_list.size()):
            p = t5_paths_list.get(i)
            if not os.path.exists(p):
                continue
            if default_base is None:
                default_base = os.path.dirname(p) if os.path.isfile(p) else p

            if os.path.isdir(p):
                for root_d, dirs, files in os.walk(p):
                    dirs[:] = [d for d in dirs
                               if d not in ('.git', '.runs', '.srcs',
                                            '.sim', '.cache', '.gen', '.hw')]
                    for f in files:
                        if f.endswith(('.v', '.sv', '.vhd', '.vhdl')):
                            fp = os.path.join(root_d, f)
                            if fp not in seen:
                                seen.add(fp)
                                results.append(fp)
            elif os.path.isfile(p) and p.endswith(('.v', '.sv', '.vhd', '.vhdl')):
                if p not in seen:
                    seen.add(p)
                    results.append(p)
        return results, default_base or ''

    def t5_execute():
        hdl_files, base = _t5_collect_hdl_files()
        if not hdl_files:
            if t5_paths_list.size() == 0:
                messagebox.showerror('错误', '请先添加目录或 HDL 文件')
            else:
                _log(t5_log, '  未找到 HDL 文件', C['yellow'])
            return

        _log(t5_log, f'扫描: {len(hdl_files)} 个 HDL 文件', C['fg'])
        _log(t5_log, '', '')

        def _bg():
            changed = 0
            total = len(hdl_files)
            for fp in hdl_files:
                try:
                    rel = os.path.relpath(fp, base)
                except ValueError:
                    rel = os.path.basename(fp)  # 不同盘符
                if len(rel) > 60:
                    rel = '...' + rel[-57:]
                try:
                    old_sz, new_sz, c, _ = format_fpga_code(fp, dry_run=False)
                    if c:
                        _log(t5_log, f'  \u2714 {rel}', C['green'])
                        changed += 1
                    else:
                        _log(t5_log, f'  - {rel} (无变化)', C['sub'])
                except Exception as e:
                    _log(t5_log, f'  \u2718 {rel}: {e}', C['red'])

            _log(t5_log, '', '')
            _log(t5_log, f'\u2714 整理完成! {changed}/{total} 个文件已格式化', C['green'])

        threading.Thread(target=_bg, daemon=True).start()

    t5_btn.config(command=t5_execute)

    # --- 文件列表生成 ---
    t5_list_btn = ttk.Button(t5_btn_row, text='📋 生成文件列表 (.f)',
                             style='Normal.TButton')
    t5_list_btn.pack(side='left', padx=6)

    # --- 仿真产物清理 ---
    _SIM_TRASH_DIRS  = ['work', 'xsim.dir', 'obj_dir', 'sim_build',
                        'INCA_libs', 'csrc']
    _SIM_TRASH_FILES = ['transcript', 'vsim.wlf', '*.vcd', '*.vpd',
                        '*.wlf', '*.vvp', '*.lxt', '*.lxt2', '*.fst',
                        '*.ghw', '*.cf', '*.backup.jou', '*.backup.log',
                        'webtalk*', '*.stacktrace', '*.pb', '*.jou']

    t5_clean_btn = ttk.Button(t5_btn_row, text='🧹 清理仿真产物',
                              style='Normal.TButton')
    t5_clean_btn.pack(side='left', padx=6)

    def t5_gen_filelist():
        hdl_files, base = _t5_collect_hdl_files()
        if not hdl_files:
            messagebox.showerror('错误', '未找到 HDL 文件')
            return

        _log(t5_log, f'📋 收集到 {len(hdl_files)} 个 HDL 文件', C['blue'])

        # 生成 .f 文件列表 (放在第一个目录下)
        out_dir = base if os.path.isdir(base) else os.path.dirname(base)
        flist_path = os.path.join(out_dir, 'filelist.f')
        with open(flist_path, 'w', encoding='utf-8') as f:
            for fp in sorted(hdl_files):
                try:
                    rel = os.path.relpath(fp, out_dir).replace('\\', '/')
                except ValueError:
                    rel = os.path.basename(fp)
                f.write(f'{rel}\n')

        _log(t5_log, f'  已保存: {flist_path}', C['green'])
        _log(t5_log, f'  已保存: {flist_path}', C['green'])

        # 预览前 20 行
        for fp in sorted(hdl_files)[:20]:
            try:
                rel = os.path.relpath(fp, out_dir)
            except ValueError:
                rel = os.path.basename(fp)
            _log(t5_log, f'    {rel}', C['sub'])
        if len(hdl_files) > 20:
            _log(t5_log, f'    ... 还有 {len(hdl_files) - 20} 个文件', C['sub'])

    def t5_clean_sim():
        # 收集所有目录项（跳过单独文件）
        dirs_to_clean = []
        for i in range(t5_paths_list.size()):
            p = t5_paths_list.get(i)
            if os.path.isdir(p):
                dirs_to_clean.append(p)
        if not dirs_to_clean:
            messagebox.showerror('错误', '列表中没有目录, 无法清理仿真产物')
            return
        if not messagebox.askyesno('确认清理',
            f'将递归清理 {len(dirs_to_clean)} 个目录下的仿真产物。\n\n'
            '包括: ModelSim/Questa, Vivado XSim, Verilator,\n'
            '      GHDL, Icarus Verilog 的临时文件。\n\n'
            '确认继续?'):
            return

        deleted = []
        for p in dirs_to_clean:
            for root_d, dirs, files in os.walk(p, topdown=True):
                dirs[:] = [d for d in dirs if d != '.git']
                for d in dirs[:]:
                    if d in _SIM_TRASH_DIRS or d.startswith('work.'):
                        full = os.path.join(root_d, d)
                        try:
                            import shutil as _sh5
                            _sh5.rmtree(full, ignore_errors=True)
                            deleted.append(('dir', os.path.relpath(full, p)))
                            dirs.remove(d)
                        except OSError:
                            pass
                import fnmatch as _fn5
                for f in files:
                    for pat in _SIM_TRASH_FILES:
                        if _fn5.fnmatch(f, pat):
                            full = os.path.join(root_d, f)
                            try:
                                os.unlink(full)
                                deleted.append(('file', os.path.relpath(full, p)))
                            except OSError:
                                pass
                            break

        d_dirs  = sum(1 for t, _ in deleted if t == 'dir')
        d_files = sum(1 for t, _ in deleted if t == 'file')
        _log(t5_log, '', '')
        _log(t5_log, f'🧹 清理完成: {d_dirs} 个目录 + {d_files} 个文件', C['green'])
        for t, path in deleted[:30]:
            _log(t5_log, f'  {"[目录]" if t=="dir" else "[文件]"} {path}', C['sub'])
        if len(deleted) > 30:
            _log(t5_log, f'  ... 还有 {len(deleted) - 30} 个', C['sub'])

    t5_list_btn.config(command=t5_gen_filelist)
    t5_clean_btn.config(command=t5_clean_sim)

    # ══════════════════════════════════════
    # TAB 6 — IP 文档查寻 (离线 PDF, 搜 ip_docs/ 文件夹)
    # ══════════════════════════════════════
    from urllib.parse import quote_plus

    _DOC_DIR = os.path.join(_PROJECT_ROOT, 'ip_docs')

    def _scan_pdfs(root_dir):
        docs = []
        if not root_dir or not os.path.isdir(root_dir):
            return docs
        for dirpath, _d, filenames in os.walk(root_dir):
            for f in filenames:
                if f.lower().endswith('.pdf'):
                    fp = os.path.join(dirpath, f)
                    rel = os.path.relpath(dirpath, root_dir)
                    docs.append({'name': f, 'path': fp, 'dir': rel, 'ext': '.pdf'})
        return sorted(docs, key=lambda x: x['name'].lower())

    # --- UI ---
    t6 = ttk.Frame(nb, style='TFrame')
    nb.add(t6, text='📚 IP文档')
    t6.grid_rowconfigure(1, weight=0)   # 搜索栏(固定)
    t6.grid_rowconfigure(2, weight=0)   # DocNav状态条(固定)
    t6.grid_rowconfigure(3, weight=1)   # PanedWindow(可拖拽分隔)
    t6.grid_columnconfigure(0, weight=1)

    # DocNav 路径 — 来自 ⚙设置 Tab 全局配置 (对齐 Tab2 Vivado 样式)
    from src.app_config import get_xdocs_xml_path as _app_xdocs
    from src.amd_doc_downloader import find_xdocs_xml

    t6_dn_frame = ttk.LabelFrame(t6, text=' DocNav 路径 (来自 ⚙设置 Tab) ', )
    t6_dn_frame.grid(row=2, column=0, sticky='ew', padx=12, pady=(4, 4))
    t6_dn_frame.grid_columnconfigure(0, weight=1)

    _t6_dn_var = tk.StringVar()
    _t6_dn_label = ttk.Label(t6_dn_frame, textvariable=_t6_dn_var,
                              font=(F, 9), foreground=C['sub'])
    _t6_dn_label.grid(row=0, column=0, sticky='w', padx=14, pady=(8, 4))

    _t6_dn_btn = ttk.Button(t6_dn_frame, text='⚙ 打开设置',
                             command=lambda: nb.select(t16), style='Small.TButton')

    def _t6_refresh_dn():
        """刷新 DocNav 路径状态 (设置Tab变更时感知)"""
        from src.app_config import get_valid_docnav_dirs as _t6_dn_dirs
        dirs = _t6_dn_dirs()
        if dirs:
            _t6_dn_var.set(f'✔ 已配置 {len(dirs)} 个 DocNav 路径')
            _t6_dn_label.config(foreground=C['green'])
            _t6_dn_btn.grid_remove()
        else:
            _t6_dn_var.set('✘ 未配置 DocNav 路径 — 请到 ⚙设置 Tab 添加')
            _t6_dn_label.config(foreground=C['red'])
            _t6_dn_btn.grid(row=0, column=1, padx=(0, 14), pady=(8, 4))

    _t6_refresh_dn()

    # PanedWindow: 上=离线搜索 / 下=DocNav下载
    t6_pane = ttk.PanedWindow(t6, orient='vertical')
    t6_pane.grid(row=3, column=0, sticky='nsew', padx=8, pady=(2, 4))

    # 上半: 离线搜索区
    t6_offline = ttk.Frame(t6_pane, style='TFrame')
    t6_offline.grid_rowconfigure(0, weight=1)  # tree
    t6_offline.grid_rowconfigure(1, weight=0)  # status
    t6_offline.grid_columnconfigure(0, weight=1)
    t6_offline.grid_columnconfigure(1, weight=0)  # scrollbar
    t6_pane.add(t6_offline, weight=1)

    # 下半: DocNav 下载区 (延后添加, 等 t6_dl_frame 创建)
    # t6_pane.add(t6_dl_frame, weight=1)  -- 在创建 t6_dl_frame 后执行

    # 提示条
    fc6 = ttk.LabelFrame(t6, text=' 搜索 IP 文档 (PDF) ', )
    fc6.grid(row=0, column=0, sticky='ew', padx=12, pady=(10, 4))
    fc6.grid_columnconfigure(0, weight=1)

    ttk.Label(fc6, text=f'📂 文档文件夹: {_DOC_DIR}    将下载的 IP PDF 丢进去即可',
              foreground=C['sub'], font=(F, 8)).grid(
        row=0, column=0, sticky='w', padx=14, pady=(8, 2))
    ttk.Button(fc6, text='📂 打开文件夹', command=lambda: os.startfile(_DOC_DIR) if _sys.platform == 'win32' else None,
               style='Small.TButton').grid(row=0, column=1, padx=(0, 14), pady=(8, 2))

    # 搜索行
    fc6s = ttk.LabelFrame(t6, text=' 关键词 ', )
    fc6s.grid(row=1, column=0, sticky='ew', padx=12, pady=(4, 4))
    fc6s.grid_columnconfigure(0, weight=1)

    t6_search = tk.StringVar()
    t6_search_entry = ttk.Entry(fc6s, textvariable=t6_search, font=(M, 12))
    t6_search_entry.grid(row=0, column=0, sticky='ew', padx=14, pady=(8, 8))

    def _t6_do_search():
        t6_tree.delete(*t6_tree.get_children())
        kw = t6_search.get().strip().lower()
        if not os.path.isdir(_DOC_DIR):
            t6_status.set('📂 ip_docs 文件夹不存在, 请先放入 PDF 文件')
            return
        # 每次搜索都重新扫描, 确保删除/新增文件即时生效
        t6_status.set('⏳ 正在扫描 ip_docs ...')
        def _bg():
            nonlocal t6_all_docs
            t6_all_docs = _scan_pdfs(_DOC_DIR)
            root.after(0, lambda: _t6_do_filter())
        threading.Thread(target=_bg, daemon=True).start()

    def _t6_do_filter():
        kw = t6_search.get().strip().lower()
        t6_tree.delete(*t6_tree.get_children())
        if not kw:
            filtered = t6_all_docs
        else:
            kws = kw.split()
            # 用 IP 映射表反向扩展: "can" → ["PG096","UG765","PG223","PG050"]
            expanded = set()
            for k in kws:
                for key, ids in _IP_DOC_MAP.items():
                    if k in key or key in k:
                        expanded.update([id.lower() for id in ids])
            title_map = _t6_load_title_map()
            filtered = [d for d in t6_all_docs
                        if all(k in d['name'].lower() or k in d['dir'].lower()
                               or k in title_map.get(os.path.splitext(d['name'])[0], '').lower()
                               or os.path.splitext(d['name'])[0].lower() in expanded
                               for k in kws)]
        for d in filtered:
            name, title = _t6_display_name(d)
            t6_tree.insert('', 'end', values=(name, title), iid=d['path'])
        t6_status.set(f'找到 {len(filtered)} / {len(t6_all_docs)} 个 PDF')

    t6_search_entry.bind('<Return>', lambda e: _t6_do_search())
    ttk.Button(fc6s, text='🔍  搜索', command=_t6_do_search,
               style='Accent.TButton').grid(row=0, column=1,
               padx=(4, 2), pady=(8, 8))
    ttk.Button(fc6s, text='刷新', command=_t6_do_search,
               style='Small.TButton').grid(row=0, column=2,
               padx=(0, 2), pady=(8, 8))
    ttk.Button(fc6s, text='删除', command=lambda: _t6_delete_pdf(),
               style='Normal.TButton').grid(row=0, column=3,
               padx=(0, 14), pady=(8, 8))

    # 结果列表 (在 PanedWindow 上半 pane 中)
    t6_tree = ttk.Treeview(t6_offline, columns=('name', 'title'), show='headings',
                           selectmode='extended', style='Treeview')
    t6_tree.heading('name', text='文档')
    t6_tree.heading('title', text='标题 / 路径')
    t6_tree.column('name', width=110, minwidth=80)
    t6_tree.column('title', width=650, minwidth=350)
    t6_tree.grid(row=0, column=0, sticky='nsew', padx=(6, 0), pady=(2, 0))

    t6_scroll = ttk.Scrollbar(t6_offline, orient='vertical', command=t6_tree.yview)
    t6_scroll.grid(row=0, column=1, sticky='ns', pady=(2, 0))
    t6_tree.configure(yscrollcommand=t6_scroll.set)

    def _t6_open(event=None):
        sel = t6_tree.selection()
        if not sel or not os.path.isfile(sel[0]):
            return
        try:
            os.startfile(sel[0]) if _sys.platform == 'win32' else subprocess.Popen(['xdg-open', sel[0]])
        except Exception as e:
            messagebox.showerror('打开失败', str(e))
    t6_tree.bind('<Double-Button-1>', _t6_open)

    def _t6_delete_pdf():
        sel = t6_tree.selection()
        if not sel:
            messagebox.showinfo('提示', '请先在列表中选中要删除的 PDF (支持Ctrl多选)')
            return
        paths = [s for s in sel if os.path.isfile(s)]
        if not paths:
            return
        names = [os.path.basename(p) for p in paths]
        if not messagebox.askyesno('确认删除', f'确定要删除 {len(paths)} 个文件吗？\n\n' + '\n'.join(names) + '\n\n此操作不可撤销'):
            return
        deleted = 0
        for path in paths:
            try:
                os.remove(path); deleted += 1
            except OSError:
                pass
        t6_status.set(f'✔ 已删除 {deleted}/{len(paths)} 个文件')
        _t6_do_search()  # 刷新列表

    t6_status = tk.StringVar(value='就绪 - 在 ip_docs 文件夹放入 PDF, 输入关键词搜索')
    ttk.Label(t6_offline, textvariable=t6_status, foreground=C['sub'],
              font=(F, 9)).grid(row=1, column=0, sticky='w', padx=10, pady=(0, 4))

    t6_all_docs = _scan_pdfs(_DOC_DIR)
    if t6_all_docs:
        t6_status.set(f'已索引 {len(t6_all_docs)} 个 PDF  ·  输入关键词搜索')

    # ══════════════════════════════════════
    # DocNav 搜索下载 (解析 xdocs.xml → 搜索 → 直链下载 PDF 到 ip_docs/)
    # ══════════════════════════════════════
    from src.amd_doc_downloader import (
        find_xdocs_xml, parse_xdocs, search_docs, batch_download,
    )
    from src.app_config import get_xdocs_xml_path as _app_xdocs

    # ====== 常用 IP → 文档 ID 映射表 (后台智能搜索) ======
    _IP_DOC_MAP = {
        # CAN
        'can': ['PG096', 'UG765', 'PG223', 'PG050'],
        'canfd': ['PG223', 'PG050'],
        # Ethernet
        'ethernet': ['PG051', 'UG470', 'PG082', 'PG138'],
        'tri mode ethernet': ['PG051'],
        'axi ethernet': ['PG138'],
        'gige': ['PG051', 'UG470'],
        # PCIe
        'pcie': ['PG054', 'PG213', 'PG023'],
        'axi pcie': ['PG054'],
        # DMA
        'dma': ['PG021', 'PG085', 'PG195'],
        'axi dma': ['PG021'],
        'vdma': ['PG085'],
        'cdma': ['PG195'],
        # Memory/DDR
        'ddr': ['PG150', 'UG586'],
        'mig': ['PG150'],
        'axi bram': ['PG058'],
        'block memory': ['PG058'],
        # FIFO
        'fifo': ['PG057', 'PG073'],
        'fifo generator': ['PG057'],
        'axi fifo': ['PG073'],
        # UART
        'uart': ['PG142'],
        'axi uart': ['PG142'],
        'uartlite': ['PG142'],
        # I2C
        'i2c': ['PG090'],
        'axi i2c': ['PG090'],
        # SPI
        'spi': ['PG153'],
        'axi spi': ['PG153'],
        'qspi': ['PG153'],
        # GPIO
        'gpio': ['PG144'],
        'axi gpio': ['PG144'],
        # Timer
        'timer': ['PG086', 'PG128'],
        'axi timer': ['PG086'],
        'watchdog': ['PG128'],
        # Interrupt
        'interrupt': ['PG099'],
        'axi intc': ['PG099'],
        # Video
        'video': ['PG103', 'PG231', 'PG236'],
        'vtc': ['PG103'],
        'hdmi': ['PG236'],
        'displayport': ['PG233'],
        # Transceiver
        'gtx': ['UG476'],
        'gth': ['UG476'],
        'gtp': ['UG482'],
        'transceiver': ['UG476', 'UG482'],
        'aurora': ['PG074', 'PG221'],
        # Zynq
        'zynq': ['UG585', 'UG821'],
        'zynq7000': ['UG585'],
        'zynq ultra': ['UG1085'],
        'ps': ['UG585'],
        # Clock
        'clock': ['PG065'],
        'clocking wizard': ['PG065'],
        'pll': ['PG065'],
        'mmcm': ['PG065'],
        # AXI
        'axi': ['PG059', 'PG087', 'PG247'],
        'axi interconnect': ['PG059'],
        'axi crossbar': ['PG059'],
        'axi smartconnect': ['PG247'],
        # Crypto/Math
        'dsp': ['PG151'],
        'fir': ['PG149'],
        'fftw': ['PG109'],
        'cordic': ['PG105'],
        # Misc
        'jtag': ['PG153'],
        'mdio': ['PG047'],
        'xadc': ['UG480'],
        'sysmon': ['UG480'],
        'icap': ['PG134'],
        'sem': ['PG036'],
        # Vivado/Zynq基础
        'trm': ['UG585'],
        'ug585': ['UG585'],
        'ug476': ['UG476'],
        'ug470': ['UG470'],
    }

    def _t6_expand_keyword(kw):
        """关键词智能扩展: 映射表匹配 + 子串匹配"""
        kw_lower = kw.lower().strip()
        expanded = set()
        expanded.add(kw)  # 保留原词
        for key, doc_ids in _IP_DOC_MAP.items():
            if kw_lower == key or kw_lower in key or key in kw_lower:
                expanded.update(doc_ids)
        return list(expanded)

    # ── 搜索 + 下载行 (PanedWindow 下半 pane) ──
    t6_dl_frame = ttk.LabelFrame(t6_pane, text=' 🔍 搜索 IP 文档 (DocNav) → 下载到 ip_docs/docnav/ ')
    t6_pane.add(t6_dl_frame, weight=1)
    t6_dl_frame.grid_rowconfigure(1, weight=1)  # Treeview 行可拉伸
    t6_dl_frame.grid_columnconfigure(0, weight=1)
    t6_dl_frame.grid_columnconfigure(1, weight=1)

    t6_dl_kw = tk.StringVar()
    t6_dl_entry = ttk.Entry(t6_dl_frame, textvariable=t6_dl_kw, font=(M, 12))
    t6_dl_entry.grid(row=0, column=0, sticky='ew', padx=14, pady=(8, 4))
    t6_dl_entry.bind('<Return>', lambda e: _t6_do_docnav_search())

    t6_dl_btn_frame = ttk.Frame(t6_dl_frame, style='TFrame')
    t6_dl_btn_frame.grid(row=0, column=1, sticky='ew', padx=(0, 14), pady=(8, 4))

    ttk.Button(t6_dl_btn_frame, text='🔍 搜索 DocNav',
               command=lambda: _t6_do_docnav_search(),
               style='Accent.TButton').pack(side='left', padx=(0, 8))
    ttk.Button(t6_dl_btn_frame, text='📥 下载选中',
               command=lambda: _t6_download_selected(),
               style='Success.TButton').pack(side='left', padx=(0, 8))
    ttk.Button(t6_dl_btn_frame, text='📥 下载全部',
               command=lambda: _t6_download_all(),
               style='Normal.TButton').pack(side='left')

    # ── 搜索结果列表 (docID + title) ──
    t6_dl_tree = ttk.Treeview(t6_dl_frame, columns=('docid', 'title'),
                              show='headings', selectmode='extended',
                              style='Treeview')
    t6_dl_tree.heading('docid', text='Doc ID')
    t6_dl_tree.heading('title', text='文档标题')
    t6_dl_tree.column('docid', width=120, minwidth=80)
    t6_dl_tree.column('title', width=520, minwidth=200)
    t6_dl_tree.grid(row=1, column=0, columnspan=2, sticky='nsew',
                    padx=14, pady=(4, 4))

    t6_dl_scroll = ttk.Scrollbar(t6_dl_frame, orient='vertical',
                                 command=t6_dl_tree.yview)
    t6_dl_scroll.grid(row=1, column=2, sticky='ns', pady=(4, 4))
    t6_dl_tree.configure(yscrollcommand=t6_dl_scroll.set)

    # ── 下载状态 ──
    t6_dl_progress = ttk.Progressbar(t6_dl_frame, mode='determinate', length=400)
    t6_dl_progress.grid(row=2, column=0, columnspan=2, sticky='ew',
                        padx=14, pady=(4, 2))
    t6_dl_progress.grid_remove()  # 默认隐藏

    t6_dl_status = tk.StringVar(value='输入关键词搜索 DocNav 数据库')
    ttk.Label(t6_dl_frame, textvariable=t6_dl_status,
              foreground=C['sub'], font=(F, 9)).grid(
        row=3, column=0, columnspan=2, sticky='w', padx=16, pady=(2, 8))

    # ── 内部状态 ──
    t6_dl_results = []  # 当前搜索结果 [{docID, title, downloadURL}, ...]

    def _t6_do_docnav_search():
        """智能搜索: 关键词 → IP映射表扩展 → DocNav 数据库"""
        kw = t6_dl_kw.get().strip()
        t6_dl_tree.delete(*t6_dl_tree.get_children())
        t6_dl_results.clear()

        if not kw:
            t6_dl_status.set('请输入 IP 名称，如 CAN / Ethernet / PCIe / AXI / DMA ...')
            return

        # 优先从 app_config 读取 DocNav 路径
        xml_path = _app_xdocs() or find_xdocs_xml()[0]
        if not xml_path:
            t6_dl_status.set('❌ 未找到 DocNav 数据库, 请到 ⚙设置 Tab 添加 DocNav 路径')
            messagebox.showwarning('未找到 DocNav',
                '请先安装 AMD/Xilinx DocNav 工具,\n'
                '或到 ⚙设置 Tab 添加 DocNav 安装目录\n'
                '(包含 resources/xdocs.xml 的文件夹)')
            return

        # 智能关键词扩展
        expanded_kws = _t6_expand_keyword(kw)
        t6_dl_status.set(f'⏳ 正在搜索 "{kw}" (扩展: {",".join(expanded_kws[:6])}) ...')

        def _bg():
            nonlocal t6_dl_results
            try:
                all_docs = {}
                for ekw in expanded_kws:
                    results = search_docs(ekw, xml_path)
                    for d in results:
                        if d['docID'] not in all_docs:
                            all_docs[d['docID']] = d
                t6_dl_results = list(all_docs.values())
            except Exception as e:
                root.after(0, lambda: t6_dl_status.set(f'❌ 搜索出错: {str(e)[:60]}'))
                return
            root.after(0, _fill_results)

        threading.Thread(target=_bg, daemon=True).start()

    def _fill_results():
        tree = t6_dl_tree
        for d in t6_dl_results:
            title_text = d['title'][:100] if d['title'] else ''
            tree.insert('', 'end', values=(d['docID'], f'{title_text} ({d["docID"]})' if title_text else d['docID']),
                       iid=d['docID'])
        n = len(t6_dl_results)
        t6_dl_status.set(f'找到 {n} 篇文档  ·  勾选后点击"下载选中"或"下载全部"')

    def _t6_download_all():
        """下载全部搜索结果"""
        if not t6_dl_results:
            messagebox.showinfo('提示', '请先搜索文档')
            return
        _t6_do_download(t6_dl_results)

    def _t6_download_selected():
        """下载选中的文档"""
        sel = t6_dl_tree.selection()
        if not sel:
            messagebox.showinfo('提示', '请先勾选要下载的文档 (可多选)')
            return
        to_dl = [d for d in t6_dl_results if d['docID'] in sel]
        if not to_dl:
            return
        _t6_do_download(to_dl)

    def _t6_do_download(docs_to_dl):
        """执行下载, 同时保存标题映射到 _meta.json"""
        output_dir = os.path.join(_PROJECT_ROOT, 'ip_docs', 'docnav')
        os.makedirs(output_dir, exist_ok=True)

        total = len(docs_to_dl)
        t6_dl_progress.config(maximum=total, value=0)
        t6_dl_progress.grid()
        t6_dl_status.set(f'⏳ 下载中 0/{total} ...')

        success_count = [0]
        fail_count = [0]
        title_map = {}  # {docID: title}

        def _bg():
            # 先加载已有映射
            meta_file = os.path.join(output_dir, '_meta.json')
            if os.path.exists(meta_file):
                try:
                    with open(meta_file, 'r', encoding='utf-8') as f:
                        title_map.update(json.load(f))
                except Exception:
                    pass

            for i, doc in enumerate(docs_to_dl):
                key = doc['docID']
                url = doc['downloadURL']
                title_map[key] = doc.get('title', '')
                output_path = os.path.join(output_dir, f'{key}.pdf')

                from src.amd_doc_downloader import download_pdf as _dl_pdf
                ok, msg = _dl_pdf(url, output_path)
                if ok:
                    success_count[0] += 1
                else:
                    fail_count[0] += 1

                # 每下载一个就即时保存映射 (防中途崩溃丢失)
                try:
                    with open(meta_file, 'w', encoding='utf-8') as f:
                        json.dump(title_map, f, indent=2, ensure_ascii=False)
                except Exception:
                    pass

                # 更新 UI
                cur = i + 1
                root.after(0, lambda c=cur, s=success_count[0], f=fail_count[0],
                                  m=msg, k=key: (
                    t6_dl_progress.config(value=c),
                    t6_dl_status.set(
                        f'⏳ [{c}/{total}]  {k}  {"✓" if "成功" in m or "已存在" in m else "✗"} {m}'),
                ))

            # 完成
            root.after(0, lambda: (
                t6_dl_progress.config(value=total),
                t6_dl_status.set(
                    f'✅ 下载完成! 成功 {success_count[0]}, 失败 {fail_count[0]}  '
                    f'→ {output_dir}'),
                t6_dl_progress.grid_remove(),
                _t6_refresh_offline(),
            ))

        threading.Thread(target=_bg, daemon=True).start()

    # ── 标题映射加载 ──
    def _t6_load_title_map():
        """从 ip_docs/docnav/_meta.json 加载 {docID: title}"""
        meta_file = os.path.join(_DOC_DIR, 'docnav', '_meta.json')
        if os.path.exists(meta_file):
            try:
                with open(meta_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception:
                pass
        return {}

    def _t6_display_name(doc):
        """组装显示名: '文档'列=docID, '标题'列='CAN LogiCORE IP Product Guide (PG096)'"""
        stem = os.path.splitext(doc['name'])[0]
        title_map = _t6_load_title_map()
        display_title = title_map.get(stem, '') or title_map.get(doc['name'], '')
        if display_title:
            return (stem, f'{display_title} ({stem})')
        # 无标题映射: 显示 docID + 路径
        return (stem, f'{stem} ({doc["dir"]})')

    def _t6_refresh_offline():
        """刷新离线 PDF 搜索列表"""
        nonlocal t6_all_docs
        t6_all_docs = _scan_pdfs(_DOC_DIR)
        t6_status.set(f'已索引 {len(t6_all_docs)} 个 PDF  ·  输入关键词搜索')
        t6_tree.delete(*t6_tree.get_children())
        for d in t6_all_docs:
            name, title = _t6_display_name(d)
            t6_tree.insert('', 'end', values=(name, title), iid=d['path'])

    def _t6_open_dl(event=None):
        """双击搜索结果: 如果已下载则打开, 否则下载"""
        sel = t6_dl_tree.selection()
        if not sel:
            return
        doc_id = sel[0]
        output_path = os.path.join(_PROJECT_ROOT, 'ip_docs', 'docnav', f'{doc_id}.pdf')
        if os.path.isfile(output_path) and os.path.getsize(output_path) > 0:
            try:
                os.startfile(output_path) if _sys.platform == 'win32' else subprocess.Popen(['xdg-open', output_path])
            except Exception as e:
                messagebox.showerror('打开失败', str(e))
        else:
            # 未下载, 自动下载
            to_dl = [d for d in t6_dl_results if d['docID'] == doc_id]
            if to_dl:
                _t6_do_download(to_dl)
    t6_dl_tree.bind('<Double-Button-1>', _t6_open_dl)

    # ══════════════════════════════════════
    # TAB 7 — FPGA 计算器

    # ══════════════════════════════════════
    # TAB 7 — FPGA 计算器
    # ══════════════════════════════════════
    t7 = ttk.Frame(nb, style='TFrame')
    nb.add(t7, text='🧮 计算器')
    t7.grid_rowconfigure(1, weight=1)
    t7.grid_columnconfigure(0, weight=1)

    calc_nb = ttk.Notebook(t7)
    calc_nb.grid(row=1, column=0, sticky='nsew', padx=8, pady=(4, 8))

    # --- 公共 UI 工厂 ---
    def _mk_in(parent, text, var, r, c, unit='', w=12):
        # 列分配: c=0 -> 0,1,2,3   c=1 -> 4,5,6,7
        #        (label, entry, spacer, unit) 避免与下一组列冲突
        col = c * 4
        ttk.Label(parent, text=text, font=(F, 10)).grid(
            row=r, column=col, sticky='w', padx=(14, 4), pady=3)
        e = ttk.Entry(parent, textvariable=var, font=(M, 10), width=w)
        e.grid(row=r, column=col + 1, sticky='w', padx=(0, 4), pady=3)
        if unit:
            # 单位 label 与 Entry 之间加 12px 间隔 (用户要求: 数字与单位 M/bit 不要贴一起)
            ttk.Label(parent, text=unit, foreground=C['sub'],
                      font=(F, 9)).grid(row=r, column=col + 2,
                      sticky='w', padx=(12, 14), pady=3)
        return e

    def _mk_out(parent, text, var, r, c=0):
        col = c * 4
        ttk.Label(parent, text=text, font=(F, 10)).grid(
            row=r, column=col, sticky='w', padx=(14, 4), pady=3)
        # 结果值与名称之间也加 12px 间隔, 保持视觉一致
        ttk.Label(parent, textvariable=var, font=(M, 10, 'bold'),
                  foreground=C['blue']).grid(
            row=r, column=col + 1, sticky='w', padx=(12, 14), pady=3)

    # ═══════════════════════════════════
    # 子页 1 — FIFO 深度计算
    # ═══════════════════════════════════
    c1 = ttk.Frame(calc_nb, style='TFrame')
    calc_nb.add(c1, text='  FIFO 深度  ')
    c1.grid_columnconfigure(0, weight=1)

    f1i = ttk.LabelFrame(c1, text=' 输入参数 ', )
    f1i.grid(row=0, column=0, sticky='ew', padx=10, pady=(8, 4))

    v_fw, v_fr = tk.StringVar(value='100'), tk.StringVar(value='80')
    v_ww, v_rw = tk.StringVar(value='32'), tk.StringVar(value='32')
    v_burst = tk.StringVar(value='128')

    _mk_in(f1i, '写时钟频率',   v_fw,    0, 0, 'MHz')
    _mk_in(f1i, '读时钟频率',   v_fr,    1, 0, 'MHz')
    _mk_in(f1i, '写数据位宽',   v_ww,    0, 1, 'bit')
    _mk_in(f1i, '读数据位宽',   v_rw,    1, 1, 'bit')
    _mk_in(f1i, '突发长度',     v_burst, 2, 0, '次写')

    f1o = ttk.LabelFrame(c1, text=' 结果 ', )
    f1o.grid(row=1, column=0, sticky='ew', padx=10, pady=4)
    v_dmin, v_dsafe, v_rate = tk.StringVar(value='—'), tk.StringVar(value='—'), tk.StringVar(value='—')
    _mk_out(f1o, '最小深度',       v_dmin,  0)
    _mk_out(f1o, '推荐深度 (+20%)', v_dsafe, 1)
    _mk_out(f1o, '读写速率比',     v_rate,  2)

    def _calc_fifo_depth(*_):
        try:
            fw, fr = float(v_fw.get()), float(v_fr.get())
            ww, rw = float(v_ww.get()), float(v_rw.get())
            burst = float(v_burst.get())
            if min(fw, fr, ww, rw, burst) <= 0: raise ValueError
        except ValueError:
            v_dmin.set('输入无效'); v_dsafe.set('—'); v_rate.set('—'); return
        r = (fw * ww) / (fr * rw)
        if r <= 1:
            v_dmin.set('1'); v_rate.set(f'{r:.2f} (读 ≥ 写，无积压)')
        else:
            d = max(1, int(burst * (1 - 1 / r)) + 1)
            v_dmin.set(str(d)); v_rate.set(f'{r:.2f} (写快于读)')
        v_dsafe.set(str(max(4, int(int(v_dmin.get()) * 1.2 + 0.5))))

    for sv in (v_fw, v_fr, v_ww, v_rw, v_burst):
        sv.trace_add('write', _calc_fifo_depth)
    _calc_fifo_depth()

    # ═══════════════════════════════════
    # 子页 2 — FIFO 用量计算
    # ═══════════════════════════════════
    c2 = ttk.Frame(calc_nb, style='TFrame')
    calc_nb.add(c2, text='  FIFO 用量  ')
    c2.grid_columnconfigure(0, weight=1)

    f2i = ttk.LabelFrame(c2, text=' 输入参数 ', )
    f2i.grid(row=0, column=0, sticky='ew', padx=10, pady=(8, 4))

    v_depth  = tk.StringVar(value='512')
    v_wcnt   = tk.StringVar(value='0')
    v_rcnt   = tk.StringVar(value='0')

    _mk_in(f2i, 'FIFO 总深度', v_depth, 0, 0, '条目')
    _mk_in(f2i, '已写入计数', v_wcnt,  1, 0, '次')
    _mk_in(f2i, '已读出计数', v_rcnt,  2, 0, '次')

    f2o = ttk.LabelFrame(c2, text=' 结果 ', )
    f2o.grid(row=1, column=0, sticky='ew', padx=10, pady=4)
    v_used, v_free, v_pct, v_flag = (tk.StringVar(value='—') for _ in range(4))
    _mk_out(f2o, '已用条目',   v_used, 0)
    _mk_out(f2o, '空闲条目',   v_free, 1)
    _mk_out(f2o, '使用率',     v_pct,  2)
    _mk_out(f2o, '状态标志',   v_flag, 3)

    def _calc_fifo_usage(*_):
        try:
            d = int(v_depth.get())
            w = int(v_wcnt.get())
            r = int(v_rcnt.get())
            if d <= 0: raise ValueError
        except ValueError:
            v_used.set('输入无效'); v_free.set('—'); v_pct.set('—'); v_flag.set('—'); return
        used = max(0, w - r)
        if used > d: used = d
        free = d - used
        pct = used / d * 100
        v_used.set(str(used)); v_free.set(str(free)); v_pct.set(f'{pct:.1f} %')
        if used == 0:       v_flag.set('Empty (空)')
        elif used == d:     v_flag.set('Full (满)')
        elif pct >= 90:     v_flag.set(f'Almost-Full (将满, {pct:.0f}%)')
        elif pct <= 10:     v_flag.set(f'Almost-Empty (将空, {pct:.0f}%)')
        else:               v_flag.set('正常')

    for sv in (v_depth, v_wcnt, v_rcnt):
        sv.trace_add('write', _calc_fifo_usage)
    _calc_fifo_usage()

    # ═══════════════════════════════════
    # 子页 3 — 时钟换算
    # ═══════════════════════════════════
    c3 = ttk.Frame(calc_nb, style='TFrame')
    calc_nb.add(c3, text='  时钟换算  ')
    c3.grid_columnconfigure(0, weight=1)

    f3i = ttk.LabelFrame(c3, text=' 输入 (二选一) ', )
    f3i.grid(row=0, column=0, sticky='ew', padx=10, pady=(8, 4))

    v_fmhz   = tk.StringVar(value='100')
    v_period = tk.StringVar(value='10.000')

    _mk_in(f3i, '频率',      v_fmhz,   0, 0, 'MHz')
    _mk_in(f3i, '周期',      v_period, 1, 0, 'ns')

    f3o = ttk.LabelFrame(c3, text=' 结果 ', )
    f3o.grid(row=1, column=0, sticky='ew', padx=10, pady=4)
    v_fkhz, v_perps, v_perus = tk.StringVar(value='—'), tk.StringVar(value='—'), tk.StringVar(value='—')
    _mk_out(f3o, '频率',       v_fkhz,   0)
    _mk_out(f3o, '周期',       v_perps,  1)
    _mk_out(f3o, '周期',       v_perus,  2)

    def _calc_clk_from_freq(*_):
        try:
            fmhz = float(v_fmhz.get())
            if fmhz <= 0: raise ValueError
            ps = 1e6 / fmhz
            v_period.set(f'{ps / 1000:.3f}')
            v_fkhz.set(f'{fmhz:.3f} MHz  =  {fmhz * 1e3:.1f} KHz  =  {fmhz / 1e3:.4f} GHz')
            v_perps.set(f'{ps:.3f} ps')
            v_perus.set(f'{ps / 1e6:.6f} μs')
        except ValueError:
            v_fkhz.set('—'); v_perps.set('—'); v_perus.set('—')

    def _calc_clk_from_period(*_):
        try:
            pns = float(v_period.get())
            if pns <= 0: raise ValueError
            fmhz = 1e3 / pns
            v_fmhz.set(f'{fmhz:.3f}')
            v_fkhz.set(f'{fmhz:.3f} MHz  =  {fmhz * 1e3:.1f} KHz  =  {fmhz / 1e3:.4f} GHz')
            v_perps.set(f'{pns * 1e3:.3f} ps')
            v_perus.set(f'{pns / 1e3:.6f} μs')
        except ValueError:
            v_fkhz.set('—'); v_perps.set('—'); v_perus.set('—')

    v_fmhz.trace_add('write', lambda *_: _calc_clk_from_freq())
    v_period.trace_add('write', lambda *_: _calc_clk_from_period())
    _calc_clk_from_freq()

    # 常见频率参考
    f3r = ttk.LabelFrame(c3, text=' 常见 FPGA 时钟参考 ', )
    f3r.grid(row=2, column=0, sticky='nsew', padx=10, pady=4)
    refs = [
        ('100 MHz  (AXI4-Lite / 常用)',  '10.000 ns'),
        ('125 MHz  (Gigabit Ethernet)',   '8.000 ns'),
        ('150 MHz  (AXI4 常用)',          '6.667 ns'),
        ('200 MHz  (DDR3/4 控制器)',      '5.000 ns'),
        ('250 MHz  (高速 AXI4-Stream)',   '4.000 ns'),
        ('300 MHz  (UltraScale 典型)',     '3.333 ns'),
        ('27 MHz   (视频像素时钟)',        '37.037 ns'),
        ('148.5 MHz (1080p60 像素时钟)',   '6.734 ns'),
    ]
    for i, (f, p) in enumerate(refs):
        ttk.Label(f3r, text=f'{f:35s}', font=(M, 9),
                  foreground=C['sub']).grid(row=i, column=0, sticky='w',
                  padx=(14, 4), pady=1)
        ttk.Label(f3r, text=p, font=(M, 9),
                  foreground=C['sub']).grid(row=i, column=1, sticky='w',
                  padx=(0, 14), pady=1)

    # ═══════════════════════════════════
    # 子页 4 — 位宽 / 定点数换算
    # ═══════════════════════════════════
    c4 = ttk.Frame(calc_nb, style='TFrame')
    calc_nb.add(c4, text='  位宽/定点  ')
    c4.grid_columnconfigure(0, weight=1)

    f4i = ttk.LabelFrame(c4, text=' 位宽与符号 ', )
    f4i.grid(row=0, column=0, sticky='ew', padx=10, pady=(8, 4))

    v_bits  = tk.StringVar(value='16')
    v_signed = tk.BooleanVar(value=True)
    _mk_in(f4i, '位宽', v_bits, 0, 0, 'bit')
    ttk.Checkbutton(f4i, text='有符号 (signed)', variable=v_signed,
                    style='TCheckbutton').grid(row=0, column=3,
                    sticky='w', padx=(0, 14), pady=3)

    f4o = ttk.LabelFrame(c4, text=' 整数范围 ', )
    f4o.grid(row=1, column=0, sticky='ew', padx=10, pady=4)
    v_min, v_max, v_hexmin, v_hexmax = (tk.StringVar(value='—') for _ in range(4))
    _mk_out(f4o, 'Min (十进制)', v_min,    0)
    _mk_out(f4o, 'Max (十进制)', v_max,    1)
    _mk_out(f4o, 'Min (Hex)',   v_hexmin, 2)
    _mk_out(f4o, 'Max (Hex)',   v_hexmax, 3)

    def _calc_bitwidth(*_):
        try:
            n = int(v_bits.get())
            if n <= 0 or n > 64: raise ValueError
        except ValueError:
            v_min.set('输入无效'); v_max.set('—'); v_hexmin.set('—'); v_hexmax.set('—'); return
        if v_signed.get():
            lo, hi = -(1 << (n - 1)), (1 << (n - 1)) - 1
        else:
            lo, hi = 0, (1 << n) - 1
        v_min.set(str(lo)); v_max.set(str(hi))
        v_hexmin.set(hex(lo & ((1 << n) - 1))); v_hexmax.set(hex(hi))

    v_bits.trace_add('write', _calc_bitwidth)
    v_signed.trace_add('write', lambda *_: _calc_bitwidth())
    _calc_bitwidth()

    # 定点数 Qm.n 换算
    f4q = ttk.LabelFrame(c4, text=' 定点数 Qm.n 换算 ', )
    f4q.grid(row=2, column=0, sticky='nsew', padx=10, pady=4)

    v_qm = tk.StringVar(value='1')
    v_qn = tk.StringVar(value='15')
    v_qval = tk.StringVar(value='0')
    _mk_in(f4q, '整数位 m',  v_qm,   0, 0)
    _mk_in(f4q, '小数位 n',  v_qn,   0, 1)
    _mk_in(f4q, '定点整数值', v_qval, 1, 0, 'Q 格式整数')

    v_float, v_res = tk.StringVar(value='—'), tk.StringVar(value='—')
    _mk_out(f4q, '浮点值',      v_float, 2)
    _mk_out(f4q, '分辨率 (LSB)', v_res,   3)

    def _calc_fixedpt(*_):
        try:
            m = int(v_qm.get())
            n = int(v_qn.get())
            ival = int(v_qval.get())
            if m < 1 or n < 1: raise ValueError
        except ValueError:
            v_float.set('输入无效'); v_res.set('—'); return
        total = m + n
        fval = ival / (1 << n)
        lsb = 1.0 / (1 << n)
        v_float.set(f'{fval:.9f}')
        v_res.set(f'1 / 2^{n} = {lsb:.9f}')

    for sv in (v_qm, v_qn, v_qval):
        sv.trace_add('write', _calc_fixedpt)
    _calc_fixedpt()

    # ═══════════════════════════════════
    # 子页 5 — AXI DDR 带宽估算
    # ═══════════════════════════════════
    c5 = ttk.Frame(calc_nb, style='TFrame')
    calc_nb.add(c5, text='  DDR 带宽  ')
    c5.grid_columnconfigure(0, weight=1)

    f5i = ttk.LabelFrame(c5, text=' 输入参数 ', )
    f5i.grid(row=0, column=0, sticky='ew', padx=10, pady=(8, 4))

    v_bw   = tk.StringVar(value='64')
    v_dclk = tk.StringVar(value='200')
    v_eff  = tk.StringVar(value='80')

    _mk_in(f5i, '数据位宽',   v_bw,   0, 0, 'bit')
    _mk_in(f5i, 'DDR 时钟频率', v_dclk, 1, 0, 'MHz')
    _mk_in(f5i, '效率系数',   v_eff,  2, 0, '%')

    f5o = ttk.LabelFrame(c5, text=' 结果 ', )
    f5o.grid(row=1, column=0, sticky='ew', padx=10, pady=4)
    v_raw, v_effbw, v_gb = tk.StringVar(value='—'), tk.StringVar(value='—'), tk.StringVar(value='—')
    _mk_out(f5o, '理论峰值带宽',   v_raw,   0)
    _mk_out(f5o, '有效带宽',       v_effbw, 1)
    _mk_out(f5o, '有效带宽',       v_gb,    2)

    def _calc_ddr_bw(*_):
        try:
            bw  = float(v_bw.get())
            clk = float(v_dclk.get())
            eff = float(v_eff.get())
            if min(bw, clk, eff) <= 0: raise ValueError
        except ValueError:
            v_raw.set('输入无效'); v_effbw.set('—'); v_gb.set('—'); return
        raw_mb = bw / 8 * clk * 2  # DDR: ×2
        eff_mb = raw_mb * eff / 100
        v_raw.set(f'{raw_mb:.1f} MB/s')
        v_effbw.set(f'{eff_mb:.1f} MB/s')
        v_gb.set(f'{eff_mb / 1024:.2f} GB/s')

    for sv in (v_bw, v_dclk, v_eff):
        sv.trace_add('write', _calc_ddr_bw)
    _calc_ddr_bw()

    # ═══════════════════════════════════
    # 子页 6 — BRAM / DSP / LUT 估算
    # ═══════════════════════════════════
    c6 = ttk.Frame(calc_nb, style='TFrame')
    calc_nb.add(c6, text='  资源估算  ')
    c6.grid_columnconfigure(0, weight=1)

    # BRAM
    f6b = ttk.LabelFrame(c6, text=' BRAM 估算 (7 系列) ', )
    f6b.grid(row=0, column=0, sticky='ew', padx=10, pady=(8, 4))

    v_mdepth = tk.StringVar(value='1024')
    v_mwidth = tk.StringVar(value='32')
    _mk_in(f6b, '存储深度', v_mdepth, 0, 0, 'words')
    _mk_in(f6b, '数据位宽', v_mwidth, 0, 1, 'bit')

    v_b36, v_b18, v_bits_used = tk.StringVar(value='—'), tk.StringVar(value='—'), tk.StringVar(value='—')
    _mk_out(f6b, 'BRAM 36Kb',  v_b36, 1)
    _mk_out(f6b, 'BRAM 18Kb',  v_b18, 2)
    _mk_out(f6b, '总存储量',   v_bits_used, 3)

    def _calc_bram(*_):
        try:
            d = int(v_mdepth.get())
            w = int(v_mwidth.get())
            if d <= 0 or w <= 0: raise ValueError
        except ValueError:
            v_b36.set('—'); v_b18.set('—'); v_bits_used.set('—'); return
        total_bits = d * w
        # 按 36Kb 优先
        primitives = []
        for pbits, name in [(36864, '36Kb'), (18432, '18Kb')]:
            for cfg_w in [72, 36, 18, 9, 4, 2, 1]:
                cfg_d = pbits // cfg_w
                if cfg_d >= min(512, d) and cfg_w >= w:
                    n_prims = (d + cfg_d - 1) // cfg_d
                    primitives.append((n_prims, name, pbits))
                    break
        if not primitives:
            v_b36.set('无法实现'); v_b18.set('—'); v_bits_used.set(f'{total_bits} bits'); return

        n36 = sum(p[0] for p in primitives if '36' in p[1])
        n18 = sum(p[0] for p in primitives if '18' in p[1])
        v_b36.set(str(n36)); v_b18.set(str(n18))
        v_bits_used.set(f'{total_bits} bits  ({total_bits / 1024:.1f} Kb)')

    v_mdepth.trace_add('write', _calc_bram)
    v_mwidth.trace_add('write', _calc_bram)
    _calc_bram()

    # DSP
    f6d = ttk.LabelFrame(c6, text=' DSP48 估算 (7 系列) ', )
    f6d.grid(row=1, column=0, sticky='ew', padx=10, pady=4)

    v_ma = tk.StringVar(value='25')
    v_mb = tk.StringVar(value='18')
    _mk_in(f6d, '乘数 A 位宽', v_ma, 0, 0, 'bit')
    _mk_in(f6d, '乘数 B 位宽', v_mb, 0, 1, 'bit')

    v_dsp_cnt = tk.StringVar(value='—')
    _mk_out(f6d, '所需 DSP48E1', v_dsp_cnt, 1)

    def _calc_dsp(*_):
        try:
            a = int(v_ma.get()); b = int(v_mb.get())
            if a <= 0 or b <= 0: raise ValueError
        except ValueError:
            v_dsp_cnt.set('—'); return
        # DSP48E1: 25×18 乘法器
        n_a = (a + 24) // 25
        n_b = (b + 17) // 18
        v_dsp_cnt.set(str(n_a * n_b))

    v_ma.trace_add('write', _calc_dsp)
    v_mb.trace_add('write', _calc_dsp)
    _calc_dsp()

    # LUT 粗略估算
    f6l = ttk.LabelFrame(c6, text=' 粗略 LUT 估算 ', )
    f6l.grid(row=2, column=0, sticky='nsew', padx=10, pady=4)

    v_lut_bram  = tk.StringVar(value='0')
    v_lut_dsp   = tk.StringVar(value='0')
    v_lut_regs  = tk.StringVar(value='1000')
    _mk_in(f6l, 'BRAM 数量',  v_lut_bram, 0, 0)
    _mk_in(f6l, 'DSP 数量',   v_lut_dsp,  0, 1)
    _mk_in(f6l, '寄存器数',   v_lut_regs, 1, 0, 'approx')

    v_lut_total = tk.StringVar(value='—')
    _mk_out(f6l, '预估 LUT 总数', v_lut_total, 2)

    def _calc_lut(*_):
        try:
            nb = int(v_lut_bram.get()); nd = int(v_lut_dsp.get()); nr = int(v_lut_regs.get())
            if min(nb, nd, nr) < 0: raise ValueError
        except ValueError:
            v_lut_total.set('—'); return
        # BRAM: 每个约消耗 10 LUT (控制逻辑)
        # DSP:  每个约消耗 150 LUT (旁路逻辑 + 互连)
        # 寄存器: ~0.7 LUT / reg (粗略，一个 LUT 可做 2 个 FF 的简单逻辑)
        total = nb * 10 + nd * 150 + nr * 0.7
        v_lut_total.set(f'{int(total + 0.5):,}  LUT')

    for sv in (v_lut_bram, v_lut_dsp, v_lut_regs):
        sv.trace_add('write', _calc_lut)
    _calc_lut()

    # ══════════════════════════════════════
    # TAB 8 — 约束生成器 (管脚 / 时序 / 其他)
    # ══════════════════════════════════════
    t8 = ttk.Frame(nb, style='TFrame')
    nb.add(t8, text='🔌 约束生成')
    t8.grid_rowconfigure(0, weight=1)
    t8.grid_columnconfigure(0, weight=1)

    # 子页 Notebook
    t8_nb = ttk.Notebook(t8)
    t8_nb.grid(row=0, column=0, sticky='nsew', padx=4, pady=(4, 4))


    # ══════════════════════════════════════
    # 子页 4 — 名称映射 (管脚表 Port ↔ 顶层信号)
    # ══════════════════════════════════════
    t8_a4 = ttk.Frame(t8_nb, style='TFrame')
    t8_nb.add(t8_a4, text='  🔗 名称映射  ')
    t8_a4.grid_columnconfigure(0, weight=1)
    t8_a4.grid_rowconfigure(2, weight=1)  # PanedWindow 可扩展

    # ── 管脚表 ──
    fm4_pin = ttk.LabelFrame(t8_a4, text=' 管脚表 (CSV/Excel) ')
    fm4_pin.grid(row=0, column=0, sticky='ew', padx=12, pady=(8, 2))
    fm4_pin.grid_columnconfigure(1, weight=1)
    m4_pin_file = tk.StringVar()
    ttk.Entry(fm4_pin, textvariable=m4_pin_file, font=(F, 10)).grid(row=0, column=0, sticky='ew', padx=(14, 4), pady=(6, 4), columnspan=2)
    ttk.Button(fm4_pin, text='浏览', command=lambda: (f := filedialog.askopenfilename(
        title='选择管脚表', filetypes=[('表格文件','*.csv *.xlsx'),('CSV','*.csv'),('Excel','*.xlsx'),('All','*.*')])) and m4_pin_file.set(f),
        style='Normal.TButton').grid(row=0, column=2, padx=(4, 4), pady=(6, 4))
    ttk.Button(fm4_pin, text='解析', command=lambda: _m4_parse_pin(),
               style='Accent.TButton').grid(row=0, column=3, padx=(0, 14), pady=(6, 4))

    # ── 顶层文件 ──
    fm4_top = ttk.LabelFrame(t8_a4, text=' 顶层文件 (.v/.sv/.vhd) ')
    fm4_top.grid(row=1, column=0, sticky='ew', padx=12, pady=2)
    fm4_top.grid_columnconfigure(1, weight=1)
    m4_top_file = tk.StringVar()
    ttk.Entry(fm4_top, textvariable=m4_top_file, font=(F, 10)).grid(row=0, column=0, sticky='ew', padx=(14, 4), pady=(6, 4), columnspan=2)
    ttk.Button(fm4_top, text='浏览', command=lambda: (f := filedialog.askopenfilename(
        title='选择顶层文件', filetypes=[('Verilog/VHDL','*.v *.sv *.vhd *.vhdl'),('All','*.*')])) and m4_top_file.set(f),
        style='Normal.TButton').grid(row=0, column=2, padx=(4, 4), pady=(6, 4))
    ttk.Button(fm4_top, text='解析', command=lambda: _m4_parse_top(),
               style='Accent.TButton').grid(row=0, column=3, padx=(0, 14), pady=(6, 4))

    # ── 可拖拽分隔: 上=端口匹配区, 下=已匹配列表 ──
    m4_pane = ttk.PanedWindow(t8_a4, orient='vertical')
    m4_pane.grid(row=2, column=0, sticky='nsew', padx=12, pady=(0, 0))

    # ── 上区: 主映射区 (左列表 | 匹配按钮 | 右列表) ──
    m4_map_frame = ttk.Frame(m4_pane, style='TFrame')
    m4_map_frame.grid_columnconfigure(0, weight=1)  # 左列表
    m4_map_frame.grid_columnconfigure(2, weight=1)  # 右列表
    m4_map_frame.grid_rowconfigure(0, weight=1)

    # 左: 管脚表 Port 列表
    m4_left_frame = ttk.Frame(m4_map_frame, style='TFrame')
    m4_left_frame.grid(row=0, column=0, sticky='nsew', padx=(8, 2), pady=(6, 4))
    m4_left_frame.grid_columnconfigure(0, weight=1)
    m4_left_frame.grid_rowconfigure(0, weight=0)
    m4_left_frame.grid_rowconfigure(1, weight=1)

    ttk.Label(m4_left_frame, text='管脚表 Port', font=(F, 9, 'bold'),
              foreground=C['sub']).grid(row=0, column=0, sticky='w', padx=2, pady=(0, 2))
    m4_left_list = tk.Listbox(m4_left_frame, selectmode='extended', exportselection=False,
                               font=(M, 10), bg=C['ebg'], fg=C['fg'], relief='flat',
                               highlightthickness=0, activestyle='none')
    m4_left_scroll = ttk.Scrollbar(m4_left_frame, orient='vertical', command=m4_left_list.yview)
    m4_left_list.configure(yscrollcommand=m4_left_scroll.set)
    m4_left_list.grid(row=1, column=0, sticky='nsew')
    m4_left_scroll.grid(row=1, column=1, sticky='ns')

    # 中: 匹配按钮 (在左右列表之间居中)
    m4_mid_frame = ttk.Frame(m4_map_frame, style='TFrame')
    m4_mid_frame.grid(row=0, column=1, sticky='', padx=8, pady=(6, 4))
    ttk.Button(m4_mid_frame, text='→  匹配  →', command=lambda: _m4_match(),
               style='Accent.TButton').pack(pady=(0, 4))
    ttk.Button(m4_mid_frame, text='← 取消匹配 ←', command=lambda: _m4_unmatch(),
               style='Normal.TButton').pack()
    # 同名自动匹配
    ttk.Button(m4_mid_frame, text='⚡ 自动匹配', command=lambda: _m4_auto_match(),
               style='Normal.TButton').pack(pady=(12, 0))

    # 右: 顶层信号列表
    m4_right_frame = ttk.Frame(m4_map_frame, style='TFrame')
    m4_right_frame.grid(row=0, column=2, sticky='nsew', padx=(2, 8), pady=(6, 4))
    m4_right_frame.grid_columnconfigure(0, weight=1)
    m4_right_frame.grid_rowconfigure(0, weight=0)
    m4_right_frame.grid_rowconfigure(1, weight=1)

    ttk.Label(m4_right_frame, text='顶层信号', font=(F, 9, 'bold'),
              foreground=C['sub']).grid(row=0, column=0, sticky='w', padx=2, pady=(0, 2))
    m4_right_list = tk.Listbox(m4_right_frame, selectmode='extended', exportselection=False,
                                font=(M, 10), bg=C['ebg'], fg=C['fg'], relief='flat',
                                highlightthickness=0, activestyle='none')
    m4_right_scroll = ttk.Scrollbar(m4_right_frame, orient='vertical', command=m4_right_list.yview)
    m4_right_list.configure(yscrollcommand=m4_right_scroll.set)
    m4_right_list.grid(row=1, column=0, sticky='nsew')
    m4_right_scroll.grid(row=1, column=1, sticky='ns')

    # ── 下区: 已匹配列表 ──
    m4_matched_frame = ttk.Frame(m4_pane, style='TFrame')
    m4_matched_frame.grid_columnconfigure(0, weight=1)
    m4_matched_frame.grid_rowconfigure(0, weight=1)
    # 可滚动 Canvas
    m4_matched_canvas = tk.Canvas(m4_matched_frame, bg=C['ebg'], highlightthickness=0)
    m4_matched_scroll = ttk.Scrollbar(m4_matched_frame, orient='vertical',
                                       command=m4_matched_canvas.yview)
    m4_matched_inner = ttk.Frame(m4_matched_canvas, style='TFrame')
    m4_matched_canvas.configure(yscrollcommand=m4_matched_scroll.set)
    m4_matched_canvas.create_window((0, 0), window=m4_matched_inner, anchor='nw')
    m4_matched_inner.bind('<Configure>',
        lambda e: m4_matched_canvas.configure(scrollregion=m4_matched_canvas.bbox('all')))
    m4_matched_canvas.bind('<Configure>',
        lambda e: m4_matched_canvas.itemconfig('all', width=m4_matched_canvas.winfo_width()-4) if m4_matched_canvas.find_all() else None)
    m4_matched_canvas.grid(row=0, column=0, sticky='nsew', padx=0, pady=0)
    m4_matched_scroll.grid(row=0, column=1, sticky='ns')

    # 添加两个面板到 PanedWindow
    m4_pane.add(m4_map_frame, weight=3)
    m4_pane.add(m4_matched_frame, weight=1)

    # 状态栏
    m4_status = tk.StringVar(value='分别加载管脚表和顶层文件后开始匹配')
    m4_btn_frame = ttk.Frame(t8_a4, style='TFrame')
    m4_btn_frame.grid(row=3, column=0, sticky='ew', padx=12, pady=(2, 8))
    ttk.Label(m4_btn_frame, textvariable=m4_status, foreground=C['sub'], font=(F, 8)).pack(side='left', padx=(4, 8))
    ttk.Button(m4_btn_frame, text='💾 生成新文件', command=lambda: _m4_save_csv(),
               style='Accent.TButton').pack(side='right', padx=(0, 4))
    ttk.Button(m4_btn_frame, text='❌ 清除所有匹配', command=lambda: _m4_clear_all(),
               style='Normal.TButton').pack(side='right', padx=(0, 4))

    m4_parsed_pins = []     # [{port, pin, dir, iostd, ...}, ...]  from CSV
    m4_top_signals = []     # ['sys_clk', 'cpu_rst_n', ...]  from top file
    m4_port_map = {}        # {pin_port: top_signal}  -- current matches
    m4_matched_labels = {}  # {pin_port: tk.Label widget} -- for easy removal

    # ── 解析管脚表 ──
    def _m4_parse_pin():
        import csv as _csv
        fp = m4_pin_file.get().strip()
        if not fp or not os.path.isfile(fp):
            messagebox.showerror('错误', '请选择有效的管脚表文件')
            return
        m4_parsed_pins.clear()
        m4_port_map.clear()
        _m4_refresh_all()
        ext = os.path.splitext(fp)[1].lower()
        try:
            if ext == '.csv':
                header, rows = _parse_csv(fp)
            elif ext == '.xlsx':
                full = _xlsx_fallback_via_stdlib(fp)
                if not full:
                    m4_status.set('xlsx 解析失败'); return
                header = full[0] if full else None
                rows = full[1:] if len(full) > 1 else []
            else:
                m4_status.set(f'不支持的文件格式: {ext}'); return
        except Exception as e:
            m4_status.set(f'文件读取失败: {e}'); return

        if not header:
            m4_status.set('管脚表解析失败: 无表头'); return

        col_map = _detect_columns(header)
        if 'port' not in col_map.values():
            m4_status.set('未找到 Port 列 (尝试了: port/signal/name 等)'); return

        for row in rows:
            entry = {}
            for ci, field in col_map.items():
                if ci < len(row):
                    entry[field] = row[ci].strip()
            pp = entry.get('port', '').strip()
            if not pp: continue
            # 处理多位宽: ddr_addr[14:0] → 展开为 ddr_addr[0]..ddr_addr[14]
            bus_match = re.match(r'^(\w+)\s*\[(\d+):(\d+)\]\s*$', pp)
            if bus_match:
                base = bus_match.group(1); hi = int(bus_match.group(2)); lo = int(bus_match.group(3))
                step = 1 if hi >= lo else -1
                for bit in range(hi, lo - step, -step):
                    bp = f'{base}[{bit}]'
                    e2 = dict(entry); e2['port'] = bp; e2['_bus_base'] = base
                    m4_parsed_pins.append(e2)
            else:
                entry['_bus_base'] = pp  # 单bit的bus_base就是自己
                m4_parsed_pins.append(entry)

        _m4_refresh_left_list()
        m4_status.set(f'管脚表: {len(m4_parsed_pins)} 个端口 | 顶层: {len(m4_top_signals)} 个信号')

    # ── 解析顶层 ──
    def _m4_parse_top():
        fp = m4_top_file.get().strip()
        if not fp or not os.path.isfile(fp):
            messagebox.showerror('错误', '请选择有效的顶层文件'); return
        with open(fp, 'r', encoding='utf-8', errors='replace') as f:
            content = f.read()
        import re as _re
        ports_found = []  # [(name, width), ...]
        ext = os.path.splitext(fp)[1].lower()
        if ext in ('.vhd', '.vhdl'):
            port_blocks = _re.findall(r'PORT\s*\((.*?)\)\s*;', content, _re.DOTALL | _re.IGNORECASE)
            for pb in port_blocks:
                for line in pb.split('\n'):
                    line = line.strip()
                    if not line or line.startswith('--'): continue
                    m = _re.match(r'(\w+)\s*:\s*(in|out|inout|buffer)\s+', line, _re.IGNORECASE)
                    if m:
                        nm = m.group(1); w = '1'
                        wm = _re.search(r'VECTOR\s*\(\s*(\d+)\s+downto\s+(\d+)\)', line, _re.IGNORECASE)
                        if wm: w = str(int(wm.group(1)) - int(wm.group(2)) + 1)
                        ports_found.append((nm, w))
        else:
            mods = _re.findall(r'module\s+\w+\s*(?:#\s*\(.*?\)\s*)?\s*\((.*?)\)\s*;', content, _re.DOTALL | _re.IGNORECASE)
            if not mods:
                m4_status.set('未找到 module 定义'); return
            body = mods[0]
            body = _re.sub(r'//.*', '', body)
            body = _re.sub(r'/\*.*?\*/', '', body, flags=_re.DOTALL)
            sv_kw = {'input','output','inout','wire','reg','logic','signed','integer',
                     'tri','wand','wor','supply0','supply1','tri0','tri1','triand','trior','trireg'}
            depth = 0; cur = ''
            for ch in body:
                if ch == '(': depth += 1
                elif ch == ')': depth -= 1
                if ch == ',' and depth == 0:
                    tokens = cur.strip().split()
                    # 提取宽度
                    w = '1'
                    wm = _re.search(r'\[(\d+)\s*:\s*(\d+)\]', cur)
                    if wm: w = str(abs(int(wm.group(1)) - int(wm.group(2))) + 1)
                    for tok in reversed(tokens):
                        tok_clean = tok.rstrip(',; ').rstrip(')').rstrip(']')
                        if tok_clean.lower() not in sv_kw and _re.match(r'^\w+$', tok_clean) and not tok_clean.startswith('['):
                            ports_found.append((tok_clean, w)); break
                    cur = ''
                else: cur += ch
            if cur.strip():
                tokens = cur.strip().split()
                w = '1'
                wm = _re.search(r'\[(\d+)\s*:\s*(\d+)\]', cur)
                if wm: w = str(abs(int(wm.group(1)) - int(wm.group(2))) + 1)
                for tok in reversed(tokens):
                    tok_clean = tok.rstrip(',; ').rstrip(')').rstrip(']')
                    if tok_clean.lower() not in sv_kw and _re.match(r'^\w+$', tok_clean) and not tok_clean.startswith('['):
                        ports_found.append((tok_clean, w)); break

        m4_top_signals[:] = ports_found
        _m4_refresh_right_list()
        m4_status.set(f'管脚表: {len(m4_parsed_pins)} 个端口 | 顶层: {len(m4_top_signals)} 个信号')

    # ── 刷新左列表(管脚表端口) ──
    def _m4_refresh_left_list():
        m4_left_list.delete(0, 'end')
        # 按名称排序, 同一前缀的信号聚在一起便于框选
        sorted_pins = sorted(m4_parsed_pins, key=lambda x: x.get('port', ''))
        for pin in sorted_pins:
            pp = pin.get('port', '').strip()
            if not pp: continue
            m4_left_list.insert('end', pp)
            if pp in m4_port_map and m4_port_map[pp]:
                m4_left_list.itemconfig('end', fg=C['green'], selectforeground=C['green'])

    # ── 刷新右列表(顶层信号) ──
    def _m4_refresh_right_list():
        m4_right_list.delete(0, 'end')
        used_signals = set(m4_port_map.values())  # 现在都是单 bit 如 ddr3_addr[0]
        for sig, width in m4_top_signals:
            w_int = int(width)
            if w_int > 1:
                # 展开总线为单 bit
                for bit in range(w_int):
                    entry = f'{sig}[{bit}]'
                    m4_right_list.insert('end', entry)
                    if entry in used_signals:
                        m4_right_list.itemconfig('end', fg=C['blue'], selectforeground=C['blue'])
            else:
                m4_right_list.insert('end', sig)
                if sig in used_signals:
                    m4_right_list.itemconfig('end', fg=C['blue'], selectforeground=C['blue'])

    # ── 刷新已匹配区 ──
    def _m4_refresh_matched():
        for w in m4_matched_inner.winfo_children():
            w.destroy()
        m4_matched_labels.clear()
        # 表头
        ttk.Label(m4_matched_inner, text='已匹配 (双击或点 ✕ 取消)', font=(F, 9, 'bold'),
                  foreground=C['sub']).pack(anchor='w', padx=14, pady=(6, 4))
        if not m4_port_map:
            ttk.Label(m4_matched_inner, text='  暂无匹配', foreground=C['sub'],
                      font=(F, 8, 'italic')).pack(anchor='w', padx=4)
            return
        import itertools as _it
        for i, (pin_port, top_sig) in enumerate(m4_port_map.items()):
            if not top_sig: continue
            frame = ttk.Frame(m4_matched_inner, style='TFrame')
            frame.pack(fill='x', padx=2, pady=1)
            lbl = ttk.Label(frame, text=f'{pin_port}  →  {top_sig}',
                           font=(M, 9), foreground=C['green'])
            lbl.pack(side='left', padx=4)
            btn = ttk.Button(frame, text='✕', width=3,
                            command=lambda pp=pin_port: _m4_remove_pair(pp),
                            style='Normal.TButton')
            btn.pack(side='right', padx=(8, 4))
            # 双击取消
            lbl.bind('<Double-1>', lambda e, pp=pin_port: _m4_remove_pair(pp))
            m4_matched_labels[pin_port] = frame
        _m4_update_status()

    # ── 匹配 ──
    def _m4_match():
        left_sel = [m4_left_list.get(i).strip() for i in m4_left_list.curselection()]
        right_sel = [m4_right_list.get(i).strip() for i in m4_right_list.curselection()]
        if not left_sel:
            messagebox.showinfo('提示', '请在左侧选择管脚表端口')
            return
        if not right_sel:
            messagebox.showinfo('提示', '请在右侧选择顶层信号')
            return
        # 一一对应, 按数量少的来
        count = min(len(left_sel), len(right_sel))
        for i in range(count):
            m4_port_map[left_sel[i]] = right_sel[i]
        _m4_refresh_all()

    # ── 取消匹配 ──
    def _m4_unmatch():
        left_sel = [m4_left_list.get(i).strip() for i in m4_left_list.curselection()]
        if not left_sel:
            messagebox.showinfo('提示', '请在左侧选择要取消匹配的端口')
            return
        for pp in left_sel:
            if pp in m4_port_map:
                del m4_port_map[pp]
        _m4_refresh_all()

    # ── 移除单对 ──
    def _m4_remove_pair(pp):
        if pp in m4_port_map:
            del m4_port_map[pp]
        _m4_refresh_all()

    # ── 自动匹配同名 ──
    def _m4_auto_match():
        if not m4_parsed_pins or not m4_top_signals:
            messagebox.showinfo('提示', '请先解析管脚表和顶层文件')
            return
        # 构建右列表所有展开后的单 bit 信号集合
        all_top_bits = set()
        for s, w in m4_top_signals:
            w_int = int(w)
            if w_int > 1:
                for bit in range(w_int):
                    all_top_bits.add(f'{s}[{bit}]')
            else:
                all_top_bits.add(s)
        auto = 0
        for pin in m4_parsed_pins:
            pp = pin.get('port', '').strip()
            if not pp or pp in m4_port_map:
                continue
            if pp in all_top_bits:
                m4_port_map[pp] = pp
                auto += 1
        _m4_refresh_all()
        m4_status.set(f'自动匹配 {auto} 个同名端口')

    # ── 清除所有匹配 ──
    def _m4_clear_all():
        if m4_port_map:
            if messagebox.askyesno('确认', f'确定清除所有 {len(m4_port_map)} 个匹配?'):
                m4_port_map.clear()
                _m4_refresh_all()

    # ── 刷新全部 ──
    def _m4_refresh_all():
        _m4_refresh_left_list()
        _m4_refresh_right_list()
        _m4_refresh_matched()
        _m4_update_status()

    # ── 更新状态 ──
    def _m4_update_status():
        matched = sum(1 for v in m4_port_map.values() if v)
        m4_status.set(f'已匹配 {matched}/{len(m4_parsed_pins)} 个端口 | 管脚表: {len(m4_parsed_pins)} | 顶层: {len(m4_top_signals)}')

    # ── 生成新文件 ──
    def _m4_save_csv():
        if not m4_port_map:
            messagebox.showinfo('提示', '没有匹配的端口，请先进行匹配'); return
        initial = os.path.splitext(os.path.basename(m4_pin_file.get() or 'pinout'))[0] + '_mapped.csv'
        fp = filedialog.asksaveasfilename(title='保存映射后的管脚表 (仅匹配的端口)', defaultextension='.csv',
            filetypes=[('CSV','*.csv'),('Excel','*.xlsx')], initialfile=initial)
        if not fp:
            return
        try:
            # 从原文件获取表头和列映射
            orig_fp = m4_pin_file.get()
            if os.path.isfile(orig_fp):
                hdr, _ = _parse_csv(orig_fp)
                col_map = _detect_columns(hdr)
            else:
                hdr = list(m4_parsed_pins[0].keys()) if m4_parsed_pins else []
                col_map = {i: k for i, k in enumerate(hdr)}

            # 建立反向索引: 原始header名 → entry中的key
            hdr_to_key = {}
            for i, h in enumerate(hdr):
                if i in col_map:
                    hdr_to_key[h] = col_map[i]

            ext = os.path.splitext(fp)[1].lower()
            if ext == '.csv':
                import csv as _csv
                with open(fp, 'w', newline='', encoding='utf-8-sig') as f:
                    w = _csv.writer(f)
                    w.writerow(hdr)
                    count = 0
                    for pin in m4_parsed_pins:
                        pp = pin.get('port', '').strip()
                        if pp not in m4_port_map or not m4_port_map[pp]:
                            continue
                        mapped_name = m4_port_map[pp]
                        row = []
                        for h in hdr:
                            key = hdr_to_key.get(h, h)
                            if key in ('port',) or h.lower() in ('port',):
                                row.append(mapped_name)  # 替换为映射后的名
                            else:
                                row.append(pin.get(key, pin.get(h, '')))
                        w.writerow(row)
                        count += 1
                    m4_status.set(f'已保存: {os.path.basename(fp)} — {count} 个匹配端口')
            elif ext == '.xlsx':
                _save_xlsx_simple(fp, hdr, m4_parsed_pins, m4_port_map, hdr_to_key,
                                  'Port' if 'Port' in hdr else (next((h for h in hdr if h.lower()=='port'), hdr[0])))
                count = sum(1 for pin in m4_parsed_pins
                           if pin.get('port','').strip() in m4_port_map and m4_port_map[pin.get('port','').strip()])
                m4_status.set(f'已保存: {os.path.basename(fp)} — {count} 个匹配端口')
            else:
                messagebox.showerror('格式错误', '仅支持 .csv 和 .xlsx')
        except Exception as e:
            messagebox.showerror('保存失败', str(e))

    def _save_xlsx_simple(fp, hdr, pins, port_map, hdr_to_key=None, port_col_name='Port'):
        """简单的 xlsx 保存 (不使用 openpyxl)"""
        import zipfile
        if hdr_to_key is None:
            hdr_to_key = {h: h for h in hdr}
        count = 0
        rows_xml = ['<sheetData>']
        for pin in pins:
            pp = pin.get('port', '').strip()
            if pp not in port_map or not port_map[pp]:
                continue
            mapped_name = port_map[pp]
            cells = []
            for ci, h in enumerate(hdr):
                col_letter = chr(65 + ci) if ci < 26 else 'A' + chr(65 + ci - 26)
                key = hdr_to_key.get(h, h)
                if key in ('port',) or h.lower() in ('port',):
                    val = mapped_name
                else:
                    val = pin.get(key, pin.get(h, ''))
                cells.append(f'<c r="{col_letter}{count+1}" t="inlineStr"><is><t>{_xml_esc(val)}</t></is></c>')
            rows_xml.append(f'<row r="{count+1}">{"".join(cells)}</row>')
            count += 1
        rows_xml.append('</sheetData>')

        # 表头行
        hdr_cells = []
        for ci, h in enumerate(hdr):
            col_letter = chr(65 + ci) if ci < 26 else 'A' + chr(65 + ci - 26)
            hdr_cells.append(f'<c r="{col_letter}1" t="inlineStr"><is><t>{_xml_esc(h)}</t></is></c>')
        rows_xml.insert(1, f'<row r="1">{"".join(hdr_cells)}</row>')

        sheet_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
        <worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
        {''.join(rows_xml)}
        </worksheet>'''

        with zipfile.ZipFile(fp, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr('[Content_Types].xml', '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
</Types>''')
            zf.writestr('_rels/.rels', '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>''')
            zf.writestr('xl/workbook.xml', '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
<sheets><sheet name="mapped_ports" sheetId="1" r:id="rId1"/></sheets>
</workbook>''')
            zf.writestr('xl/_rels/workbook.xml.rels', '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
</Relationships>''')
            zf.writestr('xl/worksheets/sheet1.xml', sheet_xml)

    def _xml_esc(s):
        return str(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;')
    # ── 子页 1: 管脚约束 (原 Tab 8 全部内容) ──
    t8_a1 = ttk.Frame(t8_nb, style='TFrame')
    t8_nb.add(t8_a1, text='  📌 管脚约束  ')
    t8_a1.grid_rowconfigure(3, weight=1)  # PanedWindow (管脚表 + XDC, 可拖拽分界)
    t8_a1.grid_columnconfigure(0, weight=1)

    # --- I/O 标准 → VCCO 电压映射 (7 系列) ---
    _IO_VCCO = {
        'LVCMOS33':3.3,'LVCMOS25':2.5,'LVCMOS18':1.8,'LVCMOS15':1.5,'LVCMOS12':1.2,
        'LVTTL':3.3,'PCI33_3':3.3,
        'SSTL15':1.5,'SSTL135':1.35,'SSTL18':1.8,
        'SSTL15_R':1.5,'SSTL135_R':1.35,'SSTL18_R':1.8,
        'HSTL_I':1.5,'HSTL_II':1.5,'HSTL_I_18':1.8,'HSTL_II_18':1.8,
        'HSTL_I_R':1.5,'HSTL_II_R':1.5,'HSUL_12':1.2,'POD12':1.2,
        'LVDS_25':2.5,'LVDS':2.5,'BLVDS_25':2.5,'MINI_LVDS_25':2.5,
        'PPDS_25':2.5,'RSDS_25':2.5,'TMDS_33':3.3,
        'DIFF_SSTL15':1.5,'DIFF_SSTL135':1.35,'DIFF_SSTL18':1.8,
        'DIFF_HSTL_I':1.5,'DIFF_HSTL_II':1.5,
        'DIFF_HSTL_I_18':1.8,'DIFF_HSTL_II_18':1.8,
        'DIFF_HSUL_12':1.2,'DIFF_POD12':1.2,'MOBILE_DDR':1.8,
    }

    def _get_vcco(iostd):
        """查找 I/O 标准对应的 VCCO，返回 (voltage, matched_key) 或 (None, None)
        兼容: LVCMOS33V3 / LVCMOS33V2 / LVCMOS33_BANK13 等"电平+Bank后缀"写法
        """
        if not iostd: return None, None
        k = iostd.strip().upper()
        if k in _IO_VCCO: return _IO_VCCO[k], k
        # 模糊匹配：去掉尾部 _V\d / _BANK\d / _R / _T / _D / _S 等后缀
        base = re.sub(r'(_V\d+|_BANK\d+|_(R|T|D|S))$', '', k)
        if base in _IO_VCCO: return _IO_VCCO[base], base
        return None, None

    # --- 列名自动识别 ---
    _COL_ALIASES = {
        'port': ['port','signal','name','pin_name','net','信号名','端口名','信号',
                 'adr','address','addr','fpga_name','efr_name','signal_name'],
        'pin':  ['pin','package_pin','location','loc','管脚','引脚','pin#','package',
                 'pin_number','pin#'],
        'dir':  ['dir','direction','io','mode','方向','in/out','i/o'],
        'iostd':['iostandard','io_standard','standard','lvcmos','lvds','电平标准','电平','iostd'],
        'bank': ['bank','bank#','bank号','bank编号','fpga_bank'],
        'drive':['drive','drive_strength','驱动','驱动强度','drive(ma)'],
        'slew': ['slew','slew_rate','斜率','slew_rate','sr'],
        'pull': ['pull','pulltype','pull_type','上下拉','pullup'],
    }

    def _detect_columns(header_row):
        """返回 {col_index: field_key} 映射
        匹配策略: 1) 完全相等  2) header 含别名  3) 别名含 header (短别名)
        """
        idx_map = {}
        for i, col_name in enumerate(header_row):
            cn = col_name.strip().lower().replace(' ', '_').replace('-', '_')
            if not cn:
                continue
            for field, aliases in _COL_ALIASES.items():
                if field in idx_map.values():
                    continue
                for a in aliases:
                    an = a.lower().replace(' ', '_')
                    if cn == an or an in cn or cn in an:
                        idx_map[i] = field
                        break
        return idx_map

    # --- 解析 ---
    def _parse_csv(filepath):
        import csv
        with open(filepath, 'r', encoding='utf-8-sig', errors='replace') as f:
            # 检测分隔符
            sample = f.read(8192)
            f.seek(0)
            dialect = csv.Sniffer().sniff(sample, delimiters=',;\t')
            reader = csv.reader(f, dialect)
            header = next(reader, None)
            if not header: return None, None
            rows = [row for row in reader if any(c.strip() for c in row)]
        return header, rows

    def _xlsx_fallback_via_stdlib(filepath):
        """
        纯标准库解析 .xlsx: zipfile + xml.etree.
        当 openpyxl 不可用 (便携/离线环境) 时, _parse_xlsx 会静默回退到这里.
        返回值: full_grid (list[list[str]]) 或 None (文件不是合法 .xlsx).
        设计目标: 输出与 openpyxl 分支的 full_grid 形状完全一致,
                  让后续的合并单元格传播 / 表头探测 / 行过滤逻辑直接复用.
        .xlsx = zip 容器, 内含:
          xl/sharedStrings.xml        共享字符串池
          xl/worksheets/sheet1.xml    单元格; 字符串引用 t="s" + 池下标,
                                       或 t="inlineStr" + <is><t>...</t></is>
        """
        import zipfile
        import xml.etree.ElementTree as ET
        import re as _re

        NS = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
        _CELL_REF = _re.compile(r'^([A-Za-z]+)(\d+)$')

        def _col_idx(letters):
            n = 0
            for ch in letters:
                n = n * 26 + (ord(ch.upper()) - ord('A') + 1)
            return n - 1

        def _ref_to_rc(ref):
            """'B3' -> (row_idx=2, col_idx=1) 全部 0-indexed"""
            m = _CELL_REF.match(ref)
            if not m:
                return None
            return int(m.group(2)) - 1, _col_idx(m.group(1))

        def _cell_text(c):
            """统一提取一个 <c> 元素的文本内容"""
            t = c.get('t', 'n')
            if t == 's':
                # 共享字符串池引用
                v = c.find(f'{NS}v')
                if v is None or v.text is None:
                    return ''
                try:
                    idx = int(v.text)
                except (TypeError, ValueError):
                    return ''
                return shared[idx] if 0 <= idx < len(shared) else ''
            if t == 'inlineStr':
                # <is><t>...</t></is>
                is_el = c.find(f'{NS}is')
                if is_el is None:
                    return ''
                t_el = is_el.find(f'{NS}t')
                return t_el.text if (t_el is not None and t_el.text is not None) else ''
            if t == 'str':
                # 公式字符串结果
                v = c.find(f'{NS}v')
                return v.text if (v is not None and v.text is not None) else ''
            # 数值/布尔/日期/空: 转为字符串
            v = c.find(f'{NS}v')
            if v is None or v.text is None:
                return ''
            return v.text

        try:
            with zipfile.ZipFile(filepath, 'r') as zf:
                names = set(zf.namelist())

                # 1) 共享字符串池
                shared = []
                if 'xl/sharedStrings.xml' in names:
                    root = ET.fromstring(zf.read('xl/sharedStrings.xml'))
                    for si in root.findall(f'{NS}si'):
                        # 一个 si 可能有多个 <t> (富文本拼接)
                        parts = [t_el.text or ''
                                 for t_el in si.iter(f'{NS}t')]
                        shared.append(''.join(parts).strip())

                # 2) 找第一个 sheet 的 xml 路径
                sheet_path = None
                for cand in ('xl/worksheets/sheet1.xml',
                             'xl/worksheets/sheet.xml'):
                    if cand in names:
                        sheet_path = cand
                        break
                if not sheet_path:
                    # 兜底: 扫描 worksheets/ 目录
                    for n in names:
                        if n.startswith('xl/worksheets/sheet') and n.endswith('.xml'):
                            sheet_path = n
                            break
                if not sheet_path:
                    return None

                # 3) 解析 sheet xml
                root = ET.fromstring(zf.read(sheet_path))
                sheet_data = root.find(f'{NS}sheetData')
                if sheet_data is None:
                    return None

                # 收集每个 row 的所有 cell 引用, 同时算出行/列范围
                max_row = 0
                max_col = 0
                rows_meta = []  # [(row_idx, [(col_idx, cell_text), ...]), ...]
                for row in sheet_data.findall(f'{NS}row'):
                    r_attr = row.get('r')
                    if r_attr:
                        try:
                            r_idx = int(r_attr) - 1
                        except ValueError:
                            r_idx = max_row
                    else:
                        r_idx = max_row
                    if r_idx > max_row:
                        max_row = r_idx
                    cells = []
                    for c in row.findall(f'{NS}c'):
                        ref = c.get('r')
                        if not ref:
                            continue
                        rc = _ref_to_rc(ref)
                        if rc is None:
                            continue
                        _, c_idx = rc
                        if c_idx > max_col:
                            max_col = c_idx
                        cells.append((c_idx, _cell_text(c)))
                    rows_meta.append((r_idx, cells))

                # 4) 构造全零填充的二维表
                if max_row < 0 or max_col < 0:
                    return None
                full_grid = [[''] * (max_col + 1) for _ in range(max_row + 1)]
                for r_idx, cells in rows_meta:
                    for c_idx, txt in cells:
                        full_grid[r_idx][c_idx] = (txt or '').strip()

                # 5) 合并单元格值传播 (sheet xml 里 <mergeCells><mergeCell ref="A1:B2"/>)
                merged_value = {}
                mc = root.find(f'{NS}mergeCells')
                if mc is not None:
                    for mr in mc.findall(f'{NS}mergeCell'):
                        ref = mr.get('ref', '')
                        if ':' not in ref:
                            continue
                        start, end = ref.split(':', 1)
                        s = _ref_to_rc(start)
                        e = _ref_to_rc(end)
                        if s is None or e is None:
                            continue
                        r1, c1 = s
                        r2, c2 = e
                        v = (full_grid[r1][c1] if r1 < len(full_grid) and c1 < len(full_grid[r1]) else '')
                        v = (v or '').strip()
                        if not v:
                            continue
                        for rr in range(r1, r2 + 1):
                            for cc in range(c1, c2 + 1):
                                merged_value[(rr, cc)] = v
                for (rr, cc), v in merged_value.items():
                    if 0 <= rr < len(full_grid) and 0 <= cc < len(full_grid[rr]):
                        if not full_grid[rr][cc]:
                            full_grid[rr][cc] = v

                return full_grid
        except Exception:
            # zip 损坏 / xml 非法 / 缺关键文件 → 视为不可解析, 让上层提示
            return None

    def _parse_xlsx(filepath):
        """
        解析 Excel 管脚表 (.xlsx 用 openpyxl, .xls 用 xlrd)
        当 openpyxl 不可用时, 自动回退到内置标准库解析 (_xlsx_fallback_via_stdlib),
        对用户完全透明 — 不会弹 "需安装 openpyxl" 的错误框.
        """
        ext = os.path.splitext(filepath)[1].lower()
        # 把内置提示集中到模块属性, 方便统一维护
        missing = lambda pkg, hint: (
            f'需安装 {pkg} 以读取 {hint} 文件。\n'
            f'  方法 1 (联网):  pip install {pkg}\n'
            f'  方法 2 (离线):  把对应的 .whl 放进 _wheels\\ 目录后重跑 setup_offline.bat\n'
            f'  方法 3 (兜底):  打开 Excel, 「另存为 CSV UTF-8」再用本工具导入'
        )
        if ext == '.xls':
            # .xls 是旧二进制 BIFF 格式, 纯标准库无法解析, 必须用 xlrd
            try:
                import xlrd
            except ImportError:
                return None, missing('xlrd', '.xls')
            wb = xlrd.open_workbook(filepath)
            ws = wb.sheet_by_index(0)
            full_grid = []
            for r in range(ws.nrows):
                full_grid.append([str(ws.cell_value(r, c)).strip() if ws.cell_value(r, c) != '' else ''
                                  for c in range(ws.ncols)])
            if not full_grid:
                return None, 'Excel 文件为空'
            # .xls 没有合并单元格 API, 跳过传播
            n_cols = ws.ncols
        else:
            # .xlsx 是 ZIP 容器, 优先用 openpyxl; 缺失时静默回退到内置 zipfile+xml
            full_grid = None
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
                ws = wb.active
                _ogrid = []
                for row in ws.iter_rows(values_only=True):
                    _ogrid.append([(str(v).strip() if v is not None else '') for v in row])
                if _ogrid:
                    full_grid = _ogrid
                else:
                    wb.close()
                    return None, 'Excel 文件为空'
                _max_col = max((len(r) for r in full_grid), default=0)
                for r in full_grid:
                    if len(r) < _max_col:
                        r.extend([''] * (_max_col - len(r)))
                # 用 openpyxl 的合并单元格 API 做值传播
                _merged_value = {}
                for mr in ws.merged_cells.ranges:
                    try:
                        _r1, _c1, _r2, _c2 = mr.min_row, mr.min_col, mr.max_row, mr.max_col
                    except Exception:
                        continue
                    _v = ws.cell(row=_r1, column=_c1).value
                    _v = str(_v).strip() if _v is not None else ''
                    if not _v:
                        continue
                    for _rr in range(_r1, _r2 + 1):
                        for _cc in range(_c1, _c2 + 1):
                            if 1 <= _rr <= len(full_grid) and 1 <= _cc <= _max_col:
                                _merged_value[(_rr, _cc)] = _v
                for (_rr, _cc), _v in _merged_value.items():
                    if 1 <= _rr <= len(full_grid) and 1 <= _cc <= len(full_grid[_rr - 1]):
                        if not full_grid[_rr - 1][_cc - 1]:
                            full_grid[_rr - 1][_cc - 1] = _v
                wb.close()
            except ImportError:
                # openpyxl 不在 → 静默走内置标准库, 不要弹错误框
                full_grid = _xlsx_fallback_via_stdlib(filepath)
                if full_grid is None:
                    return None, ('无法解析 .xlsx 文件 (zip 结构异常或非标准格式)。\n'
                                  '建议:  打开 Excel, 「另存为 CSV UTF-8」再导入。')
            if not full_grid:
                return None, 'Excel 文件为空'
            n_cols = max((len(r) for r in full_grid), default=0)

        # --- 3) 探测表头行 (在前 30 行内) ---
        _HDR_KW = ('port', 'pin', 'signal', 'name', '管脚', '引脚', '端口', 'iostandard', 'direction')
        header_idx = -1
        probe_limit = min(30, len(full_grid))
        for i in range(probe_limit):
            row = full_grid[i]
            non_empty = [c for c in row if c]
            if len(non_empty) < 2:
                continue  # 合并占位行, 跳过
            if len(set(non_empty)) == 1:
                continue  # 整行同一个值 (行内合并), 跳过
            low = ' '.join(row).lower()
            if any(kw in low for kw in _HDR_KW):
                header_idx = i
                break
        if header_idx < 0:
            # 兜底: 找第一个"非空值种类 >= 3"的行当表头
            for i in range(probe_limit):
                row = full_grid[i]
                non_empty = [c for c in row if c]
                if len(non_empty) >= 3 and len(set(non_empty)) >= 3:
                    header_idx = i
                    break
        if header_idx < 0:
            header_idx = 0  # 最终兜底
        header = full_grid[header_idx]

        # --- 4) 数据行: 表头之后的非空行 ---
        rows = []
        for row in full_grid[header_idx + 1:]:
            if any(c for c in row):
                rows.append(row)

        # --- 5) 排除"组标题行" (整行同一值) 与 中文描述行 ---
        filtered = []
        for row in rows:
            non_empty = [c for c in row if c]
            if not non_empty:
                continue
            unique = set(non_empty)
            if len(unique) == 1:
                # 整行同一个值 → 合并单元格组标题, 跳过
                continue
            # 首列含中文 + 不像英文标识符 → 描述/分组行, 跳过
            first = non_empty[0]
            if re.search(r'[\u4e00-\u9fff]', first) and not re.match(r'^[A-Za-z_][\w\[\]:]*$', first):
                continue
            filtered.append(row)
        rows = filtered

        if ext != '.xls':
            wb.close()
        return header, rows

    # --- XDC 生成 ---
    def _gen_xdc(pins_data, opts):
        """
        pins_data: list[{port, pin, dir, iostd, bank, drive, slew, pull}]
        opts: {drive, slew, pull}
        返回 (xdc_str, bank_warnings)
        """
        lines = []
        lines.append('# ====== Auto-generated by FPGA Toolbox ======')
        lines.append(f'# {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
        lines.append('')

        # 按方向分组
        groups = {'IN': [], 'OUT': [], 'INOUT': [], '': []}
        for p in pins_data:
            d_raw = (p.get('dir') or '').strip().upper()
            # 归一化: input→IN, output→OUT, bidir/inout/io→INOUT
            d_norm = {'INPUT': 'IN', 'OUTPUT': 'OUT',
                      'INOUT': 'INOUT', 'BIDIR': 'INOUT', 'BIDIRECTIONAL': 'INOUT',
                      'BI': 'INOUT', 'IO': 'INOUT', 'I/O': 'INOUT', 'IN/OUT': 'INOUT',
                      'I': 'IN', 'O': 'OUT'}.get(d_raw, d_raw)
            if d_norm in groups:
                groups[d_norm].append(p)
            else:
                groups[''].append(p)

        for grp_name in ('IN', 'OUT', 'INOUT', ''):
            grp = groups[grp_name]
            if not grp: continue
            section = {'IN':'Input Ports','OUT':'Output Ports',
                       'INOUT':'Inout Ports','':'Other Ports'}[grp_name]
            lines.append(f'# --- {section} ---')
            lines.append('')
            for p in grp:
                port = p.get('port','')
                pin  = p.get('pin','')
                iostd = p.get('iostd','')
                if not port or not pin: continue
                lines.append(f'# {port}  (dir={grp_name or "?"})')
                lines.append(f'set_property PACKAGE_PIN {pin} [get_ports {{{port}}}]')
                if iostd:
                    lines.append(f'set_property IOSTANDARD {iostd} [get_ports {{{port}}}]')
                if opts.get('drive') and p.get('drive'):
                    lines.append(f'set_property DRIVE {p["drive"]} [get_ports {{{port}}}]')
                if opts.get('slew') and p.get('slew'):
                    lines.append(f'set_property SLEW {p["slew"]} [get_ports {{{port}}}]')
                if opts.get('pull') and p.get('pull'):
                    lines.append(f'set_property PULLTYPE {p["pull"]} [get_ports {{{port}}}]')
                lines.append('')

        # Bank 电压检查
        bank_warnings = []
        bank_map = {}  # bank → [(port, iostd, vcco)]
        for p in pins_data:
            bank = p.get('bank','').strip()
            iostd = p.get('iostd','').strip()
            if not bank or not iostd: continue
            v, _ = _get_vcco(iostd)
            bank_map.setdefault(bank, []).append((p.get('port',''), iostd, v))

        for bank, entries in sorted(bank_map.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 999):
            vccos = set(e[2] for e in entries if e[2] is not None)
            if not vccos: continue
            if len(vccos) > 1:
                detail = ', '.join(f'{p}({s}→{v}V)' for p, s, v in entries[:8])
                bank_warnings.append(f'⚠ Bank {bank}: 电压冲突! {detail}')
            else:
                v = list(vccos)[0]
                bank_warnings.append(f'✅ Bank {bank}: 统一 {v}V')

        return '\n'.join(lines), bank_warnings

    # --- Tab 8 UI (子页 1: 管脚约束) ---
    # Row 0 — 文件选择
    fc8 = ttk.LabelFrame(t8_a1, text=' 管脚表文件 ', )
    fc8.grid(row=0, column=0, sticky='ew', padx=12, pady=(10, 4))
    fc8.grid_columnconfigure(1, weight=1)

    t8_file = tk.StringVar()
    ttk.Entry(fc8, textvariable=t8_file, font=(F, 10)).grid(
        row=0, column=0, sticky='ew', padx=(14, 4), pady=(8, 4), columnspan=2)
    ttk.Button(fc8, text='浏览',
               command=lambda: (f := filedialog.askopenfilename(
                   title='选择管脚表',
                   filetypes=[('表格文件','*.csv *.xlsx'),
                              ('CSV','*.csv'),('Excel','*.xlsx'),
                              ('All','*.*')])) and t8_file.set(f),
               style='Normal.TButton').grid(row=0, column=2, padx=(4, 4), pady=(8, 4))
    ttk.Button(fc8, text='解析', command=lambda: _t8_parse(),
               style='Accent.TButton').grid(row=0, column=3,
               padx=(0, 14), pady=(8, 4))

    # Row 1 — 选项
    fc8o = ttk.LabelFrame(t8_a1, text=' 生成选项 ', )
    fc8o.grid(row=1, column=0, sticky='ew', padx=12, pady=4)

    v8_drive = tk.BooleanVar(value=False)
    v8_slew  = tk.BooleanVar(value=False)
    v8_pull  = tk.BooleanVar(value=False)
    v8_bankcheck = tk.BooleanVar(value=False)
    _chk_row = ttk.Frame(fc8o, style='TFrame')
    _chk_row.pack(fill='x', padx=14, pady=(6, 6))
    for var, lbl in [(v8_drive,'DRIVE'),(v8_slew,'SLEW'),(v8_pull,'PULLTYPE'),
                     (v8_bankcheck,'Bank 电压检查')]:
        ttk.Checkbutton(_chk_row, text=lbl, variable=var,
                        style='TCheckbutton').pack(side='left', padx=(0, 10))


    # Row 2 — 预览表头
    ttk.Label(t8_a1, text='管脚表预览 (解析后自动显示)', foreground=C['sub'],
              font=(F, 9, 'bold')).grid(row=2, column=0, sticky='w',
              padx=16, pady=(6, 2))

    _t8_all_checked = True  # 初始全选
    # _t8_toggle_all 在后面按钮区定义 (需要 t8_tree 先存在)

    # Row 3 — PanedWindow (可拖拽调整预览/XDC分界)
    t8_pane = ttk.PanedWindow(t8_a1, orient='vertical')
    t8_pane.grid(row=3, column=0, sticky='nsew', padx=12, pady=(0, 4))
    # 上半: 管脚表预览 + Bank 日志 + 按钮
    t8_top = ttk.Frame(t8_pane, style='TFrame')
    t8_top.grid_rowconfigure(0, weight=1)  # tree
    t8_top.grid_rowconfigure(1, weight=0)  # bank
    t8_top.grid_rowconfigure(2, weight=0)  # buttons
    t8_top.grid_columnconfigure(0, weight=1)
    t8_pane.add(t8_top, weight=2)
    # 下半: XDC 预览
    t8_bot = ttk.Frame(t8_pane, style='TFrame')
    t8_bot.grid_rowconfigure(0, weight=1)
    t8_bot.grid_columnconfigure(0, weight=1)
    t8_pane.add(t8_bot, weight=1)

    # ── 上半: Treeview ──
    _T8_COLS = ('_sel','port','pin','dir','iostd','bank','drive','slew','pull')
    _T8_NAMES = {'_sel':'选择','port':'Port','pin':'Pin','dir':'Dir','iostd':'IOSTD',
                 'bank':'Bank','drive':'Drive','slew':'Slew','pull':'Pull'}
    t8_tree = ttk.Treeview(t8_top, columns=_T8_COLS, show='headings',
                           selectmode='browse', style='Treeview')
    for c in _T8_COLS:
        t8_tree.heading(c, text=_T8_NAMES.get(c, c))
        t8_tree.column(c, width=90, minwidth=60)
    t8_tree.column('_sel', width=44, minwidth=40)
    t8_tree.column('port', width=160)
    t8_tree.column('pin', width=100)
    t8_tree.column('iostd', width=120)
    t8_tree.grid(row=0, column=0, sticky='nsew')

    # ── 列宽拖拽状态 ──
    _t8_drag_col = None
    _t8_drag_start_x = 0
    _t8_drag_start_w = 0
    def _t8_get_col_edges():
        edges = {}
        x = 0
        for c in t8_tree['columns']:
            w = t8_tree.column(c, 'width')
            edges[c] = (x, x + w)
            x += w
        return edges
    def _t8_col_press(event):
        nonlocal _t8_drag_col, _t8_drag_start_x, _t8_drag_start_w
        if t8_tree.identify_region(event.x, event.y) != 'heading':
            return
        col = t8_tree.identify_column(event.x)
        if not col:
            return
        edges = _t8_get_col_edges()
        if col not in edges:
            return
        left, right = edges[col]
        if abs(event.x - right) > 6:
            return
        _t8_drag_col = col
        _t8_drag_start_x = event.x_root
        _t8_drag_start_w = t8_tree.column(col, 'width')
        t8_tree.configure(cursor='sb_h_double_arrow')
    def _t8_col_move(event):
        nonlocal _t8_drag_col, _t8_drag_start_w
        if _t8_drag_col:
            new_w = max(20, _t8_drag_start_w + event.x_root - _t8_drag_start_x)
            t8_tree.column(_t8_drag_col, width=new_w, minwidth=20)
    def _t8_col_release(event):
        nonlocal _t8_drag_col
        _t8_drag_col = None
        t8_tree.configure(cursor='')

    # 复选框切换 — release 避免和拖拽冲突
    def _t8_check_toggle(event):
        nonlocal _t8_drag_col
        if _t8_drag_col:
            return
        if t8_tree.identify_column(event.x) != '#1':
            return
        item = t8_tree.identify_row(event.y)
        if not item:
            return
        # locate pin by item ID
        children = t8_tree.get_children()
        if item not in children:
            return
        idx = children.index(item)
        if idx < len(t8_parsed_pins):
            pin = t8_parsed_pins[idx]
            pin['_sel'] = not pin.get('_sel', True)
            t8_tree.set(item, '_sel', '☑' if pin['_sel'] else '☐')
            # 更新状态栏选中计数
            sel_cnt = sum(1 for p in t8_parsed_pins if p.get('_sel', True))
            total = len(t8_parsed_pins)
            t8_status.set(f'已解析 {total} 个引脚（已选 {sel_cnt}/{total}）')
    t8_tree.bind('<ButtonRelease-1>', _t8_check_toggle)
    t8_tree.bind('<ButtonPress-1>', _t8_col_press, add='+')
    t8_tree.bind('<B1-Motion>', _t8_col_move, add='+')
    t8_tree.bind('<ButtonRelease-1>', _t8_col_release, add='+')

    t8_scroll = ttk.Scrollbar(t8_top, orient='vertical', command=t8_tree.yview)
    t8_scroll.grid(row=0, column=1, sticky='ns')
    t8_tree.configure(yscrollcommand=t8_scroll.set)

    # Row 1 — Bank 检查日志
    t8_bank_lf, t8_bank_log = _log_widget(t8_top, 4)
    t8_bank_lf.grid(row=1, column=0, sticky='nsew', pady=(2, 2))

    # ── 全选/取消全选 (需在 t8_tree 和 t8_parsed_pins 之后, f8btn 之前) ──
    def _t8_toggle_all():
        nonlocal _t8_all_checked
        _t8_all_checked = not _t8_all_checked
        mark = '☑' if _t8_all_checked else '☐'
        _t8_select_all_btn.config(text='☐ 取消全选' if _t8_all_checked else '☑ 全选')
        children = t8_tree.get_children()
        for i, pin in enumerate(t8_parsed_pins):
            pin['_sel'] = _t8_all_checked
            if i < len(children):
                t8_tree.set(children[i], '_sel', mark)
        total = len(t8_parsed_pins)
        sel_cnt = total if _t8_all_checked else 0
        t8_status.set(f'已解析 {total} 个引脚（已选 {sel_cnt}/{total}）')

    # Row 2 — 操作按钮 (全选在最前)
    f8btn = ttk.Frame(t8_top, style='TFrame')
    f8btn.grid(row=2, column=0, sticky='nsew', pady=4)
    _t8_select_all_btn = ttk.Button(f8btn, text='☐ 取消全选',
        command=_t8_toggle_all, style='Small.TButton')
    _t8_select_all_btn.pack(side='left', padx=(0, 10))
    ttk.Separator(f8btn, orient='vertical').pack(side='left', fill='y', padx=4)
    ttk.Button(f8btn, text='⚙  生成 XDC', command=lambda: _t8_gen_xdc(),
               style='Accent.TButton').pack(side='left', padx=(6, 6))
    ttk.Button(f8btn, text='💾  保存为 .xdc', command=lambda: _t8_save_xdc(),
               style='Success.TButton').pack(side='left', padx=(0, 6))
    ttk.Button(f8btn, text='📋  复制', command=lambda: _t8_copy_xdc(),
               style='Info.TButton').pack(side='left', padx=(0, 6))
    ttk.Button(f8btn, text='📤  导出更多...', command=lambda: _t8_export_menu(),
               style='Normal.TButton').pack(side='left', padx=(0, 6))
    t8_status = tk.StringVar(value='就绪')
    ttk.Label(f8btn, textvariable=t8_status, foreground=C['sub'],
              font=(F, 9)).pack(side='right', padx=(20, 0))

    # ── 下半: XDC 预览 ──
    t8_xdc_frame = ttk.LabelFrame(t8_bot, text=' XDC 预览 ', )
    t8_xdc_frame.grid(row=0, column=0, sticky='nsew')
    t8_xdc_frame.grid_rowconfigure(0, weight=1)
    t8_xdc_frame.grid_columnconfigure(0, weight=1)
    t8_xdc_text = tk.Text(t8_xdc_frame, font=(M, 9), bg=C['ebg'], fg=C['fg'],
                          insertbackground=C['fg'], relief='flat',
                          padx=10, pady=8, wrap='none')
    t8_xdc_text.grid(row=0, column=0, sticky='nsew')
    t8_xh_scroll = ttk.Scrollbar(t8_xdc_frame, orient='horizontal',
                                 command=t8_xdc_text.xview)
    t8_xh_scroll.grid(row=1, column=0, sticky='ew')
    t8_xdc_text.configure(xscrollcommand=t8_xh_scroll.set)
    # 阻止编辑但允许选择
    t8_xdc_text.bind('<Key>', lambda e: 'break')

    # --- 数据 ---
    t8_parsed_pins = []   # 解析后的引脚数据 [{port,pin,dir,iostd,...}, ...]
    t8_active_cols = list(_T8_COLS)  # 当前输入文件中实际存在的列
    t8_xdc_output = ''    # 生成的 XDC 文本

    def _t8_parse():
        """解析管脚表文件"""
        nonlocal t8_xdc_output, t8_active_cols, _t8_all_checked
        t8_parsed_pins.clear()
        t8_active_cols = list(_T8_COLS)
        t8_tree.delete(*t8_tree.get_children())
        t8_xdc_text.delete('1.0', 'end')
        t8_xdc_output = ''
        t8_bank_log.delete('1.0', 'end')

        fp = t8_file.get().strip()
        if not fp or not os.path.isfile(fp):
            messagebox.showerror('错误', '请选择有效的管脚表文件')
            return
        ext = os.path.splitext(fp)[1].lower()
        try:
            if ext == '.csv':
                header, rows = _parse_csv(fp)
                if header is None:
                    messagebox.showerror('错误', 'CSV 文件为空或格式不正确')
                    return
                if isinstance(rows, str):
                    messagebox.showerror('错误', rows)
                    return
            elif ext in ('.xlsx', '.xls'):
                result = _parse_xlsx(fp)
                if isinstance(result, tuple) and len(result) == 2 and result[0] is None:
                    messagebox.showerror('Excel 读取失败', result[1])
                    return
                header, rows = result
                if header is None:
                    messagebox.showerror('错误', 'Excel 文件为空或格式不正确')
                    return
            else:
                messagebox.showerror('错误', f'不支持的文件类型: {ext}\n请使用 .csv 或 .xlsx')
                return
        except Exception as e:
            messagebox.showerror('解析失败', str(e))
            return

        # 检测列
        col_map = _detect_columns(header)
        if 'port' not in col_map.values() or 'pin' not in col_map.values():
            t8_status.set(f'解析失败: 未识别到 Port/Pin 列。表头: {", ".join(header[:6])}')
            return

        # ── 根据输入表实际列, 动态识别活动列 ──
        t8_active_cols = ['_sel'] + [f for f in _T8_COLS if f in col_map.values() and f != '_sel']
        active_cols = t8_active_cols

        # 动态配置 Treeview 列 (按文件实际列, _sel 固定在首列)
        t8_tree.configure(columns=active_cols)
        for c in active_cols:
            t8_tree.heading(c, text=_T8_NAMES.get(c, c))
            t8_tree.column(c, width=90, minwidth=60)
        _T8_WIDE = {'port': 160, 'pin': 100, 'iostd': 120}
        for c, w in _T8_WIDE.items():
            if c in active_cols:
                t8_tree.column(c, width=w)
        # _sel 列窄
        if '_sel' in active_cols:
            t8_tree.column('_sel', width=44, minwidth=40)

        # 重置全选状态
        nonlocal _t8_all_checked
        _t8_all_checked = True
        _t8_select_all_btn.config(text='☐ 取消全选')

        # 映射行数据 (组标题行已在 _parse_xlsx 内被排除)
        for row in rows:
            entry = {f: '' for f in active_cols if f != '_sel'}
            for ci, field in col_map.items():
                if ci < len(row) and field in entry:
                    entry[field] = row[ci].strip()
            if entry['port'] and entry['pin']:
                entry['_sel'] = True  # 默认选中
                t8_parsed_pins.append(entry)
                t8_tree.insert('', 'end', values=['☑'] + [entry.get(c, '') for c in active_cols if c != '_sel'])

        t8_status.set(f'已解析 {len(t8_parsed_pins)} 个引脚（已选 {len(t8_parsed_pins)}/{len(t8_parsed_pins)}）')
        t8_bank_log.delete('1.0', 'end')

    def _t8_gen_xdc():
        """生成 XDC + Bank 检查 (仅生成勾选的引脚)"""
        nonlocal t8_xdc_output
        t8_xdc_output = ''
        t8_xdc_text.delete('1.0', 'end')
        t8_bank_log.delete('1.0', 'end')

        if not t8_parsed_pins:
            messagebox.showinfo('提示', '请先解析管脚表')
            return

        # 只生成勾选的引脚
        selected = [p for p in t8_parsed_pins if p.get('_sel', True)]
        if not selected:
            messagebox.showinfo('提示', '没有选中任何引脚，请先在左侧勾选')
            return

        opts = {'drive': v8_drive.get(), 'slew': v8_slew.get(),
                'pull': v8_pull.get()}
        xdc, warnings = _gen_xdc(selected, opts)
        t8_xdc_output = xdc
        t8_xdc_text.insert('1.0', xdc)

        if v8_bankcheck.get() and warnings:
            for w in warnings:
                color = C['red'] if '⚠' in w else C['green']
                _log(t8_bank_log, w, color)
        elif v8_bankcheck.get():
            _log(t8_bank_log, '✅ 无 Bank 电压冲突（或缺少 Bank 列信息）', C['green'])

        t8_status.set(f'XDC 已生成 ({len(selected)}/{len(t8_parsed_pins)} 个端口)')

    def _t8_save_xdc():
        if not t8_xdc_output:
            messagebox.showinfo('提示', '请先生成 XDC')
            return
        fp = filedialog.asksaveasfilename(
            title='保存 XDC 文件',
            defaultextension='.xdc',
            filetypes=[('XDC 文件','*.xdc'),('All','*.*')],
            initialfile=os.path.splitext(os.path.basename(
                t8_file.get() or 'pinout'))[0] + '.xdc')
        if fp:
            try:
                with open(fp, 'w', encoding='utf-8') as f:
                    f.write(t8_xdc_output)
                t8_status.set(f'已保存: {fp}')
            except OSError as e:
                messagebox.showerror('保存失败', str(e))

    def _t8_copy_xdc():
        """一键复制 XDC 到剪贴板"""
        if not t8_xdc_output:
            messagebox.showinfo('提示', '请先生成 XDC')
            return
        root.clipboard_clear()
        root.clipboard_append(t8_xdc_output)
        root.update()
        t8_status.set(f'✔ 已复制 {len(t8_xdc_output)} 字符到剪贴板')

    def _t8_export_menu():
        """导出更多选项弹窗"""
        if not t8_xdc_output:
            messagebox.showinfo('提示', '请先生成 XDC')
            return
        if not t8_parsed_pins:
            messagebox.showinfo('提示', '请先解析管脚表')
            return

        dlg = tk.Toplevel(root)
        dlg.title('导出选项')
        dlg.resizable(False, False)
        dlg.transient(root)
        dlg.grab_set()
        dlg.configure(bg=C['card'])
        dlg.geometry('540x440')

        # 标题
        header = tk.Frame(dlg, bg=C['blue'])
        header.pack(fill='x')
        tk.Label(header, text='📤  导出管脚约束', bg=C['blue'], fg='#ffffff',
                 font=(F, 12, 'bold'),
                 padx=20, pady=12).pack(anchor='w')

        body = tk.Frame(dlg, bg=C['card'])
        body.pack(fill='both', expand=True, padx=24, pady=16)

        tk.Label(body, text='选择导出格式:', bg=C['card'], fg=C['fg'],
                 font=(F, 10, 'bold')).pack(anchor='w', pady=(0, 10))

        def _opt_btn(text, desc, cmd, primary=False):
            f = tk.Frame(body, bg=C['card'])
            f.pack(fill='x', pady=4)
            btn = ttk.Button(f, text=text, command=cmd,
                             style='Accent.TButton' if primary else 'Normal.TButton')
            btn.pack(side='left', padx=(0, 12))
            tk.Label(f, text=desc, bg=C['card'], fg=C['sub'],
                     font=(F, 9)).pack(side='left')

        _opt_btn('📄  保存为 .xdc',
                 'Vivado 约束文件, 标准格式',
                 lambda: (dlg.destroy(), _t8_save_xdc()),
                 primary=True)
        _opt_btn('📜  保存为 .tcl',
                 'Vivado Tcl 脚本 (source xdc.tcl 可加载)',
                 lambda: (dlg.destroy(), _t8_save_tcl()))
        _opt_btn('📊  保存为 .csv',
                 '管脚表 (回写, 便于工程归档)',
                 lambda: (dlg.destroy(), _t8_save_csv()))
        _opt_btn('🏗  导出到 Vivado 工程',
                 '直接写入工程的 src/constrs 目录',
                 lambda: (dlg.destroy(), _t8_export_to_vivado()))
        _opt_btn('📋  生成完整报告 (HTML)',
                 '管脚表 + Bank 检查 + XDC 一体化',
                 lambda: (dlg.destroy(), _t8_export_html_report()))
        _opt_btn('📑  生成完整报告 (Markdown)',
                 '适合贴入 GitLab / 文档',
                 lambda: (dlg.destroy(), _t8_export_md_report()))

        # 底部关闭
        ttk.Button(dlg, text='关闭', command=dlg.destroy,
                   style='Normal.TButton').pack(pady=(0, 14))

        dlg.update_idletasks()
        x = root.winfo_x() + (root.winfo_width() - dlg.winfo_width()) // 2
        y = root.winfo_y() + (root.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f'+{x}+{y}')

    def _t8_save_tcl():
        """导出为 Vivado Tcl 脚本 (可直接 source)"""
        fp = filedialog.asksaveasfilename(
            title='保存 TCL 脚本',
            defaultextension='.tcl',
            filetypes=[('Tcl 脚本', '*.tcl'), ('All', '*.*')],
            initialfile=os.path.splitext(os.path.basename(
                t8_file.get() or 'pinout'))[0] + '.tcl')
        if not fp:
            return
        try:
            tcl = _gen_tcl_wrapper(t8_xdc_output, t8_parsed_pins)
            with open(fp, 'w', encoding='utf-8') as f:
                f.write(tcl)
            t8_status.set(f'✔ TCL 已保存: {os.path.basename(fp)}')
        except OSError as e:
            messagebox.showerror('保存失败', str(e))

    def _t8_save_csv():
        """导出回 CSV 格式 (规范化)"""
        fp = filedialog.asksaveasfilename(
            title='保存管脚表 (CSV)',
            defaultextension='.csv',
            filetypes=[('CSV', '*.csv'), ('All', '*.*')],
            initialfile=os.path.splitext(os.path.basename(
                t8_file.get() or 'pinout'))[0] + '_export.csv')
        if not fp:
            return
        try:
            with open(fp, 'w', encoding='utf-8-sig', newline='') as f:
                w = csv.writer(f)
                w.writerow(t8_active_cols)
                for p in t8_parsed_pins:
                    w.writerow([p.get(c, '') for c in t8_active_cols])
            t8_status.set(f'✔ CSV 已保存: {os.path.basename(fp)}')
        except OSError as e:
            messagebox.showerror('保存失败', str(e))

    def _t8_export_to_vivado():
        """导出到 Vivado 工程的 src/constrs 目录"""
        # 1. 选 Vivado 工程根目录
        proj_dir = filedialog.askdirectory(title='选择 Vivado 工程根目录 (含 .xpr)')
        if not proj_dir:
            return
        # 2. 智能识别 conctr/constraints 目录
        cand_dirs = []
        for sub in ['src', 'constrs', 'constraints', 'sources/constrs']:
            d = os.path.join(proj_dir, sub)
            if os.path.isdir(d):
                cand_dirs.append(d)
        # 3. 让用户选
        target_dir = None
        if len(cand_dirs) == 1:
            target_dir = cand_dirs[0]
        elif len(cand_dirs) > 1:
            dlg = tk.Toplevel(root)
            dlg.title('选择约束目录')
            dlg.resizable(False, False)
            dlg.transient(root)
            dlg.grab_set()
            dlg.configure(bg=C['card'])
            tk.Label(dlg, text='找到多个约束目录, 选择写入位置:',
                     bg=C['card'], fg=C['fg'], font=(F, 10),
                     padx=20, pady=(14, 8)).pack(anchor='w')
            sel = tk.StringVar(value=cand_dirs[0])
            for d in cand_dirs:
                rel = os.path.relpath(d, proj_dir)
                ttk.Radiobutton(dlg, text=rel, variable=sel, value=d).pack(
                    anchor='w', padx=28, pady=2)
            ttk.Button(dlg, text='确定',
                       command=lambda: (setattr(dlg, '_sel', sel.get()),
                                        dlg.destroy()),
                       style='Accent.TButton').pack(pady=12)
            dlg.update_idletasks()
            x = root.winfo_x() + (root.winfo_width() - dlg.winfo_width()) // 2
            y = root.winfo_y() + (root.winfo_height() - dlg.winfo_height()) // 2
            dlg.geometry(f'+{x}+{y}')
            dlg.wait_window()
            target_dir = getattr(dlg, '_sel', None)
        else:
            # 找不到, 直接用工程根
            target_dir = proj_dir
        if not target_dir:
            return
        # 4. 写入
        bn = os.path.splitext(os.path.basename(
            t8_file.get() or 'pinout'))[0] + '.xdc'
        fp = os.path.join(target_dir, bn)
        try:
            with open(fp, 'w', encoding='utf-8') as f:
                f.write(t8_xdc_output)
            t8_status.set(f'✔ 已写入 Vivado 工程: {os.path.relpath(fp, proj_dir)}')
            messagebox.showinfo('导出成功',
                                f'XDC 文件已写入:\n{fp}\n\n'
                                f'提示: 在 Vivado 中执行:\n'
                                f'  source {bn}\n'
                                f'或直接 add_files 添加到工程。')
        except OSError as e:
            messagebox.showerror('保存失败', str(e))

    def _t8_export_html_report():
        """导出完整 HTML 报告"""
        fp = filedialog.asksaveasfilename(
            title='保存 HTML 报告',
            defaultextension='.html',
            filetypes=[('HTML', '*.html'), ('All', '*.*')],
            initialfile=os.path.splitext(os.path.basename(
                t8_file.get() or 'pinout'))[0] + '_report.html')
        if not fp:
            return
        try:
            html = _build_html_report(
                t8_parsed_pins, t8_xdc_output,
                t8_bank_log.get('1.0', 'end'),
                t8_file.get())
            with open(fp, 'w', encoding='utf-8') as f:
                f.write(html)
            t8_status.set(f'✔ HTML 报告已保存: {os.path.basename(fp)}')
            if messagebox.askyesno('打开报告', '已生成 HTML 报告, 是否打开?'):
                try:
                    os.startfile(fp)  # Windows
                except AttributeError:
                    import subprocess
                    subprocess.Popen(['xdg-open', fp])  # Linux/Mac
        except OSError as e:
            messagebox.showerror('保存失败', str(e))

    def _t8_export_md_report():
        """导出 Markdown 报告"""
        fp = filedialog.asksaveasfilename(
            title='保存 Markdown 报告',
            defaultextension='.md',
            filetypes=[('Markdown', '*.md'), ('All', '*.*')],
            initialfile=os.path.splitext(os.path.basename(
                t8_file.get() or 'pinout'))[0] + '_report.md')
        if not fp:
            return
        try:
            md = _build_md_report(
                t8_parsed_pins, t8_xdc_output,
                t8_bank_log.get('1.0', 'end'),
                t8_file.get())
            with open(fp, 'w', encoding='utf-8') as f:
                f.write(md)
            t8_status.set(f'✔ MD 报告已保存: {os.path.basename(fp)}')
        except OSError as e:
            messagebox.showerror('保存失败', str(e))

    # ═══════════════════════════════════
    # 子页 2 — 时序约束 (主时钟 / 衍生时钟 / 异步)
    # ═══════════════════════════════════
    t8_a2 = ttk.Frame(t8_nb, style='TFrame')
    t8_nb.add(t8_a2, text='  ⏱ 时序约束  ')
    t8_a2.grid_columnconfigure(0, weight=1)
    # 与子页1对称: 按钮 row (2) 之后, 预览 row (3) 占剩余空间
    t8_a2.grid_rowconfigure(0, weight=0, minsize=0)  # 主时钟 (按内容)
    t8_a2.grid_rowconfigure(1, weight=0, minsize=0)  # 衍生时钟 (按内容)
    t8_a2.grid_rowconfigure(2, weight=0, minsize=0)  # 按钮 (按内容, 不被拉伸)
    t8_a2.grid_rowconfigure(3, weight=1)              # 预览 (主要, 占剩余空间)

    # 主时钟列表 (可多行添加)
    t8_clocks = []  # [{name, period_ns, waveform}, ...]

    fc2a = ttk.LabelFrame(t8_a2, text=' 主时钟配置 (输入多个 create_clock) ')
    fc2a.grid(row=0, column=0, sticky='ew', padx=12, pady=(10, 4))
    fc2a.grid_columnconfigure(0, weight=1)

    # 时钟列表 (Treeview)
    t8_clk_list = ttk.Treeview(fc2a,
                                columns=('name', 'period', 'waveform', 'source'),
                                show='headings', height=2)
    t8_clk_list.heading('name',     text='时钟名')
    t8_clk_list.heading('period',   text='周期 (ns)')
    t8_clk_list.heading('waveform', text='波形 (0,5)')
    t8_clk_list.heading('source',   text='源端 / 对象')
    t8_clk_list.column('name',     width=120, anchor='w')
    t8_clk_list.column('period',   width=90,  anchor='center')
    t8_clk_list.column('waveform', width=120, anchor='center')
    t8_clk_list.column('source',   width=180, anchor='w')
    t8_clk_list.grid(row=0, column=0, sticky='nsew', padx=(8, 4), pady=(8, 4))

    t8_clk_scr = ttk.Scrollbar(fc2a, orient='vertical', command=t8_clk_list.yview)
    t8_clk_scr.grid(row=0, column=1, sticky='ns', pady=(8, 4))
    t8_clk_list.configure(yscrollcommand=t8_clk_scr.set)

    # 输入区
    fi2 = ttk.Frame(fc2a, style='TFrame')
    fi2.grid(row=1, column=0, columnspan=2, sticky='ew', padx=8, pady=(0, 8))
    ttk.Label(fi2, text='名称').grid(row=0, column=0, padx=(0, 4), pady=4)
    t8_clk_name = tk.StringVar(value='sys_clk')
    ttk.Entry(fi2, textvariable=t8_clk_name, width=14).grid(row=0, column=1, padx=(0, 8))
    ttk.Label(fi2, text='周期(ns)').grid(row=0, column=2, padx=(0, 4))
    t8_clk_period = tk.StringVar(value='10.0')
    ttk.Entry(fi2, textvariable=t8_clk_period, width=8).grid(row=0, column=3, padx=(0, 8))
    ttk.Label(fi2, text='波形').grid(row=0, column=4, padx=(0, 4))
    t8_clk_wave = tk.StringVar(value='0,5')
    ttk.Entry(fi2, textvariable=t8_clk_wave, width=8).grid(row=0, column=5, padx=(0, 8))
    ttk.Label(fi2, text='源端').grid(row=0, column=6, padx=(0, 4))
    t8_clk_src = tk.StringVar(value='[get_ports clk]')
    ttk.Entry(fi2, textvariable=t8_clk_src, width=22).grid(row=0, column=7, padx=(0, 8))

    def _t8_add_clk(*_):
        nm = t8_clk_name.get().strip()
        pd_ = t8_clk_period.get().strip()
        wv = t8_clk_wave.get().strip()
        sc = t8_clk_src.get().strip()
        if not nm or not pd_:
            messagebox.showerror('错误', '时钟名和周期不能为空')
            return
        t8_clocks.append({'name': nm, 'period_ns': pd_, 'waveform': wv, 'source': sc})
        t8_clk_list.insert('', 'end', values=(nm, pd_, wv, sc))
        t8_clk_name.set(f'clk{len(t8_clocks) + 1}')

    def _t8_del_clk():
        sel = t8_clk_list.selection()
        if not sel:
            return
        idx = t8_clk_list.index(sel[0])
        t8_clk_list.delete(sel[0])
        if 0 <= idx < len(t8_clocks):
            t8_clocks.pop(idx)

    ttk.Button(fi2, text='➕ 添加', command=_t8_add_clk, style='Accent.TButton').grid(
        row=0, column=8, padx=(8, 4))
    ttk.Button(fi2, text='🗑 删除', command=_t8_del_clk, style='Normal.TButton').grid(
        row=0, column=9, padx=(0, 4))

    # 衍生时钟 + 异步选项
    fc2b = ttk.LabelFrame(t8_a2, text=' 衍生时钟 / 异步组 / 跨时钟域 ')
    fc2b.grid(row=1, column=0, sticky='ew', padx=12, pady=(4, 4))
    fc2b.grid_columnconfigure(1, weight=1)

    v8_gen_dcm = tk.BooleanVar(value=True)   # generate_clock

    ttk.Checkbutton(fc2b, text='启用 generate_clock (衍生时钟)', variable=v8_gen_dcm,
                    style='TCheckbutton').grid(row=0, column=0, columnspan=2, sticky='w',
                                               padx=(8, 4), pady=(6, 4))

    # ── 衍生时钟参数配置 ──
    # 来源主时钟 / 衍生名 / multiply / divide / source pin / edges
    t8_gen_list = ttk.Treeview(fc2b,
                               columns=('name', 'src_clk', 'mult', 'div', 'pin', 'edges'),
                               show='headings', height=1)
    t8_gen_list.heading('name', text='衍生名')
    t8_gen_list.heading('src_clk', text='源时钟')
    t8_gen_list.heading('mult', text='×')
    t8_gen_list.heading('div', text='÷')
    t8_gen_list.heading('pin', text='源引脚')
    t8_gen_list.heading('edges', text='边沿')
    for c, w in [('name', 100), ('src_clk', 90), ('mult', 50), ('div', 50),
                 ('pin', 130), ('edges', 80)]:
        t8_gen_list.column(c, width=w, anchor='center')
    t8_gen_list.grid(row=1, column=0, columnspan=2, sticky='ew', padx=8, pady=(2, 2))

    # 默认示例: sys_clk → sys_clk_div2 (multiply=1, divide=2)
    t8_gen_clocks = []  # [{name, src, mult, div, pin, edges}, ...]

    fi2b = ttk.Frame(fc2b, style='TFrame')
    fi2b.grid(row=2, column=0, columnspan=2, sticky='ew', padx=8, pady=(2, 2))
    for i in range(10):
        fi2b.grid_columnconfigure(i, weight=0)
    ttk.Label(fi2b, text='名:').grid(row=0, column=0, padx=(0, 2))
    t8_gen_name = tk.StringVar(value='sys_clk_div2')
    ttk.Entry(fi2b, textvariable=t8_gen_name, width=12).grid(row=0, column=1, padx=(0, 6))
    ttk.Label(fi2b, text='源:').grid(row=0, column=2, padx=(0, 2))
    t8_gen_src = tk.StringVar(value='sys_clk')
    ttk.Entry(fi2b, textvariable=t8_gen_src, width=10).grid(row=0, column=3, padx=(0, 6))
    ttk.Label(fi2b, text='×').grid(row=0, column=4, padx=(0, 2))
    t8_gen_mult = tk.StringVar(value='1')
    ttk.Entry(fi2b, textvariable=t8_gen_mult, width=4).grid(row=0, column=5, padx=(0, 6))
    ttk.Label(fi2b, text='÷').grid(row=0, column=6, padx=(0, 2))
    t8_gen_div = tk.StringVar(value='2')
    ttk.Entry(fi2b, textvariable=t8_gen_div, width=4).grid(row=0, column=7, padx=(0, 6))
    ttk.Label(fi2b, text='源引脚:').grid(row=0, column=8, padx=(0, 2))
    t8_gen_pin = tk.StringVar(value='[get_pins u_pll/CLK_OUT]')
    ttk.Entry(fi2b, textvariable=t8_gen_pin, width=20).grid(row=0, column=9, padx=(0, 6), sticky='ew')
    fi2b.grid_columnconfigure(9, weight=1)
    ttk.Label(fi2b, text='边沿:').grid(row=0, column=10, padx=(0, 2))
    t8_gen_edges = tk.StringVar(value='')
    ttk.Entry(fi2b, textvariable=t8_gen_edges, width=10).grid(row=0, column=11, padx=(0, 6))

    fi2c = ttk.Frame(fc2b, style='TFrame')
    fi2c.grid(row=3, column=0, columnspan=2, sticky='ew', padx=8, pady=(2, 6))

    def _t8_add_gen(*_):
        nm = t8_gen_name.get().strip()
        sc = t8_gen_src.get().strip()
        if not nm or not sc:
            messagebox.showerror('错误', '衍生名和源时钟不能为空')
            return
        try:
            m = int(t8_gen_mult.get() or '1')
            d = int(t8_gen_div.get() or '1')
        except ValueError:
            messagebox.showerror('错误', 'multiply/divide 必须为整数')
            return
        if m < 1 or d < 1:
            messagebox.showerror('错误', 'multiply/divide 必须 ≥ 1')
            return
        pin = t8_gen_pin.get().strip() or '[get_pins <divider>/Q]'
        edges = t8_gen_edges.get().strip()
        t8_gen_clocks.append({'name': nm, 'src': sc, 'mult': m, 'div': d,
                              'pin': pin, 'edges': edges})
        t8_gen_list.insert('', 'end', values=(nm, sc, m, d, pin, edges))
        # 自动递增衍生名
        if nm.endswith('_div2'):
            t8_gen_name.set(nm[:-1] + '3')
        else:
            t8_gen_name.set(nm + '_new')

    def _t8_del_gen():
        sel = t8_gen_list.selection()
        if not sel:
            return
        idx = t8_gen_list.index(sel[0])
        t8_gen_list.delete(sel[0])
        if 0 <= idx < len(t8_gen_clocks):
            t8_gen_clocks.pop(idx)

    ttk.Button(fi2c, text='➕ 添加衍生', command=_t8_add_gen,
               style='Accent.TButton').pack(side='left', padx=(0, 6))
    ttk.Button(fi2c, text='🗑 删除', command=_t8_del_gen,
               style='Normal.TButton').pack(side='left', padx=(0, 6))

    # 预填一个示例: sys_clk_div2 (1/2)
    t8_gen_clocks.append({'name': 'sys_clk_div2', 'src': 'sys_clk', 'mult': 1,
                          'div': 2, 'pin': '[get_pins u_pll/CLK_OUT]', 'edges': ''})
    t8_gen_list.insert('', 'end', values=('sys_clk_div2', 'sys_clk', 1, 2,
                                          '[get_pins u_pll/CLK_OUT]', ''))

    # 预填主时钟示例 (与衍生时钟 sys_clk_div2 配套, 用户一点"生成"就能看到 XDC)
    t8_clocks.append({'name': 'sys_clk', 'period_ns': '10.0', 'waveform': '0,5',
                      'source': '[get_ports clk]'})
    t8_clk_list.insert('', 'end', values=('sys_clk', '10.0', '0,5', '[get_ports clk]'))

    # 生成 / 保存按钮
    fc2c = ttk.Frame(t8_a2, style='TFrame')
    fc2c.grid(row=2, column=0, sticky='ew', padx=12, pady=4)
    ttk.Button(fc2c, text='⚙  生成时序约束', command=lambda: _t8_gen_timing_xdc(),
               style='Accent.TButton').pack(side='left', padx=(0, 6))
    ttk.Button(fc2c, text='💾  保存 .xdc', command=lambda: _t8_save_timing_xdc(),
               style='Success.TButton').pack(side='left', padx=(0, 6))
    ttk.Button(fc2c, text='📋  复制', command=lambda: _t8_copy_timing_xdc(),
               style='Info.TButton').pack(side='left', padx=(0, 6))
    ttk.Button(fc2c, text='📤  导出更多...', command=lambda: _t8_export_timing_menu(),
               style='Normal.TButton').pack(side='left', padx=(0, 6))
    t8_timing_status = tk.StringVar(value='就绪')
    ttk.Label(fc2c, textvariable=t8_timing_status, foreground=C['sub'],
              font=(F, 9)).pack(side='right', padx=(20, 0))

    # 预览
    t8_timing_frame = ttk.LabelFrame(t8_a2, text=' 时序约束预览 ')
    t8_timing_frame.grid(row=3, column=0, sticky='nsew', padx=12, pady=(4, 8))
    t8_timing_frame.grid_rowconfigure(0, weight=1)
    t8_timing_frame.grid_columnconfigure(0, weight=1)
    t8_timing_text = tk.Text(t8_timing_frame, font=(M, 9), bg=C['ebg'], fg=C['fg'],
                             relief='flat', padx=10, pady=8, wrap='none', height=14)
    t8_timing_text.grid(row=0, column=0, sticky='nsew')
    t8_timing_text.bind('<Key>', lambda e: 'break')
    t8_timing_xh = ttk.Scrollbar(t8_timing_frame, orient='horizontal',
                                  command=t8_timing_text.xview)
    t8_timing_xh.grid(row=1, column=0, sticky='ew')
    t8_timing_text.configure(xscrollcommand=t8_timing_xh.set)

    t8_timing_output = ''

    def _build_timing_xdc():
        lines = ['# ====== Auto-generated Timing Constraints ======',
                 f'# {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', '']
        # 主时钟
        for c in t8_clocks:
            wv = c['waveform'] or '0,5'
            wv_args = f' -waveform {{{wv}}}' if wv else ''
            lines.append(f'create_clock -name {c["name"]} -period {c["period_ns"]}'
                         f'{wv_args} {c["source"]}')
        lines.append('')
        # 衍生时钟 (使用 t8_gen_clocks 中实际配置的 multiply/divide/pin/edges)
        if v8_gen_dcm.get() and t8_gen_clocks:
            lines.append('# === 衍生时钟 (create_generated_clock) ===')
            for g in t8_gen_clocks:
                # 校验源时钟必须存在
                src_names = {c['name'] for c in t8_clocks}
                if g['src'] not in src_names:
                    lines.append(f'# ⚠ 源时钟 {g["src"]} 未在主时钟列表中, 已跳过 {g["name"]}')
                    continue
                src_clk = next(c for c in t8_clocks if c['name'] == g['src'])
                # 计算新周期 (ns) 用于波形默认值
                new_period = float(src_clk['period_ns']) * g['div'] / g['mult']
                half = new_period / 2
                edges = g['edges']
                if edges:
                    # 边沿 1 2 3 → 1,half+1,half+1+half... 用简单均分替代
                    edges_arg = f' -edge {{ {edges} }}'
                else:
                    edges_arg = f' -waveform {{0 {half}}}'
                lines.append(f'create_generated_clock -name {g["name"]} '
                             f'-source {src_clk["source"]} \\')
                lines.append(f'    -multiply_by {g["mult"]} -divide_by {g["div"]}'
                             f'{edges_arg} {g["pin"]}')
        return '\n'.join(lines) + '\n'

    def _t8_gen_timing_xdc():
        nonlocal t8_timing_output
        if not t8_clocks:
            messagebox.showinfo('提示', '请先添加至少一个主时钟')
            return
        t8_timing_output = _build_timing_xdc()
        t8_timing_text.delete('1.0', 'end')
        t8_timing_text.insert('1.0', t8_timing_output)
        t8_timing_status.set(f'✔ 已生成 {len(t8_clocks)} 个时钟约束')

    def _t8_save_timing_xdc():
        nonlocal t8_timing_output
        if not t8_timing_output:
            _t8_gen_timing_xdc()
        if not t8_timing_output:
            return
        fp = filedialog.asksaveasfilename(
            title='保存时序约束',
            defaultextension='.xdc',
            filetypes=[('XDC', '*.xdc'), ('All', '*.*')],
            initialfile='timing_constraints.xdc')
        if fp:
            try:
                with open(fp, 'w', encoding='utf-8') as f:
                    f.write(t8_timing_output)
                t8_timing_status.set(f'✔ 已保存: {os.path.basename(fp)}')
            except OSError as e:
                messagebox.showerror('保存失败', str(e))

    def _t8_copy_timing_xdc():
        """复制时序约束 XDC 到系统剪贴板"""
        nonlocal t8_timing_output
        if not t8_timing_output:
            _t8_gen_timing_xdc()
        if not t8_timing_output:
            return
        try:
            root.clipboard_clear()
            root.clipboard_append(t8_timing_output)
            root.update()
            t8_timing_status.set('✔ 已复制到剪贴板')
        except tk.TclError as e:
            messagebox.showerror('复制失败', str(e))

    # ── 时序约束的"导出更多"弹窗 ──
    # 时序约束只有纯 XDC 文本, 没有管脚表, 因此 .csv/导出到 Vivado 不可用
    def _t8_save_timing_tcl():
        """把时序约束包装为 Vivado Tcl 脚本"""
        fp = filedialog.asksaveasfilename(
            title='保存 TCL 脚本',
            defaultextension='.tcl',
            filetypes=[('Tcl 脚本', '*.tcl'), ('All', '*.*')],
            initialfile='timing_constraints.tcl')
        if not fp:
            return
        try:
            tcl = _gen_tcl_wrapper(t8_timing_output, [])
            with open(fp, 'w', encoding='utf-8') as f:
                f.write(tcl)
            t8_timing_status.set(f'✔ TCL 已保存: {os.path.basename(fp)}')
        except OSError as e:
            messagebox.showerror('保存失败', str(e))

    def _t8_save_timing_html():
        """导出时序约束 HTML 报告"""
        fp = filedialog.asksaveasfilename(
            title='保存时序约束报告 (HTML)',
            defaultextension='.html',
            filetypes=[('HTML', '*.html'), ('All', '*.*')],
            initialfile='timing_constraints_report.html')
        if not fp:
            return
        try:
            html = f'''<!DOCTYPE html>
<html lang="zh-CN"><head><meta charset="utf-8">
<title>时序约束报告</title>
<style>
body {{ font-family: -apple-system, "Microsoft YaHei", sans-serif;
        margin: 24px; background: #f4f6f9; color: #1f2329; }}
.box {{ max-width: 960px; margin: 0 auto; background: #fff;
        padding: 24px; border-radius: 8px;
        box-shadow: 0 1px 3px rgba(0,0,0,.08); }}
h1 {{ color: #2b6cb0; }}
pre {{ background: #1e1e1e; color: #d4d4d4; padding: 16px;
       border-radius: 6px; overflow-x: auto; font-size: 12px; line-height: 1.5; }}
.meta {{ color: #6c757d; font-size: 13px; margin-bottom: 16px; }}
</style></head><body>
<div class="box">
<h1>⏱ 时序约束报告</h1>
<div class="meta">生成时间: {datetime.datetime.now():%Y-%m-%d %H:%M:%S}</div>
<pre>{html_mod.escape(t8_timing_output)}</pre>
</div></body></html>'''
            with open(fp, 'w', encoding='utf-8') as f:
                f.write(html)
            t8_timing_status.set(f'✔ HTML 已保存: {os.path.basename(fp)}')
        except OSError as e:
            messagebox.showerror('保存失败', str(e))

    def _t8_export_timing_menu():
        """时序约束的导出更多弹窗 (.tcl / .html)"""
        if not t8_timing_output:
            _t8_gen_timing_xdc()
        if not t8_timing_output:
            return
        dlg = tk.Toplevel(root)
        dlg.title('导出选项')
        dlg.resizable(False, False)
        dlg.transient(root)
        dlg.grab_set()
        dlg.configure(bg=C['card'])
        dlg.geometry('540x300')
        header = tk.Frame(dlg, bg=C['blue'])
        header.pack(fill='x')
        tk.Label(header, text='📤  导出时序约束', bg=C['blue'], fg='#ffffff',
                 font=(F, 12, 'bold'), padx=20, pady=12).pack(anchor='w')
        body = tk.Frame(dlg, bg=C['card'])
        body.pack(fill='both', expand=True, padx=24, pady=16)
        tk.Label(body, text='选择导出格式:', bg=C['card'], fg=C['fg'],
                 font=(F, 10, 'bold')).pack(anchor='w', pady=(0, 10))
        def _opt_btn(text, desc, cmd, primary=False):
            f = tk.Frame(body, bg=C['card']); f.pack(fill='x', pady=4)
            ttk.Button(f, text=text, command=cmd,
                       style='Accent.TButton' if primary else 'Normal.TButton'
                       ).pack(side='left', padx=(0, 12))
            tk.Label(f, text=desc, bg=C['card'], fg=C['sub'],
                     font=(F, 9)).pack(side='left')
        _opt_btn('📄  保存为 .xdc',
                 'Vivado 约束文件, 标准格式',
                 lambda: (dlg.destroy(), _t8_save_timing_xdc()),
                 primary=True)
        _opt_btn('📜  保存为 .tcl',
                 'Vivado Tcl 脚本 (source 可加载)',
                 lambda: (dlg.destroy(), _t8_save_timing_tcl()))
        _opt_btn('📋  生成完整报告 (HTML)',
                 '时序约束可读性报告',
                 lambda: (dlg.destroy(), _t8_save_timing_html()))
        ttk.Button(dlg, text='关闭', command=dlg.destroy,
                   style='Normal.TButton').pack(pady=(0, 14))
        dlg.update_idletasks()
        x = root.winfo_x() + (root.winfo_width() - dlg.winfo_width()) // 2
        y = root.winfo_y() + (root.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f'+{x}+{y}')

    # 预填一个示例
    t8_clocks.append({'name': 'sys_clk', 'period_ns': '10.0',
                      'waveform': '0,5', 'source': '[get_ports clk]'})
    t8_clk_list.insert('', 'end', values=('sys_clk', '10.0', '0,5', '[get_ports clk]'))


    # ═══════════════════════════════════
    # 子页 3 — 其他约束 (工程压缩 / 未用管脚 / Flash / 跳过)
    # ═══════════════════════════════════
    t8_a3 = ttk.Frame(t8_nb, style='TFrame')
    t8_nb.add(t8_a3, text='  🔧 其他约束  ')
    t8_a3.grid_columnconfigure(0, weight=1)
    # 不给非预览行 weight, 紧凑排列
    t8_a3.grid_rowconfigure(0, weight=0, minsize=0)  # 工程配置
    t8_a3.grid_rowconfigure(1, weight=0, minsize=0)  # 未用管脚
    t8_a3.grid_rowconfigure(2, weight=0, minsize=0)  # Flash / 配置选项
    t8_a3.grid_rowconfigure(3, weight=0, minsize=0)  # 按钮 (按内容, 不被拉伸)
    t8_a3.grid_rowconfigure(4, weight=1)              # 预览行 (占用剩余空间)

    # 复选框组
    fc3a = ttk.LabelFrame(t8_a3, text=' 工程配置 ')
    fc3a.grid(row=0, column=0, sticky='ew', padx=12, pady=(10, 4))
    fc3a.grid_columnconfigure(0, weight=1)

    v8_compress = tk.BooleanVar(value=False)
    v8_no_compress = tk.BooleanVar(value=True)

    def _on_compress_toggle():
        """压缩 / 不压缩 互斥: 选中一个时自动取消另一个"""
        if v8_compress.get() and v8_no_compress.get():
            # 当前被勾的就是触发者; 通过 _state_aware 区分
            pass
        # 使用 trace 模式更可靠:
        #  - 当 v8_compress 变成 True 时, v8_no_compress = False
        #  - 当 v8_no_compress 变成 True 时, v8_compress = False
    def _trace_compress(*_):
        if v8_compress.get():
            v8_no_compress.set(False)
    def _trace_no_compress(*_):
        if v8_no_compress.get():
            v8_compress.set(False)
    v8_compress.trace_add('write', _trace_compress)
    v8_no_compress.trace_add('write', _trace_no_compress)

    ttk.Checkbutton(fc3a, text='启用工程压缩 (set_property BITSTREAM.GENERAL.COMPRESS TRUE)',
                    variable=v8_compress, style='TCheckbutton').grid(
        row=0, column=0, sticky='w', padx=8, pady=(6, 2))
    ttk.Checkbutton(fc3a, text='不压缩位流 (set_property BITSTREAM.GENERAL.COMPRESS FALSE)',
                    variable=v8_no_compress, style='TCheckbutton').grid(
        row=1, column=0, sticky='w', padx=8, pady=(2, 6))

    fc3b = ttk.LabelFrame(t8_a3, text=' 未用管脚 (UNUSED PINS) ')
    fc3b.grid(row=1, column=0, sticky='ew', padx=12, pady=(4, 4))
    fc3b.grid_columnconfigure(0, weight=1)

    t8_unused_mode = tk.StringVar(value='pullup')
    rb_frame = ttk.Frame(fc3b, style='TFrame')
    rb_frame.grid(row=0, column=0, sticky='w', padx=8, pady=6)
    for i, (val, lbl) in enumerate([
        ('pullup',  '上拉 (PULLUP)'),
        ('pulldown','下拉 (PULLDOWN)'),
        ('float',   '悬空 (FLOAT)'),
    ]):
        ttk.Radiobutton(rb_frame, text=lbl, variable=t8_unused_mode, value=val,
                        style='TRadiobutton').grid(row=0, column=i, padx=(0, 14), pady=2)

    # Flash 启动速度
    fc3c = ttk.LabelFrame(t8_a3, text=' Flash / 配置选项 ')
    fc3c.grid(row=2, column=0, sticky='ew', padx=12, pady=(4, 4))
    fc3c.grid_columnconfigure(0, weight=1)
    # 不再给 row 2 设 weight, 避免把单个复选框撑成弹簧

    # SPI Flash 速度 (单行靠左, 单位 MHz 紧随下拉)
    fl_frame = ttk.Frame(fc3c, style='TFrame')
    fl_frame.grid(row=0, column=0, sticky='w', padx=8, pady=(8, 4))
    ttk.Label(fl_frame, text='SPI Flash 速度:').pack(side='left', padx=(0, 6))
    t8_flash_speed = tk.StringVar(value='50')
    ttk.Combobox(fl_frame, textvariable=t8_flash_speed, width=8, state='readonly',
                 values=['3', '6', '10', '12', '16', '20', '25', '33', '40', '50',
                         '66', '80', '100', '125', '133']).pack(side='left', padx=(0, 4))
    ttk.Label(fl_frame, text='MHz').pack(side='left', padx=(0, 4))

    # ── 4 个附加选项, 用同一行 pack 横向排, 紧凑不拉伸 ──
    v8_skip_unused = tk.BooleanVar(value=False)  # 不使用管脚约束 (抑制 DRC 警告)
    opt_frame = ttk.Frame(fc3c, style='TFrame')
    opt_frame.grid(row=1, column=0, sticky='w', padx=8, pady=(2, 8))
    ttk.Checkbutton(opt_frame, text='CONFIG_VOLTAGE = 3.3',
                    variable=tk.BooleanVar(value=True),
                    style='TCheckbutton').pack(side='left', padx=(0, 16))
    ttk.Checkbutton(opt_frame, text='CFGBVS VCCO (Bank 0/14/15)',
                    variable=tk.BooleanVar(value=True),
                    style='TCheckbutton').pack(side='left', padx=(0, 16))
    ttk.Checkbutton(opt_frame, text='USR_ACCESS TIMESTAMP',
                    variable=tk.BooleanVar(value=False),
                    style='TCheckbutton').pack(side='left', padx=(0, 16))
    ttk.Checkbutton(opt_frame, text='不使用管脚约束 (抑制 NSTD/RTSTAT/UCIO DRC 警告)',
                    variable=v8_skip_unused,
                    style='TCheckbutton').pack(side='left', padx=(0, 16))

    # 生成 / 保存 / 复制
    fc3d = ttk.Frame(t8_a3, style='TFrame')
    fc3d.grid(row=3, column=0, sticky='ew', padx=12, pady=(4, 4))
    ttk.Button(fc3d, text='⚙  生成其他约束', command=lambda: _t8_gen_misc_xdc(),
               style='Accent.TButton').pack(side='left', padx=(0, 6))
    ttk.Button(fc3d, text='💾  保存 .xdc', command=lambda: _t8_save_misc_xdc(),
               style='Success.TButton').pack(side='left', padx=(0, 6))
    ttk.Button(fc3d, text='📋  复制', command=lambda: _t8_copy_misc_xdc(),
               style='Info.TButton').pack(side='left', padx=(0, 6))
    ttk.Button(fc3d, text='📤  导出更多...', command=lambda: _t8_export_misc_menu(),
               style='Normal.TButton').pack(side='left', padx=(0, 6))
    t8_misc_status = tk.StringVar(value='就绪')
    ttk.Label(fc3d, textvariable=t8_misc_status, foreground=C['sub'],
              font=(F, 9)).pack(side='right', padx=(20, 0))

    t8_misc_frame = ttk.LabelFrame(t8_a3, text=' 其他约束预览 ')
    t8_misc_frame.grid(row=4, column=0, sticky='nsew', padx=12, pady=(4, 8))
    t8_misc_frame.grid_rowconfigure(0, weight=1)
    t8_misc_frame.grid_columnconfigure(0, weight=1)
    t8_misc_text = tk.Text(t8_misc_frame, font=(M, 9), bg=C['ebg'], fg=C['fg'],
                           relief='flat', padx=10, pady=8, wrap='none', height=12)
    t8_misc_text.grid(row=0, column=0, sticky='nsew')
    t8_misc_text.bind('<Key>', lambda e: 'break')
    t8_misc_xh = ttk.Scrollbar(t8_misc_frame, orient='horizontal',
                                command=t8_misc_text.xview)
    t8_misc_xh.grid(row=1, column=0, sticky='ew')
    t8_misc_text.configure(xscrollcommand=t8_misc_xh.set)

    t8_misc_output = ''

    def _build_misc_xdc():
        lines = ['# ====== Other Constraints ======',
                 f'# {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', '']
        # 工程压缩
        if v8_compress.get():
            lines.append('set_property BITSTREAM.GENERAL.COMPRESS TRUE [current_design]')
        if v8_no_compress.get():
            lines.append('set_property BITSTREAM.GENERAL.COMPRESS FALSE [current_design]')
        # 未用管脚
        mode = t8_unused_mode.get()
        if mode == 'pullup':
            lines.append('set_property CFGBVS VCCO [current_design]')
            lines.append('set_property CONFIG_VOLTAGE 3.3 [current_design]')
            lines.append('set_property BITSTREAM.CONFIG.UNUSEDPIN Pullup [current_design]')
        elif mode == 'pulldown':
            lines.append('set_property BITSTREAM.CONFIG.UNUSEDPIN Pulldown [current_design]')
        elif mode == 'float':
            lines.append('set_property BITSTREAM.CONFIG.UNUSEDPIN Float [current_design]')
        # Flash 速度
        sp = t8_flash_speed.get()
        lines.append(f'set_property BITSTREAM.CONFIG.SPI_BUSWIDTH 4 [current_design]')
        lines.append(f'set_property BITSTREAM.CONFIG.SPI_FREQUENCY {sp} [current_design]')
        # 可选
        lines.append('set_property BITSTREAM.CONFIG.CONFIGRATE {0} [current_design]'.format(sp))
        # 跳过未使用管脚 — 仅抑制 DRC 警告, 不修改 BitGen 配置
        if v8_skip_unused.get():
            lines.append('')
            lines.append('# --- 抑制未使用管脚相关 DRC 警告 (Vivado 默认 Error) ---')
            lines.append('set_property SEVERITY {Warning} [get_drc_checks NSTD-1]')
            lines.append('set_property SEVERITY {Warning} [get_drc_checks RTSTAT-1]')
            lines.append('set_property SEVERITY {Warning} [get_drc_checks UCIO-1]')
        return '\n'.join(lines) + '\n'

    def _t8_gen_misc_xdc():
        import sys as _dbg_sys
        nonlocal t8_misc_output
        t8_misc_output = _build_misc_xdc()
        print(f'[DEBUG _t8_gen_misc_xdc] output len={len(t8_misc_output)}', file=_dbg_sys.stderr)
        print(f'[DEBUG _t8_gen_misc_xdc] output preview={t8_misc_output[:200]!r}', file=_dbg_sys.stderr)
        t8_misc_text.delete('1.0', 'end')
        t8_misc_text.insert('1.0', t8_misc_output)
        print(f'[DEBUG _t8_gen_misc_xdc] after insert text widget content={t8_misc_text.get("1.0", "end")[:200]!r}', file=_dbg_sys.stderr)
        t8_misc_status.set('✔ 其他约束已生成')

    def _t8_save_misc_xdc():
        nonlocal t8_misc_output
        if not t8_misc_output:
            _t8_gen_misc_xdc()
        if not t8_misc_output:
            return
        fp = filedialog.asksaveasfilename(
            title='保存其他约束',
            defaultextension='.xdc',
            filetypes=[('XDC', '*.xdc'), ('All', '*.*')],
            initialfile='misc_constraints.xdc')
        if fp:
            try:
                with open(fp, 'w', encoding='utf-8') as f:
                    f.write(t8_misc_output)
                t8_misc_status.set(f'✔ 已保存: {os.path.basename(fp)}')
            except OSError as e:
                messagebox.showerror('保存失败', str(e))

    def _t8_copy_misc_xdc():
        """复制其他约束 XDC 到系统剪贴板"""
        nonlocal t8_misc_output
        if not t8_misc_output:
            _t8_gen_misc_xdc()
        if not t8_misc_output:
            return
        try:
            root.clipboard_clear()
            root.clipboard_append(t8_misc_output)
            root.update()  # 确保剪贴板内容保留
            t8_misc_status.set('✔ 已复制到剪贴板')
        except tk.TclError as e:
            messagebox.showerror('复制失败', str(e))

    # ── 其他约束的"导出更多"弹窗 ──
    # 其他约束只有纯 XDC 文本, 没有管脚表, 因此 .csv/导出到 Vivado 不可用
    def _t8_save_misc_tcl():
        """把其他约束包装为 Vivado Tcl 脚本"""
        fp = filedialog.asksaveasfilename(
            title='保存 TCL 脚本',
            defaultextension='.tcl',
            filetypes=[('Tcl 脚本', '*.tcl'), ('All', '*.*')],
            initialfile='misc_constraints.tcl')
        if not fp:
            return
        try:
            tcl = _gen_tcl_wrapper(t8_misc_output, [])
            with open(fp, 'w', encoding='utf-8') as f:
                f.write(tcl)
            t8_misc_status.set(f'✔ TCL 已保存: {os.path.basename(fp)}')
        except OSError as e:
            messagebox.showerror('保存失败', str(e))

    def _t8_save_misc_html():
        """导出其他约束 HTML 报告"""
        fp = filedialog.asksaveasfilename(
            title='保存其他约束报告 (HTML)',
            defaultextension='.html',
            filetypes=[('HTML', '*.html'), ('All', '*.*')],
            initialfile='misc_constraints_report.html')
        if not fp:
            return
        try:
            html = f'''<!DOCTYPE html>
<html lang="zh-CN"><head><meta charset="utf-8">
<title>其他约束报告</title>
<style>
body {{ font-family: -apple-system, "Microsoft YaHei", sans-serif;
        margin: 24px; background: #f4f6f9; color: #1f2329; }}
.box {{ max-width: 960px; margin: 0 auto; background: #fff;
        padding: 24px; border-radius: 8px;
        box-shadow: 0 1px 3px rgba(0,0,0,.08); }}
h1 {{ color: #2b6cb0; }}
pre {{ background: #1e1e1e; color: #d4d4d4; padding: 16px;
       border-radius: 6px; overflow-x: auto; font-size: 12px; line-height: 1.5; }}
.meta {{ color: #6c757d; font-size: 13px; margin-bottom: 16px; }}
</style></head><body>
<div class="box">
<h1>🔧 其他约束报告</h1>
<div class="meta">生成时间: {datetime.datetime.now():%Y-%m-%d %H:%M:%S}</div>
<pre>{html_mod.escape(t8_misc_output)}</pre>
</div></body></html>'''
            with open(fp, 'w', encoding='utf-8') as f:
                f.write(html)
            t8_misc_status.set(f'✔ HTML 已保存: {os.path.basename(fp)}')
        except OSError as e:
            messagebox.showerror('保存失败', str(e))

    def _t8_export_misc_menu():
        """其他约束的导出更多弹窗 (.tcl / .html)"""
        if not t8_misc_output:
            _t8_gen_misc_xdc()
        if not t8_misc_output:
            return
        dlg = tk.Toplevel(root)
        dlg.title('导出选项')
        dlg.resizable(False, False)
        dlg.transient(root)
        dlg.grab_set()
        dlg.configure(bg=C['card'])
        dlg.geometry('540x300')
        header = tk.Frame(dlg, bg=C['blue'])
        header.pack(fill='x')
        tk.Label(header, text='📤  导出其他约束', bg=C['blue'], fg='#ffffff',
                 font=(F, 12, 'bold'), padx=20, pady=12).pack(anchor='w')
        body = tk.Frame(dlg, bg=C['card'])
        body.pack(fill='both', expand=True, padx=24, pady=16)
        tk.Label(body, text='选择导出格式:', bg=C['card'], fg=C['fg'],
                 font=(F, 10, 'bold')).pack(anchor='w', pady=(0, 10))
        def _opt_btn(text, desc, cmd, primary=False):
            f = tk.Frame(body, bg=C['card']); f.pack(fill='x', pady=4)
            ttk.Button(f, text=text, command=cmd,
                       style='Accent.TButton' if primary else 'Normal.TButton'
                       ).pack(side='left', padx=(0, 12))
            tk.Label(f, text=desc, bg=C['card'], fg=C['sub'],
                     font=(F, 9)).pack(side='left')
        _opt_btn('📄  保存为 .xdc',
                 'Vivado 约束文件, 标准格式',
                 lambda: (dlg.destroy(), _t8_save_misc_xdc()),
                 primary=True)
        _opt_btn('📜  保存为 .tcl',
                 'Vivado Tcl 脚本 (source 可加载)',
                 lambda: (dlg.destroy(), _t8_save_misc_tcl()))
        _opt_btn('📋  生成完整报告 (HTML)',
                 '其他约束可读性报告',
                 lambda: (dlg.destroy(), _t8_save_misc_html()))
        ttk.Button(dlg, text='关闭', command=dlg.destroy,
                   style='Normal.TButton').pack(pady=(0, 14))
        dlg.update_idletasks()
        x = root.winfo_x() + (root.winfo_width() - dlg.winfo_width()) // 2
        y = root.winfo_y() + (root.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f'+{x}+{y}')


    # ══════════════════════════════════════
    # TAB 9 — 分析辅助 (CDC / 时序 / Log)
    # ══════════════════════════════════════
    from src.app_config import get_vivado_bin_dirs as _t9_viv_dirs

    t9 = ttk.Frame(nb, style='TFrame')
    nb.add(t9, text='🔍 分析辅助')
    t9.grid_rowconfigure(1, weight=1)
    t9.grid_columnconfigure(0, weight=1)

    # ── Vivado 路径历史 (两子页共享, 持久化到 ~/.fpga_tool/vivado_paths.json) ──
    _VIVADO_HIST_DIR = os.path.join(os.path.expanduser('~'), '.fpga_tool')
    _VIVADO_HIST_FILE = os.path.join(_VIVADO_HIST_DIR, 'vivado_paths.json')
    _VIVADO_HIST_MAX = 10   # 最多保留 10 条

    def _vivado_hist_load():
        try:
            if os.path.isfile(_VIVADO_HIST_FILE):
                with open(_VIVADO_HIST_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                if isinstance(data, list):
                    return [d for d in data if isinstance(d, str) and d]
        except Exception:
            pass
        return []

    def _vivado_hist_save(items):
        try:
            os.makedirs(_VIVADO_HIST_DIR, exist_ok=True)
            with open(_VIVADO_HIST_FILE, 'w', encoding='utf-8') as f:
                json.dump(items[:_VIVADO_HIST_MAX], f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _vivado_hist_add(path, combobox, var):
        """把 path 加进历史, 刷新 combobox 内容"""
        if not path:
            return
        p = os.path.abspath(path)
        items = _vivado_hist_load()
        # 移除旧的同名项, 头插
        items = [x for x in items if os.path.abspath(x) != p]
        items.insert(0, p)
        items = items[:_VIVADO_HIST_MAX]
        _vivado_hist_save(items)
        combobox['values'] = items
        var.set(p)

    def _vivado_hist_del(path, combobox, var):
        if not path:
            return
        p = os.path.abspath(path)
        items = _vivado_hist_load()
        items = [x for x in items if os.path.abspath(x) != p]
        _vivado_hist_save(items)
        combobox['values'] = items
        if var.get() and os.path.abspath(var.get()) == p:
            var.set('')

    aux_nb = ttk.Notebook(t9)
    aux_nb.grid(row=1, column=0, sticky='nsew', padx=8, pady=(4, 8))

    # ═══════════════════════════════════
    # 子页 1 — CDC 检查提示
    # ═══════════════════════════════════
    a1 = ttk.Frame(aux_nb, style='TFrame')
    aux_nb.add(a1, text='  CDC 检查  ')
    a1.grid_columnconfigure(0, weight=1)
    a1.grid_rowconfigure(3, weight=1)

    # CDC 场景数据库
    _CDC_SCENARIOS = [
        ('single_slow2fast', '单 bit 信号 — 慢时钟域 → 快时钟域',
         '推荐方案：双寄存器同步器 (2-FF Synchronizer)',
         '适用条件：目的时钟频率 ≥ 源时钟频率 × 1.5，单 bit 信号。\n'
         'MTBF 取决于时钟频率差和寄存器翻转速率。需添加 ASYNC_REG 约束。\n'
         'Verilog 需使用 `(* ASYNC_REG = "TRUE" *)` 或 XDC 约束。',
         'reg sync_ff1, sync_ff2;\n'
         'always @(posedge clk_dst) begin\n'
         '    {sync_ff2, sync_ff1} <= {sync_ff1, sig_src};\n'
         'end\n'
         'assign sig_dst = sync_ff2;',
         'signal sync_ff1, sync_ff2 : std_logic;\n'
         'process(clk_dst)\n'
         'begin\n'
         '    if rising_edge(clk_dst) then\n'
         '        sync_ff2 <= sync_ff1;\n'
         '        sync_ff1 <= sig_src;\n'
         '    end if;\n'
         'end process;\n'
         'sig_dst <= sync_ff2;'),
        ('single_fast2slow', '单 bit 信号 — 快时钟域 → 慢时钟域',
         '推荐方案：脉冲展宽 + 双寄存器同步器',
         '适用条件：目的时钟频率 < 源时钟频率，脉冲宽度可能 < 目的时钟周期。\n'
         '需要将脉冲展宽到 ≥ 目的时钟周期的 1.5 倍。\n'
         '在源域中检测边沿并翻转 toggle，在目的域同步后检测 toggle 变化。',
         '// 源域：边沿检测 → toggle\n'
         'reg pulse_toggle;\n'
         'always @(posedge clk_src) begin\n'
         '    if (pulse_in) pulse_toggle <= ~pulse_toggle;\n'
         'end\n'
         '// 目的域：2-FF 同步 + 边沿检测\n'
         'reg [2:0] sync_ff;\n'
         'always @(posedge clk_dst) begin\n'
         '    sync_ff <= {sync_ff[1:0], pulse_toggle};\n'
         'end\n'
         'assign pulse_out = sync_ff[2] ^ sync_ff[1];',
         'signal pulse_toggle : std_logic := \'0\';\n'
         '-- 源域\n'
         'process(clk_src)\n'
         'begin\n'
         '    if rising_edge(clk_src) then\n'
         '        if pulse_in = \'1\' then\n'
         '            pulse_toggle <= not pulse_toggle;\n'
         '        end if;\n'
         '    end if;\n'
         'end process;\n'
         '-- 目的域: 2-FF sync + edge detect\n'
         'signal sync_ff : std_logic_vector(2 downto 0);\n'
         'process(clk_dst)\n'
         'begin\n'
         '    if rising_edge(clk_dst) then\n'
         '        sync_ff <= sync_ff(1 downto 0) & pulse_toggle;\n'
         '    end if;\n'
         'end process;\n'
         'pulse_out <= sync_ff(2) xor sync_ff(1);'),
        ('multi_lowbw', '多 bit 数据 — 低速传输',
         '推荐方案：握手协议 (Req/Ack Handshake)',
         '适用条件：数据带宽低，延迟不敏感。\n'
         '发送端置 req，接收端采样后置 ack，发送端见 ack 后撤 req。\n'
         '每条数据都需要 req/ack 来回，效率约 1/3~1/5 时钟周期。',
         '// 发送端\n'
         'reg req, data_valid;\n'
         'always @(posedge clk_src) begin\n'
         '    if (send && !req) begin\n'
         '        data_buf <= data_in;\n'
         '        req <= 1\'b1;\n'
         '    end else if (req && ack_sync) begin\n'
         '        req <= 1\'b0;\n'
         '    end\n'
         'end\n'
         '// 接收端\n'
         'reg [1:0] req_ff;\n'
         'always @(posedge clk_dst) begin\n'
         '    req_ff <= {req_ff[0], req};\n'
         '    if (req_ff[1]) begin\n'
         '        data_out <= data_buf;\n'
         '        ack <= 1\'b1;\n'
         '    end else ack <= 1\'b0;\n'
         'end\n'
         '// ack 同步回源域 (2-FF)...',
         '-- 发送端\n'
         'process(clk_src)\n'
         'begin\n'
         '    if rising_edge(clk_src) then\n'
         '        if send = \'1\' and req = \'0\' then\n'
         '            data_buf <= data_in;\n'
         '            req <= \'1\';\n'
         '        elsif req = \'1\' and ack_sync = \'1\' then\n'
         '            req <= \'0\';\n'
         '        end if;\n'
         '    end if;\n'
         'end process;\n'
         '-- ack 同步回源域需额外 2-FF...'),
        ('multi_highbw', '多 bit 数据 — 高速传输',
         '推荐方案：异步 FIFO (含格雷码指针)',
         '适用条件：持续高速数据流，对延迟敏感。\n'
         '经典结构：双端口 RAM + 读写指针 + 格雷码转换 + 2-FF 同步。\n'
         '需配合 set_max_delay / set_false_path 约束。',
         '// 写指针 → 格雷码\n'
         'wire [ADDR_W:0] wptr_gray = wptr ^ (wptr >> 1);\n'
         '// 同步到读域\n'
         'reg [ADDR_W:0] rq1_wptr, rq2_wptr;\n'
         'always @(posedge rclk) rq1_wptr <= wptr_gray;\n'
         'always @(posedge rclk) rq2_wptr <= rq1_wptr;\n'
         '// 格雷码 → 二进制\n'
         'function [ADDR_W:0] gray2bin;\n'
         '    input [ADDR_W:0] g;\n'
         '    integer i;\n'
         '    begin\n'
         '        gray2bin = g;\n'
         '        for (i=ADDR_W-1; i>=0; i=i-1)\n'
         '            gray2bin[i] = gray2bin[i+1] ^ g[i];\n'
         '    end\n'
         'endfunction\n'
         '// Empty/Full 判断...',
         '-- 写指针 → 格雷码\n'
         'wptr_gray <= wptr xor (\'0\' & wptr(ADDR_W downto 1));\n'
         '-- 同步到读域 (2-FF)\n'
         'process(rclk)\n'
         'begin\n'
         '    if rising_edge(rclk) then\n'
         '        rq1_wptr <= wptr_gray;\n'
         '        rq2_wptr <= rq1_wptr;\n'
         '    end if;\n'
         'end process;\n'
         '-- Gray→Bin (XOR chain)...'),
        ('reset_sync', '复位同步 — 异步复位同步释放',
         '推荐方案：异步复位同步释放电路',
         '适用条件：外部异步复位输入，内部需同步释放避免亚稳态。\n'
         '复位有效时立即复位 (异步)，释放时被目的时钟同步 (同步释放)。',
         'reg rst_sync1, rst_sync2;\n'
         'always @(posedge clk or negedge rst_n_async) begin\n'
         '    if (!rst_n_async) begin\n'
         '        rst_sync1 <= 1\'b0;\n'
         '        rst_sync2 <= 1\'b0;\n'
         '    end else begin\n'
         '        rst_sync1 <= 1\'b1;\n'
         '        rst_sync2 <= rst_sync1;\n'
         '    end\n'
         'end\n'
         'assign rst_n_sync = rst_sync2;',
         'process(clk, rst_n_async)\n'
         'begin\n'
         '    if rst_n_async = \'0\' then\n'
         '        rst_sync1 <= \'0\';\n'
         '        rst_sync2 <= \'0\';\n'
         '    elsif rising_edge(clk) then\n'
         '        rst_sync1 <= \'1\';\n'
         '        rst_sync2 <= rst_sync1;\n'
         '    end if;\n'
         'end process;\n'
         'rst_n_sync <= rst_sync2;'),
    ]

    # CDC 选择器
    fa1 = ttk.LabelFrame(a1, text=' CDC 场景 ', )
    fa1.grid(row=0, column=0, sticky='ew', padx=10, pady=(8, 4))
    fa1.grid_columnconfigure(1, weight=1)

    a1_scenario = tk.StringVar(value='single_slow2fast')
    a1_combo = ttk.Combobox(fa1, textvariable=a1_scenario,
                            values=[s[0] for s in _CDC_SCENARIOS],
                            state='readonly', font=(F, 10), width=40)
    a1_combo.grid(row=0, column=0, sticky='w', padx=(14, 8), pady=(8, 6))
    ttk.Label(fa1, text=f'← 共 {len(_CDC_SCENARIOS)} 种场景', font=(F, 9),
              foreground=C['sub']).grid(row=0, column=1, sticky='w',
              padx=(0, 14), pady=(8, 6))

    # 推荐说明
    fa1o = ttk.LabelFrame(a1, text=' 推荐方案 ', )
    fa1o.grid(row=1, column=0, sticky='ew', padx=10, pady=4)
    a1_title  = tk.StringVar()
    a1_cond   = tk.StringVar()
    ttk.Label(fa1o, textvariable=a1_title, font=(F, 11, 'bold'),
              foreground=C['blue']).grid(row=0, column=0, sticky='w',
              padx=14, pady=(8, 2))
    ttk.Label(fa1o, textvariable=a1_cond, font=(F, 9),
              foreground=C['sub'], wraplength=700).grid(row=1, column=0,
              sticky='w', padx=14, pady=(0, 8))

    # 代码面板 (Verilog / VHDL)
    a1_nb = ttk.Notebook(a1)
    a1_nb.grid(row=3, column=0, sticky='nsew', padx=10, pady=(4, 8))

    a1_vf = ttk.Frame(a1_nb, style='TFrame')
    a1_vf.grid_rowconfigure(0, weight=1)
    a1_vf.grid_columnconfigure(0, weight=1)
    a1_nb.add(a1_vf, text='  Verilog  ')
    a1_v_text = tk.Text(a1_vf, font=(M, 9), bg=C['ebg'], fg=C['fg'],
                        relief='flat', padx=12, pady=10, wrap='none')
    a1_v_text.grid(row=0, column=0, sticky='nsew')
    a1_v_text.bind('<Key>', lambda e: 'break')
    a1_vs = ttk.Scrollbar(a1_vf, orient='vertical', command=a1_v_text.yview)
    a1_vs.grid(row=0, column=1, sticky='ns')
    a1_v_text.configure(yscrollcommand=a1_vs.set)

    a1_hf = ttk.Frame(a1_nb, style='TFrame')
    a1_hf.grid_rowconfigure(0, weight=1)
    a1_hf.grid_columnconfigure(0, weight=1)
    a1_nb.add(a1_hf, text='  VHDL  ')
    a1_h_text = tk.Text(a1_hf, font=(M, 9), bg=C['ebg'], fg=C['fg'],
                        relief='flat', padx=12, pady=10, wrap='none')
    a1_h_text.grid(row=0, column=0, sticky='nsew')
    a1_h_text.bind('<Key>', lambda e: 'break')
    a1_hs = ttk.Scrollbar(a1_hf, orient='vertical', command=a1_h_text.yview)
    a1_hs.grid(row=0, column=1, sticky='ns')
    a1_h_text.configure(yscrollcommand=a1_hs.set)

    def _cdc_update(*_):
        key = a1_scenario.get()
        for s in _CDC_SCENARIOS:
            if s[0] == key:
                a1_title.set(s[2])
                a1_cond.set(s[3])
                a1_v_text.delete('1.0', 'end')
                a1_v_text.insert('1.0', s[4])
                a1_h_text.delete('1.0', 'end')
                a1_h_text.insert('1.0', s[5])
                break

    a1_combo.bind('<<ComboboxSelected>>', _cdc_update)
    _cdc_update()

    # ═══════════════════════════════════
    # 子页 2 — 时序报告解析
    # ═══════════════════════════════════
    a2 = ttk.Frame(aux_nb, style='TFrame')
    aux_nb.add(a2, text='  时序报告  ')
    a2.grid_columnconfigure(0, weight=1)
    a2.grid_rowconfigure(5, weight=1)

    # ── 工程根 + Vivado 路径 (可缺省, 后面手动浏览报告也行) ──
    fa2p = ttk.LabelFrame(a2, text=' 工程信息 (缺报告时自动调 Vivado 生成) ')
    fa2p.grid(row=0, column=0, sticky='ew', padx=10, pady=(8, 4))
    fa2p.grid_columnconfigure(1, weight=1)

    a2_proj = tk.StringVar()                 # Vivado 工程根 (含 .xpr)
    a2_vivado = tk.StringVar()               # vivado.bat / vivado 完整路径
    a2_ver = tk.StringVar()                  # Vivado 版本号 (用于估计 tcl 命令兼容性)

    ttk.Label(fa2p, text='工程路径:', font=(F, 9)).grid(
        row=0, column=0, sticky='w', padx=(14, 4), pady=(6, 2))
    ttk.Entry(fa2p, textvariable=a2_proj, font=(F, 9)).grid(
        row=0, column=1, sticky='ew', padx=2, pady=(6, 2))
    ttk.Button(fa2p, text='浏览',
               command=lambda: (d := filedialog.askdirectory(
                   title='选择 Vivado 工程根目录')) and a2_proj.set(d),
               style='Normal.TButton').grid(row=0, column=2, padx=(4, 14), pady=(6, 2))

    # Vivado 路径 — 来自 ⚙设置 Tab
    ttk.Label(fa2p, text='Vivado 路径:', font=(F, 9)).grid(
        row=1, column=0, sticky='w', padx=(14, 4), pady=(2, 2))
    _a2_viv_var = tk.StringVar()
    _a2_viv_label = ttk.Label(fa2p, textvariable=_a2_viv_var, font=(F, 8))
    _a2_viv_label.grid(row=1, column=1, sticky='w', padx=2, pady=(2, 2))
    _a2_viv_btn = ttk.Button(fa2p, text='⚙ 打开设置', command=lambda: nb.select(t16),
                              style='Small.TButton')

    # 从 Vivado 路径提取版本号 (a2/a3共享)
    def _t9_extract_versions():
        vers = []
        for d in _t9_viv_dirs():
            d = d.replace('\\', '/')
            parts = d.split('/')
            for p in parts:
                if re.match(r'^\d{4}\.\d$', p): vers.append(p); break
        return vers

    def _a2_refresh_viv():
        dirs = _t9_viv_dirs()
        if dirs:
            _a2_viv_var.set(f'✔ 已配置 {len(dirs)} 个 Vivado 路径')
            _a2_viv_label.config(foreground=C['green'])
            a2_vivado.set('')  # 由 _a2_resolve_vivado 按版本匹配
            _a2_viv_btn.grid_remove()
        else:
            _a2_viv_var.set('✘ 未配置 — 请到 ⚙设置 Tab 添加')
            _a2_viv_label.config(foreground=C['red'])
            a2_vivado.set('')
            _a2_viv_btn.grid(row=1, column=2, padx=(4, 14), pady=(2, 2))
        a2_ver_combo['values'] = _t9_extract_versions()

    ttk.Label(fa2p, text='版本:', font=(F, 9)).grid(
        row=2, column=0, sticky='w', padx=(14, 4), pady=(2, 6))
    a2_ver_combo = ttk.Combobox(fa2p, textvariable=a2_ver, font=(F, 9),
                                 width=10, values=_t9_extract_versions())
    a2_ver_combo.grid(row=2, column=1, sticky='w', padx=2, pady=(2, 6))
    _a2_refresh_viv()  # Combobox 创建后调用
    ttk.Label(fa2p, text='  ⓘ 留空时按 PATH 自动找 vivado',
              foreground=C['sub'], font=(F, 8)).grid(
        row=2, column=2, sticky='w', padx=(4, 14), pady=(2, 6))

    ttk.Button(fa2p, text='🔍 检测 / 自动生成报告',
               command=lambda: _a2_detect_and_gen(),
               style='Accent.TButton').grid(row=3, column=0, columnspan=3,
                                            sticky='ew', padx=14, pady=(2, 8))
    a2_status = tk.StringVar(value='点击上面的按钮自动检测/生成报告, 或直接在下方选已有报告')
    ttk.Label(fa2p, textvariable=a2_status, foreground=C['sub'],
              font=(F, 8)).grid(row=4, column=0, columnspan=3,
                                sticky='w', padx=14, pady=(0, 6))

    # ── 报告路径 (手动选或自动填) ──
    fa2 = ttk.LabelFrame(a2, text=' 报告路径 ', )
    fa2.grid(row=1, column=0, sticky='ew', padx=10, pady=4)
    fa2.grid_columnconfigure(1, weight=1)

    a2_path = tk.StringVar()
    ttk.Entry(fa2, textvariable=a2_path, font=(F, 10)).grid(
        row=0, column=0, sticky='ew', padx=(14, 4), pady=(8, 4), columnspan=2)
    ttk.Button(fa2, text='浏览',
               command=lambda: (f := filedialog.askopenfilename(
                   title='选择时序报告',
                   filetypes=[('报告文件','*.rpt;*.txt'),('All','*.*')],
                   initialdir=a2_path.get() or None)) and a2_path.set(f),
               style='Normal.TButton').grid(row=0, column=2, padx=(4, 4), pady=(8, 4))
    ttk.Button(fa2, text='解析', command=lambda: _a2_parse(),
               style='Accent.TButton').grid(row=0, column=3,
               padx=(0, 14), pady=(8, 4))

    # 摘要行
    fa2s = ttk.LabelFrame(a2, text=' 汇总 ', )
    fa2s.grid(row=2, column=0, sticky='ew', padx=10, pady=4)
    a2_wns = tk.StringVar(value='—')
    a2_tns = tk.StringVar(value='—')
    a2_fail = tk.StringVar(value='—')
    for i, (lbl, var) in enumerate([('WNS (最差负 Slack)', a2_wns),
                                     ('TNS (总负 Slack)', a2_tns),
                                     ('违例路径数', a2_fail)]):
        ttk.Label(fa2s, text=lbl, font=(F, 9),
                  foreground=C['sub']).grid(row=0, column=i*2, sticky='w',
                  padx=(14 if i==0 else 8, 2), pady=(6, 6))
        ttk.Label(fa2s, textvariable=var, font=(M, 10, 'bold'),
                  foreground=C['red']).grid(row=0, column=i*2+1, sticky='w',
                  padx=(0, 14), pady=(6, 6))

    # 路径列表
    a2_tree = ttk.Treeview(a2, columns=('slack','src','dst','logic'),
                          show='headings', selectmode='browse', style='Treeview')
    a2_tree.heading('slack', text='Slack')
    a2_tree.heading('src',   text='Source')
    a2_tree.heading('dst',   text='Destination')
    a2_tree.heading('logic', text='Logic Levels')
    a2_tree.column('slack', width=120)
    a2_tree.column('src',   width=280)
    a2_tree.column('dst',   width=280)
    a2_tree.column('logic', width=160)
    a2_tree.grid(row=5, column=0, sticky='nsew', padx=10, pady=(0, 8))
    a2_sc = ttk.Scrollbar(a2, orient='vertical', command=a2_tree.yview)
    a2_sc.grid(row=5, column=1, sticky='ns', pady=(0, 8))
    a2_tree.configure(yscrollcommand=a2_sc.set)

    def _a2_parse():
        a2_tree.delete(*a2_tree.get_children())
        fp = a2_path.get().strip()
        if not fp or not os.path.isfile(fp):
            messagebox.showerror('错误', '请选择有效的时序报告文件')
            return
        try:
            with open(fp, 'r', encoding='utf-8', errors='replace') as f:
                text = f.read()
        except OSError as e:
            messagebox.showerror('读取失败', str(e))
            return

        # 提取 WNS / TNS
        wns = tns = None
        m_wns = re.search(r'WNS[:\s]*([+-]?\d+\.?\d*)\s*(ns|ps)?', text, re.I)
        m_tns = re.search(r'TNS[:\s]*([+-]?\d+\.?\d*)\s*(ns|ps)?', text, re.I)
        if m_wns: wns = float(m_wns.group(1))
        if m_tns: tns = float(m_tns.group(1))

        a2_wns.set(f'{wns:.3f} ns' if wns is not None else '未找到')
        a2_tns.set(f'{tns:.3f} ns' if tns is not None else '未找到')
        if wns is not None:
            a2_wns.set(f'{wns:.3f} ns {"⚠" if wns < 0 else "✅"}')

        # 提取违例路径
        # 匹配格式: Slack (VIOLATED) : -0.123ns
        #            Source: ... Destination: ... Logic Levels: ...
        paths = []
        # 方式1: 结构化解析
        pattern = re.compile(
            r'Slack\s*\(?VIOLATED\)?\s*:\s*([+-]?\d+\.?\d*)\s*(ns|ps)?.*?'
            r'Source:\s*(.+?)\n.*?'
            r'Destination:\s*(.+?)\n.*?'
            r'(?:.*?\n)*?'
            r'Logic Levels:\s*(.+?)(?:\n|$)',
            re.IGNORECASE | re.DOTALL)

        for m in pattern.finditer(text):
            slack = float(m.group(1))
            src  = m.group(3).strip()
            dst  = m.group(4).strip()
            logic= m.group(5).strip()
            paths.append((slack, src, dst, logic))

        # 方式2: 如果方式1没找到，尝试更宽松的模式
        if not paths:
            for m in re.finditer(
                r'Slack\s*\(?(VIOLATED|MET)\)?\s*:\s*([+-]?\d+\.?\d*)\s*(ns|ps)?',
                text, re.I):
                slack = float(m.group(2))
                if slack < 0:
                    # 尝试提取上下文
                    ctx_start = max(0, m.start() - 200)
                    ctx = text[ctx_start:m.end() + 300]
                    src_m = re.search(r'Source:\s*(.+?)\n', ctx)
                    dst_m = re.search(r'Destination:\s*(.+?)\n', ctx)
                    logic_m = re.search(r'Logic Levels:\s*(.+?)(?:\n|$)', ctx)
                    paths.append((
                        slack,
                        src_m.group(1).strip() if src_m else '?',
                        dst_m.group(1).strip() if dst_m else '?',
                        logic_m.group(1).strip() if logic_m else '?'
                    ))

        a2_fail.set(str(len(paths)))

        for slack, src, dst, logic in paths:
            tag = 'violated' if slack < 0 else 'met'
            a2_tree.insert('', 'end',
                          values=(f'{slack:.3f} ns', src, dst, logic),
                          tags=(tag,))

        a2_tree.tag_configure('violated', foreground=C['red'])
        a2_tree.tag_configure('met', foreground=C['green'])

    # ══════════════════════════════════════════════════════════
    # 自动检测 / 生成 时序报告
    # ══════════════════════════════════════════════════════════
    def _find_xpr_recursive(project_dir, max_depth=3):
        """递归查找 .xpr 文件 (支持粗略路径)"""
        if not project_dir or not os.path.isdir(project_dir):
            return None
        try:
            for f in os.listdir(project_dir):
                if f.lower().endswith('.xpr'):
                    return os.path.join(project_dir, f)
            if max_depth > 0:
                for item in os.listdir(project_dir):
                    sub = os.path.join(project_dir, item)
                    if os.path.isdir(sub) and not item.startswith('.') and item not in ('.git', '__pycache__', '.Xil'):
                        result = _find_xpr_recursive(sub, max_depth - 1)
                        if result:
                            return result
        except (OSError, PermissionError):
            pass
        return None

    def _a2_find_xpr(project_dir):
        return _find_xpr_recursive(project_dir)

    def _a2_resolve_vivado(vivado_path):
        """解析 Vivado 可执行路径: 优先版本匹配 → 全局第一个 → PATH"""
        if vivado_path and os.path.isfile(vivado_path):
            return vivado_path
        dirs = _t9_viv_dirs()
        # 按版本匹配
        ver = a2_ver.get().strip()
        if ver and dirs:
            for d in dirs:
                if ver in d.replace('\\', '/'):
                    for exe in ('vivado.exe','vivado.bat','vivado'):
                        fp = os.path.join(d, exe)
                        if os.path.isfile(fp): return fp
        # 回退第一个
        if dirs:
            for exe in ('vivado.exe','vivado.bat','vivado'):
                fp = os.path.join(dirs[0], exe)
                if os.path.isfile(fp): return fp
        import shutil
        return shutil.which('vivado') or shutil.which('vivado.bat')

    def _a2_scan_reports(project_dir):
        """扫描工程下所有可能的时序报告文件"""
        cands = []
        if not project_dir or not os.path.isdir(project_dir):
            return cands
        # 优先: *.runs/impl_1/*_timing_summary_routed.rpt (Vivado 默认)
        runs = os.path.join(project_dir, f'{os.path.basename(project_dir)}.runs')
        if not os.path.isdir(runs):
            # 找 .runs 目录
            for d in os.listdir(project_dir):
                full = os.path.join(project_dir, d)
                if d.endswith('.runs') and os.path.isdir(full):
                    runs = full
                    break
        if os.path.isdir(runs):
            for step in ('impl_1', 'synth_1'):
                sd = os.path.join(runs, step)
                if os.path.isdir(sd):
                    for f in os.listdir(sd):
                        if f.endswith('.rpt') and ('timing' in f.lower() or 'utilization' in f.lower() or 'route' in f.lower() or 'opt' in f.lower()):
                            cands.append(os.path.join(sd, f))
        # 次选: 工程根下所有 .rpt
        for f in os.listdir(project_dir):
            if f.endswith('.rpt'):
                cands.append(os.path.join(project_dir, f))
        return cands

    def _a2_detect_and_gen():
        """主入口: 检测报告存在性, 缺哪个用 Vivado 生成哪个"""
        proj = a2_proj.get().strip()
        viv = a2_vivado.get().strip()
        if not proj:
            messagebox.showerror('错误', '请先填写 Vivado 工程路径 (含 .xpr 的目录)')
            return
        xpr = _a2_find_xpr(proj)
        if not xpr:
            messagebox.showerror('错误', f'在 {proj} 下未找到 .xpr 工程文件')
            return

        existing = _a2_scan_reports(proj)
        # 区分: 已经有时序摘要 (impl) vs 仅 synth 报告
        has_impl_timing = any(('impl' in p) and ('timing' in p.lower() or 'route' in p.lower()) for p in existing)
        has_synth = any('synth_1' in p for p in existing)

        a2_status.set(f'检测到 {len(existing)} 个已有报告' + (' (含时序)' if has_impl_timing else ''))

        if has_impl_timing:
            # 直接用最新 impl 时序报告
            best = max((p for p in existing if 'impl' in p and 'timing' in p.lower()),
                       key=os.path.getmtime, default=None)
            if best is None:
                best = max(existing, key=os.path.getmtime)
            a2_path.set(best)
            a2_status.set(f'✔ 找到时序报告: {os.path.basename(best)}  —  自动填入')
            _a2_parse()
            return

        # 缺时序报告 → 跑 Vivado 生成
        if not has_synth:
            # 先 synth, 再 impl 才能出时序
            a2_status.set('⏳ 工程还未综合, 先跑 synth_1 + impl_1 ... (可能 5-30 分钟)')
            _a2_run_vivado(xpr, viv,
                [
                  'open_project "{}"'.format(xpr),
                  'launch_runs synth_1 -to_step write_synth_checkpoint -jobs 4',
                  'wait_on_run synth_1',
                  'launch_runs impl_1 -to_step write_bitstream -jobs 4',
                  'wait_on_run impl_1',
                  'open_run impl_1',
                  'report_timing_summary -delay_type max -max_paths 10 -file [get_property DIRECTORY [current_run]]/impl_1_timing_summary_routed.rpt',
                  'report_timing_summary -delay_type min -max_paths 10 -file [get_property DIRECTORY [current_run]]/impl_1_timing_summary_routed_min.rpt',
                  'exit',
                ],
                on_done=lambda: _a2_after_gen(xpr))
        else:
            # 已有 synth, 跑 impl 出时序
            a2_status.set('⏳ 工程已综合, 跑 impl_1 生成时序报告 ... (可能 5-15 分钟)')
            _a2_run_vivado(xpr, viv,
                [
                  'open_project "{}"'.format(xpr),
                  'launch_runs impl_1 -to_step write_bitstream -jobs 4',
                  'wait_on_run impl_1',
                  'open_run impl_1',
                  'report_timing_summary -delay_type max -max_paths 10 -file [get_property DIRECTORY [current_run]]/impl_1_timing_summary_routed.rpt',
                  'report_timing_summary -delay_type min -max_paths 10 -file [get_property DIRECTORY [current_run]]/impl_1_timing_summary_routed_min.rpt',
                  'exit',
                ],
                on_done=lambda: _a2_after_gen(xpr))

    def _a2_after_gen(xpr):
        """Vivado 跑完后重新扫描, 填入最新报告并解析"""
        proj = a2_proj.get().strip()
        reports = _a2_scan_reports(proj)
        # 找带 timing 的
        timing_reports = [p for p in reports if 'timing' in p.lower()]
        if timing_reports:
            best = max(timing_reports, key=os.path.getmtime)
            a2_path.set(best)
            a2_status.set(f'✔ Vivado 已生成: {os.path.basename(best)}  —  自动解析')
            _a2_parse()
        else:
            a2_status.set('⚠ Vivado 跑完但未找到 timing 报告, 请检查工程')

    def _a2_run_vivado(xpr, vivado_path, tcl_cmds, on_done=None):
        """在后台线程跑 vivado -mode batch -source <tmp.tcl>"""
        viv_bin = _a2_resolve_vivado(vivado_path)
        if not viv_bin:
            messagebox.showerror('错误', '找不到 vivado 可执行文件, 请手动指定 Vivado 路径')
            a2_status.set('✘ 找不到 vivado, 请在"Vivado 路径"字段指定')
            return
        # 写临时 tcl
        tmp_tcl = os.path.join(os.path.dirname(xpr), '_a2_gen_reports.tcl')
        try:
            with open(tmp_tcl, 'w', encoding='utf-8') as f:
                f.write('\n'.join(tcl_cmds))
        except OSError as e:
            messagebox.showerror('错误', f'写 tcl 失败: {e}')
            return

        a2_status.set('⏳ Vivado 后台运行中 ... (可继续浏览工具其他 Tab)')

        def _worker():
            try:
                # Vivado 在 Windows 上需要 bat, Unix 直接可执行
                cmd = [viv_bin, '-mode', 'batch', '-source', tmp_tcl]
                if os.name == 'nt' and not viv_bin.lower().endswith('.bat'):
                    cmd = ['cmd', '/c', viv_bin] + cmd[1:]
                proc = subprocess.run(cmd, capture_output=True, text=True,
                                      timeout=60*30, encoding='utf-8', errors='replace')
                rc = proc.returncode
                if rc == 0:
                    root.after(0, lambda: a2_status.set(
                        '✔ Vivado 执行完成 (rc=0), 正在重新扫描报告 ...'))
                    if on_done:
                        root.after(0, on_done)
                else:
                    err_tail = (proc.stdout or '')[-500:] + '\n' + (proc.stderr or '')[-500:]
                    root.after(0, lambda: a2_status.set(
                        f'✘ Vivado 返回 rc={rc} — 末尾输出:\n{err_tail[:300]}'))
            except subprocess.TimeoutExpired:
                root.after(0, lambda: a2_status.set('✘ Vivado 超过 30 分钟超时'))
            except Exception as e:
                root.after(0, lambda: a2_status.set(f'✘ Vivado 调用失败: {e}'))
            finally:
                # 清理临时 tcl
                try:
                    os.remove(tmp_tcl)
                except OSError:
                    pass

        threading.Thread(target=_worker, daemon=True).start()

    # ═══════════════════════════════════
    # 子页 3 — Log/报告清洗
    # ═══════════════════════════════════
    a3 = ttk.Frame(aux_nb, style='TFrame')
    aux_nb.add(a3, text='  Log 清洗  ')
    a3.grid_columnconfigure(0, weight=1)
    a3.grid_rowconfigure(5, weight=1)

    # ── 工程根 + Vivado 路径 (无 log 时自动跑 synth 生成 vivado.log) ──
    fa3p = ttk.LabelFrame(a3, text=' 工程信息 (缺 log 时自动调 Vivado 综合) ')
    fa3p.grid(row=0, column=0, sticky='ew', padx=10, pady=(8, 4))
    fa3p.grid_columnconfigure(1, weight=1)

    a3_proj = tk.StringVar()
    a3_vivado = tk.StringVar()
    a3_ver = tk.StringVar()

    ttk.Label(fa3p, text='工程路径:', font=(F, 9)).grid(
        row=0, column=0, sticky='w', padx=(14, 4), pady=(6, 2))
    ttk.Entry(fa3p, textvariable=a3_proj, font=(F, 9)).grid(
        row=0, column=1, sticky='ew', padx=2, pady=(6, 2))
    ttk.Button(fa3p, text='浏览',
               command=lambda: (d := filedialog.askdirectory(
                   title='选择 Vivado 工程根目录')) and a3_proj.set(d),
               style='Normal.TButton').grid(row=0, column=2, padx=(4, 14), pady=(6, 2))

    # Vivado 路径 — 来自 ⚙设置 Tab
    ttk.Label(fa3p, text='Vivado 路径:', font=(F, 9)).grid(
        row=1, column=0, sticky='w', padx=(14, 4), pady=(2, 2))
    _a3_viv_var = tk.StringVar()
    _a3_viv_label = ttk.Label(fa3p, textvariable=_a3_viv_var, font=(F, 8))
    _a3_viv_label.grid(row=1, column=1, sticky='w', padx=2, pady=(2, 2))
    _a3_viv_btn = ttk.Button(fa3p, text='⚙ 打开设置', command=lambda: nb.select(t16),
                              style='Small.TButton')

    def _a3_refresh_viv():
        dirs = _t9_viv_dirs()
        if dirs:
            _a3_viv_var.set(f'✔ 已配置 {len(dirs)} 个 Vivado 路径')
            _a3_viv_label.config(foreground=C['green'])
            a3_vivado.set('')  # 由 _a3_resolve_vivado 按版本匹配
            _a3_viv_btn.grid_remove()
        else:
            _a3_viv_var.set('✘ 未配置 — 请到 ⚙设置 Tab 添加')
            _a3_viv_label.config(foreground=C['red'])
            a3_vivado.set('')
            _a3_viv_btn.grid(row=1, column=2, padx=(4, 14), pady=(2, 2))
        a3_ver_combo['values'] = _t9_extract_versions()

    ttk.Label(fa3p, text='版本:', font=(F, 9)).grid(
        row=2, column=0, sticky='w', padx=(14, 4), pady=(2, 6))
    a3_ver_combo = ttk.Combobox(fa3p, textvariable=a3_ver, font=(F, 9),
                                 width=10, values=_t9_extract_versions())
    a3_ver_combo.grid(row=2, column=1, sticky='w', padx=2, pady=(2, 6))
    _a3_refresh_viv()
    ttk.Label(fa3p, text='  ⓘ 留空时按 PATH 自动找 vivado',
              foreground=C['sub'], font=(F, 8)).grid(
        row=2, column=2, sticky='w', padx=(4, 14), pady=(2, 6))

    ttk.Button(fa3p, text='🔍 检测 / 自动生成 vivado.log',
               command=lambda: _a3_detect_and_gen(),
               style='Accent.TButton').grid(row=3, column=0, columnspan=3,
                                            sticky='ew', padx=14, pady=(2, 8))
    a3_status = tk.StringVar(value='点击上面按钮自动检测/生成 vivado.log, 或直接在下方选已有 log')
    ttk.Label(fa3p, textvariable=a3_status, foreground=C['sub'],
              font=(F, 8)).grid(row=4, column=0, columnspan=3,
                                sticky='w', padx=14, pady=(0, 6))

    # ── Log 路径 (手动选或自动填) ──
    fa3 = ttk.LabelFrame(a3, text=' Log / 报告文件 ', )
    fa3.grid(row=1, column=0, sticky='ew', padx=10, pady=4)
    fa3.grid_columnconfigure(1, weight=1)

    a3_path = tk.StringVar()
    ttk.Entry(fa3, textvariable=a3_path, font=(F, 10)).grid(
        row=0, column=0, sticky='ew', padx=(14, 4), pady=(8, 4), columnspan=2)
    ttk.Button(fa3, text='浏览',
               command=lambda: (f := filedialog.askopenfilename(
                   title='选择 Log 文件',
                   filetypes=[('Log/报告','*.log;*.rpt;*.txt'),('All','*.*')])) and a3_path.set(f),
               style='Normal.TButton').grid(row=0, column=2, padx=(4, 4), pady=(8, 4))
    ttk.Button(fa3, text='清洗', command=lambda: _a3_clean(),
               style='Accent.TButton').grid(row=0, column=3,
               padx=(0, 14), pady=(8, 4))

    # 过滤选项
    fa3o = ttk.LabelFrame(a3, text=' 关注类型 ', )
    fa3o.grid(row=2, column=0, sticky='ew', padx=10, pady=4)

    _WARN_TYPES = [
        ('latch',     'Latch 推断',        r'inferring latch|inferred latch|LATCH'),
        ('unconnected','端口未连接',        r'unconnected port|has no connect|is not connected'),
        ('undriven',  '未驱动',             r'undriven|has no driver|without a driver'),
        ('multidrv',  '多驱动',             r'multi[- ]driven|multiple drivers'),
        ('comb_loop', '组合环',             r'combinational loop|feedback loop'),
        ('width_mis', '位宽不匹配',         r'width mismatch|width.*truncat|size mismatch'),
        ('unused',    '未使用寄存器/信号',   r'unused sequential|unused.*removed|removed'),
        ('blackbox',  '黑盒/空模块',         r'black box|empty module|unresolved'),
        ('timing',    '时序违例',            r'timing.*not met|timing violation|VIOLATED'),
        ('dsp_bram',  'DSP/BRAM 级联警告',  r'(DSP|BRAM).*cascade|cascade.*limit'),
        ('clk_skew',  '时钟歪斜/不稳定',     r'clock.*skew|phase.*mismatch|clock.*unstable'),
        ('other',     '其他 Warning',       r'^(WARNING|CRITICAL WARNING|Warning)'),
    ]

    a3_filters = {}
    for i, (key, label, _) in enumerate(_WARN_TYPES):
        var = tk.BooleanVar(value=True)
        a3_filters[key] = var
        ttk.Checkbutton(fa3o, text=label, variable=var,
                        style='TCheckbutton').grid(
            row=i // 4, column=i % 4, sticky='w',
            padx=(14 if i%4==0 else 4, 4), pady=3)

    # 统计 + 详情
    a3_tree = ttk.Treeview(a3, columns=('type','line','message'),
                           show='headings', selectmode='browse', style='Treeview')
    a3_tree.heading('type',    text='类型')
    a3_tree.heading('line',    text='行号')
    a3_tree.heading('message', text='消息')
    a3_tree.column('type',    width=130)
    a3_tree.column('line',    width=60)
    a3_tree.column('message', width=600)
    a3_tree.grid(row=5, column=0, sticky='nsew', padx=10, pady=(0, 8))
    a3_sc = ttk.Scrollbar(a3, orient='vertical', command=a3_tree.yview)
    a3_sc.grid(row=5, column=1, sticky='ns', pady=(0, 8))
    a3_tree.configure(yscrollcommand=a3_sc.set)

    # 统计标签
    a3_stats = tk.StringVar(value='就绪')
    ttk.Label(a3, textvariable=a3_stats, foreground=C['sub'],
              font=(F, 9)).grid(row=6, column=0, sticky='w', padx=16,
              pady=(0, 8))

    def _a3_clean():
        a3_tree.delete(*a3_tree.get_children())
        fp = a3_path.get().strip()
        if not fp or not os.path.isfile(fp):
            messagebox.showerror('错误', '请选择有效的 Log 文件')
            return
        try:
            with open(fp, 'r', encoding='utf-8', errors='replace') as f:
                lines = f.readlines()
        except OSError as e:
            messagebox.showerror('读取失败', str(e))
            return

        # 按类型分类
        categorized = {k: [] for k in a3_filters}
        for lineno, line in enumerate(lines, 1):
            for key, var in a3_filters.items():
                if not var.get():
                    continue
                _, _, pattern = _WARN_TYPES[
                    next(i for i, w in enumerate(_WARN_TYPES) if w[0] == key)]
                try:
                    if re.search(pattern, line, re.IGNORECASE):
                        categorized[key].append((lineno, line.strip()))
                        break  # 每条消息只归入第一个匹配类型
                except re.error:
                    pass

        # 显示
        type_labels = {w[0]: w[1] for w in _WARN_TYPES}
        total = 0
        stats_parts = []
        for key in categorized:
            cat = categorized[key]
            if not a3_filters[key].get():
                continue
            total += len(cat)
            if cat:
                stats_parts.append(f'{type_labels[key]}: {len(cat)}')
            # 最多显示 500 条
            for lineno, msg in cat[:500]:
                a3_tree.insert('', 'end',
                              values=(type_labels[key], str(lineno), msg[:200]))

        a3_stats.set('  |  '.join(stats_parts) if stats_parts else '未匹配到警告')
        if total > 500:
            a3_stats.set(a3_stats.get() + f'  (显示前500 / 共{total})')

    # ══════════════════════════════════════════════════════════
    # 自动检测 / 生成 vivado.log
    # ══════════════════════════════════════════════════════════
    def _a3_find_xpr(project_dir):
        return _find_xpr_recursive(project_dir)

    def _a3_resolve_vivado(vivado_path):
        # 1) 用户指定的路径
        if vivado_path and os.path.isfile(vivado_path):
            return vivado_path
        # 2) 全局设置 (按版本匹配, 只找 vivado)
        dirs = _t9_viv_dirs()
        ver = a3_ver.get().strip()
        if ver and dirs:
            for d in dirs:
                if ver in d.replace('\\', '/'):
                    for exe in ('vivado.exe','vivado.bat','vivado'):
                        fp = os.path.join(d, exe)
                        if os.path.isfile(fp): return fp
        # 3) 回退第一个 / PATH
        if dirs:
            for exe in ('vivado.exe','vivado.bat','vivado'):
                fp = os.path.join(dirs[0], exe)
                if os.path.isfile(fp): return fp
        import shutil
        return shutil.which('vivado') or shutil.which('vivado.bat')

    def _a3_scan_logs(project_dir):
        """扫描工程下 vivado.log / vivado.jou / synth_1/vivado.log"""
        cands = []
        if not project_dir or not os.path.isdir(project_dir):
            return cands
        # 工程根
        for f in ('vivado.log', 'vivado.jou', 'vivado.pb'):
            p = os.path.join(project_dir, f)
            if os.path.isfile(p):
                cands.append(p)
        # runs 子目录
        for d in os.listdir(project_dir):
            if not d.endswith('.runs'):
                continue
            runs = os.path.join(project_dir, d)
            for step in os.listdir(runs):
                sd = os.path.join(runs, step)
                if not os.path.isdir(sd):
                    continue
                for f in ('vivado.log', 'vivado.pb', 'runme.log'):
                    p = os.path.join(sd, f)
                    if os.path.isfile(p):
                        cands.append(p)
        return cands

    def _a3_detect_and_gen():
        proj = a3_proj.get().strip()
        viv = a3_vivado.get().strip()
        if not proj:
            messagebox.showerror('错误', '请先填写 Vivado 工程路径 (含 .xpr 的目录)')
            return
        xpr = _a3_find_xpr(proj)
        if not xpr:
            messagebox.showerror('错误', f'在 {proj} 下未找到 .xpr 工程文件')
            return

        existing = _a3_scan_logs(proj)
        # 优先选最大的 vivado.log (synth 阶段 log)
        vivado_logs = [p for p in existing if p.endswith('vivado.log')]
        if vivado_logs:
            best = max(vivado_logs, key=os.path.getsize)
            a3_path.set(best)
            a3_status.set(f'✔ 找到 log: {os.path.basename(best)} ({os.path.getsize(best)//1024} KB) — 自动清洗')
            _a3_clean()
            return

        # 没有 log → 跑 synth 生成
        a3_status.set('⏳ 工程无 log, 调 Vivado 跑 synth_1 生成 vivado.log ... (3-15 分钟)')
        _a3_run_vivado(xpr, viv,
            [
              'open_project "{}"'.format(xpr),
              'launch_runs synth_1 -to_step write_synth_checkpoint -jobs 4',
              'wait_on_run synth_1',
              'exit',
            ],
            on_done=lambda: _a3_after_gen(xpr))

    def _a3_after_gen(xpr):
        proj = a3_proj.get().strip()
        logs = _a3_scan_logs(proj)
        vivado_logs = [p for p in logs if p.endswith('vivado.log')]
        if vivado_logs:
            best = max(vivado_logs, key=os.path.getsize)
            a3_path.set(best)
            a3_status.set(f'✔ Vivado 已生成: {os.path.basename(best)} ({os.path.getsize(best)//1024} KB) — 自动清洗')
            _a3_clean()
        else:
            a3_status.set('⚠ Vivado 跑完但未找到 vivado.log, 请检查工程')

    def _a3_run_vivado(xpr, vivado_path, tcl_cmds, on_done=None):
        viv_bin = _a3_resolve_vivado(vivado_path)
        if not viv_bin:
            messagebox.showerror('错误', '找不到 vivado 可执行文件, 请手动指定 Vivado 路径')
            a3_status.set('✘ 找不到 vivado, 请在"Vivado 路径"字段指定')
            return
        tmp_tcl = os.path.join(os.path.dirname(xpr), '_a3_gen_log.tcl')
        try:
            with open(tmp_tcl, 'w', encoding='utf-8') as f:
                f.write('\n'.join(tcl_cmds))
        except OSError as e:
            messagebox.showerror('错误', f'写 tcl 失败: {e}')
            return

        a3_status.set('⏳ Vivado 后台运行中 ... (可继续浏览工具其他 Tab)')

        def _worker():
            try:
                cmd = [viv_bin, '-mode', 'batch', '-source', tmp_tcl]
                if os.name == 'nt' and not viv_bin.lower().endswith('.bat'):
                    cmd = ['cmd', '/c', viv_bin] + cmd[1:]
                proc = subprocess.run(cmd, capture_output=True, text=True,
                                      timeout=60*30, encoding='utf-8', errors='replace')
                rc = proc.returncode
                if rc == 0:
                    root.after(0, lambda: a3_status.set(
                        '✔ Vivado 执行完成 (rc=0), 正在重新扫描 log ...'))
                    if on_done:
                        root.after(0, on_done)
                else:
                    err_tail = (proc.stdout or '')[-500:] + '\n' + (proc.stderr or '')[-500:]
                    root.after(0, lambda: a3_status.set(
                        f'✘ Vivado 返回 rc={rc} — 末尾输出:\n{err_tail[:300]}'))
            except subprocess.TimeoutExpired:
                root.after(0, lambda: a3_status.set('✘ Vivado 超过 30 分钟超时'))
            except Exception as e:
                root.after(0, lambda: a3_status.set(f'✘ Vivado 调用失败: {e}'))
            finally:
                try:
                    os.remove(tmp_tcl)
                except OSError:
                    pass

        threading.Thread(target=_worker, daemon=True).start()

    # ═══════════════════════════════════
    # 子页 4 — 代码模板 (Verilog / VHDL 常用片段)
    # ═══════════════════════════════════
    a4 = ttk.Frame(aux_nb, style='TFrame')
    aux_nb.add(a4, text='  代码模板  ')
    a4.grid_columnconfigure(0, weight=1)
    a4.grid_rowconfigure(2, weight=1)

    try:
        from _t9_templates import TEMPLATES as _T9_TPL
    except Exception as _e:
        _T9_TPL = [('err', f'加载失败: {_e}', '-- 模板库未找到 --', '-- 模板库未找到 --')]

    # 上: 选择行
    fa4 = ttk.Frame(a4, style='TFrame')
    fa4.grid(row=0, column=0, sticky='ew', padx=10, pady=(8, 4))
    ttk.Label(fa4, text='模板:', font=(F, 10)).pack(side='left', padx=(4, 6))
    a4_tpl_var = tk.StringVar()
    a4_combo = ttk.Combobox(fa4, textvariable=a4_tpl_var,
                            values=[f'{name}  ({key})' for key, name, _, _ in _T9_TPL],
                            state='readonly', width=42, font=(F, 10))
    a4_combo.pack(side='left', padx=(0, 8), pady=4)
    if _T9_TPL:
        a4_combo.current(0)
    ttk.Label(fa4, text='Verilog / VHDL 都提供, 可直接复制',
              foreground=C['sub'], font=(F, 9)).pack(side='left', padx=(6, 0))

    # 中: 代码区 (Notebook 切 Verilog / VHDL)
    a4_nb = ttk.Notebook(a4)
    a4_nb.grid(row=2, column=0, sticky='nsew', padx=10, pady=(4, 4))

    a4_vf = ttk.Frame(a4_nb, style='TFrame')
    a4_vf.grid_rowconfigure(0, weight=1)
    a4_vf.grid_columnconfigure(0, weight=1)
    a4_nb.add(a4_vf, text='  Verilog  ')
    a4_v_text = tk.Text(a4_vf, font=(M, 9), bg=C['ebg'], fg=C['fg'],
                        relief='flat', padx=12, pady=10, wrap='none')
    a4_v_text.grid(row=0, column=0, sticky='nsew')
    a4_vs = ttk.Scrollbar(a4_vf, orient='vertical', command=a4_v_text.yview)
    a4_vs.grid(row=0, column=1, sticky='ns')
    a4_v_text.configure(yscrollcommand=a4_vs.set)

    a4_hf = ttk.Frame(a4_nb, style='TFrame')
    a4_hf.grid_rowconfigure(0, weight=1)
    a4_hf.grid_columnconfigure(0, weight=1)
    a4_nb.add(a4_hf, text='  VHDL  ')
    a4_h_text = tk.Text(a4_hf, font=(M, 9), bg=C['ebg'], fg=C['fg'],
                        relief='flat', padx=12, pady=10, wrap='none')
    a4_h_text.grid(row=0, column=0, sticky='nsew')
    a4_hs = ttk.Scrollbar(a4_hf, orient='vertical', command=a4_h_text.yview)
    a4_hs.grid(row=0, column=1, sticky='ns')
    a4_h_text.configure(yscrollcommand=a4_hs.set)

    # 下: 操作按钮
    fa4b = ttk.Frame(a4, style='TFrame')
    fa4b.grid(row=3, column=0, sticky='ew', padx=10, pady=(4, 8))
    ttk.Button(fa4b, text='📋  复制 Verilog',
               command=lambda: (_t9_clip(a4_v_text.get('1.0', 'end')), None)[1],
               style='Normal.TButton').pack(side='left', padx=(0, 6))
    ttk.Button(fa4b, text='📋  复制 VHDL',
               command=lambda: (_t9_clip(a4_h_text.get('1.0', 'end')), None)[1],
               style='Normal.TButton').pack(side='left', padx=(0, 6))
    ttk.Button(fa4b, text='💾  保存为 .v',
               command=lambda: _t9_save(a4_v_text.get('1.0', 'end'), '.v'),
               style='Normal.TButton').pack(side='left', padx=(0, 6))
    ttk.Button(fa4b, text='💾  保存为 .vhd',
               command=lambda: _t9_save(a4_h_text.get('1.0', 'end'), '.vhd'),
               style='Normal.TButton').pack(side='left', padx=(0, 6))

    def _t9_clip(txt):
        try:
            root.clipboard_clear()
            root.clipboard_append(txt.rstrip())
            root.update()
        except Exception:
            pass

    def _t9_save(content, ext):
        from tkinter import filedialog as _fd
        fp = _fd.asksaveasfilename(
            title='保存模板代码',
            defaultextension=ext,
            filetypes=[(f'{ext} 文件', f'*{ext}'), ('All', '*.*')])
        if fp:
            try:
                with open(fp, 'w', encoding='utf-8') as f:
                    f.write(content.rstrip() + '\n')
                messagebox.showinfo('已保存', f'模板已保存到:\n{fp}')
            except OSError as e:
                messagebox.showerror('保存失败', str(e))

    def _a4_update(*_):
        sel = a4_tpl_var.get()
        # 从显示串里找 key
        for key, name, v_code, h_code in _T9_TPL:
            if sel.startswith(name):
                a4_v_text.delete('1.0', 'end')
                a4_v_text.insert('1.0', v_code)
                a4_h_text.delete('1.0', 'end')
                a4_h_text.insert('1.0', h_code)
                break

    a4_combo.bind('<<ComboboxSelected>>', _a4_update)
    _a4_update()  # 初始化显示第一个模板

    # ═══════════════════════════════════════
    # 子页 5 — 仿真报告分析 (xvlog 日志)
    # ═══════════════════════════════════════
    a5 = ttk.Frame(aux_nb, style='TFrame')
    aux_nb.add(a5, text='  仿真报告  ')
    a5.grid_columnconfigure(0, weight=1)
    a5.grid_rowconfigure(2, weight=1)

    # Row 0 — 工程路径 + 扫描
    fc5 = ttk.LabelFrame(a5, text=' 工程路径 ')
    fc5.grid(row=0, column=0, sticky='ew', padx=12, pady=(10, 4))
    fc5.grid_columnconfigure(0, weight=1)

    a5_path = tk.StringVar()
    ttk.Entry(fc5, textvariable=a5_path, font=(F, 10)).grid(
        row=0, column=0, sticky='ew', padx=(14, 4), pady=(8, 4))
    ttk.Button(fc5, text='浏览', command=lambda: (
        d := filedialog.askdirectory(title='选择工程目录')) and a5_path.set(d),
        style='Normal.TButton').grid(row=0, column=1, padx=4, pady=(8, 4))
    ttk.Button(fc5, text='🔍 扫描', command=lambda: _a5_scan(),
               style='Accent.TButton').grid(row=0, column=2,
               padx=(0, 14), pady=(8, 4))

    # Row 1 — 文件列表
    a5_list_frame = ttk.LabelFrame(a5, text=' xvlog 文件列表 ')
    a5_list_frame.grid(row=1, column=0, sticky='ew', padx=12, pady=(2, 4))

    a5_file_var = tk.StringVar(value='')
    a5_file_combo = ttk.Combobox(a5_list_frame, textvariable=a5_file_var,
                                  state='readonly', font=(F, 10))
    a5_file_combo.pack(fill='x', padx=14, pady=(6, 4))
    a5_file_combo.bind('<<ComboboxSelected>>', lambda e: _a5_load_file())

    a5_status = tk.StringVar(value='请选择工程路径后点击扫描')
    ttk.Label(a5_list_frame, textvariable=a5_status, font=(F, 9),
              foreground=C['sub']).pack(fill='x', padx=14, pady=(0, 6))

    # Row 2 — 日志内容
    a5_content_frame = ttk.LabelFrame(a5, text=' 日志内容 ')
    a5_content_frame.grid(row=2, column=0, sticky='nsew', padx=12, pady=(2, 8))
    a5_content_frame.grid_rowconfigure(0, weight=1)
    a5_content_frame.grid_columnconfigure(0, weight=1)

    a5_text = tk.Text(a5_content_frame, font=(M, 9), bg='#fafbfc', fg=C['fg'],
                      relief='flat', padx=10, pady=6, wrap='none')
    a5_text.grid(row=0, column=0, sticky='nsew')
    a5_scroll_y = ttk.Scrollbar(a5_content_frame, orient='vertical',
                                 command=a5_text.yview)
    a5_scroll_y.grid(row=0, column=1, sticky='ns')
    a5_text.configure(yscrollcommand=a5_scroll_y.set)
    a5_scroll_x = ttk.Scrollbar(a5_content_frame, orient='horizontal',
                                 command=a5_text.xview)
    a5_scroll_x.grid(row=1, column=0, sticky='ew')
    a5_text.configure(xscrollcommand=a5_scroll_x.set)
    a5_text.bind('<Key>', lambda e: 'break')  # 只读

    # --- 数据 ---
    a5_files = []  # [(rel_path, full_path), ...]

    def _a5_scan():
        nonlocal a5_files
        a5_files = []
        a5_text.delete('1.0', 'end')
        a5_file_combo['values'] = []
        a5_file_var.set('')
        root_p = a5_path.get().strip()
        if not root_p or not os.path.isdir(root_p):
            a5_status.set('⚠ 请选择有效的工程目录')
            return
        # 递归扫描 xvlog*.log
        import fnmatch
        found = []
        for dirpath, _, filenames in os.walk(root_p):
            for fn in filenames:
                if fnmatch.fnmatch(fn.lower(), 'xvlog*.log'):
                    full = os.path.join(dirpath, fn)
                    rel = os.path.relpath(full, root_p)
                    found.append((rel, full))
        a5_files = sorted(found, key=lambda x: x[0])
        names = [f[0] for f in a5_files]
        a5_file_combo['values'] = names
        if names:
            a5_file_var.set(names[0])
            a5_status.set(f'找到 {len(a5_files)} 个 xvlog 文件')
            _a5_load_file()
        else:
            a5_status.set('未找到 xvlog 文件')

    def _a5_load_file():
        sel = a5_file_var.get()
        for rel, full in a5_files:
            if rel == sel:
                a5_text.delete('1.0', 'end')
                try:
                    with open(full, 'r', encoding='utf-8', errors='replace') as f:
                        a5_text.insert('1.0', f.read())
                    a5_text.see('1.0')
                    a5_status.set(f'已加载: {rel}')
                except Exception as e:
                    a5_text.insert('1.0', f'[读取失败] {e}')
                    a5_status.set(f'⚠ 读取失败: {rel}')
                return

    # ═══════════════════════════════════
    # 子页 6 — Debug 屏蔽 (递归 VHDL, 注释 attribute debug)
    # ═══════════════════════════════════
    a6 = ttk.Frame(aux_nb, style='TFrame')
    aux_nb.add(a6, text='  Debug 屏蔽  ')
    a6.grid_columnconfigure(0, weight=1)
    a6.grid_rowconfigure(4, weight=1)

    fa6 = ttk.LabelFrame(a6, text=' 工程信息 ')
    fa6.grid(row=0, column=0, sticky='ew', padx=10, pady=(8, 4))
    fa6.grid_columnconfigure(1, weight=1)
    a6_path = tk.StringVar()
    ttk.Label(fa6, text='工程路径:', font=(F, 9)).grid(row=0, column=0, sticky='w', padx=(14, 4), pady=(8, 4))
    ttk.Entry(fa6, textvariable=a6_path, font=(F, 9)).grid(row=0, column=1, sticky='ew', padx=2, pady=(8, 4))
    ttk.Button(fa6, text='浏览',
               command=lambda: (d := filedialog.askdirectory(title='选择 FPGA 工程根目录')) and a6_path.set(d),
               style='Normal.TButton').grid(row=0, column=2, padx=(4, 14), pady=(8, 4))

    a6_btn_frame = ttk.Frame(a6, style='TFrame')
    a6_btn_frame.grid(row=1, column=0, sticky='ew', padx=10, pady=(2, 4))
    ttk.Button(a6_btn_frame, text='🔄 扫描并屏蔽', command=lambda: _a6_execute(),
               style='Accent.TButton').pack(side='left', padx=(0, 6))

    a6_status = tk.StringVar(value='选择工程路径，递归扫描 VHDL 文件，屏蔽 attribute debug')
    ttk.Label(a6, textvariable=a6_status, foreground=C['sub'], font=(F, 8)).grid(
        row=2, column=0, sticky='w', padx=16, pady=(2, 2))

    a6_lf, a6_log = _log_widget(a6, 12)
    a6_lf.grid(row=4, column=0, sticky='nsew', padx=10, pady=(2, 8))

    def _a6_execute():
        root_p = a6_path.get().strip()
        if not root_p or not os.path.isdir(root_p):
            messagebox.showerror('错误', '请选择有效的工程目录')
            return

        def _bg():
            def l(msg, color=None):
                root.after(0, lambda: _log(a6_log, msg, color or C['fg']))
            l(f'扫描目录: {root_p}', C['blue'])
            import fnmatch
            vhdl_files = []
            for dirpath, _, filenames in os.walk(root_p):
                for fn in filenames:
                    if fnmatch.fnmatch(fn.lower(), '*.vhd') or fnmatch.fnmatch(fn.lower(), '*.vhdl'):
                        vhdl_files.append(os.path.join(dirpath, fn))
            l(f'找到 {len(vhdl_files)} 个 VHDL 文件', C['sub'])
            total = len(vhdl_files)
            masked = 0
            modified_files = 0

            # 需要屏蔽的调试属性关键词
            debug_patterns = [
                'attribute mark_debug',
                'attribute dont_touch',
                'attribute keep',
                'attribute syn_keep',
                'attribute debug',
            ]

            for i, fp in enumerate(vhdl_files):
                rel = os.path.relpath(fp, root_p)
                try:
                    with open(fp, 'r', encoding='utf-8', errors='replace') as f:
                        lines = f.readlines()
                except OSError as e:
                    l(f'  ✘ {rel}: 读取失败 {e}', C['red'])
                    continue

                changed = False
                for j, line in enumerate(lines):
                    stripped = line.lstrip().lower()
                    for pat in debug_patterns:
                        if pat in stripped and not stripped.startswith('--'):
                            lines[j] = '-- ' + line
                            masked += 1
                            changed = True
                            break

                if changed:
                    try:
                        with open(fp, 'w', encoding='utf-8') as f:
                            f.writelines(lines)
                        l(f'  ✔ {rel}', C['green'])
                        modified_files += 1
                    except OSError as e:
                        l(f'  ✘ {rel}: 写入失败 {e}', C['red'])
                else:
                    l(f'  · {rel} (无变更)', C['sub'])

            t = f'完成: 扫描 {total} 文件, 修改 {modified_files} 文件, 屏蔽 {masked} 行'
            root.after(0, lambda: a6_status.set(t))
            l(t, C['green' if masked else 'sub'])

        threading.Thread(target=_bg, daemon=True).start()

    # ══════════════════════════════════════════════════════════════
    # TAB 10 — 串口助手 (XCOM/SSCOM 风格: 简洁清爽, 单数据框显示收发)
    # ══════════════════════════════════════════════════════════════
    t10 = ttk.Frame(nb, style='TFrame')
    nb.add(t10, text='📡 串口助手')
    # 整体行布局:
    #   row 0: 顶部数据区 (单框, [Rx]/[Tx] 前缀区分收发) + 极简工具栏
    #   row 1: 接收/发送 设置行 (横向流式, 含自动换行/编码/转义/HEX/定时)
    #   row 2: 发送输入区 (3 行: 提示文字 / 输入框 / 文件操作)
    #   row 3: 状态栏 (最右: ⚙ 串口设置 按钮)
    t10.grid_rowconfigure(0, weight=4)   # 数据区
    t10.grid_rowconfigure(1, weight=0)   # 设置行
    t10.grid_rowconfigure(2, weight=2, minsize=100)  # 发送区 (3行)
    t10.grid_rowconfigure(3, weight=0)   # 状态栏
    t10.grid_columnconfigure(0, weight=1)

    # --- pyserial 检测 ---
    # 离线/便携环境: 优先从两条静态路径找, 找不到再尝试 pip 装.
    #   1) _pyserial_lib/serial/    — 仓库自带副本 (开发/源码运行)
    #   2) runtime/Lib/site-packages/ — build_portable_runtime.py 装进去的 (便携)
    #   3) 兜底: 尝试 pip 装 (会失败于无网环境, 不影响启动)
    _HAS_SERIAL = False
    _SERIAL_INSTALL_ERR = ''
    _SERIAL_HINT = ''   # 离线环境友好的"手动解决"提示

    def _try_import_serial_from(p):
        """把目录 p 临时塞进 sys.path 头, 试 import serial,
        成功返回 serial 模块, 失败返回 None. 不改全局状态."""
        if not p or not os.path.isdir(p):
            return None
        if p not in _sys.path:
            _sys.path.insert(0, p)
        try:
            import serial as _s
            import serial.tools.list_ports  # noqa: F401
            return _s
        except Exception:
            return None

    # 候选路径 (按优先级)
    _serial_search_paths = [
        os.path.join(_PROJECT_ROOT, '_pyserial_lib'),                       # 仓库自带
        os.path.join(_PROJECT_ROOT, 'runtime', 'Lib', 'site-packages'),     # 便携 runtime
        os.path.join(_PROJECT_ROOT, 'runtime', 'python', 'Lib', 'site-packages'),
    ]
    _serial_pkg = None
    for _p in _serial_search_paths:
        _serial_pkg = _try_import_serial_from(_p)
        if _serial_pkg is not None:
            break

    if _serial_pkg is not None:
        _HAS_SERIAL = True
        # 拿到真正的 serial 模块后, 让后续代码用的 'serial' 就是它
        serial = _serial_pkg
    else:
        # 兜底: 再试一次"系统已装 pyserial"的情况
        try:
            import serial
            import serial.tools.list_ports
            _HAS_SERIAL = True
        except ImportError:
            pass

    if not _HAS_SERIAL:
        # 离线环境友好提示 — 不调用任何 subprocess, 不联网, 不强求用户懂 pip
        # 列出已经搜过的路径, 告诉用户把 serial/ 放到哪个位置就能用
        tried = '\n'.join(f'    • {p}' for p in _serial_search_paths)
        _SERIAL_HINT = (
            f'未在以下位置找到 pyserial:\n{tried}\n\n'
            f'解决方法 (二选一):\n'
            f'  A) 重新跑一次 setup_offline.bat (会装 pyserial 到 runtime\\)\n'
            f'  B) 把 _pyserial_lib\\serial\\ 整个目录复制到:\n'
            f'     {os.path.join(_PROJECT_ROOT, "runtime", "Lib", "site-packages")}\\serial\\\n\n'
            f'完成后重启工具即可。'
        )
        _SERIAL_INSTALL_ERR = _SERIAL_HINT

    # --- 串口状态变量 ---
    t10_ser = None
    t10_reader_job = None
    t10_rx_buf = b''
    t10_rx_total = 0
    t10_tx_total = 0
    t10_packet_pending = False
    t10_timed_job = None

    # --- 配置变量 (放底部 Tab) ---
    t10_port    = tk.StringVar(value='')
    t10_baud    = tk.StringVar(value='115200')
    t10_data    = tk.StringVar(value='8')
    t10_stop    = tk.StringVar(value='1')
    t10_parity  = tk.StringVar(value='None')
    t10_flow    = tk.StringVar(value='None')
    t10_dtr_on  = tk.BooleanVar(value=False)
    t10_rts_on  = tk.BooleanVar(value=False)

    # --- 显示选项 ---
    t10_hex_disp   = tk.BooleanVar(value=False)
    t10_latest     = tk.BooleanVar(value=True)
    t10_timestamp  = tk.BooleanVar(value=False)
    t10_autoscroll = tk.BooleanVar(value=True)

    # --- 自动记录日志 (勾选后弹文件选择, 后续 [Rx]/[Tx] 自动写入文件) ---
    t10_auto_log   = tk.BooleanVar(value=False)
    t10_log_fp     = None   # 当前打开的日志文件句柄
    t10_log_path   = ''    # 当前日志文件路径 (用于状态提示)

    # --- 发送选项 ---
    t10_hex_send    = tk.BooleanVar(value=False)
    t10_append_rn   = tk.BooleanVar(value=True)
    t10_suffix      = tk.StringVar(value='\\r\\n')
    t10_escape      = tk.BooleanVar(value=True)
    t10_encoding    = tk.StringVar(value='UTF-8')
    t10_timed_enable = tk.BooleanVar(value=False)
    t10_timed_ms    = tk.StringVar(value='1000')

    # --- 信号指示灯 ---
    t10_cts_var = tk.StringVar(value='CTS ○')
    t10_dsr_var = tk.StringVar(value='DSR ○')
    t10_dcd_var = tk.StringVar(value='DCD ○')
    t10_ri_var  = tk.StringVar(value='RI  ○')
    t10_status  = tk.StringVar(value='● 未连接')

    # --- 占位 widget (挂在独立的隐藏容器, 不占 t10 布局)
    #     实际显示在 ⚙ 弹窗里, 这里只是占位以保证
    #     _t10_open/_t10_close/_t10_refresh_ports 等函数不会因
    #     弹窗未打开/已销毁而崩溃 ---
    _hidden_holder = tk.Frame(root)  # 独立容器, 不显示
    t10_port_cb = ttk.Combobox(_hidden_holder, textvariable=t10_port,
                               state='readonly' if _HAS_SERIAL else 'disabled',
                               font=(F, 10), width=10)
    t10_open_btn = ttk.Button(_hidden_holder, text='⛔ 打开串口',
                              command=lambda: _t10_toggle(),
                              style='Accent.TButton', width=12)

    # =================================================================
    # Row 0: 顶部数据区 (单数据框, 通过前缀 [Rx]/[Tx] 区分)
    # =================================================================
    f10rx = ttk.LabelFrame(t10, text=' 数据 ')
    f10rx.grid(row=0, column=0, sticky='nsew', padx=12, pady=(10, 4))
    f10rx.grid_rowconfigure(0, weight=1)   # 文本框
    f10rx.grid_columnconfigure(0, weight=1)

    # 数据文本框 (终端模式: 历史只读, 末行可输入)
    t10_rx_text = tk.Text(f10rx, font=(M, 10), bg='#fafbfc', fg=C['fg'],
                          insertbackground=C['fg'], relief='flat',
                          padx=10, pady=6, wrap='word')
    t10_rx_text.grid(row=0, column=0, sticky='nsew', padx=10, pady=(8, 8))

    # ── 终端输入追踪: Rx 数据写入后, 光标之后才是用户输入区 ──
    t10_input_start_mark = None  # 用户输入起始位置 (Tkinter index 或 None)
    t10_password_mode = False   # 收到密码提示后自动隐藏下次输入

    def _t10_send_from_box(event):
        """Enter 发送当前行中用户输入的部分 (跳过接收到的提示符)"""
        nonlocal t10_input_start_mark, t10_tx_total, t10_password_mode
        if not t10_ser or not t10_ser.is_open:
            return 'break'
        cur = t10_rx_text.index('insert')
        line_start = cur.split('.')[0] + '.0'
        # ── 只取用户输入部分 (input_start 之后到光标) ──
        if t10_input_start_mark:
            try:
                # 确保 mark 在当前行 (防止跨行错位)
                start_ls = t10_rx_text.index(f'{t10_input_start_mark} linestart')
                cur_ls = t10_rx_text.index(f'{cur} linestart')
                if start_ls == cur_ls:
                    user_text = t10_rx_text.get(t10_input_start_mark, cur).strip()
                else:
                    user_text = ''  # input_start 不在当前行, 发送空行
            except tk.TclError:
                user_text = ''
        else:
            # 无 input_start (空行直接打字), 取整行
            user_text = t10_rx_text.get(line_start, cur).strip()
        if not user_text:
            # 空行也发送: 很多终端协议需要空行确认
            if t10_append_rn.get():
                t10_ser.write(b'\r\n')
                t10_tx_total += 2
                _t10_update_counts()
            t10_input_start_mark = None
            t10_password_mode = False
            # 不手动插 \n — 设备回显自带换行
            t10_rx_text.see('end')
            _t10_lock_history()
            return 'break'
        # 编码发送
        raw = _t10_apply_escape(user_text)
        try:
            if t10_hex_send.get():
                hex_str = re.sub(r'[\s,;]+', ' ', raw).strip()
                data = bytes.fromhex(hex_str.replace(' ', ''))
            else:
                data = raw.encode(t10_encoding.get(), errors='replace')
        except ValueError:
            messagebox.showerror('格式错误', 'Hex 格式错误')
            return 'break'
        if t10_append_rn.get():
            data += b'\r\n'
        try:
            t10_ser.write(data)
            t10_tx_total += len(data)
            _t10_update_counts()
            # ── 删除用户本地输入 (仅当前行), 等设备回显换行 ──
            if t10_input_start_mark:
                line_idx = t10_input_start_mark.split('.')[0]
                t10_rx_text.delete(t10_input_start_mark, f'{line_idx}.end')
                # 密码模式: 用 * 替代明文显示
                if t10_password_mode:
                    masked = '*' * len(user_text) if user_text else '***'
                    t10_rx_text.insert(f'{line_idx}.end', masked)
            # 不手动插 \n — 设备回显自带换行, 手动插会导致双换行
            t10_rx_text.see('end')
            t10_input_start_mark = None
            t10_password_mode = False
            _t10_lock_history()
            _t10_write_log('Tx', data)
        except (serial.SerialException, OSError) as e:
            messagebox.showerror('发送失败', str(e))
            _t10_close()
        return 'break'  # 阻止默认换行

    # Shift+Enter 正常换行, 纯 Enter 触发发送
    t10_rx_text.bind('<Return>', _t10_send_from_box)
    t10_rx_text.bind('<Shift-Return>', lambda e: None)  # Shift+Enter 正常换行
    t10_rx_sc = ttk.Scrollbar(f10rx, orient='vertical',
                              command=t10_rx_text.yview)
    t10_rx_sc.grid(row=0, column=1, sticky='ns', padx=(0, 4), pady=(8, 8))
    t10_rx_text.configure(yscrollcommand=t10_rx_sc.set)

    # =================================================================
    # 终端化: 让接收框像 Xshell 那样 — 历史只读, 只能在最后一行输入
    #   • hist_lock  tag 覆盖 "已存在的" 所有内容
    #   • 每次 _t10_print_text() 追加新内容前, 自动给光标之前的所有行打 tag
    #   • 任何对 hist_lock 区域的修改 (Backspace / Delete / Cut / 粘贴 /
    #     光标移动 / 鼠标选择 / 输入) 都被拦截并强制钳制到末行
    #   • 5000 行上限, 溢出时自动从顶部删除最老行
    # =================================================================
    t10_rx_text.tag_configure('hist_lock', foreground=C['fg'])  # 不改颜色, 只用于范围
    t10_rx_text.tag_raise('hist_lock')  # 确保不被其它 tag 覆盖 (避免 tag 优先级冲突)

    _T10_MAX_LINES = 5000

    def _t10_lock_history():
        """把当前 'end-1c 之前' 的所有内容打上 hist_lock tag.
        调用时机: 任何准备"切换到下一行" 之前 (即用户按 Enter 发送后, 或新数据追加前)."""
        try:
            # 'end - 1 char' 即最后一行的行末; 之前的内容都属于历史
            cutoff = t10_rx_text.index('end-1c')
            t10_rx_text.tag_add('hist_lock', '1.0', cutoff)
            # 5000 行上限: 多出来的从最顶删
            line_count = int(t10_rx_text.index('end-1c').split('.')[0])
            if line_count > _T10_MAX_LINES:
                excess = line_count - _T10_MAX_LINES
                t10_rx_text.delete('1.0', f'{excess + 1}.0')
                # 删除后 hist_lock 范围要相应更新 (1.0 起点变了)
                t10_rx_text.tag_remove('hist_lock', '1.0', 'end')
                t10_rx_text.tag_add('hist_lock', '1.0', t10_rx_text.index('end-1c'))
        except tk.TclError:
            pass

    def _t10_clamp_cursor_to_input():
        """强制把光标钳制在 '输入区' (= 末行).
        只在光标离开末行时纠正 — 末行是合法输入区, 不能动.
        """
        try:
            insert_idx = t10_rx_text.index('insert')
            end_idx = t10_rx_text.index('end-1c')
            # 末行 = linestart(end) == linestart(insert)
            ins_ls = t10_rx_text.index(f'{insert_idx} linestart')
            end_ls = t10_rx_text.index(f'{end_idx} linestart')
            if ins_ls != end_ls:
                # 光标在历史区, 强制钳到末行
                t10_rx_text.mark_set('insert', 'end-1c')
                t10_rx_text.see('insert')
        except tk.TclError:
            pass

    def _t10_block_in_history(event):
        """拦截对历史区 (= 非末行) 的修改操作.
        末行是合法输入区, 不拦截. 仅当光标真的在历史区才 return 'break'.
        """
        try:
            insert_idx = t10_rx_text.index('insert')
            end_idx = t10_rx_text.index('end-1c')
            ins_ls = t10_rx_text.index(f'{insert_idx} linestart')
            end_ls = t10_rx_text.index(f'{end_idx} linestart')
            if ins_ls != end_ls:
                # 光标在历史区 (非末行), 钳制到末行后阻止默认行为
                t10_rx_text.mark_set('insert', 'end-1c')
                t10_rx_text.see('insert')
                return 'break'
        except tk.TclError:
            pass
        return None  # 放行 (末行正常编辑)

    # ── 绑定: 拦截历史区的所有修改 ──
    # 1) Backspace / Delete
    t10_rx_text.bind('<BackSpace>', _t10_block_in_history)
    t10_rx_text.bind('<Delete>', _t10_block_in_history)
    # 2) 剪切 (Ctrl+X) — 末行内允许剪切, 跨历史区禁止
    t10_rx_text.bind('<Control-x>', _t10_block_in_history)
    t10_rx_text.bind('<Control-X>', _t10_block_in_history)
    # 3) 鼠标点击 — 先钳制光标, 再放行
    t10_rx_text.bind('<Button-1>',
                     lambda e: (t10_rx_text.after_idle(_t10_clamp_cursor_to_input), None)[1])
    # 4) 方向键 / Home — 历史区钳制后放行, 末行放行
    for _k in ('<Left>', '<Up>', '<Home>'):
        t10_rx_text.bind(_k, _t10_block_in_history)
    # 5) 任意按键后钳制 (兜底, 防止上面漏掉的组合键把光标带到历史区)
    t10_rx_text.bind('<KeyRelease>',
                     lambda e: t10_rx_text.after_idle(_t10_clamp_cursor_to_input))
    # 6) Ctrl+V 粘贴 — 强制粘贴到末行
    def _t10_paste_in_history(event=None):
        try:
            insert_idx = t10_rx_text.index('insert')
            end_idx = t10_rx_text.index('end-1c')
            ins_ls = t10_rx_text.index(f'{insert_idx} linestart')
            end_ls = t10_rx_text.index(f'{end_idx} linestart')
            if ins_ls != end_ls:
                t10_rx_text.mark_set('insert', 'end-1c')
        except tk.TclError:
            pass
        return None  # 放行粘贴
    t10_rx_text.bind('<Control-v>', _t10_paste_in_history)
    t10_rx_text.bind('<Control-V>', _t10_paste_in_history)

    # =================================================================
    # Row 1: 设置行 (选项: HEX显示 / 时间戳 / 自动换行 / 编码 / 转义 / HEX发送 / 定时)
    # =================================================================
    hint = ttk.Frame(t10, style='TFrame')
    hint.grid(row=1, column=0, sticky='ew', padx=14, pady=2)

    # 接收组
    ttk.Label(hint, text='接收:', font=(F, 9, 'bold'),
              foreground=C['blue']).pack(side='left', padx=(0, 4))
    ttk.Checkbutton(hint, text='HEX显示', variable=t10_hex_disp,
                    style='TCheckbutton').pack(side='left', padx=(0, 4))
    ttk.Checkbutton(hint, text='时间戳', variable=t10_timestamp,
                    style='TCheckbutton').pack(side='left', padx=(0, 4))

    ttk.Separator(hint, orient='vertical').pack(
        side='left', fill='y', padx=6)

    # 发送组
    ttk.Label(hint, text='发送:', font=(F, 9, 'bold'),
              foreground=C['blue']).pack(side='left', padx=(0, 4))
    ttk.Checkbutton(hint, text='自动换行(\\r\\n)', variable=t10_append_rn,
                    style='TCheckbutton').pack(side='left', padx=(0, 4))
    ttk.Label(hint, text='编码', font=(F, 9),
              foreground=C['sub']).pack(side='left', padx=(0, 1))
    ttk.Combobox(hint, textvariable=t10_encoding,
                 values=['UTF-8', 'GBK', 'GB2312', 'ASCII', 'Latin-1', 'UTF-16'],
                 state='readonly', font=(M, 9), width=7).pack(
        side='left', padx=(0, 4))
    ttk.Checkbutton(hint, text='转义', variable=t10_escape,
                    style='TCheckbutton').pack(side='left', padx=(0, 4))
    ttk.Checkbutton(hint, text='HEX发送', variable=t10_hex_send,
                    style='TCheckbutton').pack(side='left', padx=(0, 4))
    ttk.Checkbutton(hint, text='定时', variable=t10_timed_enable,
                    command=lambda: _t10_toggle_timer(),
                    style='TCheckbutton').pack(side='left', padx=(8, 1))
    ttk.Entry(hint, textvariable=t10_timed_ms,
              font=(M, 9), width=5, justify='center').pack(
        side='left', padx=(0, 1))
    ttk.Label(hint, text='ms', font=(F, 9),
              foreground=C['sub']).pack(side='left', padx=(0, 4))

    # 最右: 自动记录日志复选框 (右边距 12, 避免贴边)
    ttk.Checkbutton(hint, text='自动记录日志', variable=t10_auto_log,
                    command=lambda: _t10_toggle_auto_log(),
                    style='TCheckbutton').pack(
        side='right', padx=(4, 12))

    # =================================================================
    # Row 2: 发送输入行 (提示 + 输入框 + 发送按钮, 高度优先保证可见)
    # =================================================================
    send_row = ttk.Frame(t10, style='TFrame')
    send_row.grid(row=2, column=0, sticky='nsew', padx=12, pady=(2, 2))
    send_row.grid_columnconfigure(0, weight=1)
    send_row.grid_rowconfigure(0, weight=1, minsize=40)  # 输入框

    # 输入框: tk.Text(height=1) 填满高度 + 文字左上角
    t10_send_entry = tk.Text(send_row, font=(M, 11), bg='white', fg=C['fg'],
                              insertbackground=C['fg'], relief='solid', bd=1,
                              height=1, wrap='none', undo=False)
    t10_send_entry.grid(row=0, column=0, sticky='nsew', padx=(4, 6), pady=4)
    t10_send_entry.bind('<Return>', lambda e: _t10_send())
    # Shift+Enter 换行 (单行一般不需要, 但保留)
    t10_send_entry.bind('<Shift-Return>', lambda e: None)

    # 发送按钮 (右侧醒目大按钮)
    t10_send_btn = ttk.Button(send_row, text='▶ 发送',
                              command=lambda: _t10_send(),
                              style='Accent.TButton')
    t10_send_btn.grid(row=0, column=1, sticky='nse', padx=(0, 4), pady=4,
                       ipady=4)

    # 第 2 行: (文件操作按钮已移至状态栏最右侧, 此行不再使用)
    # (留空, 仅保留 grid 行配置以防布局错位)

    # =================================================================
    # (Row 3 已删除 - 文件/导出/清屏 已合并到 send_row 第 3 行)
    # (Row 3 状态栏在下面)
    # =================================================================

    # =================================================================
    # Row 3: 状态栏 (右侧带 ⚙ 串口设置 按钮)
    # =================================================================
    sbar = tk.Frame(t10, bg=C['card'], highlightbackground=C['bd'],
                    highlightthickness=1, bd=0)
    sbar.grid(row=3, column=0, sticky='ew', padx=12, pady=(0, 8))

    dot = tk.Canvas(sbar, width=10, height=10, bg=C['card'],
                    highlightthickness=0, bd=0)
    dot.create_oval(2, 2, 8, 8, fill=C['red'], outline='')
    dot.pack(side='left', padx=(14, 6), pady=8)

    ttk.Label(sbar, textvariable=t10_status, font=(F, 9, 'bold'),
              foreground=C['fg']).pack(side='left', padx=(0, 16), pady=8)

    ttk.Separator(sbar, orient='vertical').pack(side='left', fill='y', padx=4)

    info_text = tk.StringVar(value='端口: --  波特率: 115200  '
                                   '数据位: 8  停止位: 1  校验: None')
    ttk.Label(sbar, textvariable=info_text, font=(F, 9),
              foreground=C['sub']).pack(side='left', padx=(4, 16), pady=8)

    ttk.Separator(sbar, orient='vertical').pack(side='left', fill='y', padx=4)

    t10_rx_label = tk.StringVar(value='Rx: 0')
    t10_tx_label = tk.StringVar(value='Tx: 0')
    ttk.Label(sbar, textvariable=t10_rx_label, font=(F, 9),
              foreground=C['sub']).pack(side='left', padx=(4, 8), pady=8)
    ttk.Label(sbar, textvariable=t10_tx_label, font=(F, 9),
              foreground=C['sub']).pack(side='left', padx=(0, 8), pady=8)

    ttk.Separator(sbar, orient='vertical').pack(side='left', fill='y', padx=4)

    for _v in [t10_cts_var, t10_dsr_var, t10_dcd_var, t10_ri_var]:
        ttk.Label(sbar, textvariable=_v, font=(F, 9),
                  foreground=C['sub']).pack(side='left', padx=(0, 6), pady=8)

    # 最右: 操作按钮区 (串口设置 / 清屏 / 发送文件)
    # 顺序从右往左: 串口设置 | 清屏 | 发送文件
    # (⚙ 串口设置 放最右, 方便一键访问; 📁 发送文件 放最左)
    ttk.Separator(sbar, orient='vertical').pack(side='right', fill='y', padx=4)
    ttk.Button(sbar, text='⚙ 串口设置', style='Accent.TButton',
               command=lambda: _t10_open_settings()).pack(
        side='right', padx=(0, 4), pady=6)
    ttk.Button(sbar, text='🧹 清屏', style='Small.TButton',
               command=lambda: _t10_clear_rx()).pack(
        side='right', padx=(0, 4), pady=6)
    ttk.Button(sbar, text='📁 发送文件', style='Small.TButton',
               command=lambda: _t10_send_file()).pack(
        side='right', padx=(0, 4), pady=6)


    def _goto_con_tab():
        # 不再使用 (已无底部 Tab)
        pass

    # (已移除"串口设置"跳转按钮, 所有参数已直接显示在主界面)

    # =================================================================
    # 工具函数
    # =================================================================
    def _t10_copy_to_clip(text):
        try:
            root.clipboard_clear()
            root.clipboard_append(text)
            root.update()
        except Exception:
            pass

    def _t10_toast(msg):
        old = t10_status.get()
        t10_status.set(f'✓ {msg}')
        root.after(2000, lambda: t10_status.set(old))

    def _t10_reset_defaults():
        t10_baud.set('115200')
        t10_data.set('8')
        t10_stop.set('1')
        t10_parity.set('None')
        t10_flow.set('None')
        t10_hex_disp.set(False)
        t10_latest.set(True)
        t10_timestamp.set(False)
        t10_autoscroll.set(True)
        t10_hex_send.set(False)
        t10_append_rn.set(True)
        t10_escape.set(True)
        t10_encoding.set('UTF-8')
        t10_timed_enable.set(False)
        t10_timed_ms.set('1000')
        _t10_update_info()
        _t10_toast('已重置为默认设置')

    def _t10_update_info():
        if t10_ser and t10_ser.is_open:
            port = t10_port.get()
        else:
            port = '--'
        info_text.set(
            f'端口: {port}  波特率: {t10_baud.get()}  '
            f'数据位: {t10_data.get()}  '
            f'停止位: {t10_stop.get()}  校验: {t10_parity.get()}')

    # =================================================================
    # 串口设置弹窗
    # =================================================================
    _t10_settings_win = None  # 单例: 避免重复打开

    def _t10_open_settings():
        nonlocal _t10_settings_win
        if _t10_settings_win is not None and _t10_settings_win.winfo_exists():
            _t10_settings_win.deiconify()
            _t10_settings_win.lift()
            _t10_settings_win.focus_force()
            return

        win = tk.Toplevel(root)
        win.title('串口设置')
        # 在 geometry 之前先禁用调整 (Windows 上更可靠)
        win.resizable(False, False)
        win.geometry('600x520')
        win.transient(root)
        # 二次确认: 通过 wm_resizable 强制设置, 避免主题/重绘时重置
        try:
            win.wm_resizable(False, False)
        except Exception:
            pass
        # 关闭时清理引用
        def _on_close():
            nonlocal _t10_settings_win
            try:
                win.destroy()
            finally:
                _t10_settings_win = None
        win.protocol('WM_DELETE_WINDOW', _on_close)
        _t10_settings_win = win

        # 弹窗主容器
        main = ttk.Frame(win, style='TFrame', padding=12)
        main.pack(fill='both', expand=True)

        # ── 分组 0: 串口连接 (Port / Baud / 打开串口) ──
        gp_conn = ttk.LabelFrame(main, text=' 串口连接 ')
        gp_conn.pack(fill='x', pady=(0, 8))
        for c in range(8):
            gp_conn.grid_columnconfigure(c, weight=0)

        ttk.Label(gp_conn, text='端口', font=(F, 9, 'bold'),
                  foreground=C['fg']).grid(row=0, column=0, padx=(10, 4),
                                            pady=8, sticky='w')
        # 注意: 用 _win_port_cb 本地名, 不要覆盖外层 t10_port_cb (会破坏外层函数引用)
        _win_port_cb = ttk.Combobox(gp_conn, textvariable=t10_port,
                                    state='readonly' if _HAS_SERIAL else 'disabled',
                                    font=(F, 10), width=10)
        _win_port_cb.grid(row=0, column=1, padx=(0, 4))
        ttk.Button(gp_conn, text='↻', width=2, style='Small.TButton',
                   command=lambda: _t10_refresh_ports()).grid(
            row=0, column=2, padx=(0, 12))

        ttk.Label(gp_conn, text='波特率', font=(F, 9, 'bold'),
                  foreground=C['fg']).grid(row=0, column=3, padx=(0, 4),
                                            pady=8, sticky='w')
        ttk.Combobox(gp_conn, textvariable=t10_baud,
                     values=['9600','19200','38400','57600','115200',
                             '230400','460800','921600','1000000',
                             '2000000','3000000', '自定义'],
                     state='readonly', font=(F, 10), width=10).grid(
            row=0, column=4, padx=(0, 12))

        # 打开串口按钮 (放在弹窗里, 用本地名避免覆盖外层 t10_open_btn)
        _win_open_btn = ttk.Button(gp_conn, text='⛔ 打开串口',
                                   command=lambda: _t10_toggle(),
                                   style='Accent.TButton', width=12)
        _win_open_btn.grid(row=0, column=5, padx=(0, 10), sticky='w')

        # ── 分组 1: 串口参数 (数据/停止/校验/流控) ──
        gp1 = ttk.LabelFrame(main, text=' 串口参数 ')
        gp1.pack(fill='x', pady=(0, 8))
        for c in range(8):
            gp1.grid_columnconfigure(c, weight=0)

        ttk.Label(gp1, text='数据位', font=(F, 9),
                  foreground=C['sub']).grid(row=0, column=0, padx=(10, 4),
                                             pady=8, sticky='w')
        ttk.Combobox(gp1, textvariable=t10_data, values=['5','6','7','8'],
                     state='readonly', font=(F, 10), width=5).grid(
            row=0, column=1, padx=(0, 12))
        ttk.Label(gp1, text='停止位', font=(F, 9),
                  foreground=C['sub']).grid(row=0, column=2, padx=(0, 4),
                                             pady=8, sticky='w')
        ttk.Combobox(gp1, textvariable=t10_stop, values=['1','1.5','2'],
                     state='readonly', font=(F, 10), width=5).grid(
            row=0, column=3, padx=(0, 12))
        ttk.Label(gp1, text='校验', font=(F, 9),
                  foreground=C['sub']).grid(row=0, column=4, padx=(0, 4),
                                             pady=8, sticky='w')
        ttk.Combobox(gp1, textvariable=t10_parity,
                     values=['None','Even','Odd','Mark','Space'],
                     state='readonly', font=(F, 10), width=7).grid(
            row=0, column=5, padx=(0, 12))
        ttk.Label(gp1, text='流控', font=(F, 9),
                  foreground=C['sub']).grid(row=0, column=6, padx=(0, 4),
                                             pady=8, sticky='w')
        ttk.Combobox(gp1, textvariable=t10_flow,
                     values=['None','RTS/CTS','XON/XOFF'],
                     state='readonly', font=(F, 10), width=10).grid(
            row=0, column=7, padx=(0, 10), sticky='ew')

        # ── 分组 2: 驱动安装 ──
        gp2 = ttk.LabelFrame(main, text=' 驱动安装 (USB-串口芯片) ')
        gp2.pack(fill='x', pady=(0, 8))
        ttk.Label(gp2, text='将对应驱动 .inf/.exe 放入子目录后点击安装:',
                  font=(F, 9), foreground=C['sub']).grid(
            row=0, column=0, columnspan=3, sticky='w', padx=10, pady=(8, 4))
        ttk.Button(gp2, text='⚙ 安装 CP210x', width=14,
                   style='Small.TButton',
                   command=lambda: _t10_auto_install_driver('CP210x', 'cp210x')
                   ).grid(row=1, column=0, padx=(10, 6), pady=(0, 8))
        ttk.Button(gp2, text='⚙ 安装 CH340/CH341', width=14,
                   style='Small.TButton',
                   command=lambda: _t10_auto_install_driver('CH340/CH341', 'ch340')
                   ).grid(row=1, column=1, padx=(0, 6), pady=(0, 8))
        ttk.Button(gp2, text='⚙ 安装 FT232', width=14,
                   style='Small.TButton',
                   command=lambda: _t10_auto_install_driver('FT232', 'ft232')
                   ).grid(row=1, column=2, padx=(0, 10), pady=(0, 8))

        # ── 分组 3: DTR / RTS 手动控制 ──
        gp3 = ttk.LabelFrame(main, text=' DTR / RTS 手动控制 ')
        gp3.pack(fill='x', pady=(0, 8))
        ttk.Label(gp3, text='仅在已连接时生效 (用于复位/进入 Bootloader)',
                  font=(F, 9), foreground=C['sub']).grid(
            row=0, column=0, columnspan=4, sticky='w', padx=10, pady=(6, 4))
        ttk.Checkbutton(gp3, text='DTR', variable=t10_dtr_on,
                        command=lambda: _t10_set_dtr_rts(),
                        style='TCheckbutton').grid(
            row=1, column=0, padx=(10, 12), pady=(0, 8), sticky='w')
        ttk.Checkbutton(gp3, text='RTS', variable=t10_rts_on,
                        command=lambda: _t10_set_dtr_rts(),
                        style='TCheckbutton').grid(
            row=1, column=1, padx=(0, 12), pady=(0, 8), sticky='w')
        ttk.Label(gp3, text='提示: 切换 RTS/DTR 可触发 ESP32/STM32 进入下载模式',
                  font=(F, 8), foreground=C['sub']).grid(
            row=1, column=2, columnspan=2, sticky='w', padx=(4, 10), pady=(0, 8))

        # ── 分组 4: 当前参数摘要 ──
        gp4 = ttk.LabelFrame(main, text=' 当前参数 ')
        gp4.pack(fill='x', pady=(0, 8))
        _summary_text = tk.StringVar()
        def _refresh_summary():
            _summary_text.set(
                f'端口: {t10_port.get() or "--"}    '
                f'波特率: {t10_baud.get()}    '
                f'数据位: {t10_data.get()}    '
                f'停止位: {t10_stop.get()}    '
                f'校验: {t10_parity.get()}    '
                f'流控: {t10_flow.get()}')
        _refresh_summary()
        ttk.Label(gp4, textvariable=_summary_text, font=(F, 9),
                  foreground=C['fg']).pack(padx=10, pady=8, anchor='w')
        # 主界面变量变化时同步
        for _v in (t10_port, t10_baud, t10_data, t10_stop, t10_parity, t10_flow):
            _v.trace_add('write', lambda *a: _refresh_summary())

        # ── 底部按钮 ──
        btn_bar = ttk.Frame(main, style='TFrame')
        btn_bar.pack(fill='x', pady=(8, 0))
        ttk.Button(btn_bar, text='🔄 恢复默认',
                   command=lambda: _t10_reset_defaults(),
                   style='Small.TButton').pack(side='left', padx=(0, 6))
        ttk.Button(btn_bar, text='关闭',
                   command=_on_close,
                   style='Accent.TButton').pack(side='right')

        # ── 终极防调整: 布局完成后再次禁用 resizable
        # (某些 ttk 主题在子控件初始化时会重置窗口属性)
        win.update_idletasks()
        win.resizable(False, False)
        try:
            win.wm_resizable(False, False)
        except Exception:
            pass
        # 强制固定窗口尺寸 (用户硬拖拽时也会被拉回)
        win.minsize(600, 520)
        win.maxsize(600, 520)

    # =================================================================
    # 驱动安装
    # =================================================================
    _DRV_DIR = (os.path.join(getattr(_sys, '_MEIPASS', ''), 'serial_drivers')
                if getattr(_sys, '_MEIPASS', '')
                else os.path.join(_PROJECT_ROOT, 'serial_drivers'))

    def _t10_auto_install_driver(chip_name, subdir):
        d = os.path.join(_DRV_DIR, subdir)
        if not os.path.isdir(d):
            messagebox.showinfo('未找到驱动',
                f'请将 {chip_name} 驱动文件(.inf 或 .exe)放到:\n'
                f'  {d}\n\n然后重新点击安装。')
            return
        files = os.listdir(d)
        inf_file = next((f for f in files if f.lower().endswith('.inf')), None)
        exe_file = next((f for f in files
                        if f.lower().endswith(('.exe', '.msi'))), None)
        if not inf_file and not exe_file:
            messagebox.showinfo('未找到驱动',
                f'目录存在但无 .inf 或 .exe:\n  {d}\n\n'
                f'请放入 {chip_name} 驱动文件后重试。')
            return
        if inf_file:
            inf_path = os.path.join(d, inf_file)
            try:
                r = subprocess.run(
                    ['pnputil', '/add-driver', inf_path, '/install'],
                    capture_output=True, text=True, timeout=60)
                if r.returncode == 0:
                    messagebox.showinfo('安装成功',
                        f'{chip_name} 驱动安装成功!\n请重新插拔 USB 设备。')
                else:
                    if exe_file:
                        subprocess.Popen([os.path.join(d, exe_file)], shell=True)
                        messagebox.showinfo('提示', 'pnputil 失败，已启动安装程序。')
                    else:
                        messagebox.showerror('安装失败',
                            f'pnputil 返回:\n{(r.stderr or r.stdout)[:300]}')
            except FileNotFoundError:
                os.startfile(d) if _sys.platform == 'win32' else None
                messagebox.showinfo('提示', f'请手动安装:\n{inf_path}')
            except Exception as e:
                messagebox.showerror('错误', str(e))
        elif exe_file:
            exe_path = os.path.join(d, exe_file)
            try:
                subprocess.Popen([exe_path], shell=True)
                messagebox.showinfo('提示',
                    f'已启动 {chip_name} 驱动安装程序。\n'
                    f'请按向导完成安装后重新插拔 USB。')
            except Exception as e:
                messagebox.showerror('错误', str(e))

    # =================================================================
    # 收发逻辑
    # =================================================================
    def _t10_safe_widget(name):
        """安全访问可能已被销毁的 widget, 返回 None 如果不存在"""
        try:
            w = eval(name)
            if w and hasattr(w, 'winfo_exists') and w.winfo_exists():
                return w
        except Exception:
            pass
        return None

    def _t10_refresh_ports():
        if not _HAS_SERIAL:
            return
        try:
            ports = [p.device for p in serial.tools.list_ports.comports()]
        except Exception:
            ports = []
        # ── 添加虚拟测试端口 ──
        ports.append('[测试] 虚拟设备 (socket://localhost:9876)')
        cb = _t10_safe_widget('t10_port_cb')
        if cb is not None:
            try:
                cb['values'] = ports or ['(无可用串口)']
            except Exception:
                pass
        if ports and not t10_port.get():
            t10_port.set(ports[0])

    def _t10_toggle():
        if t10_ser and t10_ser.is_open:
            _t10_close()
        else:
            _t10_open()

    def _t10_open():
        nonlocal t10_ser, t10_reader_job
        if not _HAS_SERIAL:
            messagebox.showerror('缺少依赖',
                '需要 pyserial 库。请执行:\n    pip install pyserial')
            return
        port_raw = t10_port.get().strip()
        if not port_raw:
            messagebox.showerror('错误', '请选择串口')
            return
        # ── 提取实际 port (虚拟端口格式: [测试] 虚拟设备 (socket://host:port)) ──
        import re as _re
        m = _re.search(r'\((socket://[^)]+)\)', port_raw)
        if m:
            port = m.group(1)
        else:
            port = port_raw

        try:
            parity_map = {'None': 'N', 'Even': 'E', 'Odd': 'O',
                          'Mark': 'M', 'Space': 'S'}
            if port.startswith('socket://'):
                t10_ser = serial.serial_for_url(port, do_not_open=True)
                t10_ser.baudrate = int(t10_baud.get())
                t10_ser.timeout = 0.05
                t10_ser.open()
            else:
                t10_ser = serial.Serial(
                    port=port,
                    baudrate=int(t10_baud.get()),
                bytesize=int(t10_data.get()),
                stopbits=float(t10_stop.get()),
                parity=parity_map.get(t10_parity.get(), 'N'),
                timeout=0.05,
            )
            t10_ser.reset_input_buffer()
            t10_ser.reset_output_buffer()
            t10_ser.dtr = t10_dtr_on.get()
            t10_ser.rts = t10_rts_on.get()
        except Exception as e:
            messagebox.showerror('打开失败', str(e))
            t10_ser = None
            return

        # 安全更新 UI (控件可能存在于弹窗或主界面, 任一位置)
        btn = _t10_safe_widget('t10_open_btn')
        if btn is not None:
            try:
                btn.config(text='⛔ 关闭串口', style='Success.TButton')
            except Exception:
                pass
        cb = _t10_safe_widget('t10_port_cb')
        if cb is not None:
            try:
                cb.config(state='disabled')
            except Exception:
                pass
        t10_status.set(f'● 已连接 {port} @ {t10_baud.get()}')
        try:
            dot.delete('all')
            dot.create_oval(2, 2, 8, 8, fill=C['green'], outline='')
        except Exception:
            pass
        t10_rx_buf = b''
        t10_rx_total = 0
        t10_tx_total = 0
        _t10_update_counts()
        _t10_update_signals()
        _t10_update_info()
        _t10_start_reader()
        # 打开串口后: 接收区光标归零, 输入起点设在1.0
        t10_rx_text.mark_set('insert', '1.0')
        t10_input_start_mark = '1.0'
        t10_rx_text.focus_set()

    def _t10_close():
        nonlocal t10_ser, t10_reader_job
        if t10_reader_job:
            root.after_cancel(t10_reader_job)
            t10_reader_job = None
        if t10_ser and t10_ser.is_open:
            try:
                t10_ser.close()
            except Exception:
                pass
        t10_ser = None
        # 安全更新 UI
        btn = _t10_safe_widget('t10_open_btn')
        if btn is not None:
            try:
                btn.config(text='⛔ 打开串口', style='Accent.TButton')
            except Exception:
                pass
        cb = _t10_safe_widget('t10_port_cb')
        if cb is not None and _HAS_SERIAL:
            try:
                cb.config(state='readonly')
            except Exception:
                pass
        t10_status.set('● 未连接')
        try:
            dot.delete('all')
            dot.create_oval(2, 2, 8, 8, fill=C['red'], outline='')
        except Exception:
            pass
        t10_dtr_on.set(False)
        t10_rts_on.set(False)
        t10_cts_var.set('CTS ○')
        t10_dsr_var.set('DSR ○')
        t10_dcd_var.set('DCD ○')
        t10_ri_var.set('RI  ○')
        _t10_stop_timer()
        t10_timed_enable.set(False)
        _t10_update_info()

    def _t10_set_dtr_rts():
        if t10_ser and t10_ser.is_open:
            try:
                t10_ser.dtr = t10_dtr_on.get()
                t10_ser.rts = t10_rts_on.get()
            except Exception:
                pass

    def _t10_update_signals():
        if not t10_ser or not t10_ser.is_open:
            return
        try:
            t10_cts_var.set(f'CTS ●' if t10_ser.cts else 'CTS ○')
            t10_dsr_var.set(f'DSR ●' if t10_ser.dsr else 'DSR ○')
            t10_dcd_var.set(f'DCD ●' if t10_ser.cd  else 'DCD ○')
            t10_ri_var.set(f'RI  ●' if t10_ser.ri  else 'RI  ○')
        except Exception:
            pass

    def _t10_start_reader():
        nonlocal t10_reader_job
        _sig_counter = 0

        def _read_loop():
            nonlocal t10_reader_job, _sig_counter
            try:
                if t10_ser and t10_ser.is_open:
                    data = t10_ser.read(t10_ser.in_waiting or 1)
                    if data:
                        root.after(0, lambda d=data: _t10_on_rx(d))
                    _sig_counter += 1
                    if _sig_counter >= 25:
                        _sig_counter = 0
                        root.after(0, lambda: _t10_update_signals())
            except (serial.SerialException, OSError):
                root.after(0, lambda: _t10_handle_error())
                return
            t10_reader_job = root.after(20, _read_loop)

        t10_reader_job = root.after(20, _read_loop)

    def _t10_on_rx(data):
        nonlocal t10_rx_buf, t10_rx_total, t10_packet_pending
        t10_rx_buf += data
        t10_rx_total += len(data)
        _t10_update_counts()

        if not t10_packet_pending:
            t10_packet_pending = True
            root.after(80, _t10_flush_packet)

    def _t10_flush_packet():
        nonlocal t10_rx_buf, t10_packet_pending, t10_input_start_mark, t10_password_mode
        t10_packet_pending = False
        if not t10_rx_buf:
            return

        data = t10_rx_buf
        t10_rx_buf = b''

        if t10_timestamp.get():
            now = datetime.datetime.now()
            ts = now.strftime('[%H:%M:%S.') + f'{now.microsecond // 1000:03d}]'
            t10_rx_text.insert('end', ts + ' ', 'ts')

        decoded_text = ''
        if t10_hex_disp.get():
            hex_str = ' '.join(f'{b:02X}' for b in data)
            t10_rx_text.insert('end', hex_str + '\n', 'hex')
        else:
            try:
                text = data.decode(t10_encoding.get(), errors='replace')
            except Exception:
                text = data.decode('latin-1', errors='replace')
            decoded_text = text  # 保存解码文本用于密码检测
            # ── 逐字符处理控制字符: \r=换行, \n=换行, \r\n=一个换行, \b=退格 ──
            i = 0
            n = len(text)
            _buf = []
            def _flush():
                nonlocal _buf
                if _buf:
                    t10_rx_text.insert('insert', ''.join(_buf))
                    _buf = []
            while i < n:
                ch = text[i]
                if ch == '\r':
                    _flush()
                    # \r 作为换行: 插入新行, 如果后面紧跟 \n 就跳过它
                    t10_rx_text.insert('insert', '\n')
                    if i + 1 < n and text[i + 1] == '\n':
                        i += 1  # skip \n after \r
                    i += 1
                elif ch == '\n':
                    _flush()
                    t10_rx_text.insert('insert', '\n')
                    i += 1
                elif ch == '\b':
                    _flush()
                    try:
                        t10_rx_text.delete('insert -1c', 'insert')
                    except tk.TclError:
                        pass
                    i += 1
                else:
                    _buf.append(ch)
                    i += 1
            _flush()
            # ── 自动识别密码提示: 下次输入用 * 隐藏 ──
            if re.search(r'(?i)(password|passwd|passcode|密钥|密码|PIN)\s*[:：]', decoded_text):
                t10_password_mode = True

        t10_rx_text.see('end')  # 自动滚动
        # ── 标记输入起点: 最后一行起始位置 ──
        last_nl = t10_rx_text.index('end-1l linestart')
        if t10_rx_text.get(last_nl, 'end-1c').strip() == '':
            t10_input_start_mark = last_nl  # 最后一行为空行, 这就是输入区
        else:
            t10_input_start_mark = t10_rx_text.index('end-1c')
        t10_rx_text.mark_set('insert', t10_input_start_mark)
        _t10_lock_history()
        _t10_write_log('Rx', data)  # 写入日志 (如已开启)

    def _t10_clear_rx():
        nonlocal t10_rx_total, t10_input_start_mark
        t10_rx_text.delete('1.0', 'end')
        t10_rx_text.tag_remove('hist_lock', '1.0', 'end')
        t10_input_start_mark = '1.0'
        t10_rx_text.mark_set('insert', '1.0')
        t10_rx_text.focus_set()
        t10_rx_total = 0
        _t10_update_counts()

    def _t10_toggle_auto_log():
        """复选框回调: 勾选时弹文件选择, 取消勾选时关闭文件句柄"""
        nonlocal t10_log_fp, t10_log_path
        if t10_auto_log.get():
            # 勾选 -> 让用户选保存位置
            default_name = ('serial_log_' +
                            datetime.datetime.now().strftime('%Y%m%d_%H%M%S') +
                            '.log')
            fp = filedialog.asksaveasfilename(
                title='选择日志保存路径',
                defaultextension='.log',
                initialfile=default_name,
                filetypes=[('日志文件', '*.log'),
                           ('文本文件', '*.txt'),
                           ('所有文件', '*.*')])
            if not fp:
                # 用户取消 -> 回退到不勾选
                t10_auto_log.set(False)
                return
            try:
                t10_log_fp = open(fp, 'a', encoding='utf-8')
                t10_log_path = fp
                # 写文件头 (注释行, 方便后期回溯会话)
                t10_log_fp.write(
                    f'# FPGA Toolbox Serial Log\n'
                    f'# Started: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n'
                    f'# Port: {t10_port.get() or "--"}  '
                    f'Baud: {t10_baud.get()}  '
                    f'Data: {t10_data.get()}{t10_stop.get()}  '
                    f'Parity: {t10_parity.get()}\n'
                    f'# ---\n')
                t10_log_fp.flush()
                _t10_toast(f'📝 已开启日志: {os.path.basename(fp)}')
            except Exception as e:
                messagebox.showerror('打开日志失败', str(e))
                t10_log_fp = None
                t10_log_path = ''
                t10_auto_log.set(False)
        else:
            # 取消勾选 -> 关闭文件
            if t10_log_fp:
                try:
                    t10_log_fp.write(
                        f'# Closed: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
                    t10_log_fp.close()
                except Exception:
                    pass
                t10_log_fp = None
                t10_log_path = ''
                _t10_toast('📝 日志已关闭')

    def _t10_write_log(direction, data):
        """把一条 [Rx]/[Tx] 写入日志文件 (已开启时)"""
        if not t10_log_fp or t10_log_fp.closed:
            return
        try:
            ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S') + \
                 f'.{datetime.datetime.now().microsecond // 1000:03d}'
            if direction == 'Tx' and t10_hex_send.get():
                content = ' '.join(f'{b:02X}' for b in data)
            elif t10_hex_disp.get():
                content = ' '.join(f'{b:02X}' for b in data)
            else:
                try:
                    content = data.decode(t10_encoding.get(), errors='replace')
                except Exception:
                    content = data.decode('latin-1', errors='replace')
            # 一行一条, \n 替换为空格避免日志里多行
            content = content.replace('\r', '').replace('\n', '\\n')
            t10_log_fp.write(f'[{ts}] [{direction}] {content}\n')
            t10_log_fp.flush()
        except Exception:
            # 写失败 -> 静默关闭, 避免刷屏
            try:
                t10_log_fp.close()
            except Exception:
                pass



    def _t10_apply_escape(text):
        if not t10_escape.get():
            return text
        text = text.replace('\\r', '\r').replace('\\n', '\n')\
                   .replace('\\t', '\t').replace('\\0', '\0')
        def _hex_repl(m):
            try:
                return chr(int(m.group(1), 16))
            except Exception:
                return m.group(0)
        text = re.sub(r'\\x([0-9a-fA-F]{2})', _hex_repl, text)
        def _oct_repl(m):
            try:
                return chr(int(m.group(1), 8))
            except Exception:
                return m.group(0)
        text = re.sub(r'\\([0-7]{1,3})', _oct_repl, text)
        return text

    def _t10_send(silent=False):
        if not t10_ser or not t10_ser.is_open:
            if not silent:
                messagebox.showinfo('提示', '请先打开串口')
            return
        raw = t10_send_entry.get('1.0', 'end-1c')
        if not raw and not t10_append_rn.get():
            return
        raw = _t10_apply_escape(raw)
        try:
            if t10_hex_send.get():
                hex_str = re.sub(r'[\s,;]+', ' ', raw).strip()
                data = bytes.fromhex(hex_str.replace(' ', ''))
            else:
                data = raw.encode(t10_encoding.get(), errors='replace')
        except ValueError:
            messagebox.showerror('格式错误',
                'Hex 格式错误，请使用 "AA BB 0D" 格式')
            return

        # 自动换行: 勾选后固定追加 \r\n (简化: 不再提供后缀选项)
        if t10_append_rn.get():
            data += b'\r\n'

        try:
            t10_ser.write(data)
            nonlocal t10_tx_total
            t10_tx_total += len(data)
            _t10_update_counts()
            _t10_log_to_box(data, is_tx=True, is_hex=t10_hex_send.get())
            # 不清空发送框 — 保留内容方便重复发送
        except (serial.SerialException, OSError) as e:
            messagebox.showerror('发送失败', str(e))
            _t10_close()

    def _t10_log_to_box(data, is_tx=True, is_hex=False):
        nonlocal t10_input_start_mark
        """向数据框写入一条发送记录 (带 [Tx] 标签, 仅发送框触发)"""
        # 前缀 (发送框走这里, 带标签)
        prefix = '[Tx] ' if is_tx else '[Rx] '
        tag = 'tx_tag' if is_tx else 'rx_tag'
        t10_rx_text.insert('end', prefix, tag)
        # 内容
        if is_hex:
            text = ' '.join(f'{b:02X}' for b in data)
        else:
            try:
                text = data.decode(t10_encoding.get(), errors='replace')
            except Exception:
                text = data.decode('latin-1', errors='replace')
        t10_rx_text.insert('end', text)
        t10_rx_text.see('end')
        # ── 更新输入起点 + 终端化 ──
        t10_input_start_mark = t10_rx_text.index('end-1c')
        t10_rx_text.mark_set('insert', t10_input_start_mark)
        _t10_lock_history()
        # 写入日志 (如已开启)
        if is_tx:
            _t10_write_log('Tx', data)

    def _t10_toggle_timer():
        nonlocal t10_timed_job
        if t10_timed_enable.get():
            _t10_start_timer()
        else:
            _t10_stop_timer()

    def _t10_start_timer():
        nonlocal t10_timed_job
        _t10_stop_timer()
        try:
            interval = max(50, int(t10_timed_ms.get()))
        except ValueError:
            interval = 1000
            t10_timed_ms.set('1000')

        def _timed_loop():
            nonlocal t10_timed_job
            if t10_ser and t10_ser.is_open and t10_timed_enable.get():
                _t10_send(silent=True)
                t10_timed_job = root.after(interval, _timed_loop)

        t10_timed_job = root.after(interval, _timed_loop)

    def _t10_stop_timer():
        nonlocal t10_timed_job
        if t10_timed_job:
            root.after_cancel(t10_timed_job)
            t10_timed_job = None

    def _t10_update_counts():
        t10_rx_label.set(f'Rx: {t10_rx_total:,}')
        t10_tx_label.set(f'Tx: {t10_tx_total:,}')

    def _t10_handle_error():
        messagebox.showwarning('串口断开', '串口连接异常断开')
        _t10_close()
        _t10_update_counts()

    def _t10_send_file():
        if not t10_ser or not t10_ser.is_open:
            messagebox.showinfo('提示', '请先打开串口')
            return
        fp = filedialog.askopenfilename(
            title='选择要发送的文件',
            filetypes=[('所有文件', '*.*'),
                       ('二进制文件', '*.bin'),
                       ('文本文件', '*.txt'),
                       ('Hex 文件', '*.hex')])
        if not fp:
            return
        try:
            with open(fp, 'rb') as f:
                data = f.read()
        except Exception as e:
            messagebox.showerror('读取失败', str(e))
            return
        if not data:
            messagebox.showinfo('提示', '文件为空')
            return
        if len(data) > 1024 * 1024:
            if not messagebox.askyesno('确认',
                f'文件大小 {len(data):,} 字节 ({len(data)/1024/1024:.1f} MB)\n'
                f'发送可能需要较长时间，确认发送？'):
                return
        try:
            chunk_size = 4096
            sent = 0
            for i in range(0, len(data), chunk_size):
                chunk = data[i:i+chunk_size]
                t10_ser.write(chunk)
                sent += len(chunk)
            nonlocal t10_tx_total
            t10_tx_total += sent
            _t10_update_counts()
            _t10_log_to_box(
                f'📁 {os.path.basename(fp)} ({sent:,} 字节)\n'.encode(),
                is_tx=True, is_hex=False)
        except (serial.SerialException, OSError) as e:
            messagebox.showerror('发送失败', str(e))
            _t10_close()

    # --- 样式标签 ---
    t10_rx_text.tag_config('ts', foreground=C['sub'], font=(M, 8))
    t10_rx_text.tag_config('hex', foreground=C['blue'], font=(M, 10))
    t10_rx_text.tag_config('rx_tag', foreground='#16a34a',
                           font=(M, 10, 'bold'))  # 接收前缀: 绿色
    t10_rx_text.tag_config('tx_tag', foreground='#ea580c',
                           font=(M, 10, 'bold'))  # 发送前缀: 橙色

    # --- 初始化 ---
    if _HAS_SERIAL:
        _t10_refresh_ports()
        _t10_update_info()
    else:
        t10_port_cb['values'] = ['(pyserial 未安装)']
        t10_port.set('(pyserial 未安装)')
        t10_status.set('⚠ pyserial 未安装')
        hint = _SERIAL_HINT or _SERIAL_INSTALL_ERR.strip() or '(无详细信息)'
        t10_rx_text.insert('1.0',
            f'⚠ 串口助手需要 pyserial 库，但当前环境找不到它。\n\n'
            f'Python: {_sys.executable}\n\n'
            f'━━━━━━━━━━━━━━━━━━━━━━━\n'
            f'{hint}\n')
        # ── 终端化: 错误提示一次性写入, 也立即锁住 ──
        _t10_lock_history()

    # TAB 11 — 网络助手
    # ══════════════════════════════════════
    import socket as _socket
    import time as _time
    import struct as _struct

    t11 = ttk.Frame(nb, style='TFrame')
    nb.add(t11, text='🌐 网络助手')
    t11.grid_rowconfigure(0, weight=0)  # 网络配置 (网卡+流列表)
    t11.grid_rowconfigure(1, weight=1)  # 收发日志 (主显示区, 按钮行已合并到日志框内)
    t11.grid_columnconfigure(0, weight=1)

    # --- 网卡枚举 ---
    def _t11_list_ips():
        """枚举本机所有非环回网卡, 返回 '名称 — IP' 格式列表, 方便识别"""
        items = []

        if _sys.platform == 'win32':
            # ── Windows: 解析 ipconfig, 提取适配器名 + IPv4 ──
            try:
                r = subprocess.run(['ipconfig'], capture_output=True,
                                   text=True, timeout=5)
                current_adapter = ''
                for line in r.stdout.split('\n'):
                    line = line.rstrip()
                    # 检测适配器名: "以太网适配器 以太网:" 或 "无线局域网适配器 Wi-Fi:"
                    m = re.match(r'^(.+适配器)\s+(.+):$', line)
                    if m:
                        current_adapter = m.group(2).strip()
                        continue
                    # 检测 IPv4
                    m = re.search(r'IPv4[^:]*:\s*(\d+\.\d+\.\d+\.\d+)', line)
                    if m and current_adapter:
                        ip = m.group(1)
                        if not ip.startswith('127.'):
                            label = f'{current_adapter}  —  {ip}'
                            if ip not in [x.split('—')[-1].strip() for x in items]:
                                items.append(label)
            except Exception:
                pass
        else:
            # ── Linux/macOS: 解析 ip addr, 提取接口名 + inet ──
            try:
                r = subprocess.run(['ip', 'addr'], capture_output=True,
                                   text=True, timeout=5)
                current_iface = ''
                for line in r.stdout.split('\n'):
                    line = line.rstrip()
                    m = re.match(r'^\d+:\s+(\S+):', line)
                    if m:
                        current_iface = m.group(1)
                        continue
                    m = re.search(r'inet\s+(\d+\.\d+\.\d+\.\d+)', line)
                    if m and current_iface and current_iface != 'lo':
                        ip = m.group(1)
                        if not ip.startswith('127.'):
                            label = f'{current_iface}  —  {ip}'
                            if ip not in [x.split('—')[-1].strip() for x in items]:
                                items.append(label)
            except Exception:
                pass

        # ── 兜底: socket 聚合 (可能只有 IP 没有名称) ──
        if not items:
            try:
                host = _socket.gethostname()
                for info in _socket.getaddrinfo(host, None):
                    ip = info[4][0]
                    if ip not in [x.split('—')[-1].strip() for x in items] \
                       and not ip.startswith('127.'):
                        items.append(f'{host}  —  {ip}')
            except Exception:
                pass

        return items or ['0.0.0.0 (默认)']

    def _t11_extract_ip(display_str):
        """从 'Wi-Fi  —  192.168.1.1' 提取 '192.168.1.1'"""
        m = re.search(r'(\d+\.\d+\.\d+\.\d+)', display_str or '')
        return m.group(1) if m else display_str

    # --- 协议发送核心 ---
    def _t11_udp_send(src_ip, dst_ip, dst_port, data):
        sock = _socket.socket(_socket.AF_INET, _socket.SOCK_DGRAM)
        sock.setsockopt(_socket.SOL_SOCKET, _socket.SO_REUSEADDR, 1)
        try:
            sock.bind((src_ip, 0))
        except Exception:
            pass  # 绑不了就用默认
        sock.sendto(data, (dst_ip, dst_port))
        sock.close()
        return len(data)

    def _t11_tcp_send(src_ip, dst_ip, dst_port, data, timeout=3):
        sock = _socket.socket(_socket.AF_INET, _socket.SOCK_STREAM)
        sock.settimeout(timeout)
        sock.setsockopt(_socket.SOL_SOCKET, _socket.SO_REUSEADDR, 1)
        try:
            sock.bind((src_ip, 0))
        except Exception:
            pass
        try:
            sock.connect((dst_ip, dst_port))
            sock.sendall(data)
            # 尝试接收响应
            try:
                resp = sock.recv(4096)
            except _socket.timeout:
                resp = b''
            sock.close()
            return len(data), resp
        except Exception as e:
            sock.close()
            raise e

    def _t11_ping(ip, size=64, count=4):
        cmd = ['ping', '-n' if _sys.platform == 'win32' else '-c',
               str(count),
               '-l' if _sys.platform == 'win32' else '-s',
               str(size), ip]
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=count * 3 + 5)
        return r.stdout

    def _t11_arp_table():
        r = subprocess.run(['arp', '-a'], capture_output=True,
                          text=True, timeout=5)
        return r.stdout

    def _t11_gen_data(length, mode='random', custom_hex=''):
        if mode == 'random':
            return os.urandom(length)
        elif mode == 'zero':
            return b'\x00' * length
        elif mode == 'ff':
            return b'\xff' * length
        elif mode == 'aa':
            return b'\xaa' * length
        elif mode == 'inc':
            return bytes(i % 256 for i in range(length))
        elif mode == 'custom' and custom_hex:
            try:
                h = re.sub(r'[\s,;]+', '', custom_hex)
                raw = bytes.fromhex(h)
                # 重复填充到目标长度
                return (raw * ((length // len(raw)) + 1))[:length]
            except ValueError:
                return b'\x00' * length
        return b'\x00' * length

    # --- Tab 11 UI ---
    # ===== 流 (Stream) 系统: 每条流独立配置 (协议/网卡/IP/端口/长度/模式/速率/hex) =====
    t11_streams = []           # list[dict], 每个 dict 是一条流的完整配置
    t11_current_idx = 0        # 当前选中的流索引
    t11_loading_stream = False # 同步 StringVar 时防止回环

    def _t11_new_stream(name=None, copy_from=None):
        """新建一条流 (默认名 stream_N)"""
        if copy_from is not None and 0 <= copy_from < len(t11_streams):
            base = dict(t11_streams[copy_from])
        else:
            base = {
                'proto': 'UDP', 'dst_ip': '192.168.1.100',
                'dst_port': '8080', 'len': '64', 'mode': 'random',
                'rate': '100', 'custom': '',
            }
        if name is None:
            name = f'stream_{len(t11_streams) + 1}'
        base['name'] = name
        base['enabled'] = True   # 默认启用发送
        t11_streams.append(base)
        return len(t11_streams) - 1

    def _t11_save_current_to_stream():
        """把 StringVar 当前值保存到 t11_current_idx 指向的流
        (注: nic 是全局的, 不存到流 dict)"""
        if 0 <= t11_current_idx < len(t11_streams):
            s = t11_streams[t11_current_idx]
            s['proto'] = t11_proto.get()
            s['dst_ip'] = t11_dst_ip.get()
            s['dst_port'] = t11_dst_port.get()
            s['len'] = t11_len.get()
            s['mode'] = t11_mode.get()
            s['rate'] = t11_rate.get()
            s['custom'] = t11_custom.get()

    def _t11_load_stream(idx):
        """从流 idx 加载到 StringVar (使用 nonlocal 标志防止 trace 回环)
        (注: nic 是全局的, 不加载)"""
        nonlocal t11_loading_stream
        if not (0 <= idx < len(t11_streams)):
            return
        s = t11_streams[idx]
        t11_loading_stream = True
        try:
            t11_proto.set(s['proto'])
            t11_dst_ip.set(s['dst_ip'])
            t11_dst_port.set(s['dst_port'])
            t11_len.set(s['len'])
            t11_mode.set(s['mode'])
            t11_rate.set(s['rate'])
            t11_custom.set(s['custom'])
        finally:
            t11_loading_stream = False

    # 默认添加一条流
    _t11_new_stream(name='stream_1')
    t11_streams[0].update({
        'proto': 'UDP', 'dst_ip': '192.168.1.100',
        'dst_port': '8080', 'len': '64', 'mode': 'random',
        'rate': '100', 'custom': '',
    })

    # 顶部: 流管理区 (左: 流列表, 右: 框内 [网卡行 + 4个流操作按钮] 居中)
    fc11 = ttk.LabelFrame(t11, text=' 网络配置 ', )
    fc11.grid(row=0, column=0, sticky='ew', padx=12, pady=(10, 4))
    fc11.grid_columnconfigure(0, weight=1)  # 左: 流列表列 (可扩展)
    fc11.grid_columnconfigure(1, weight=0)  # 右: 框 (固定宽度)
    fc11.grid_rowconfigure(0, weight=1)

    # --- 左: 流列表 (整宽, Treeview 带复选框) ---
    stream_frame = ttk.Frame(fc11, style='TFrame')
    stream_frame.grid(row=0, column=0, sticky='nsew',
                      padx=(8, 4), pady=6)
    ttk.Label(stream_frame, text='流列表', font=(F, 9, 'bold'),
              foreground=C['blue']).pack(side='left', padx=(0, 6), anchor='n', pady=2)
    # Treeview 列: 启用 / 流名称 / 协议 (3 列, 不再有"当前"列)
    t11_stream_list = ttk.Treeview(stream_frame,
                                   columns=('en', 'name', 'proto'),
                                   show='headings', height=6,
                                   selectmode='browse')
    t11_stream_list.heading('en', text='启用')
    t11_stream_list.heading('name', text='流名称')
    t11_stream_list.heading('proto', text='协议')
    t11_stream_list.column('en', width=50, anchor='center')
    t11_stream_list.column('name', width=200, anchor='w')
    t11_stream_list.column('proto', width=80, anchor='center')
    t11_stream_list.pack(side='left', fill='both', expand=True, padx=(0, 4))
    t11_stream_list.tag_configure('cur_row', background='#dbeafe',
                                  font=(M, 10, 'bold'))
    t11_stream_list.tag_configure('dis_row', foreground='#9ca3af')
    # 点击 '启用' 列切换 enabled, 点击其他列切换当前流
    def _t11_on_tree_click(event):
        col = t11_stream_list.identify_column(event.x)  # '#1'..'#3'
        row = t11_stream_list.identify_row(event.y)
        if not row or not col:
            return
        item = t11_stream_list.identify('item', event.x, event.y)
        if not item:
            return
        if col == '#1':  # 启用列
            idx = int(item)
            t11_streams[idx]['enabled'] = not t11_streams[idx].get('enabled', True)
            _t11_refresh_stream_list()
        # 切换当前流 (任意非启用列点击都切换)
        else:
            idx = int(item)
            if idx != t11_current_idx:
                _t11_save_current_to_stream()
                t11_current_idx = idx
                _t11_load_stream(idx)
                _t11_refresh_stream_list()
    t11_stream_list.bind('<Button-1>', _t11_on_tree_click)

    # --- 右: 框 [网卡行 + 流操作按钮] (居中放置) ---
    right_box = ttk.LabelFrame(fc11, text=' 网卡与流操作 ', )
    right_box.grid(row=0, column=1, sticky='ns', padx=(4, 8), pady=6)
    # 框内 grid: 两行, 一列, 内容居中
    right_box.grid_columnconfigure(0, weight=1)
    right_box.grid_rowconfigure(0, weight=1)  # 网卡行
    right_box.grid_rowconfigure(1, weight=0)  # 4个流操作按钮

    # 内层 Frame, 用来整体居中
    right_inner = ttk.Frame(right_box, style='TFrame')
    right_inner.grid(row=0, column=0, sticky='')

    # 网卡行 (在 right_inner 内)
    ttk.Label(right_inner, text='网卡', font=(F, 9),
              foreground=C['sub']).grid(row=0, column=0, sticky='w',
              padx=(4, 4), pady=(4, 2))
    t11_nic = tk.StringVar(value='')
    t11_nic_cb = ttk.Combobox(right_inner, textvariable=t11_nic,
                              values=_t11_list_ips(), state='readonly',
                              font=(F, 10), width=24)
    t11_nic_cb.grid(row=1, column=0, sticky='ew', padx=4, pady=(0, 2))
    ttk.Button(right_inner, text='↻', command=lambda: (
        t11_nic_cb.configure(values=_t11_list_ips()) or
        t11_nic.set(_t11_list_ips()[0] if _t11_list_ips() else '')),
        style='Small.TButton', width=3).grid(row=1, column=1, sticky='w',
        padx=(2, 4), pady=(0, 2))

    # 报文编辑按钮 (放到下一行, 居中)
    ttk.Button(right_inner, text='⚙ 报文编辑', style='Accent.TButton',
               command=lambda: _t11_open_pkt_editor()).grid(
        row=2, column=0, columnspan=2, sticky='ew', padx=4, pady=(2, 4))

    # 4 个流操作按钮 (放到 right_box 第 2 行, 居中)
    stream_btn_bar = ttk.Frame(right_box, style='TFrame')
    stream_btn_bar.grid(row=1, column=0, sticky='', padx=4, pady=(4, 4))
    ttk.Button(stream_btn_bar, text='➕', width=4, style='Small.TButton',
               command=lambda: _t11_add_stream()).pack(
        side='left', padx=(0, 4))
    ttk.Button(stream_btn_bar, text='📋', width=4, style='Small.TButton',
               command=lambda: _t11_copy_stream()).pack(side='left', padx=4)
    ttk.Button(stream_btn_bar, text='🗑', width=4, style='Small.TButton',
               command=lambda: _t11_del_stream()).pack(side='left', padx=4)
    ttk.Button(stream_btn_bar, text='✎', width=4, style='Small.TButton',
               command=lambda: _t11_rename_stream()).pack(side='left', padx=4)

    # 其他 StringVar (弹窗内编辑)
    t11_proto     = tk.StringVar(value='UDP')   # 协议 (弹窗内编辑)
    t11_dst_ip    = tk.StringVar(value='192.168.1.100')
    t11_dst_port  = tk.StringVar(value='8080')
    t11_len       = tk.StringVar(value='64')
    t11_mode      = tk.StringVar(value='random')
    t11_rate      = tk.StringVar(value='100')
    t11_custom    = tk.StringVar(value='')      # 用户编辑的完整 hex (含协议头)
    t11_save_log  = tk.BooleanVar(value=False)  # 勾选后自动写文件

    # 加载第一条流到 StringVar
    _t11_load_stream(0)

    def _t11_refresh_stream_list():
        # 清空旧行
        for iid in t11_stream_list.get_children():
            t11_stream_list.delete(iid)
        for i, s in enumerate(t11_streams):
            en = '☑' if s.get('enabled', True) else '☐'
            name = s.get('name', f'stream_{i+1}')
            proto = s.get('proto', '?')
            tags = ('cur_row',) if i == t11_current_idx else (
                ('dis_row',) if not s.get('enabled', True) else ())
            t11_stream_list.insert('', 'end', iid=str(i),
                                   values=(en, name, proto),
                                   tags=tags)
        # 选中并滚动到当前流
        if 0 <= t11_current_idx < len(t11_streams):
            try:
                t11_stream_list.selection_set(str(t11_current_idx))
                t11_stream_list.focus(str(t11_current_idx))
                t11_stream_list.see(str(t11_current_idx))
            except Exception:
                pass
    _t11_refresh_stream_list()

    def _t11_add_stream():
        _t11_save_current_to_stream()
        _t11_new_stream()
        t11_current_idx = len(t11_streams) - 1
        _t11_load_stream(t11_current_idx)
        _t11_refresh_stream_list()

    def _t11_copy_stream():
        if not t11_streams:
            return
        _t11_save_current_to_stream()
        new_idx = _t11_new_stream(
            name=f'{t11_streams[t11_current_idx]["name"]}_copy',
            copy_from=t11_current_idx)
        t11_current_idx = new_idx
        _t11_load_stream(t11_current_idx)
        _t11_refresh_stream_list()

    def _t11_del_stream():
        if len(t11_streams) <= 1:
            messagebox.showinfo('提示', '至少保留 1 条流')
            return
        if not messagebox.askyesno('确认',
                f'确定删除 "{t11_streams[t11_current_idx]["name"]}"?'):
            return
        del t11_streams[t11_current_idx]
        t11_current_idx = min(t11_current_idx, len(t11_streams) - 1)
        _t11_load_stream(t11_current_idx)
        _t11_refresh_stream_list()

    def _t11_rename_stream():
        if not t11_streams:
            return
        from tkinter import simpledialog
        new_name = simpledialog.askstring(
            '重命名流', '新名称:', initialvalue=t11_streams[t11_current_idx]['name'],
            parent=root)
        if new_name:
            t11_streams[t11_current_idx]['name'] = new_name
            _t11_refresh_stream_list()

    # ── 报文编辑弹窗 (单例, 可调整大小) ──
    _t11_pkt_win = None

    def _t11_open_pkt_editor():
        nonlocal _t11_pkt_win
        if _t11_pkt_win is not None and _t11_pkt_win.winfo_exists():
            _t11_pkt_win.deiconify()
            _t11_pkt_win.lift()
            _t11_pkt_win.focus_force()
            return
        win = tk.Toplevel(root)
        win.title('报文编辑')
        win.resizable(True, True)
        win.geometry('640x580')
        win.minsize(540, 480)
        win.transient(root)

        def _on_pkt_close():
            nonlocal _t11_pkt_win
            try:
                win.destroy()
            finally:
                _t11_pkt_win = None
        win.protocol('WM_DELETE_WINDOW', _on_pkt_close)
        _t11_pkt_win = win

        main = ttk.Frame(win, style='TFrame', padding=12)
        main.pack(fill='both', expand=True)

        # ── 分组 1: 协议 + 模式 + 速率 (目标IP/端口/长度已合并到报文结构) ──
        gp1 = ttk.LabelFrame(main, text=' 协议设置 ')
        gp1.pack(fill='x', pady=(0, 8))
        for c in range(6):
            gp1.grid_columnconfigure(c, weight=0)
        gp1.grid_columnconfigure(1, weight=1)
        gp1.grid_columnconfigure(3, weight=1)
        gp1.grid_columnconfigure(5, weight=1)

        ttk.Label(gp1, text='协议', font=(F, 9, 'bold'),
                  foreground=C['blue']).grid(row=0, column=0, padx=(10, 4),
                                             pady=8, sticky='w')
        ttk.Combobox(gp1, textvariable=t11_proto,
                     values=['UDP','TCP','ICMP','ARP'],
                     state='readonly', font=(F, 10), width=6).grid(
            row=0, column=1, sticky='w', padx=(0, 12))

        ttk.Label(gp1, text='模式', font=(F, 9),
                  foreground=C['sub']).grid(row=0, column=2, padx=(0, 4),
                                             pady=8, sticky='w')
        ttk.Combobox(gp1, textvariable=t11_mode,
                     values=['random','zero','ff','aa','inc','custom'],
                     state='readonly', font=(F, 10), width=8).grid(
            row=0, column=3, sticky='w', padx=(0, 12), pady=8)

        ttk.Label(gp1, text='速率', font=(F, 9),
                  foreground=C['sub']).grid(row=0, column=4, padx=(0, 4),
                                             pady=8, sticky='w')
        ttk.Entry(gp1, textvariable=t11_rate, font=(M, 10),
                  width=5).grid(row=0, column=5, sticky='w', padx=(0, 4), pady=8)
        ttk.Label(gp1, text='包/秒', font=(F, 9),
                  foreground=C['sub']).grid(row=0, column=5, sticky='e',
                                             padx=(0, 10), pady=8)

        # ── 分组 2: 报文结构 (可编辑 Treeview, 双击行编辑字段) ──
        gp2 = ttk.LabelFrame(main, text=' 报文结构 (双击行编辑字段值) ')
        gp2.pack(fill='both', expand=True, pady=(0, 8))
        gp2.grid_columnconfigure(0, weight=1)
        gp2.grid_rowconfigure(0, weight=1)

        # Treeview 字段: #0=字段名(可展开), value=hex+十进制, bytes, offset, desc
        # 分层结构: 父节点=协议层 (Ethernet/IP/UDP...), 子节点=字段
        # 列: hex / dec / 字节(基于MAC首字节的偏移) / 说明
        tree_cols = ('hex', 'dec', 'bytes', 'desc')
        tree = ttk.Treeview(gp2, columns=tree_cols, show='tree headings',
                            height=8, selectmode='browse')
        tree.heading('#0', text='报文结构 (▶ 点击行 → 自定义hex高亮该字段)')
        tree.heading('hex', text='值 (hex)')
        tree.heading('dec', text='值 (dec/IP/MAC)')
        tree.heading('bytes', text='字节 (基于MAC首字节偏移)')
        tree.heading('desc', text='说明')
        tree.column('#0', width=180, anchor='w', stretch=True)
        tree.column('hex', width=200, anchor='w')
        tree.column('dec', width=180, anchor='w')
        tree.column('bytes', width=150, anchor='w')
        tree.column('desc', width=180, anchor='w')
        # 行样式
        tree.tag_configure('layer', foreground=C['blue'],
                           font=(M, 9, 'bold'))
        tree.tag_configure('field', foreground=C['fg'])
        tree.tag_configure('edit', foreground='#16a34a',
                           font=(M, 9, 'bold'))
        tree.grid(row=0, column=0, sticky='nsew', padx=4, pady=4)
        tree_sc = ttk.Scrollbar(gp2, orient='vertical', command=tree.yview)
        tree_sc.grid(row=0, column=1, sticky='ns', padx=(0, 4), pady=4)
        tree.configure(yscrollcommand=tree_sc.set)

        # ── 分组 3: 自定义Hex (直接编辑完整报文) ──
        gp3 = ttk.LabelFrame(main, text=' 自定义Hex (完整报文, 可直接编辑) ')
        gp3.pack(fill='x', pady=(0, 8))
        gp3.grid_columnconfigure(0, weight=1)
        gp3.grid_rowconfigure(0, weight=1)

        pkt_hex_text = tk.Text(gp3, font=(M, 10), height=4, wrap='word',
                               relief='flat', padx=8, pady=4, bg='#fafbfc',
                               undo=True)
        pkt_hex_text.grid(row=0, column=0, sticky='nsew', padx=4, pady=4)
        pkt_hex_sc = ttk.Scrollbar(gp3, orient='vertical',
                                  command=pkt_hex_text.yview)
        pkt_hex_sc.grid(row=0, column=1, sticky='ns', padx=(0, 4), pady=4)
        pkt_hex_text.configure(yscrollcommand=pkt_hex_sc.set)

        # 弹窗内 Text <-> StringVar 双向同步
        _t11_pkt_guard = [False]
        def _sync_var_to_pkt(*_):
            if _t11_pkt_guard[0]:
                return
            cur = t11_custom.get()
            if pkt_hex_text.get('1.0', 'end-1c') != cur:
                _t11_pkt_guard[0] = True
                try:
                    pkt_hex_text.delete('1.0', 'end')
                    pkt_hex_text.insert('1.0', cur)
                finally:
                    _t11_pkt_guard[0] = False
        def _sync_pkt_to_var(*_):
            if _t11_pkt_guard[0]:
                return
            _t11_pkt_guard[0] = True
            try:
                t11_custom.set(pkt_hex_text.get('1.0', 'end-1c'))
            finally:
                _t11_pkt_guard[0] = False
        t11_custom.trace_add('write', _sync_var_to_pkt)
        def _on_pkt_modified(event=None):
            # 必须重置 modified 标志, 否则后续修改不再触发
            pkt_hex_text.edit_modified(False)
            root.after_idle(_sync_pkt_to_var)
        pkt_hex_text.bind('<<Modified>>', _on_pkt_modified)
        _sync_var_to_pkt()

        # === 报文结构定义 (按协议类型) ===
        # 字段定义: (name, size, desc, parser, builder)
        # parser(text) -> int  (从用户输入解析为整数)
        # builder(int)  -> bytes
        def _u8(s):
            v = int(s, 0) & 0xff
            return v
        def _u16(s):
            v = int(s, 0) & 0xffff
            return v
        def _u32(s):
            v = int(s, 0) & 0xffffffff
            return v
        def _u48(s):
            # 6 字节 MAC / SHA
            s2 = re.sub(r'[\s:-]', '', s)
            if len(s2) == 12 and all(c in '0123456789abcdefABCDEF' for c in s2):
                return int(s2, 16)
            v = int(s2, 0) & ((1 << 48) - 1)
            return v
        def _ip(s):
            parts = s.strip().split('.')
            if len(parts) == 4:
                return (int(parts[0]) << 24) | (int(parts[1]) << 16) | \
                       (int(parts[2]) << 8) | int(parts[3])
            return int(s, 0) & 0xffffffff
        def _str(s):
            return s.encode('latin-1', errors='replace')
        def _hex(s):
            # 任意字节 hex 字符串 (空格分隔或连续)
            s2 = re.sub(r'[\s,;]+', '', s)
            return int(s2, 16) if s2 else 0

        # 协议模板: 嵌套结构, 每层是一个 dict:
        #   {'name': 层名, 'fields': [...], 'editable': True/False}
        # fields: (name, size, desc, parser, builder)
        #   size=None 表示 data 字段, 占剩余字节
        #   parser(text) -> int
        #   builder(int)  -> bytes
        ETH_FIELDS = [
            ('dmac', 6, '目标 MAC', _u48, lambda v: v.to_bytes(6, 'big')),
            ('smac', 6, '源 MAC',   _u48, lambda v: v.to_bytes(6, 'big')),
            ('etype', 2, 'EtherType', _u16, lambda v: v.to_bytes(2, 'big')),
        ]
        IPV4_FIELDS = [
            ('ver_ihl', 1, '版本+IHL',     _u8,  lambda v: v.to_bytes(1, 'big')),
            ('tos',     1, 'TOS/DSCP',     _u8,  lambda v: v.to_bytes(1, 'big')),
            ('tlen',    2, '总长度',       _u16, lambda v: v.to_bytes(2, 'big')),
            ('ident',   2, '标识',         _u16, lambda v: v.to_bytes(2, 'big')),
            ('flags',   2, '标志+片偏移',   _u16, lambda v: v.to_bytes(2, 'big')),
            ('ttl',     1, 'TTL',          _u8,  lambda v: v.to_bytes(1, 'big')),
            ('proto',   1, '协议号',       _u8,  lambda v: v.to_bytes(1, 'big')),
            ('csum',    2, 'IP 校验和',    _u16, lambda v: v.to_bytes(2, 'big')),
            ('sip',     4, '源 IP',        _ip,  lambda v: v.to_bytes(4, 'big')),
            ('dip',     4, '目标 IP',      _ip,  lambda v: v.to_bytes(4, 'big')),
        ]
        UDP_FIELDS = [
            ('sport', 2, '源端口',     _u16, lambda v: v.to_bytes(2, 'big')),
            ('dport', 2, '目标端口',   _u16, lambda v: v.to_bytes(2, 'big')),
            ('len',   2, '长度(含头)', _u16, lambda v: v.to_bytes(2, 'big')),
            ('csum',  2, '校验和',     _u16, lambda v: v.to_bytes(2, 'big')),
            ('data',  None, '数据(剩余字节)', _hex, None),
        ]
        TCP_FIELDS = [
            ('sport', 2, '源端口',     _u16, lambda v: v.to_bytes(2, 'big')),
            ('dport', 2, '目标端口',   _u16, lambda v: v.to_bytes(2, 'big')),
            ('seq',   4, '序列号',     _u32, lambda v: v.to_bytes(4, 'big')),
            ('ack',   4, '确认号',     _u32, lambda v: v.to_bytes(4, 'big')),
            ('flags', 2, 'DataOff+Flags', _u16, lambda v: v.to_bytes(2, 'big')),
            ('win',   2, '窗口大小',   _u16, lambda v: v.to_bytes(2, 'big')),
            ('csum',  2, '校验和',     _u16, lambda v: v.to_bytes(2, 'big')),
            ('urg',   2, '紧急指针',   _u16, lambda v: v.to_bytes(2, 'big')),
            ('data',  None, '数据(剩余字节)', _hex, None),
        ]
        ICMP_FIELDS = [
            ('type',  1, '类型',       _u8,  lambda v: v.to_bytes(1, 'big')),
            ('code',  1, '代码',       _u8,  lambda v: v.to_bytes(1, 'big')),
            ('csum',  2, '校验和',     _u16, lambda v: v.to_bytes(2, 'big')),
            ('id',    2, '标识',       _u16, lambda v: v.to_bytes(2, 'big')),
            ('seq',   2, '序列号',     _u16, lambda v: v.to_bytes(2, 'big')),
            ('data',  None, '数据(剩余字节)', _hex, None),
        ]
        ARP_FIELDS = [
            ('htype', 2, '硬件类型',   _u16, lambda v: v.to_bytes(2, 'big')),
            ('ptype', 2, '协议类型',   _u16, lambda v: v.to_bytes(2, 'big')),
            ('hlen',  1, '硬件地址长度', _u8, lambda v: v.to_bytes(1, 'big')),
            ('plen',  1, '协议地址长度', _u8, lambda v: v.to_bytes(1, 'big')),
            ('oper',  2, '操作码',     _u16, lambda v: v.to_bytes(2, 'big')),
            ('sha',   6, '发送方 MAC', _u48, lambda v: v.to_bytes(6, 'big')),
            ('spa',   4, '发送方 IP',  _ip,  lambda v: v.to_bytes(4, 'big')),
            ('tha',   6, '目标 MAC',   _u48, lambda v: v.to_bytes(6, 'big')),
            ('tpa',   4, '目标 IP',    _ip,  lambda v: v.to_bytes(4, 'big')),
        ]

        # 协议模板 (分层): 每条流显示完整包结构: Ethernet II + (IP) + 传输层
        PROTO_TEMPLATES = {
            'UDP': [
                {'name': 'Ethernet II', 'fields': ETH_FIELDS, 'editable': True},
                {'name': 'IPv4',        'fields': IPV4_FIELDS, 'editable': True},
                {'name': 'UDP',         'fields': UDP_FIELDS,  'editable': True},
            ],
            'TCP': [
                {'name': 'Ethernet II', 'fields': ETH_FIELDS, 'editable': True},
                {'name': 'IPv4',        'fields': IPV4_FIELDS, 'editable': True},
                {'name': 'TCP',         'fields': TCP_FIELDS,  'editable': True},
            ],
            'ICMP': [
                {'name': 'Ethernet II', 'fields': ETH_FIELDS, 'editable': True},
                {'name': 'IPv4',        'fields': IPV4_FIELDS, 'editable': True},
                {'name': 'ICMP',        'fields': ICMP_FIELDS, 'editable': True},
            ],
            'ARP': [
                {'name': 'Ethernet II', 'fields': ETH_FIELDS, 'editable': True},
                {'name': 'ARP',         'fields': ARP_FIELDS,  'editable': True},
            ],
        }

        def _get_field_size(template, idx, total_len):
            """计算字段实际占用字节数"""
            name, size, desc, parser, builder = template[idx]
            if size is not None:
                return size
            # data 字段: 剩余字节
            used = sum((t[1] or 0) for t in template[:idx] if t[1])
            return max(0, total_len - used)

        def _format_hex(raw_bytes):
            """只格式化为 hex 字符串: "AA BB CC DD" """
            if not raw_bytes:
                return '(空)'
            return ' '.join(f'{b:02X}' for b in raw_bytes)

        def _format_dec(raw_bytes, name='', size=None):
            """格式化为十进制/IP/MAC/文本表示, 单独一列"""
            if not raw_bytes:
                return ''
            n = len(raw_bytes)
            lname = name.lower()
            if n == 1:
                return f'{raw_bytes[0]}'
            elif n == 2 and 'port' in lname:
                val = (raw_bytes[0] << 8) | raw_bytes[1]
                return f'{val}'
            elif n == 2 and ('len' in lname or 'tlen' == lname):
                val = (raw_bytes[0] << 8) | raw_bytes[1]
                return f'{val}'
            elif n == 2 and lname in ('type', 'code'):
                val = (raw_bytes[0] << 8) | raw_bytes[1]
                return f'{val}'
            elif n == 2 and lname == 'flags':
                val = (raw_bytes[0] << 8) | raw_bytes[1]
                return f'0x{val:04X}'
            elif n == 2:
                val = (raw_bytes[0] << 8) | raw_bytes[1]
                return f'{val}'
            elif n == 4 and 'ip' in lname:
                return f'{raw_bytes[0]}.{raw_bytes[1]}.' \
                       f'{raw_bytes[2]}.{raw_bytes[3]}'
            elif n == 4 and lname in ('seq', 'ack', 'ident'):
                val = int.from_bytes(raw_bytes, 'big')
                return f'{val}'
            elif n == 4:
                val = int.from_bytes(raw_bytes, 'big')
                return f'0x{val:08X}'
            elif n == 6:
                return f'{raw_bytes[0]:02X}:{raw_bytes[1]:02X}:' \
                       f'{raw_bytes[2]:02X}:{raw_bytes[3]:02X}:' \
                       f'{raw_bytes[4]:02X}:{raw_bytes[5]:02X}'
            return ''

        # 记录每行对应的字节范围 (offset, size), 用于 hex 高亮
        _field_ranges = {}  # iid -> (offset, size)

        def _refresh_tree(*_):
            """根据 t11_custom 当前 hex 重生成 Treeview 字段值
            分层结构: 父节点=协议层 (可展开/折叠), 子节点=字段"""
            tree.delete(*tree.get_children())
            _field_ranges.clear()
            proto = t11_proto.get()
            layers = PROTO_TEMPLATES.get(proto, [])
            h = re.sub(r'[\s,;]+', '', t11_custom.get())
            try:
                pkt = bytes.fromhex(h) if h else b''
            except ValueError:
                pkt = b''
            total_len = len(pkt)
            global_offset = 0
            for layer_idx, layer in enumerate(layers):
                lname = layer['name']
                fields = layer['fields']
                # 估算本层总字节数
                layer_bytes = 0
                for fname, fsize, fdesc, fp, fb in fields:
                    if fsize is not None:
                        layer_bytes += fsize
                layer_end = min(global_offset + layer_bytes, total_len)
                # 插入父节点 (协议层), 可展开
                layer_iid = f'L{layer_idx}'
                tree.insert('', 'end', iid=layer_iid,
                            text=f'  ▼ {lname}  [{global_offset:#06x}-{layer_end-1:#06x}]',
                            values=(f'{layer_bytes} 字节', f'0x{global_offset:04X}', ''),
                            tags=('layer',), open=True)
                # 插入子节点 (字段)
                field_offset = global_offset
                for field_idx, (fname, fsize, fdesc, fp, fb) in enumerate(fields):
                    # 实际占字节数
                    if fsize is None:
                        # data 字段 = 剩余字节
                        used = sum(f[1] for f in fields[:field_idx] if f[1])
                        actual_size = max(0, total_len - field_offset)
                    else:
                        actual_size = fsize
                    fdata = pkt[field_offset:field_offset + actual_size]
                    hex_str = _format_hex(fdata)
                    dec_str = _format_dec(fdata, fname, fsize)
                    tag = 'edit' if fdata else 'field'
                    child_iid = f'L{layer_idx}.{field_idx}'
                    # "字节"列: 显示该字段第一个字节的偏移 (基于 MAC 首字节 0)
                    # 例如 dmac 起始偏移 0x0000, smac 起始偏移 0x0006, sip 起始偏移 0x001A
                    # 后面附"4 字节"表示该字段长度
                    bytes_str = f'@{field_offset:#06x} ({actual_size}字节)'
                    tree.insert(layer_iid, 'end', iid=child_iid,
                                text=f'  {fname}',
                                values=(hex_str, dec_str, bytes_str, fdesc),
                                tags=(tag,))
                    _field_ranges[child_iid] = (field_offset, actual_size)
                    field_offset += actual_size
                global_offset = field_offset
            # 剩余未解析字节
            if global_offset < total_len:
                rest = pkt[global_offset:]
                rest_bytes = f'@{global_offset:#06x} ({len(rest)}字节)'
                tree.insert('', 'end', iid='rest',
                            text='  extra (未解析)',
                            values=(_format_hex(rest), '', rest_bytes,
                                    '未解析字节'),
                            tags=('field',))
                _field_ranges['rest'] = (global_offset, len(rest))

        def _on_tree_double_click(event):
            """双击行 -> 弹输入框编辑该字段值 (子节点/字段行才能编辑)"""
            item = tree.identify_row(event.y)
            if not item or item == 'rest':
                return
            # 只允许编辑子节点 (字段行, e.g. 'L0.1')
            if '.' not in item:
                return  # 父节点 (协议层) 不直接编辑
            try:
                layer_idx, field_idx = item.split('.')
                layer_idx = int(layer_idx[1:])
                field_idx = int(field_idx)
            except (ValueError, IndexError):
                return
            proto = t11_proto.get()
            layers = PROTO_TEMPLATES.get(proto, [])
            if layer_idx >= len(layers):
                return
            fields = layers[layer_idx]['fields']
            if field_idx >= len(fields):
                return
            name, size, desc, parser, builder = fields[field_idx]
            # 当前值 (hex 列在第 0 列)
            cur_vals = tree.item(item, 'values')
            cur_hex = cur_vals[0] if cur_vals else ''
            cur_dec = cur_vals[1] if len(cur_vals) > 1 else ''
            # 弹输入框
            from tkinter import simpledialog
            prompt = f'编辑 {layers[layer_idx]["name"]} / {name} ({desc})\n\n' \
                     f'支持格式: 十进制 123 / 十六进制 0x7B / MAC aa:bb:cc:dd:ee:ff / IP 1.2.3.4 / hex "AA BB CC"\n' \
                     f'当前 hex: {cur_hex}\n当前 dec: {cur_dec}'
            new_text = simpledialog.askstring(
                f'编辑 {name}', prompt,
                initialvalue=cur_hex.replace(' ', ''),
                parent=win)
            if new_text is None:
                return
            try:
                parsed = parser(new_text)
            except (ValueError, TypeError) as e:
                messagebox.showerror('格式错误', f'无法解析 "{new_text}":\n{e}',
                                     parent=win)
                return
            # 重新拼接 hex: 找到此字段在包中的 offset
            h = re.sub(r'[\s,;]+', '', t11_custom.get())
            try:
                pkt = bytearray(bytes.fromhex(h)) if h else bytearray()
            except ValueError:
                pkt = bytearray()
            # 计算字段 offset (遍历所有前置层 + 字段)
            offset = 0
            for li in range(layer_idx):
                for fn, fs, fd, fp, fb in layers[li]['fields']:
                    if fs is None:
                        offset += max(0, len(pkt) - offset)
                    else:
                        offset += fs
            for fi in range(field_idx):
                fn, fs, fd, fp, fb = fields[fi]
                if fs is None:
                    offset += max(0, len(pkt) - offset)
                else:
                    offset += fs
            if size is not None:
                new_bytes = builder(parsed)
                fsize = size
                pkt[offset:offset + fsize] = new_bytes
            else:
                # data 字段: 用 hex 字符串直接替换
                s2 = re.sub(r'[\s,;]+', '', new_text)
                try:
                    new_bytes = bytes.fromhex(s2)
                except ValueError:
                    messagebox.showerror('格式错误',
                        'data 字段需为 hex 字符串 (如 "AA BB CC DD")',
                        parent=win)
                    return
                pkt[offset:offset + len(new_bytes)] = new_bytes
            t11_custom.set(' '.join(f'{b:02X}' for b in pkt))
            # 同步字段到 StringVar, 避免下次重生成覆盖用户编辑
            # (sip/dip 是 4 字节 IP, sport/dport 是端口, dmac/smac 是 MAC, tlen/len 是长度)
            if name == 'sip' and size == 4:
                t11_dst_ip.set(f'{new_bytes[0]}.{new_bytes[1]}.'
                               f'{new_bytes[2]}.{new_bytes[3]}')
                t11_loading_stream = True
                t11_loading_stream = False
            elif name == 'dip' and size == 4:
                t11_dst_ip.set(f'{new_bytes[0]}.{new_bytes[1]}.'
                               f'{new_bytes[2]}.{new_bytes[3]}')
            elif name in ('dport', 'sport') and size == 2:
                val = (new_bytes[0] << 8) | new_bytes[1]
                t11_dst_port.set(str(val))
            # Treeview 通过 trace 自动刷新

        def _on_tree_select(event=None):
            """选中行 -> 在 自定义hex 框高亮对应字节范围"""
            sel = tree.selection()
            if not sel:
                return
            item = sel[0]
            if item not in _field_ranges:
                return
            off, sz = _field_ranges[item]
            try:
                h = re.sub(r'[\s,;]+', '', t11_custom.get())
                pkt = bytes.fromhex(h) if h else b''
            except ValueError:
                return
            if off >= len(pkt):
                return
            # hex 显示格式固定为 "AA BB CC DD ..." (每字节 2 个 hex + 1 个空格)
            # 字节 i 的字符位置 = i*3 (起始), i*3+2 (结束)
            c1 = off * 3
            c2 = min(off + sz, len(pkt)) * 3
            # 转换到 Text 索引 (1-based 行, 0-based 列)
            start_idx = f'1.0+{c1}c'
            end_idx = f'1.0+{c2}c'
            pkt_hex_text.tag_remove('hl', '1.0', 'end')
            pkt_hex_text.tag_add('hl', start_idx, end_idx)
            # 滚动到可见
            pkt_hex_text.see(start_idx)
        pkt_hex_text.tag_config('hl', background='#fef08a',
                                foreground='#7c2d12')

        tree.bind('<Double-1>', _on_tree_double_click)
        tree.bind('<<TreeviewSelect>>', _on_tree_select)

        # 当协议变化或 hex 变化, 刷新 Treeview
        def _on_proto_or_hex_change(*_):
            _refresh_tree()
        t11_proto.trace_add('write', _on_proto_or_hex_change)
        t11_custom.trace_add('write', _on_proto_or_hex_change)
        # 初次填充: 确保有内容可显示, 若 hex 为空则生成默认报文
        if not t11_custom.get().strip():
            _t11_preview_hex()
        else:
            root.after(50, _refresh_tree)

        # ── 底部按钮 ──
        btn_bar = ttk.Frame(main, style='TFrame')
        btn_bar.pack(fill='x', pady=(8, 0))

        def _t11_load_default_pkt():
            """加载默认报文: 完整 Ethernet II + IPv4 + UDP + payload"""
            t11_proto.set('UDP')
            t11_mode.set('random')
            _t11_preview_hex()

        ttk.Button(btn_bar, text='📋 默认报文',
                   command=_t11_load_default_pkt,
                   style='Small.TButton').pack(side='left', padx=(0, 6))
        ttk.Button(btn_bar, text='↻ 重置模式',
                   command=lambda: _t11_apply_mode(),
                   style='Small.TButton').pack(side='left', padx=(0, 6))

        # 最大化切换 (记录原始几何用于还原)
        _pkt_geo_before_max = [None]
        def _toggle_pkt_max():
            try:
                if _pkt_geo_before_max[0] is None:
                    # 当前是普通状态 -> 最大化
                    _pkt_geo_before_max[0] = win.geometry()
                    # 获取屏幕大小
                    sw = win.winfo_screenwidth()
                    sh = win.winfo_screenheight()
                    win.geometry(f'{sw}x{sh}+0+0')
                    _pkt_max_btn.config(text='🗗 还原')
                else:
                    # 还原
                    win.geometry(_pkt_geo_before_max[0])
                    _pkt_geo_before_max[0] = None
                    _pkt_max_btn.config(text='⛶ 最大化')
            except Exception:
                pass
        _pkt_max_btn = ttk.Button(btn_bar, text='⛶ 最大化',
                                  command=_toggle_pkt_max,
                                  style='Small.TButton')
        _pkt_max_btn.pack(side='right', padx=(6, 6))
        ttk.Button(btn_bar, text='关闭',
                   command=_on_pkt_close,
                   style='Accent.TButton').pack(side='right')

        # 拦截窗口自身的最大化按钮 (双击标题栏), 同步我们的按钮文字
        def _on_win_state(event=None):
            try:
                if win.state() == 'zoomed':
                    _pkt_max_btn.config(text='🗗 还原')
                    if _pkt_geo_before_max[0] is None:
                        _pkt_geo_before_max[0] = '640x580'
                else:
                    _pkt_max_btn.config(text='⛶ 最大化')
                    _pkt_geo_before_max[0] = None
            except Exception:
                pass
        win.bind('<Configure>', _on_win_state)

        def _t11_apply_mode():
            """按当前模式重新生成 hex (覆盖整个报文, UDP 头+payload)"""
            mode = t11_mode.get()
            proto = t11_proto.get()
            try:
                length = max(0, min(int(t11_len.get()), 256))
            except ValueError:
                length = 16
            import random as _r
            src_port = _r.randint(1024, 65535)
            try:
                dst_port = int(t11_dst_port.get()) if t11_dst_port.get() else 0
            except ValueError:
                dst_port = 0
            if mode == 'random':
                payload = os.urandom(length)
            elif mode == 'zero':
                payload = b'\x00' * length
            elif mode == 'ff':
                payload = b'\xff' * length
            elif mode == 'aa':
                payload = b'\xaa' * length
            elif mode == 'inc':
                payload = bytes(i % 256 for i in range(length))
            else:
                return  # 'custom' 不自动生成
            if proto == 'UDP':
                hdr = (src_port.to_bytes(2, 'big') +
                       dst_port.to_bytes(2, 'big') +
                       (8 + length).to_bytes(2, 'big') + b'\x00\x00')
            elif proto == 'TCP':
                hdr = (src_port.to_bytes(2, 'big') +
                       dst_port.to_bytes(2, 'big') +
                       _r.randint(0, 0xffffffff).to_bytes(4, 'big') +
                       (0).to_bytes(4, 'big') +
                       (0x5002).to_bytes(2, 'big') +  # data_offset=5, SYN
                       (65535).to_bytes(2, 'big') + b'\x00\x00\x00\x00')
            elif proto == 'ICMP':
                hdr = bytes([0x08, 0x00, 0x00, 0x00]) + \
                      _r.randint(0, 0xffff).to_bytes(2, 'big') + \
                      _r.randint(0, 0xffff).to_bytes(2, 'big')
            else:
                hdr = b''
            pkt = hdr + payload
            t11_custom.set(' '.join(f'{b:02X}' for b in pkt))

    # 模式变化时自动生成对应 hex 预览到 自定义Hex 框
    # 用户要求: 预览只显示"数据前 N 字节" (即 payload, 不含协议头)
    PREVIEW_PAYLOAD_LEN = 16  # 默认 payload 预览长度

    # ── 校验和算法 ──
    def _ip_checksum(header: bytes) -> int:
        """计算 IP/ICMP 校验和 (反码求和)"""
        if len(header) % 2:
            header += b'\x00'
        s = 0
        for i in range(0, len(header), 2):
            s += (header[i] << 8) | header[i+1]
        while s >> 16:
            s = (s & 0xffff) + (s >> 16)
        return (~s) & 0xffff

    def _udp_checksum(src_ip: str, dst_ip: str, udp_hdr: bytes, payload: bytes) -> int:
        """计算 UDP 校验和 (含伪头: src_ip + dst_ip + 0 + 17 + udp_len)"""
        s_ip = _ip_to_bytes(src_ip)
        d_ip = _ip_to_bytes(dst_ip)
        pseudo = s_ip + d_ip + b'\x00\x11' + len(udp_hdr + payload).to_bytes(2, 'big')
        data = pseudo + udp_hdr + payload
        if len(data) % 2:
            data += b'\x00'
        s = 0
        for i in range(0, len(data), 2):
            s += (data[i] << 8) | data[i+1]
        while s >> 16:
            s = (s & 0xffff) + (s >> 16)
        csum = (~s) & 0xffff
        return 0 if csum == 0 else csum  # 0 等同于无校验 (RFC 768)

    def _ip_to_bytes(ip: str) -> bytes:
        try:
            parts = ip.strip().split('.')
            if len(parts) == 4:
                return bytes(int(p) & 0xff for p in parts)
        except Exception:
            pass
        return b'\x00\x00\x00\x00'

    def _mac_to_bytes(mac: str) -> bytes:
        try:
            s = re.sub(r'[\s:-]', '', mac)
            if len(s) == 12:
                return bytes.fromhex(s)
        except Exception:
            pass
        return b'\x00' * 6

    # ── 默认 MAC (从网卡选中的条目提取; 没解析到则用全 0) ──
    def _t11_get_smac() -> bytes:
        nic = t11_nic.get() if 't11_nic' in dir() else ''
        # 简单返回 0, 实际硬件 MAC 需要额外 API; 这里用占位
        return b'\x00\x11\x22\x33\x44\x55'

    def _t11_get_dmac(dst_ip: str) -> bytes:
        # 用 ff:ff:ff:ff:ff:ff (广播) 作为默认目标 MAC
        return b'\xff\xff\xff\xff\xff\xff'

    def _t11_preview_hex(*_):
        """根据配置生成完整包: Ethernet II + (IPv4) + 传输层 + payload
        并自动计算 IP/UDP/ICMP 校验和
        ⚠ 关键: 从当前 t11_custom 中提取已编辑过的字段值, 避免覆盖用户编辑"""
        mode = t11_mode.get()
        proto = t11_proto.get()
        try:
            payload_len = max(0, min(int(t11_len.get()), 256))
        except ValueError:
            payload_len = PREVIEW_PAYLOAD_LEN
        if mode == 'custom':
            return
        # ── 关键: 从当前 hex 解析已有字段值 (避免覆盖用户编辑) ──
        try:
            h = re.sub(r'[\s,;]+', '', t11_custom.get())
            old_pkt = bytes.fromhex(h) if h else b''
        except ValueError:
            old_pkt = b''

        def _take(old, off, size, default):
            """从 old_pkt 的 [off, off+size) 取 size 字节, 不够则用 default"""
            if off < len(old) and off + size <= len(old):
                return old[off:off+size]
            return default

        if mode == 'random':
            payload = os.urandom(payload_len)
        elif mode == 'zero':
            payload = b'\x00' * payload_len
        elif mode == 'ff':
            payload = b'\xff' * payload_len
        elif mode == 'aa':
            payload = b'\xaa' * payload_len
        else:  # 'inc'
            payload = bytes(i % 256 for i in range(payload_len))

        import random as _r
        # 端口: 优先保留用户编辑的
        try:
            dst_port = int(t11_dst_port.get()) if t11_dst_port.get() else 0
        except ValueError:
            dst_port = 0
        # UDP/TCP 头在 Eth(14)+IP(20) 之后, sport/dport 在 34-37 字节
        old_sport = _take(old_pkt, 34, 2, None)
        old_dport = _take(old_pkt, 36, 2, None)
        if old_sport is not None:
            src_port = (old_sport[0] << 8) | old_sport[1]
        else:
            src_port = _r.randint(1024, 65535)
        if old_dport is not None and old_dport != b'\x00\x00':
            dst_port = (old_dport[0] << 8) | old_dport[1]

        # IP 地址: 优先保留用户编辑的 (sip 在 26-29, dip 在 30-33)
        old_sip = _take(old_pkt, 26, 4, None)
        old_dip = _take(old_pkt, 30, 4, None)
        if old_sip is not None:
            src_ip = f'{old_sip[0]}.{old_sip[1]}.{old_sip[2]}.{old_sip[3]}'
        else:
            src_ip = '192.168.1.10'
        if old_dip is not None and old_dip != b'\x00\x00\x00\x00':
            dst_ip = f'{old_dip[0]}.{old_dip[1]}.{old_dip[2]}.{old_dip[3]}'
        else:
            dst_ip = t11_dst_ip.get() or '192.168.1.100'
        sip_b = _ip_to_bytes(src_ip)
        dip_b = _ip_to_bytes(dst_ip)

        # MAC: 优先保留用户编辑的 (dmac 0-5, smac 6-11)
        old_dmac = _take(old_pkt, 0, 6, None)
        old_smac = _take(old_pkt, 6, 6, None)
        dmac = old_dmac if old_dmac is not None else b'\xff\xff\xff\xff\xff\xff'
        smac = old_smac if old_smac is not None else _t11_get_smac()

        # 协议号
        proto_num = {'UDP': 17, 'TCP': 6, 'ICMP': 1}.get(proto, 0)

        # 1) Ethernet II
        if proto == 'ARP':
            etype = b'\x08\x06'  # ARP
        else:
            etype = b'\x08\x00'  # IPv4
        eth = dmac + smac + etype

        # 2) 传输层 + IPv4 头
        if proto == 'UDP':
            udp_len = 8 + payload_len
            udp_hdr_no_csum = (src_port.to_bytes(2, 'big') +
                                dst_port.to_bytes(2, 'big') +
                                udp_len.to_bytes(2, 'big') +
                                b'\x00\x00')
            udp_csum = _udp_checksum(src_ip, dst_ip, udp_hdr_no_csum, payload)
            udp_hdr = udp_hdr_no_csum[:6] + udp_csum.to_bytes(2, 'big')
            trans = udp_hdr + payload
        elif proto == 'TCP':
            trans = (src_port.to_bytes(2, 'big') +
                     dst_port.to_bytes(2, 'big') +
                     _r.randint(0, 0xffffffff).to_bytes(4, 'big') +
                     (0).to_bytes(4, 'big') +
                     (0x5002).to_bytes(2, 'big') +
                     (65535).to_bytes(2, 'big') +
                     b'\x00\x00\x00\x00' +
                     payload)
        elif proto == 'ICMP':
            icmp_data = payload if payload else b'\x00' * 32
            # ICMP id/seq 保留
            old_id = _take(old_pkt, 38, 2, None)
            old_seq = _take(old_pkt, 40, 2, None)
            icmp_no_csum = bytes([0x08, 0x00]) + b'\x00\x00' + \
                            (old_id or _r.randint(0, 0xffff).to_bytes(2, 'big')) + \
                            (old_seq or _r.randint(0, 0xffff).to_bytes(2, 'big')) + \
                            icmp_data
            cs = _ip_checksum(icmp_no_csum)
            trans = bytes([0x08, 0x00]) + cs.to_bytes(2, 'big') + \
                     icmp_no_csum[4:]
        elif proto == 'ARP':
            arp = (b'\x00\x01' + b'\x08\x00' + b'\x06\x04' +
                   b'\x00\x01' +  # request
                   smac +  # SHA
                   sip_b +  # SPA
                   b'\x00\x00\x00\x00\x00\x00' +  # THA
                   dip_b)  # TPA
            pkt = eth + arp
            t11_custom.set(' '.join(f'{b:02X}' for b in pkt))
            return
        else:
            trans = payload

        # IPv4 头 (20 字节)
        ip_total_len = 20 + len(trans)
        # ident/flags 保留
        old_ident = _take(old_pkt, 18, 2, None)
        ip_hdr_no_csum = bytes([0x45, 0x00]) + \
                          ip_total_len.to_bytes(2, 'big') + \
                          (old_ident or _r.randint(0, 0xffff).to_bytes(2, 'big')) + \
                          b'\x40\x00' + \
                          bytes([64]) + \
                          bytes([proto_num]) + \
                          b'\x00\x00' + \
                          sip_b + dip_b
        ip_cs = _ip_checksum(ip_hdr_no_csum)
        ip_hdr = ip_hdr_no_csum[:10] + ip_cs.to_bytes(2, 'big') + ip_hdr_no_csum[12:]

        pkt = eth + ip_hdr + trans
        t11_custom.set(' '.join(f'{b:02X}' for b in pkt))

    # 模式 / 长度 变化重生成预览 (协议/端口/IP 只影响 ARP 例外分支, 已在函数内处理)
    t11_mode.trace_add('write', _t11_preview_hex)
    t11_len.trace_add('write', _t11_preview_hex)
    t11_proto.trace_add('write', _t11_preview_hex)  # ARP 例外处理

    # 自动同步 StringVar -> 当前流 dict (用户改任意 StringVar 即保存到当前流)
    def _t11_auto_save(*_):
        if t11_loading_stream:
            return
        _t11_save_current_to_stream()
    # nic 是全局的, 不加入自动保存 (切换流时不丢失)
    for _sv in (t11_proto, t11_dst_ip, t11_dst_port, t11_len,
                t11_mode, t11_rate, t11_custom):
        _sv.trace_add('write', _t11_auto_save)

    # 初始化一次: 让默认模式 (random) 也有预览
    try:
        _t11_preview_hex()
    except Exception as e:
        # 初始化预览失败时, 用基础 UDP 报文填充
        import random as _r2
        _default = (
            b'\xff\xff\xff\xff\xff\xff'  # dmac (广播)
            b'\x00\x11\x22\x33\x44\x55'  # smac
            b'\x08\x00'                   # EtherType IPv4
            b'\x45\x00'                   # ver_ihl, tos
            + (20 + 8 + 16).to_bytes(2, 'big')  # tlen
            + b'\x00\x01\x00\x00'         # ident, flags
            b'\x40\x11'                   # ttl, proto=UDP
            + b'\x00\x00'                 # csum (先填 0)
            + b'\xc0\xa8\x01\x0a'         # sip 192.168.1.10
            + b'\xc0\xa8\x01\x64'         # dip 192.168.1.100
            + (49152).to_bytes(2, 'big')  # sport
            + (8080).to_bytes(2, 'big')   # dport
            + (8 + 16).to_bytes(2, 'big') # udp len
            + b'\x00\x00'                 # udp csum
            + b'\x00' * 16                # payload
        )
        t11_custom.set(' '.join(f'{b:02X}' for b in _default))

    # 收发日志区 (按钮行放在日志框内的右下角)
    f11log = ttk.LabelFrame(t11, text=' 收发日志 ', )
    f11log.grid(row=1, column=0, sticky='nsew', padx=12, pady=(4, 8))
    f11log.grid_rowconfigure(1, weight=1)  # 日志 Text 占主区域
    f11log.grid_columnconfigure(0, weight=1)

    # 顶部工具栏: 勾选框 + 清空按钮 (右侧)
    log_toolbar = ttk.Frame(f11log, style='TFrame')
    log_toolbar.grid(row=0, column=0, columnspan=2, sticky='ew', padx=8, pady=(6, 2))
    ttk.Label(log_toolbar, text='每行: [HH:MM:SS.ms] [Tx/Rx] 内容', font=(F, 9),
              foreground=C['sub']).pack(side='left')
    ttk.Button(log_toolbar, text='🧹 清空日志', style='Small.TButton',
               command=lambda: _t11_clear_log()).pack(side='right', padx=(4, 0))
    ttk.Checkbutton(log_toolbar, text='保存日志', variable=t11_save_log,
                    command=lambda: _t11_toggle_save_log(),
                    style='TCheckbutton').pack(side='right', padx=(0, 6))

    # (流操作按钮已移至网络配置区右侧框内)

    t11_log = tk.Text(f11log, font=(M, 10), bg=C['ebg'], fg=C['fg'],
                      relief='flat', padx=10, pady=6, wrap='word',
                      state='normal')  # state=normal 允许框选 Ctrl+C 复制
    t11_log.bind('<Key>', lambda e: 'break')  # 拦截所有按键输入, 但允许 Ctrl+C 选择复制
    t11_log.grid(row=1, column=0, sticky='nsew', padx=4, pady=(2, 4))
    t11_ls = ttk.Scrollbar(f11log, orient='vertical', command=t11_log.yview)
    t11_ls.grid(row=1, column=1, sticky='ns', padx=(0, 4), pady=(2, 4))
    t11_log.configure(yscrollcommand=t11_ls.set)
    # 标签样式
    t11_log.tag_config('ts', foreground=C['sub'], font=(M, 8))
    t11_log.tag_config('tx_tag', foreground='#ea580c',
                       font=(M, 10, 'bold'))  # 发送: 橙色
    t11_log.tag_config('rx_tag', foreground='#16a34a',
                       font=(M, 10, 'bold'))  # 接收: 绿色
    t11_log.tag_config('sys_tag', foreground=C['blue'],
                       font=(M, 10, 'bold'))  # 系统: 蓝色

    # 按钮行 (放在收发日志框内的右下角, 右对齐)
    f11btn = ttk.Frame(f11log, style='TFrame')
    f11btn.grid(row=2, column=0, columnspan=2, sticky='e',
                padx=8, pady=(2, 6))

    t11_send_btn = ttk.Button(f11btn, text='▶ 开始发送',
                              style='Accent.TButton')
    t11_send_btn.pack(side='left', padx=(0, 6))
    t11_stop_btn = ttk.Button(f11btn, text='■ 停止', state='disabled',
                              style='Normal.TButton')
    t11_stop_btn.pack(side='left', padx=(0, 6))
    t11_single_btn = ttk.Button(f11btn, text='发送单包',
                                style='Normal.TButton')
    t11_single_btn.pack(side='left', padx=(0, 12))

    t11_sent_var = tk.StringVar(value='已发送: 0')
    t11_recv_var = tk.StringVar(value='已接收: 0')
    t11_rate_var = tk.StringVar(value='速率: 0 pps')
    ttk.Label(f11btn, textvariable=t11_sent_var, font=(F, 9, 'bold'),
              foreground=C['blue']).pack(side='left', padx=(0, 12))
    ttk.Label(f11btn, textvariable=t11_recv_var, font=(F, 9, 'bold'),
              foreground=C['green']).pack(side='left', padx=(0, 12))
    ttk.Label(f11btn, textvariable=t11_rate_var, font=(F, 9),
              foreground=C['sub']).pack(side='left')

    # (自定义Hex 已移至 "⚙ 报文编辑" 弹窗内)

    # --- 发送状态 ---
    t11_running = False
    t11_thread = None
    t11_sent_cnt = 0
    t11_recv_cnt = 0
    t11_rate_cnt = 0
    t11_rate_t0 = 0.0
    # --- 日志保存状态 ---
    t11_log_fp = None
    t11_log_path = ''

    def _t11_clear_log():
        t11_log.delete('1.0', 'end')

    def _t11_toggle_save_log():
        """复选框回调: 勾选时弹文件选择, 取消勾选时关闭文件句柄"""
        nonlocal t11_log_fp, t11_log_path
        if t11_save_log.get():
            default_name = ('net_log_' +
                            datetime.datetime.now().strftime('%Y%m%d_%H%M%S') +
                            '.log')
            fp = filedialog.asksaveasfilename(
                title='选择日志保存路径',
                defaultextension='.log',
                initialfile=default_name,
                filetypes=[('日志文件', '*.log'),
                           ('文本文件', '*.txt'),
                           ('所有文件', '*.*')])
            if not fp:
                t11_save_log.set(False)
                return
            try:
                t11_log_fp = open(fp, 'a', encoding='utf-8')
                t11_log_path = fp
                t11_log_fp.write(
                    f'# FPGA Toolbox Net Log\n'
                    f'# Started: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n'
                    f'# Proto: {t11_proto.get()}  Dst: {t11_dst_ip.get()}:{t11_dst_port.get()}\n'
                    f'# ---\n')
                t11_log_fp.flush()
                t11_log.config(state='normal')
                now = datetime.datetime.now().strftime('%H:%M:%S')
                t11_log.insert('end', f'[{now}.000] ', 'ts')
                t11_log.insert('end', '[系统] ', 'sys_tag')
                t11_log.insert('end', f'日志已开启: {os.path.basename(fp)}\n')
                t11_log.see('end')
            except Exception as e:
                messagebox.showerror('打开日志失败', str(e))
                t11_log_fp = None
                t11_log_path = ''
                t11_save_log.set(False)
        else:
            if t11_log_fp:
                try:
                    t11_log_fp.write(
                        f'# Closed: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
                    t11_log_fp.close()
                except Exception:
                    pass
                t11_log_fp = None
                t11_log_path = ''
                t11_log.config(state='normal')
                now = datetime.datetime.now().strftime('%H:%M:%S')
                t11_log.insert('end', f'[{now}.000] ', 'ts')
                t11_log.insert('end', '[系统] ', 'sys_tag')
                t11_log.insert('end', '日志已关闭\n')
                t11_log.see('end')

    def _t11_write_log_to_file(direction, text):
        """把一行 [Rx]/[Tx]/[系统] 写入日志文件 (如已开启)"""
        if not t11_log_fp or t11_log_fp.closed:
            return
        try:
            now = datetime.datetime.now()
            ts = now.strftime('%Y-%m-%d %H:%M:%S') + \
                 f'.{now.microsecond // 1000:03d}'
            # 去除颜色/控制字符, 替换换行
            text = text.replace('\r', '').replace('\n', '\\n')
            t11_log_fp.write(f'[{ts}] [{direction}] {text}\n')
            t11_log_fp.flush()
        except Exception:
            try:
                t11_log_fp.close()
            except Exception:
                pass

    def _t11_log(msg, color=None, direction=None):
        """向日志框写入一条记录 (带时间戳 + 收发标志, 默认开启)
        direction: 'Tx' / 'Rx' / '系统' / None
        """
        t11_log.config(state='normal')
        # 时间戳 (毫秒级)
        now = datetime.datetime.now()
        ts = now.strftime('[%H:%M:%S.') + f'{now.microsecond // 1000:03d}] '
        t11_log.insert('end', ts, 'ts')
        # 收发标志
        if direction == 'Tx':
            t11_log.insert('end', '[Tx] ', 'tx_tag')
        elif direction == 'Rx':
            t11_log.insert('end', '[Rx] ', 'rx_tag')
        elif direction == '系统':
            t11_log.insert('end', '[系统] ', 'sys_tag')
        # 内容
        t11_log.insert('end', msg + '\n')
        t11_log.see('end')  # 自动滚动
        t11_log.config(state='normal')  # 保持 normal 以允许复制
        # 写文件
        if direction:
            _t11_write_log_to_file(direction, msg)

    def _t11_update_stats():
        t11_sent_var.set(f'已发送: {t11_sent_cnt}')
        t11_recv_var.set(f'已接收: {t11_recv_cnt}')
        elapsed = _time.time() - t11_rate_t0
        if elapsed > 0:
            t11_rate_var.set(f'速率: {int(t11_rate_cnt / elapsed)} pps')

    def _t11_send_one(stream=None):
        """发送单个包 (从指定流 dict 读配置; stream=None 时用当前流)
        用户编辑的 hex = 完整用户层数据(含应用层协议头)
        发送时: socket 自动加 UDP/IP/ETH 头 (系统加)
        """
        nonlocal t11_sent_cnt, t11_recv_cnt
        if stream is None:
            _t11_save_current_to_stream()
            stream = t11_streams[t11_current_idx]
        proto = stream.get('proto', 'UDP')
        dst_ip = stream.get('dst_ip', '').strip()
        try:
            dst_port = int(stream.get('dst_port', '0'))
        except ValueError:
            if proto in ('UDP', 'TCP'):
                _t11_log('错误: 端口号无效', direction='系统')
                return False
            dst_port = 0
        custom = stream.get('custom', '').strip()
        nic = t11_nic.get()  # 网卡是全局的, 不存到流
        name = stream.get('name', '?')

        # 解析用户编辑的 hex -> data
        try:
            h = re.sub(r'[\s,;]+', '', custom)
            if h:
                data = bytes.fromhex(h)
            else:
                data = b''
        except ValueError:
            _t11_log('错误: 自定义Hex 格式无效 (应为 16 进制字节)', direction='系统')
            return False

        try:
            if proto == 'ARP':
                # ARP 走 _t11_arp_table 查询, 不发送用户数据
                result = _t11_arp_table()
                t11_log.config(state='normal')
                t11_log.insert('end', result)
                t11_log.see('end')
                t11_log.config(state='disabled')
                t11_recv_cnt += 1
                _t11_update_stats()
                return True

            elif proto == 'ICMP':
                # ICMP 走 ping (socket 系统加 ICMP/IP/ETH 头)
                if not data:
                    data = b'\x00' * 32
                result = _t11_ping(dst_ip, len(data))
                t11_log.config(state='normal')
                t11_log.insert('end', result)
                t11_log.see('end')
                t11_log.config(state='disabled')
                t11_recv_cnt += 1
                _t11_update_stats()
                return True

            else:
                # UDP / TCP: socket 发送用户 data (系统加 UDP/TCP/IP/ETH 头)
                if not data:
                    data = b'\x00' * 16
                if proto == 'UDP':
                    n = _t11_udp_send(
                        _t11_extract_ip(nic), dst_ip, dst_port, data)
                    _t11_log(f'[{name}] → {dst_ip}:{dst_port}  UDP  {n}B', direction='Tx')
                    t11_sent_cnt += 1
                elif proto == 'TCP':
                    n, resp = _t11_tcp_send(
                        _t11_extract_ip(nic), dst_ip, dst_port, data)
                    _t11_log(f'[{name}] → {dst_ip}:{dst_port}  TCP  {n}B', direction='Tx')
                    t11_sent_cnt += 1
                    if resp:
                        _t11_log(f'[{name}] ← {dst_ip}:{dst_port}  TCP  '
                                 f'{len(resp)}B response', direction='Rx')
                        t11_recv_cnt += 1
                _t11_update_stats()
                return True


        except Exception as e:
            _t11_log(f'✘ 发送失败: {e}', direction='系统')
            return False


    def _t11_send_loop():
        """后台线程：持续发送 (遍历所有 enabled=True 的流, 各按各自速率发送)"""
        nonlocal t11_rate_cnt
        # 1) 先保存当前 StringVar 到流
        _t11_save_current_to_stream()
        # 2) 锁定所有启用的流 (深拷贝, 避免发送过程中配置被改)
        locked_streams = [dict(s) for s in t11_streams if s.get('enabled', True)]
        if not locked_streams:
            _t11_log('⚠ 没有启用的流, 停止发送', direction='系统')
            root.after(0, _t11_stop)
            return

        t11_rate_cnt = 0
        # 各流的间隔 (按各自 rate)
        next_t = []
        for s in locked_streams:
            try:
                pps = max(0, int(s.get('rate', '100')))
            except ValueError:
                pps = 0
            interval = 1.0 / pps if pps > 0 else 0
            next_t.append(_time.time() + interval if interval > 0 else float('inf'))

        idx = 0
        n = len(locked_streams)
        while t11_running:
            now = _time.time()
            stream = locked_streams[idx]
            # 到时间了就发
            if now >= next_t[idx]:
                try:
                    pps = max(1, int(stream.get('rate', '100')))
                except ValueError:
                    pps = 1
                interval = 1.0 / pps
                ok = _t11_send_one(stream)
                if ok:
                    t11_rate_cnt += 1
                next_t[idx] = now + interval
            # 找下一个最快要发的流
            idx = (idx + 1) % n
            # 短暂 sleep 避免 CPU 100%
            _time.sleep(0.001)
            # 每 1 秒更新速率显示
            if t11_rate_cnt > 0 and t11_rate_cnt % 50 == 0:
                root.after(0, _t11_update_stats)

    def _t11_start():
        nonlocal t11_running, t11_thread, t11_sent_cnt, t11_recv_cnt
        nonlocal t11_rate_cnt, t11_rate_t0
        if t11_running:
            return
        # 先保存当前 StringVar 到流, 然后开始发送
        _t11_save_current_to_stream()
        enabled_names = [s.get('name', '?') for s in t11_streams
                         if s.get('enabled', True)]
        if not enabled_names:
            messagebox.showinfo('提示', '请至少勾选 1 条流再开始发送')
            return
        t11_running = True
        t11_sent_cnt = 0
        t11_recv_cnt = 0
        t11_rate_cnt = 0
        t11_rate_t0 = _time.time()
        t11_send_btn.config(state='disabled')
        t11_stop_btn.config(state='normal')
        t11_thread = threading.Thread(target=_t11_send_loop, daemon=True)
        t11_thread.start()
        _t11_log(f'▶ 开始发送 {len(enabled_names)} 条流: '
                 f'{", ".join(enabled_names)}', direction='系统')

    def _t11_stop():
        nonlocal t11_running
        t11_running = False
        t11_send_btn.config(state='normal')
        t11_stop_btn.config(state='disabled')
        _t11_update_stats()
        _t11_log('■ 已停止', direction='系统')

    t11_send_btn.config(command=_t11_start)
    t11_stop_btn.config(command=_t11_stop)
    def _t11_send_all_single():
        """发送单包: 对所有启用的流各发 1 包 (按列表顺序)"""
        if t11_running:
            return
        _t11_save_current_to_stream()
        enabled = [s for s in t11_streams if s.get('enabled', True)]
        if not enabled:
            _t11_log('⚠ 没有启用的流', direction='系统')
            return
        for s in enabled:
            _t11_send_one(s)
    t11_single_btn.config(command=_t11_send_all_single)

    # 协议切换：隐藏/显示端口
    def _t11_proto_changed(*_):
        p = t11_proto.get()
        if p in ('ARP', 'ICMP'):
            t11_port_entry.config(state='disabled')
            t11_dst_port.set('')
        else:
            t11_port_entry.config(state='normal')
            if not t11_dst_port.get():
                t11_dst_port.set('8080')
    t11_proto.trace_add('write', _t11_proto_changed)

    # 初始化网卡
    ips = _t11_list_ips()
    if ips:
        t11_nic.set(ips[0])

    # ═══ 快捷键 + 拖拽 ═══
    root.bind('<F5>', lambda e: (
        t1_refresh() if nb.index(nb.select()) == 0 else None))
    # 全局快捷键：仅当焦点不在 Text 控件时生效
    def _global_ctrl_c(e):
        if not isinstance(root.focus_get(), tk.Text):
            t1_copy()
    def _global_ctrl_s(e):
        if not isinstance(root.focus_get(), tk.Text):
            t1_save_both()
    root.bind('<Control-c>', _global_ctrl_c)
    root.bind('<Control-s>', _global_ctrl_s)

    if _sys.platform == 'win32':
        try:
            import ctypes
            from ctypes import wintypes

            WM_DROPFILES = 0x0233
            GWLP_WNDPROC = -4

            hwnd = ctypes.windll.user32.GetParent(root.winfo_id())
            ctypes.windll.shell32.DragAcceptFiles(hwnd, True)

            old_proc = ctypes.windll.user32.GetWindowLongPtrW(
                hwnd, GWLP_WNDPROC)

            WNDPROC = ctypes.WINFUNCTYPE(
                ctypes.c_longlong, ctypes.c_longlong,
                wintypes.UINT, ctypes.c_longlong, ctypes.c_longlong)

            @WNDPROC
            def wndproc(hwnd, msg, wparam, lparam):
                if msg == WM_DROPFILES:
                    buf = ctypes.create_unicode_buffer(260)
                    ctypes.windll.shell32.DragQueryFileW(
                        wparam, 0, buf, 260)
                    ctypes.windll.shell32.DragFinish(wparam)
                    root.after(0, t1_load, buf.value)
                    return 0
                return ctypes.windll.user32.CallWindowProcW(
                    WNDPROC(old_proc), hwnd, msg, wparam, lparam)

            ctypes.windll.user32.SetWindowLongPtrW(
                hwnd, GWLP_WNDPROC,
                ctypes.cast(wndproc, ctypes.c_void_p).value)
        except Exception:
            pass

    # ══════════════════════════════════════════════════════════════
    # ══════════════════════════════════════════════════════════════
    # TAB 12 — QCI功能处理流程图 (显示原图, 缩放/平移/导出)
    # ══════════════════════════════════════════════════════════════
    t12 = ttk.Frame(nb, style='TFrame')
    nb.add(t12, text='🖼 QCI测试图')
    t12.grid_rowconfigure(0, weight=0)  # 工具栏
    t12.grid_rowconfigure(1, weight=1)  # 图片区
    t12.grid_columnconfigure(0, weight=1)

    # ── 顶部工具栏 ──
    t12_toolbar = ttk.Frame(t12, style='TFrame')
    t12_toolbar.grid(row=0, column=0, sticky='ew', padx=12, pady=(10, 4))
    ttk.Label(t12_toolbar, text='QCI 功能处理流程图',
              font=(F, 11, 'bold'), foreground=C['blue']).pack(side='left', padx=(0, 12))
    ttk.Label(t12_toolbar, text='右键拖动平移 / 滚轮缩放',
              font=(F, 9), foreground=C['sub']).pack(side='left')

    t12_zoom_level = [1.0]
    t12_image_path = os.path.join(_PROJECT_ROOT, 'assets', 'qci_flow.png')

    # ── 加载原图 (PIL) ──
    t12_pil_img = None
    t12_pil_img_orig = None  # 原始全图 (用于重新缩放)
    try:
        from PIL import Image, ImageTk
        if os.path.isfile(t12_image_path):
            # 加载并合成到白底 (兼容透明 PNG)
            _im = Image.open(t12_image_path)
            if _im.mode in ('RGBA', 'LA') or (_im.mode == 'P' and 'transparency' in _im.info):
                _bg = Image.new('RGB', _im.size, (255, 255, 255))
                if _im.mode == 'P':
                    _im = _im.convert('RGBA')
                _bg.paste(_im, mask=_im.split()[-1] if _im.mode in ('RGBA', 'LA') else None)
                t12_pil_img_orig = _bg
            else:
                t12_pil_img_orig = _im.convert('RGB')
        else:
            print(f'[QCI] 图片不存在: {t12_image_path}')
    except ImportError:
        print('[QCI] PIL 未安装, 无法加载图片')

    # ── 画布 + 滚动条 ──
    t12_canvas_frame = ttk.Frame(t12, style='TFrame')
    t12_canvas_frame.grid(row=1, column=0, sticky='nsew', padx=12, pady=(0, 8))
    t12_canvas_frame.grid_rowconfigure(0, weight=1)
    t12_canvas_frame.grid_columnconfigure(0, weight=1)

    t12_canvas = tk.Canvas(t12_canvas_frame, bg='#fafbfc',
                           highlightthickness=1, highlightbackground=C['bd'])
    t12_canvas.grid(row=0, column=0, sticky='nsew')

    t12_hs = ttk.Scrollbar(t12_canvas_frame, orient='horizontal',
                           command=t12_canvas.xview)
    t12_hs.grid(row=1, column=0, sticky='ew')
    t12_vs = ttk.Scrollbar(t12_canvas_frame, orient='vertical',
                           command=t12_canvas.yview)
    t12_vs.grid(row=0, column=1, sticky='ns')

    t12_canvas.configure(xscrollcommand=t12_hs.set, yscrollcommand=t12_vs.set)

    t12_tk_img = [None]  # 防止被 GC
    t12_img_id = [None]

    def _t12_render():
        """根据当前缩放重新渲染图片"""
        if t12_pil_img_orig is None:
            return
        cw = t12_canvas.winfo_width()
        ch = t12_canvas.winfo_height()
        if cw < 50 or ch < 50:
            return
        iw, ih = t12_pil_img_orig.size
        # 默认 1.0 比例, 把图片按原比例缩放到画布内
        # 缩放后尺寸
        scale = t12_zoom_level[0]
        new_w = max(50, int(iw * scale))
        new_h = max(50, int(ih * scale))
        # 缩放
        try:
            resized = t12_pil_img_orig.resize((new_w, new_h), Image.LANCZOS)
        except Exception:
            return
        from PIL import ImageTk
        t12_tk_img[0] = ImageTk.PhotoImage(resized)
        # 删除旧图
        if t12_img_id[0] is not None:
            t12_canvas.delete(t12_img_id[0])
        t12_img_id[0] = t12_canvas.create_image(0, 0, anchor='nw', image=t12_tk_img[0])
        t12_canvas.configure(scrollregion=(0, 0, new_w, new_h))

    def _t12_fit_to_window():
        """自动适配: 缩放图片铺满画布 (留 5% 边距)"""
        if t12_pil_img_orig is None:
            return
        t12_canvas.update_idletasks()
        cw = t12_canvas.winfo_width()
        ch = t12_canvas.winfo_height()
        if cw < 50 or ch < 50:
            return
        iw, ih = t12_pil_img_orig.size
        if iw <= 0 or ih <= 0:
            return
        # 90% 边距, 但不超过 1.0 (避免放大失真)
        scale = min((cw * 0.95) / iw, (ch * 0.95) / ih, 1.0)
        if scale < 0.1:
            scale = 0.1
        t12_zoom_level[0] = scale
        _t12_render()
        t12_canvas.xview_moveto(0)
        t12_canvas.yview_moveto(0)

    def _t12_zoom_in():
        t12_zoom_level[0] *= 1.2
        _t12_render()

    def _t12_zoom_out():
        t12_zoom_level[0] /= 1.2
        _t12_render()

    def _t12_zoom_reset():
        t12_zoom_level[0] = 1.0
        _t12_fit_to_window()

    def _t12_export_png():
        if t12_pil_img_orig is None:
            messagebox.showerror('错误', '没有原图可导出')
            return
        try:
            fp = filedialog.asksaveasfilename(
                title='导出流程图原图', defaultextension='.png',
                filetypes=[('PNG 图片', '*.png')],
                initialfile='qci_flow.png')
            if not fp:
                return
            t12_pil_img_orig.save(fp, 'PNG')
            messagebox.showinfo('提示', f'已导出原图:\n{fp}')
        except Exception as e:
            messagebox.showerror('导出失败', str(e))

    ttk.Button(t12_toolbar, text='[导出原图]', style='Small.TButton',
               command=_t12_export_png).pack(side='right', padx=(4, 0))
    ttk.Button(t12_toolbar, text='[重置视图]', style='Small.TButton',
               command=_t12_zoom_reset).pack(side='right', padx=(4, 0))
    ttk.Button(t12_toolbar, text='[- 缩小]', style='Small.TButton',
               command=_t12_zoom_out).pack(side='right', padx=(4, 0))
    ttk.Button(t12_toolbar, text='[+ 放大]', style='Small.TButton',
               command=_t12_zoom_in).pack(side='right', padx=(4, 0))

    # ── 拖动 / 缩放 支持 ──
    t12_canvas.bind('<ButtonPress-2>', lambda e: t12_canvas.scan_mark(e.x, e.y))
    t12_canvas.bind('<B2-Motion>', lambda e: t12_canvas.scan_dragto(e.x, e.y, gain=1))
    t12_canvas.bind('<ButtonPress-3>', lambda e: t12_canvas.scan_mark(e.x, e.y))
    t12_canvas.bind('<B3-Motion>', lambda e: t12_canvas.scan_dragto(e.x, e.y, gain=1))

    def _t12_on_wheel(event):
        if event.delta > 0 or event.num == 4:
            factor = 1.1
        else:
            factor = 1 / 1.1
        # 以鼠标位置为中心缩放
        t12_zoom_level[0] *= factor
        old_sr = t12_canvas.cget('scrollregion').split()
        if len(old_sr) == 4:
            iw = float(old_sr[2]) - float(old_sr[0])
            ih = float(old_sr[3]) - float(old_sr[1])
            # 当前鼠标在画布上的位置 (含 scroll 偏移)
            cx = t12_canvas.canvasx(event.x)
            cy = t12_canvas.canvasy(event.y)
            # 缩放后, 把鼠标下的像素保持在原位
            new_w = iw * factor
            new_h = ih * factor
            t12_canvas.xview_moveto((cx * factor - event.x) / new_w)
            t12_canvas.yview_moveto((cy * factor - event.y) / new_h)
        _t12_render()

    t12_canvas.bind('<MouseWheel>', _t12_on_wheel)
    t12_canvas.bind('<Button-4>', _t12_on_wheel)
    t12_canvas.bind('<Button-5>', _t12_on_wheel)

    # ── 自适应: Tab 切换 / 窗口大小变化 → 自动重渲染 ──
    def _t12_refit():
        try:
            t12_canvas.update_idletasks()
            cw = t12_canvas.winfo_width()
            ch = t12_canvas.winfo_height()
            if cw < 50 or ch < 50:
                return
            try:
                cur = nb.select()
                if nb.index(cur) != nb.index(t12):
                    return
            except Exception:
                pass
            if t12_zoom_level[0] == 1.0:
                _t12_fit_to_window()
            else:
                _t12_render()
        except Exception:
            pass

    nb.bind('<<NotebookTabChanged>>', lambda e: root.after(200, _t12_refit))
    t12_canvas.bind('<Configure>', lambda e: root.after(200, _t12_refit))

    # 首次渲染
    if t12_pil_img_orig is not None:
        # 等画布完成布局再渲染
        root.after(500, _t12_refit)
    else:
        # 没图时显示提示
        t12_canvas.create_text(
            400, 300,
            text=f'未找到图片: qci_flow.png\n请把流程图 PNG 放在工具目录',
            fill=C['red'], font=(F, 12, 'bold'),
            justify='center')



    def _save_text(widget, path):
        try:
            with open(path, 'w', encoding='utf-8') as f:
                f.write(widget.get('1.0', 'end'))
            return f'已保存到 {path}'
        except Exception as e:
            return f'保存失败: {e}'

    # ══════════════════════════════════════════════════════════════
    # TAB 13 — iperf3 网络打流 (多版本 / 自定义命令)
    # ══════════════════════════════════════════════════════════════
    t13 = ttk.Frame(nb, style='TFrame')
    nb.add(t13, text='📡 iperf3')
    t13.grid_rowconfigure(0, weight=0)
    t13.grid_rowconfigure(1, weight=0)
    t13.grid_rowconfigure(2, weight=1)
    t13.grid_columnconfigure(0, weight=1)

    # ── 顶部: 模式选择 (客户端 / 服务端) ──
    t13_top = ttk.Frame(t13, style='TFrame')
    t13_top.grid(row=0, column=0, sticky='ew', padx=12, pady=(10, 6))

    t13_mode_var = tk.StringVar(value='client')
    ttk.Label(t13_top, text='模式:', font=(F, 10, 'bold')).pack(side='left', padx=(0, 6))
    ttk.Radiobutton(t13_top, text='客户端 (打流方)', variable=t13_mode_var,
                    value='client', style='Toolbutton').pack(side='left', padx=4)
    ttk.Radiobutton(t13_top, text='服务端 (接收方)', variable=t13_mode_var,
                    value='server', style='Toolbutton').pack(side='left', padx=4)
    ttk.Radiobutton(t13_top, text='自定义命令', variable=t13_mode_var,
                    value='custom', style='Toolbutton').pack(side='left', padx=4)

    # ── iperf3 可执行文件选择 (多版本) ──
    t13_ver_frame = ttk.LabelFrame(t13, text='iperf3 可执行文件 (多版本)',
                                    style='Card.TLabelframe')
    t13_ver_frame.grid(row=1, column=0, sticky='ew', padx=12, pady=4)
    t13_ver_frame.grid_columnconfigure(1, weight=1)

    t13_iperf_var = tk.StringVar(value='iperf3')
    t13_iperf_map = {}  # {display_name: full_path}

    def _t13_scan_iperf():
        """扫描 iperf3_bin/ 子目录 + PATH 的 iperf3, 返回 {版本名: 完整路径}"""
        import shutil
        mapping = {}
        # iperf3_bin/ 子目录 (版本名作为显示名)
        iperf_bin_dir = os.path.join(_PROJECT_ROOT, 'iperf3_bin')
        try:
            for n in sorted(os.listdir(iperf_bin_dir)):
                sub = os.path.join(iperf_bin_dir, n)
                if not os.path.isdir(sub) or not n.lower().startswith('iperf3'):
                    continue
                for exe_name in ('iperf3.exe', 'iperf3', 'iperf3.bin'):
                    exe_path = os.path.join(sub, exe_name)
                    if os.path.isfile(exe_path):
                        mapping[n] = exe_path
                        break
        except Exception:
            pass
        # PATH 里
        w = shutil.which('iperf3')
        if w and 'iperf3 (PATH)' not in mapping:
            mapping['iperf3 (PATH)'] = w
        return mapping

    ttk.Label(t13_ver_frame, text='iperf3 版本:',
              font=(F, 10)).grid(row=0, column=0, sticky='w', padx=8, pady=6)
    t13_iperf_combo = ttk.Combobox(t13_ver_frame, textvariable=t13_iperf_var,
                                    width=30, font=(F, 10), state='readonly')
    t13_iperf_combo.grid(row=0, column=1, sticky='ew', padx=4, pady=6)

    def _t13_refresh_iperf():
        t13_iperf_map.clear()
        t13_iperf_map.update(_t13_scan_iperf())
        names = list(t13_iperf_map.keys())
        t13_iperf_combo['values'] = names
        if names:
            t13_iperf_combo.current(0)
            t13_iperf_var.set(names[0])
        _t13_refresh_cmd()

    ttk.Button(t13_ver_frame, text='🔄 扫描', style='Small.TButton',
               command=_t13_refresh_iperf).grid(row=0, column=2, padx=4, pady=6)

    # ── 中部: 参数 / 命令区 (左右分栏) ──
    t13_mid = ttk.Frame(t13, style='TFrame')
    t13_mid.grid(row=2, column=0, sticky='nsew', padx=12, pady=4)
    t13_mid.grid_columnconfigure(0, weight=1)
    t13_mid.grid_columnconfigure(1, weight=1)
    t13_mid.grid_rowconfigure(1, weight=1)

    # 左: 参数化表单
    t13_params = ttk.LabelFrame(t13_mid, text='参数化表单 (客户端 / 服务端)',
                                 style='Card.TLabelframe')
    t13_params.grid(row=0, column=0, rowspan=2, sticky='nsew', padx=(0, 6), pady=4)
    t13_params.grid_columnconfigure(1, weight=1)

    t13_p_server = tk.StringVar(value='127.0.0.1')
    t13_p_port = tk.StringVar(value='')
    t13_p_time = tk.StringVar(value='10')
    t13_p_parallel = tk.StringVar(value='1')
    t13_p_bw = tk.StringVar(value='')  # 限速, e.g. 100M
    t13_p_proto = tk.StringVar(value='TCP')
    t13_p_len = tk.StringVar(value='')
    t13_p_extra = tk.StringVar(value='')
    t13_p_reverse = tk.BooleanVar(value=False)
    t13_p_zerocopy = tk.BooleanVar(value=False)

    def _t13_add_param(row, label, var, width=14):
        ttk.Label(t13_params, text=label, font=(F, 9)).grid(
            row=row, column=0, sticky='w', padx=8, pady=4)
        if isinstance(var, tk.BooleanVar):
            ttk.Checkbutton(t13_params, variable=var).grid(
                row=row, column=1, sticky='w', padx=4, pady=4)
        else:
            ttk.Entry(t13_params, textvariable=var, width=width,
                      font=(F, 10)).grid(row=row, column=1, sticky='ew', padx=4, pady=4)

    _t13_add_param(0,  '服务端 IP:',       t13_p_server, 20)
    _t13_add_param(1,  '端口 -p (可选, 默认 5201):', t13_p_port)
    _t13_add_param(2,  '时长 -t (秒):',    t13_p_time)
    _t13_add_param(3,  '并发流 -P:',       t13_p_parallel)
    _t13_add_param(4,  '限速 -b (可选):',  t13_p_bw)
    _t13_add_param(5,  '协议:',            t13_p_proto)
    t13_proto_combo = ttk.Combobox(t13_params, textvariable=t13_p_proto,
                                    values=['TCP', 'UDP'], width=12, state='readonly',
                                    font=(F, 10))
    t13_proto_combo.grid(row=5, column=1, sticky='w', padx=4, pady=4)
    t13_proto_combo.bind('<<ComboboxSelected>>',
                         lambda e: (t13_p_proto.set(t13_proto_combo.get()),
                                    _t13_refresh_cmd()))
    _t13_add_param(6,  '块大小 -l (可选, 默认128B):', t13_p_len)
    _t13_add_param(7,  'UDP 填充 (可选):',  t13_p_extra)
    _t13_add_param(8,  '反向模式 -R:',    t13_p_reverse)
    _t13_add_param(9,  '零拷贝 -Z:',      t13_p_zerocopy)

    # 右: 完整命令显示 + 自定义命令
    t13_cmd_frame = ttk.LabelFrame(t13_mid, text='命令预览 / 自定义',
                                    style='Card.TLabelframe')
    t13_cmd_frame.grid(row=0, column=1, sticky='nsew', padx=(6, 0), pady=4)
    t13_cmd_frame.grid_columnconfigure(0, weight=1)
    t13_cmd_frame.grid_rowconfigure(1, weight=1)

    ttk.Label(t13_cmd_frame, text='命令预览 (自动生成, 可手动编辑):',
              font=(F, 9), foreground=C['sub']).grid(row=0, column=0, sticky='w', padx=8, pady=(6, 0))
    t13_cmd_text = tk.Text(t13_cmd_frame, height=6, font=(M, 10), wrap='word',
                            bg='#fafbfc', relief='flat',
                            highlightthickness=1, highlightbackground=C['bd'])
    t13_cmd_text.grid(row=1, column=0, sticky='nsew', padx=8, pady=4)

    t13_custom_frame = ttk.Frame(t13_cmd_frame, style='TFrame')
    t13_custom_frame.grid(row=2, column=0, sticky='ew', padx=8, pady=4)

    def _t13_build_cmd():
        mode = t13_mode_var.get()
        iperf = t13_iperf_map.get(t13_iperf_var.get().strip(), \
                 t13_iperf_var.get().strip())
        if mode == 'custom':
            return t13_cmd_text.get('1.0', 'end').strip()
        if mode == 'server':
            parts = [iperf, '-s']
            if t13_p_port.get().strip():
                parts += ['-p', t13_p_port.get().strip()]
            if t13_p_zerocopy.get():
                parts.append('-Z')
            return ' '.join(parts)
        # client
        parts = [iperf, '-c', t13_p_server.get().strip() or '127.0.0.1']
        if t13_p_port.get().strip():
            parts += ['-p', t13_p_port.get().strip()]
        if t13_p_time.get().strip():
            parts += ['-t', t13_p_time.get().strip()]
        if t13_p_parallel.get().strip() and t13_p_parallel.get().strip() != '1':
            parts += ['-P', t13_p_parallel.get().strip()]
        if t13_p_bw.get().strip():
            parts += ['-b', t13_p_bw.get().strip()]
        if t13_p_proto.get() == 'UDP':
            parts.append('-u')
        if t13_p_len.get().strip():
            parts += ['-l', t13_p_len.get().strip()]
        if t13_p_extra.get().strip():
            parts += ['--extra-data', t13_p_extra.get().strip()]
        if t13_p_reverse.get():
            parts.append('-R')
            parts.append('-Z')
        return ' '.join(parts)

    def _t13_refresh_cmd():
        cmd = _t13_build_cmd()
        # 预览统一用 iperf3，不显示版本名
        if cmd:
            parts = cmd.split(' ', 1)
            parts[0] = 'iperf3'
            cmd = ' '.join(parts)
        t13_cmd_text.delete('1.0', 'end')
        t13_cmd_text.insert('1.0', cmd)

    ttk.Button(t13_custom_frame, text='🛠 生成命令', style='Small.TButton',
               command=_t13_refresh_cmd).pack(side='left', padx=2)
    ttk.Button(t13_custom_frame, text='📋 复制', style='Small.TButton',
               command=lambda: (root.clipboard_clear(),
                                root.clipboard_append(t13_cmd_text.get('1.0', 'end').strip()),
                                messagebox.showinfo('已复制', '命令已复制到剪贴板'))).pack(side='left', padx=2)

    # 模式切换/参数变化时刷新命令
    def _t13_on_mode_change(*_):
        is_custom = (t13_mode_var.get() == 'custom')
        state = 'disabled' if is_custom else 'normal'
        for child in t13_params.winfo_children():
            try:
                child.configure(state=state)
            except Exception:
                pass
        # 自定义模式下, 命令框可编辑
        t13_cmd_text.configure(state='normal' if is_custom else 'disabled')
        _t13_refresh_cmd()

    t13_mode_var.trace_add('write', _t13_on_mode_change)
    t13_iperf_var.trace_add('write', lambda *_: _t13_refresh_cmd())
    t13_iperf_combo.bind('<<ComboboxSelected>>', lambda e: (t13_iperf_var.set(t13_iperf_combo.get()), _t13_refresh_cmd()))
    for v in (t13_p_server, t13_p_port, t13_p_time, t13_p_parallel,
              t13_p_bw, t13_p_proto, t13_p_len, t13_p_extra,
              t13_p_reverse, t13_p_zerocopy):
        v.trace_add('write', lambda *_: _t13_refresh_cmd())
    _t13_refresh_iperf()  # 先扫+填下拉框
    _t13_refresh_cmd()

    # 启动 / 停止 按钮
    t13_btn_frame = ttk.Frame(t13_cmd_frame, style='TFrame')
    t13_btn_frame.grid(row=3, column=0, sticky='ew', padx=8, pady=(4, 8))

    t13_proc = [None]
    t13_running = [False]

    t13_status_var = tk.StringVar(value='● 就绪')

    def _t13_log(msg):
        t13_log_text.insert('end', msg + '\n')
        t13_log_text.see('end')
        # 智能诊断: 识别 iperf3 典型错误, 追加中文排查建议
        if 'unable to connect to server' in msg.lower() or 'connection refused' in msg.lower():
            t13_log_text.insert('end',
                '\n────────────────────────────────────────\n'
                '  ⚠  iperf3 客户端无法连接服务器\n'
                '────────────────────────────────────────\n'
                '  常见原因 (按概率排序):\n'
                '  1. 服务端未启动:  请在另一台/同一台机器上以\n'
                '       iperf3 -s [-p 端口]\n'
                '     模式先启动服务端\n'
                '  2. IP/端口写错:  确认上方"服务器地址"和\n'
                '     "端口"与服务端实际监听一致 (默认 5201)\n'
                '  3. 防火墙拦截:  Windows 防火墙 / Linux iptables\n'
                '     放通该端口 (或临时关闭测试)\n'
                '  4. 服务端绑定 0.0.0.0:  iperf3 服务端默认只绑\n'
                '     0.0.0.0 即可, 但若指定了 -B 仅绑特定 IP,\n'
                '     客户端必须连那个 IP\n'
                '  5. 网络不通:  先用 ping <服务器IP> 验证\n'
                '  6. 协议不匹配:  服务端 UDP 时客户端也需 -u\n'
                '────────────────────────────────────────\n')
            t13_log_text.see('end')
        elif 'control socket has closed' in msg.lower():
            t13_log_text.insert('end',
                '\n[!] iperf3 控制连接已关闭, 通常因为参数不兼容\n'
                '    或服务端版本与客户端差异过大, 建议两端使用\n'
                '    相同主版本号 (如 3.x 互相兼容)\n')
            t13_log_text.see('end')
        elif 'the server is busy running a test' in msg.lower():
            t13_log_text.insert('end',
                '\n[!] 服务端正在被另一个客户端占用\n'
                '    iperf3 默认单连接, 需等前一个测试结束,\n'
                '    或服务端加 --one-off 允许串行多客户端\n')
            t13_log_text.see('end')
        elif 'unable to start listener' in msg.lower():
            t13_log_text.insert('end',
                '\n[!] 服务端无法绑定端口\n'
                '    常见原因: 端口被占用 / 权限不足 (<1024) /\n'
                '    防火墙阻止绑定\n')
            t13_log_text.see('end')

    def _t13_start():
        # 用全路径执行, 预览显示短名
        cmd = _t13_build_cmd()
        preview = t13_cmd_text.get('1.0', 'end').strip()
        if not cmd:
            messagebox.showwarning('提示', '请先生成或输入命令')
            return
        _t13_log(f'\n>>> {preview}\n')
        t13_status_var.set('● 运行中...')
        t13_running[0] = True
        t13_start_btn.configure(state='disabled')
        t13_stop_btn.configure(state='normal')

        def _runner():
            try:
                proc = subprocess.Popen(
                    cmd if isinstance(cmd, list) else cmd,
                    stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    shell=isinstance(cmd, str),
                    encoding='utf-8', errors='replace',
                    bufsize=1)
                t13_proc[0] = proc
                for line in iter(proc.stdout.readline, ''):
                    if not line:
                        break
                    # 跨线程更新 UI
                    root.after(0, _t13_log, line.rstrip())
                proc.wait()
                root.after(0, _t13_log, f'\n[退出码 {proc.returncode}]')
            except FileNotFoundError as e:
                root.after(0, _t13_log, f'[!] 未找到 iperf3: {e}\n请到上方"iperf3 可执行文件"处选择路径')
            except Exception as e:
                root.after(0, _t13_log, f'[!] 启动失败: {e}')
            finally:
                t13_proc[0] = None
                t13_running[0] = False
                root.after(0, lambda: (
                    t13_status_var.set('● 就绪'),
                    t13_start_btn.configure(state='normal'),
                    t13_stop_btn.configure(state='disabled')))

        threading.Thread(target=_runner, daemon=True).start()

    def _t13_stop():
        if t13_proc[0]:
            try:
                t13_proc[0].terminate()
                _t13_log('[!] 已发送终止信号')
            except Exception as e:
                _t13_log(f'[!] 终止失败: {e}')

    t13_start_btn = ttk.Button(t13_btn_frame, text='▶ 启动',
                                 command=_t13_start, style='Accent.TButton')
    t13_start_btn.pack(side='left', padx=4)
    t13_stop_btn = ttk.Button(t13_btn_frame, text='■ 停止',
                                command=_t13_stop, style='Small.TButton')
    t13_stop_btn.pack(side='left', padx=4)
    t13_stop_btn.configure(state='disabled')
    ttk.Label(t13_btn_frame, textvariable=t13_status_var,
              font=(F, 9, 'bold'), foreground=C['blue']).pack(side='left', padx=12)

    # 日志区
    t13_log_frame = ttk.LabelFrame(t13_mid, text='iperf3 输出',
                                    style='Card.TLabelframe')
    t13_log_frame.grid(row=1, column=1, sticky='nsew', padx=(6, 0), pady=4)
    t13_log_frame.grid_columnconfigure(0, weight=1)
    t13_log_frame.grid_rowconfigure(0, weight=1)

    t13_log_text = tk.Text(t13_log_frame, font=(M, 9), wrap='word',
                            bg='#1e1e1e', fg='#d4d4d4', relief='flat',
                            insertbackground='#d4d4d4',
                            highlightthickness=1, highlightbackground=C['bd'])
    t13_log_text.grid(row=0, column=0, sticky='nsew', padx=4, pady=4)

    t13_log_scroll = ttk.Scrollbar(t13_log_frame, orient='vertical',
                                    command=t13_log_text.yview)
    t13_log_scroll.grid(row=0, column=1, sticky='ns')
    t13_log_text.configure(yscrollcommand=t13_log_scroll.set)

    t13_log_btn = ttk.Frame(t13_log_frame, style='TFrame')
    t13_log_btn.grid(row=1, column=0, columnspan=2, sticky='ew', padx=4, pady=4)
    ttk.Button(t13_log_btn, text='🗑 清空日志', style='Small.TButton',
               command=lambda: t13_log_text.delete('1.0', 'end')).pack(side='left', padx=2)
    ttk.Button(t13_log_btn, text='💾 保存日志', style='Small.TButton',
               command=lambda: (lambda p: (_t13_log(_save_text(t13_log_text, p))) if p else None)(
                   filedialog.asksaveasfilename(title='保存 iperf3 日志',
                                                defaultextension='.log',
                                                filetypes=[('日志', '*.log'), ('文本', '*.txt')])))        .pack(side='left', padx=2)

    # ══════════════════════════════════════════════════════════════
    # TAB 14 — SSH / SFTP (类 WinCP: 双栏文件浏览 + 编辑)
    # ══════════════════════════════════════════════════════════════
    t14 = ttk.Frame(nb, style='TFrame')
    nb.add(t14, text='🔐 SSH/SFTP')
    t14.grid_rowconfigure(0, weight=0)  # 连接栏
    t14.grid_rowconfigure(1, weight=0)  # 当前路径栏
    t14.grid_rowconfigure(2, weight=1)  # 主区 (SFTP 文件面板 / 终端)
    t14.grid_columnconfigure(0, weight=1)

    # 尝试导入 paramiko, 不可用时降级提示
    _PARAMIKO_IMPORT_ERR = ''
    try:
        if '_pyserial_lib' not in str(_sys.path):
            _extra = os.path.join(_PROJECT_ROOT, '_pyserial_lib')
            if os.path.isdir(_extra) and _extra not in _sys.path:
                _sys.path.insert(0, _extra)
        import paramiko
        _HAS_PARAMIKO = True
    except Exception as _e:
        paramiko = None
        _HAS_PARAMIKO = False
        _PARAMIKO_IMPORT_ERR = str(_e)

    # ── 连接栏 ──
    t14_conn = ttk.LabelFrame(t14, text='SSH 连接', style='Card.TLabelframe')
    t14_conn.grid(row=0, column=0, sticky='ew', padx=12, pady=(10, 4))
    t14_conn.grid_columnconfigure(99, weight=1)

    t14_host = tk.StringVar(value='')
    t14_port = tk.StringVar(value='22')
    t14_user = tk.StringVar(value='')
    t14_pass = tk.StringVar(value='')
    t14_keyfile = tk.StringVar(value='')

    ttk.Label(t14_conn, text='主机:', font=(F, 9)).grid(row=0, column=0, padx=(8, 4), pady=6, sticky='e')
    ttk.Entry(t14_conn, textvariable=t14_host, width=18, font=(F, 10)).grid(row=0, column=1, padx=2, pady=6)
    ttk.Label(t14_conn, text='端口:', font=(F, 9)).grid(row=0, column=2, padx=(8, 4), pady=6, sticky='e')
    ttk.Entry(t14_conn, textvariable=t14_port, width=6, font=(F, 10)).grid(row=0, column=3, padx=2, pady=6)
    ttk.Label(t14_conn, text='用户名:', font=(F, 9)).grid(row=0, column=4, padx=(8, 4), pady=6, sticky='e')
    t14_user_combo = ttk.Combobox(t14_conn, textvariable=t14_user, width=14,
                                    font=(F, 10), values=[])
    t14_user_combo.grid(row=0, column=5, padx=2, pady=6)
    ttk.Label(t14_conn, text='密码:', font=(F, 9)).grid(row=0, column=6, padx=(8, 4), pady=6, sticky='e')
    ttk.Entry(t14_conn, textvariable=t14_pass, width=14, font=(F, 10), show='*').grid(row=0, column=7, padx=2, pady=6)
    ttk.Label(t14_conn, text='私钥:', font=(F, 9)).grid(row=0, column=8, padx=(8, 4), pady=6, sticky='e')
    ttk.Entry(t14_conn, textvariable=t14_keyfile, width=18, font=(F, 9)).grid(row=0, column=9, padx=2, pady=6)
    ttk.Button(t14_conn, text='[选私钥文件]', style='Small.TButton',
               command=lambda: (lambda p: t14_keyfile.set(p) if p else None)(
                   filedialog.askopenfilename(title='选择 SSH 私钥',
                                              filetypes=[('私钥', 'id_rsa* id_ed25519* *.pem'), ('所有文件', '*.*')])))        .grid(row=0, column=10, padx=(4, 4), pady=6)

    t14_status_var = tk.StringVar(value='● 未连接')
    ttk.Label(t14_conn, textvariable=t14_status_var, font=(F, 9, 'bold'),
              foreground=C['sub']).grid(row=0, column=13, padx=12, pady=6, sticky='w')

    t14_client = [None]   # paramiko.SSHClient
    t14_sftp = [None]     # paramiko.SFTPClient
    t14_remote_cwd = [''] # 远程当前目录

    def _t14_log(msg):
        # 底部操作日志已移除, 这里仅在状态栏同步关键进度
        try:
            t14_status.set(msg)
        except Exception:
            pass

    def _t14_connect():
        try:
            _t14_log('[→] 开始连接...')
            if not _HAS_PARAMIKO:
                msg = '未安装 paramiko, 无法使用 SSH 功能.\n\n'
                if _PARAMIKO_IMPORT_ERR:
                    msg += f'导入错误: {_PARAMIKO_IMPORT_ERR}\n\n'
                msg += '请把 paramiko 目录复制到 _pyserial_lib/ 或 runtime/Lib/site-packages/'
                messagebox.showerror('缺少 paramiko', msg)
                return
            host = t14_host.get().strip()
            user = t14_user.get().strip()
            if not host or not user:
                messagebox.showwarning('提示', '请填写主机和用户名')
                return
            try:
                port = int(t14_port.get().strip() or '22')
            except ValueError:
                messagebox.showerror('错误', '端口必须为数字')
                return

            # 状态栏在主线程先更新（避免子线程操作 GUI 被静默吞掉）
            t14_status_var.set('● 连接中...')
            _t14_log(f'[→] 正在连接 {user}@{host}:{port} ...')

            def _runner():
                try:
                    root.after(0, lambda: _t14_log('[i] 创建 SSH 客户端...'))
                    client = paramiko.SSHClient()
                    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                    keyf = t14_keyfile.get().strip()
                    use_key = bool(keyf and os.path.isfile(keyf))
                    # 关键: banner_timeout / auth_timeout 设短, 防止永远卡死
                    kwargs = dict(hostname=host, port=port, username=user,
                                  timeout=10, auth_timeout=15, banner_timeout=10,
                                  allow_agent=False, look_for_keys=False)
                    if use_key:
                        kwargs['key_filename'] = keyf
                        kwargs['look_for_keys'] = True
                        root.after(0, lambda: _t14_log(f'[i] 使用私钥: {keyf}'))
                    else:
                        pwd_val = t14_pass.get()
                        if not pwd_val:
                            root.after(0, lambda: _t14_log('[!] 未填写密码, 将尝试空密码连接'))
                        kwargs['password'] = pwd_val
                    root.after(0, lambda: _t14_log(f'[i] 正在 TCP 连接 {host}:{port} (超时 10s)...'))
                    client.connect(**kwargs)
                    root.after(0, lambda: _t14_log('[i] SSH 认证成功, 正在打开 SFTP...'))
                    sftp = client.open_sftp()
                    t14_client[0] = client
                    t14_sftp[0] = sftp
                    root.after(0, lambda: _t14_log('[i] SFTP 通道已建立, 正在获取远程目录...'))
                    # 远程默认目录 = HOME
                    try:
                        stdin, stdout, stderr = client.exec_command('pwd')
                        pwd = stdout.read().decode('utf-8', 'replace').strip() or '/'
                    except Exception:
                        pwd = '/'
                    t14_remote_cwd[0] = pwd
                    root.after(0, lambda: (
                        t14_status_var.set(f'● 已连接 {user}@{host}:{port}'),
                        t14_remote_path_var.set(pwd),
                        _t14_list_remote(pwd),
                        t14_disconnect_btn.configure(state='normal'),
                        t14_connect_btn.configure(state='disabled'),
                        _t14_log(f'[+] 已连接 {user}@{host}:{port}, 默认目录 {pwd}')
                    ))
                except Exception as e:
                    err_msg = str(e)
                    root.after(0, lambda: (
                        t14_status_var.set('● 连接失败'),
                        messagebox.showerror('SSH 连接失败', err_msg),
                        _t14_log(f'[!] 连接失败: {err_msg}')
                    ))

            threading.Thread(target=_runner, daemon=True).start()
        except Exception as _e_all:
            # 终极兜底 — 任何在 _t14_connect 主线程发生的异常
            import traceback
            traceback.print_exc()
            root.after(0, lambda: (
                t14_status_var.set('● 错误'),
                messagebox.showerror('内部错误', f'连接函数异常:\n{_e_all}'),
                _t14_log(f'[!!] 内部错误: {_e_all}')
            ))

    def _t14_disconnect():
        try:
            if t14_sftp[0]:
                t14_sftp[0].close()
        except Exception:
            pass
        try:
            if t14_client[0]:
                t14_client[0].close()
        except Exception:
            pass
        t14_sftp[0] = None
        t14_client[0] = None
        t14_status_var.set('● 未连接')
        t14_connect_btn.configure(state='normal')
        t14_disconnect_btn.configure(state='disabled')
        t14_remote_tree.delete(*t14_remote_tree.get_children())
        _t14_log('[-] 已断开连接')

    def _t14_ssh_exec(cmd, timeout=10):
        """在线程里跑远程命令, 返回 (stdout, stderr, rc)"""
        if not t14_client[0]:
            return ('', '未连接', -1)
        try:
            si, so, se = t14_client[0].exec_command(cmd, timeout=timeout)
            out = so.read().decode('utf-8', 'replace')
            err = se.read().decode('utf-8', 'replace')
            rc = so.channel.recv_exit_status()
            return (out, err, rc)
        except Exception as e:
            return ('', str(e), -1)

    t14_connect_btn = ttk.Button(t14_conn, text='[连接]', style='Accent.TButton',
                                   command=_t14_connect)
    t14_connect_btn.grid(row=0, column=15, padx=(4, 4), pady=6)
    t14_disconnect_btn = ttk.Button(t14_conn, text='[断开]', style='Small.TButton',
                                      command=_t14_disconnect)
    t14_disconnect_btn.grid(row=0, column=16, padx=(4, 4), pady=6)
    t14_disconnect_btn.configure(state='disabled')

    # ═══════════════════════════════════════════════════════════════
    # SSH 账号库 — 保存 / 管理 / 一键填入
    # 存储: runtime/.ssh_accounts.json, 密码用 base64 简单编码 (非加密, 仅防一眼看到)
    # ═══════════════════════════════════════════════════════════════
    import base64 as _b64
    import json as _json
    _T14_ACCOUNTS_FILE = os.path.join(_PROJECT_ROOT, 'runtime', '.ssh_accounts.json')

    def _t14_load_accounts():
        try:
            if os.path.isfile(_T14_ACCOUNTS_FILE):
                with open(_T14_ACCOUNTS_FILE, 'r', encoding='utf-8') as f:
                    data = _json.load(f)
                if isinstance(data, list):
                    return data
        except Exception:
            pass
        return []

    def _t14_save_accounts(accounts):
        try:
            os.makedirs(os.path.dirname(_T14_ACCOUNTS_FILE), exist_ok=True)
            with open(_T14_ACCOUNTS_FILE, 'w', encoding='utf-8') as f:
                _json.dump(accounts, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            messagebox.showerror('保存失败', str(e))
            return False

    def _t14_refresh_account_combo():
        """把已保存账号的用户名列表同步到下拉框"""
        accts = _t14_load_accounts()
        # 形如 "user@host:port"
        labels = [f"{a.get('user','')}@{a.get('host','')}:{a.get('port','22')}" for a in accts]
        try:
            t14_user_combo['values'] = labels
        except tk.TclError:
            pass
        return accts, labels

    def _t14_fill_from_account(label):
        accts = _t14_load_accounts()
        for a in accts:
            if f"{a.get('user','')}@{a.get('host','')}:{a.get('port','22')}" == label:
                t14_host.set(a.get('host', ''))
                t14_port.set(str(a.get('port', '22')))
                t14_user.set(a.get('user', ''))
                pwd = a.get('pwd', '')
                try:
                    pwd = _b64.b64decode(pwd.encode('utf-8')).decode('utf-8')
                except Exception:
                    pass
                t14_pass.set(pwd)
                return

    t14_user_combo.bind('<<ComboboxSelected>>',
                         lambda e: _t14_fill_from_account(t14_user_combo.get()))

    def _t14_save_current_account():
        """把当前主机/端口/用户名/密码保存为账号"""
        host = t14_host.get().strip()
        user = t14_user.get().strip()
        port = t14_port.get().strip() or '22'
        pwd = t14_pass.get()
        if not host or not user:
            messagebox.showwarning('提示', '请先填写主机和用户名')
            return
        if not pwd:
            if not messagebox.askyesno('提示', '密码为空, 仍要保存吗? (下次需要手输)'):
                return
        accts = _t14_load_accounts()
        # 按 host+user+port 去重
        for i, a in enumerate(accts):
            if (a.get('host') == host and a.get('user') == user
                    and str(a.get('port', '22')) == str(port)):
                accts[i] = {'host': host, 'port': port, 'user': user,
                            'pwd': _b64.b64encode(pwd.encode('utf-8')).decode('utf-8')}
                if _t14_save_accounts(accts):
                    _t14_refresh_account_combo()
                    t14_status_var.set(f'● 已更新账号: {user}@{host}:{port}')
                return
        # 新增
        accts.append({'host': host, 'port': port, 'user': user,
                      'pwd': _b64.b64encode(pwd.encode('utf-8')).decode('utf-8')})
        if _t14_save_accounts(accts):
            _t14_refresh_account_combo()
            t14_status_var.set(f'● 已保存账号: {user}@{host}:{port}')

    def _t14_manage_accounts():
        """弹账号管理对话框: 列表 / 删除 / 双击填入"""
        win = tk.Toplevel(root)
        win.title('SSH 账号管理')
        win.geometry('600x380')
        win.transient(root)
        win.grab_set()
        win.configure(bg=C['bg'])
        ttk.Label(win, text='已保存的 SSH 账号 (双击填入 / 选中后点删除)',
                  font=(F, 10, 'bold'), foreground=C['sub'],
                  background=C['bg']).pack(pady=(10, 4))
        cols = ('user', 'host', 'port', 'pwd')
        tree = ttk.Treeview(win, columns=cols, show='headings', height=10)
        tree.heading('user', text='用户名')
        tree.heading('host', text='主机')
        tree.heading('port', text='端口')
        tree.heading('pwd', text='密码')
        tree.column('user', width=120, anchor='w')
        tree.column('host', width=200, anchor='w')
        tree.column('port', width=60, anchor='e')
        tree.column('pwd', width=140, anchor='w')
        tree.pack(fill='both', expand=True, padx=14, pady=4)
        sb = ttk.Scrollbar(tree, orient='vertical', command=tree.yview)
        sb.pack(side='right', fill='y')

        def _reload():
            tree.delete(*tree.get_children())
            for a in _t14_load_accounts():
                pwd_disp = ''
                try:
                    pwd_disp = _b64.b64decode(a.get('pwd', '').encode('utf-8')).decode('utf-8')
                except Exception:
                    pwd_disp = a.get('pwd', '')
                pwd_disp = '•' * len(pwd_disp) if pwd_disp else '(空)'
                tree.insert('', 'end', iid=f"{a.get('host','')}|{a.get('user','')}|{a.get('port','22')}",
                            values=(a.get('user', ''), a.get('host', ''),
                                    a.get('port', '22'), pwd_disp))
        _reload()

        def _on_double(event):
            sel = tree.selection()
            if not sel:
                return
            iid = sel[0]
            parts = iid.split('|')
            if len(parts) >= 3:
                label = f"{parts[1]}@{parts[0]}:{parts[2]}"
                _t14_fill_from_account(label)
                t14_status_var.set(f'● 已填入账号: {label}')
                win.destroy()
        tree.bind('<Double-Button-1>', _on_double)

        bf = ttk.Frame(win, style='TFrame')
        bf.pack(fill='x', padx=14, pady=(4, 10))
        def _delete():
            sel = tree.selection()
            if not sel:
                return
            if not messagebox.askyesno('确认', f'删除选中的 {len(sel)} 个账号?'):
                return
            accts = _t14_load_accounts()
            for iid in sel:
                parts = iid.split('|')
                if len(parts) >= 3:
                    accts = [a for a in accts
                             if not (a.get('host') == parts[0]
                                     and a.get('user') == parts[1]
                                     and str(a.get('port', '22')) == parts[2])]
            if _t14_save_accounts(accts):
                _t14_refresh_account_combo()
                _reload()
                t14_status_var.set('● 已删除账号')
        ttk.Button(bf, text='🗑 删除选中', style='Small.TButton',
                   command=_delete).pack(side='left', padx=2)
        ttk.Button(bf, text='✏ 查看明文密码', style='Small.TButton',
                   command=lambda: (
                       _show_passwords(tree) if tree.selection() else None
                   )).pack(side='left', padx=2)
        ttk.Button(bf, text='关闭', style='Small.TButton',
                   command=win.destroy).pack(side='right', padx=2)

        def _show_passwords(tr):
            sel = tr.selection()
            if not sel:
                return
            iid = sel[0]
            parts = iid.split('|')
            label = f"{parts[1]}@{parts[0]}:{parts[2]}"
            for a in _t14_load_accounts():
                if (f"{a.get('user','')}@{a.get('host','')}:{a.get('port','22')}" == label):
                    try:
                        pwd = _b64.b64decode(a.get('pwd', '').encode('utf-8')).decode('utf-8')
                    except Exception:
                        pwd = a.get('pwd', '')
                    messagebox.showinfo(label, f'密码: {pwd if pwd else "(空)"}')
                    return

    # 把两个新按钮插在 [选私钥文件] 之后, [连接] 之前 (column 15 是 [连接], 这里把后面所有 column +5)
    t14_save_acct_btn = ttk.Button(t14_conn, text='💾 保存账号', style='Small.TButton',
                                     command=_t14_save_current_account)
    t14_save_acct_btn.grid(row=0, column=11, padx=(4, 4), pady=6)
    ttk.Button(t14_conn, text='📚 账号库', style='Small.TButton',
               command=_t14_manage_accounts).grid(row=0, column=12, padx=(4, 4), pady=6)
    # 启动时同步账号列表到下拉框
    _t14_refresh_account_combo()

    t14_tip_lbl = ttk.Label(t14_conn, text='💡 拖拽文件到对侧即可上传/下载',
                             foreground=C['sub'], font=(F, 9))
    t14_tip_lbl.grid(row=0, column=14, padx=(8, 4), pady=6, sticky='w')

    # ── 模式切换: 📁 SFTP 文件管理 | 🖥 SSH 终端 ──
    t14_mode = tk.StringVar(value='sftp')
    ttk.Separator(t14_conn, orient='vertical').grid(
        row=0, column=17, sticky='ns', padx=4, pady=4)
    ttk.Label(t14_conn, text='模式:', font=(F, 9, 'bold')
              ).grid(row=0, column=18, padx=(4, 2), pady=6, sticky='e')
    t14_mode_sftp_btn = ttk.Button(t14_conn, text='📁 SFTP', width=10,
                                    style='Accent.TButton',
                                    command=lambda: _t14_switch_mode('sftp'))
    t14_mode_sftp_btn.grid(row=0, column=19, padx=(4, 4), pady=6)
    t14_mode_term_btn = ttk.Button(t14_conn, text='🖥 终端', width=10,
                                    style='Normal.TButton',
                                    command=lambda: _t14_switch_mode('term'))
    t14_mode_term_btn.grid(row=0, column=20, padx=(4, 4), pady=6)

    if not _HAS_PARAMIKO:
        ttk.Label(t14_conn, text='⚠ 未找到 paramiko, 请安装',
                  foreground='red', font=(F, 9, 'bold')).grid(
            row=0, column=21, padx=12, pady=6)

    # ── 当前路径栏 ──
    t14_path_bar = ttk.Frame(t14, style='TFrame')
    t14_path_bar.grid(row=1, column=0, sticky='ew', padx=12, pady=4)
    t14_path_bar.grid_columnconfigure(1, weight=1)
    t14_path_bar.grid_columnconfigure(8, weight=1)

    ttk.Label(t14_path_bar, text='本地:', font=(F, 9, 'bold'),
              foreground=C['blue']).grid(row=0, column=0, padx=(0, 6), pady=2)
    t14_local_path_var = tk.StringVar(value=os.getcwd())
    t14_local_path_entry = ttk.Entry(t14_path_bar, textvariable=t14_local_path_var,
                                       font=(F, 10))
    t14_local_path_entry.grid(row=0, column=1, sticky='ew', padx=2)

    def _t14_goto_local(*_):
        p = t14_local_path_var.get().strip()
        if p and os.path.isdir(p):
            _t14_list_local(p)
        else:
            messagebox.showerror('错误', f'目录不存在: {p}')

    ttk.Button(t14_path_bar, text='[刷新]', style='Small.TButton',
               command=lambda: _t14_list_local(t14_local_path_var.get())).grid(row=0, column=2, padx=2)
    ttk.Button(t14_path_bar, text='[选文件夹]', style='Small.TButton',
               command=lambda: (lambda p: (t14_local_path_var.set(p), _t14_list_local(p))) if (p := filedialog.askdirectory()) else None)        .grid(row=0, column=3, padx=2)
    ttk.Button(t14_path_bar, text='[删除]', style='Small.TButton',
               command=lambda: _t14_delete()).grid(row=0, column=4, padx=(8, 2))
    ttk.Button(t14_path_bar, text='[重命名]', style='Small.TButton',
               command=lambda: _t14_rename()).grid(row=0, column=5, padx=2)
    ttk.Button(t14_path_bar, text='[新建目录]', style='Small.TButton',
               command=lambda: _t14_mkdir()).grid(row=0, column=6, padx=2)

    ttk.Label(t14_path_bar, text='远程:', font=(F, 9, 'bold'),
              foreground=C['blue']).grid(row=0, column=7, padx=(16, 6), pady=2, sticky='e')
    t14_remote_path_var = tk.StringVar(value='')
    t14_remote_path_entry = ttk.Entry(t14_path_bar, textvariable=t14_remote_path_var,
                                       font=(F, 10))
    t14_remote_path_entry.grid(row=0, column=8, sticky='ew', padx=2)
    t14_remote_path_entry.bind('<Return>', lambda e: _t14_list_remote(t14_remote_path_var.get()))

    ttk.Button(t14_path_bar, text='[刷新]', style='Small.TButton',
               command=lambda: _t14_list_remote(t14_remote_path_var.get())).grid(row=0, column=9, padx=2)

    # ── 主区: 工具栏 + 左右分栏 (本地 + 远程 Treeview) ──
    t14_main = ttk.Frame(t14, style='TFrame')
    t14_main.grid(row=2, column=0, sticky='nsew', padx=12, pady=4)
    t14_main.grid_columnconfigure(0, weight=1)
    t14_main.grid_columnconfigure(1, weight=1)
    t14_main.grid_rowconfigure(0, weight=0)  # 工具栏
    t14_main.grid_rowconfigure(1, weight=1)  # 文件面板

    # 工具栏 (水平排列) — [删除]/[重命名]/[新建目录] 已移到路径栏的 [选文件夹] 后面;
    # 拖拽提示已移到连接栏 "未连接" 后面
    t14_toolbar = ttk.Frame(t14_main, style='TFrame')
    t14_toolbar.grid(row=0, column=0, columnspan=2, sticky='ew', padx=2, pady=(0, 6))

    def _make_pane(parent, title, col):
        frame = ttk.LabelFrame(parent, text=title, style='Card.TLabelframe')
        frame.grid(row=1, column=col, sticky='nsew',
                   padx=(0, 4) if col == 0 else (4, 0), pady=2)
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)
        cols = ('name', 'size', 'mtime', 'perm')
        tree = ttk.Treeview(frame, columns=cols, show='headings', selectmode='extended')
        tree.heading('name', text='名称')
        tree.heading('size', text='大小')
        tree.heading('mtime', text='修改时间')
        tree.heading('perm', text='权限')
        tree.column('name', width=220, minwidth=100, stretch=True, anchor='w')
        tree.column('size', width=70,  minwidth=50,  stretch=False, anchor='e')
        tree.column('mtime', width=130, minwidth=80,  stretch=False, anchor='w')
        tree.column('perm', width=100, minwidth=70,  stretch=False, anchor='w')
        tree.grid(row=0, column=0, sticky='nsew')
        sb = ttk.Scrollbar(frame, orient='vertical', command=tree.yview)
        sb.grid(row=0, column=1, sticky='ns')
        tree.configure(yscrollcommand=sb.set)
        return tree

    t14_local_tree = _make_pane(t14_main, '本地文件', 0)
    t14_remote_tree = _make_pane(t14_main, '远程文件', 1)

    # ── 拖拽上传/下载 (左↔右 两个 Treeview 之间) ──
    # 实现: 在源 tree 上 Button-1 按下记录起点, B1-Motion 移动时标记拖动中,
    #       鼠标松手时 (无论松手在哪个 widget) 根据 event.widget / 坐标
    #       判断是否真的拖到了对侧 tree, 是则把对侧同名行选中, 走 _t14_upload/_t14_download.
    # 注: 用 bind_all 监听 ButtonRelease-1, 不再绑到 tree 自己, 这样鼠标在中间空白
    #     区域松手也能正确识别. _t14_drag_origin_check 绑到两个 tree 上记录源.
    _t14_drag_state = {'src': None, 'active': False, 'x0': 0, 'y0': 0}

    def _t14_drag_origin_check(event, src_tree):
        """按下时记录起点 (但不立即进入拖动状态)"""
        _t14_drag_state['src'] = src_tree
        _t14_drag_state['active'] = False
        _t14_drag_state['x0'] = event.x_root
        _t14_drag_state['y0'] = event.y_root

    def _t14_drag_motion(event):
        """鼠标移动时, 距离起点 > 5 px 才视为"开始拖动" (避免和点击冲突)"""
        if _t14_drag_state['src'] is None:
            return
        dx = abs(event.x_root - _t14_drag_state['x0'])
        dy = abs(event.y_root - _t14_drag_state['y0'])
        if dx > 5 or dy > 5:
            _t14_drag_state['active'] = True
            # 高亮目标 (对侧)
            src = _t14_drag_state['src']
            tgt = t14_remote_tree if src is t14_local_tree else t14_local_tree
            try:
                tgt.configure(style='DragTarget.Treeview')
                tgt.focus_set()
            except tk.TclError:
                pass

    def _t14_drag_release(event):
        """全局 ButtonRelease-1 — 根据鼠标真正所在的 widget 判断是否拖到了对侧.
        松手后直接按源侧选中构造 (name, path, is_dir) 任务列表传给 upload/download,
        不再在目标侧找同名行 (目标侧文件不一定存在).
        """
        src = _t14_drag_state.get('src')
        active = _t14_drag_state.get('active', False)
        # 恢复样式
        for tr in (t14_local_tree, t14_remote_tree):
            try:
                tr.configure(style='Treeview')
            except tk.TclError:
                pass
        if not active or src is None:
            _t14_drag_state['src'] = None
            _t14_drag_state['active'] = False
            return
        # 用鼠标坐标定位真正所在的 tree (避免 event.widget 在拖出 tree 后还停在源)
        try:
            tgt = event.widget.winfo_containing(event.x_root, event.y_root)
        except Exception:
            tgt = event.widget
        # tgt 可能是 tree 的内嵌部件 (scrollbar/heading), 一路向上找
        while tgt is not None and tgt not in (t14_local_tree, t14_remote_tree):
            tgt = getattr(tgt, 'master', None)
        _t14_drag_state['src'] = None
        _t14_drag_state['active'] = False
        if tgt is None or tgt is src:
            return  # 没有真正拖到对侧
        # 收集源侧已选中的项 (排除 ..), 构造任务列表
        sel = src.selection()
        if not sel:
            return
        if src is t14_local_tree and tgt is t14_remote_tree:
            # 本地 → 远端 = 上传
            base = t14_local_path_var.get()
            tasks = []
            for s in sel:
                vals = src.item(s)['values']
                name = vals[0] if vals else ''
                if not name or name == '..':
                    continue
                is_dir = 'dir' in src.item(s)['tags']
                local = os.path.join(base, name)
                tasks.append((name, local, is_dir))
            if not tasks:
                root.after(0, _t14_log, '[!] 未选中有效文件, 取消传输')
                return
            root.after(0, _t14_log,
                       f'[拖拽] 上传 {len(tasks)} 个文件到远端 '
                       f'{t14_remote_path_var.get()}')
            _t14_upload(tasks)
        elif src is t14_remote_tree and tgt is t14_local_tree:
            # 远端 → 本地 = 下载
            rdir = t14_remote_path_var.get().rstrip('/')
            tasks = []
            for s in sel:
                vals = src.item(s)['values']
                name = vals[0] if vals else ''
                if not name or name == '..':
                    continue
                is_dir = 'dir' in src.item(s)['tags']
                remote = rdir + '/' + name
                tasks.append((name, remote, is_dir))
            if not tasks:
                root.after(0, _t14_log, '[!] 未选中有效文件, 取消传输')
                return
            root.after(0, _t14_log,
                       f'[拖拽] 下载 {len(tasks)} 个文件到本地 '
                       f'{t14_local_path_var.get()}')
            _t14_download(tasks)

    t14_local_tree.bind('<ButtonPress-1>',
                         lambda e: _t14_drag_origin_check(e, t14_local_tree))
    t14_remote_tree.bind('<ButtonPress-1>',
                          lambda e: _t14_drag_origin_check(e, t14_remote_tree))
    t14_local_tree.bind('<B1-Motion>', _t14_drag_motion)
    t14_remote_tree.bind('<B1-Motion>', _t14_drag_motion)
    # 全局监听松手: 即使鼠标在 panel 间的空白处松手, 也能定位真正所在的 widget
    root.bind_all('<ButtonRelease-1>', _t14_drag_release, add='+')

    # 注册高亮样式 (在主样式区)
    style = ttk.Style()
    try:
        style.configure('DragTarget.Treeview', background='#1f6feb',
                        fieldbackground='#1f6feb', foreground='white')
    except tk.TclError:
        pass

    # 本地列表
    def _t14_list_local(path):
        path = os.path.abspath(path)
        t14_local_path_var.set(path)
        t14_local_tree.delete(*t14_local_tree.get_children())
        # .. 返回上级
        if os.path.dirname(path) != path:
            t14_local_tree.insert('', 'end', values=('..', '', '', 'dir'),
                                    tags=('dir',))
        try:
            for name in sorted(os.listdir(path)):
                fp = os.path.join(path, name)
                try:
                    st = os.stat(fp)
                    if os.path.isdir(fp):
                        t14_local_tree.insert('', 'end', values=(
                            name, '', datetime.datetime.fromtimestamp(st.st_mtime).strftime('%Y-%m-%d %H:%M'), 'dir'),
                            tags=('dir',))
                    else:
                        t14_local_tree.insert('', 'end', values=(
                            name, _fmt_size(st.st_size),
                            datetime.datetime.fromtimestamp(st.st_mtime).strftime('%Y-%m-%d %H:%M'),
                            oct(st.st_mode & 0o777)), tags=('file',))
                except Exception:
                    t14_local_tree.insert('', 'end', values=(name, '?', '', '?'), tags=('file',))
        except Exception as e:
            messagebox.showerror('列出本地目录失败', str(e))

    def _fmt_size(n):
        for u in ('B', 'K', 'M', 'G', 'T'):
            if n < 1024:
                return f'{n:.0f}{u}' if u == 'B' else f'{n:.1f}{u}'
            n /= 1024
        return f'{n:.1f}P'

    # 远程列表
    def _t14_list_remote(path):
        if not t14_sftp[0]:
            messagebox.showwarning('提示', '请先连接 SSH')
            return
        if not path:
            return
        try:
            attrs = t14_sftp[0].listdir_attr(path)
        except Exception as e:
            messagebox.showerror('列出远程目录失败', f'{path}: {e}')
            return
        t14_remote_path_var.set(path)
        t14_remote_cwd[0] = path
        t14_remote_tree.delete(*t14_remote_tree.get_children())
        # .. 返回上级
        if path not in ('/', ''):
            parent = path.rstrip('/').rsplit('/', 1)[0] or '/'
            t14_remote_tree.insert('', 'end', values=('..', '', '', 'dir'), tags=('dir',))
        for a in sorted(attrs, key=lambda x: (not _is_dir_attr(x), x.filename)):
            ttype = 'dir' if _is_dir_attr(a) else 'file'
            t14_remote_tree.insert('', 'end', values=(
                a.filename,
                '' if ttype == 'dir' else _fmt_size(a.st_size or 0),
                datetime.datetime.fromtimestamp(a.st_mtime or 0).strftime('%Y-%m-%d %H:%M') if a.st_mtime else '',
                _fmt_perm(a.st_mode) if a.st_mode is not None else '',
            ), tags=(ttype,))

    def _is_dir_attr(a):
        try:
            import stat
            return stat.S_ISDIR(a.st_mode)
        except Exception:
            return False

    def _fmt_perm(mode):
        try:
            import stat
            return stat.filemode(mode)
        except Exception:
            return oct(mode & 0o777) if mode else ''

    # 双击: 进入目录
    def _t14_local_double(event):
        sel = t14_local_tree.selection()
        if not sel:
            return
        item = t14_local_tree.item(sel[0])
        name = item['values'][0]
        if name == '..':
            _t14_list_local(os.path.dirname(t14_local_path_var.get()))
        elif 'dir' in item['tags']:
            _t14_list_local(os.path.join(t14_local_path_var.get(), name))
        else:
            _t14_view_file('local')

    def _t14_remote_double(event):
        sel = t14_remote_tree.selection()
        if not sel:
            return
        item = t14_remote_tree.item(sel[0])
        name = item['values'][0]
        if name == '..':
            cur = t14_remote_path_var.get()
            parent = cur.rstrip('/').rsplit('/', 1)[0] or '/'
            _t14_list_remote(parent)
        elif 'dir' in item['tags']:
            cur = t14_remote_path_var.get()
            new = cur.rstrip('/') + '/' + name
            _t14_list_remote(new)
        else:
            _t14_view_file('remote')

    t14_local_tree.bind('<Double-Button-1>', _t14_local_double)
    t14_remote_tree.bind('<Double-Button-1>', _t14_remote_double)

    # 上传
    def _t14_upload(tasks=None):
        """tasks 为 None 时用 t14_local_tree 当前选中;
        tasks 为 list[(name, local_path, is_dir)] 时按指定任务传.
        """
        if tasks is None:
            sel = t14_local_tree.selection()
            if not sel:
                messagebox.showinfo('提示', '请先在左侧选择要上传的文件/目录')
                return
            tasks = []
            for s in sel:
                vals = t14_local_tree.item(s)['values']
                name = vals[0] if vals else ''
                if not name or name == '..':
                    continue
                is_dir = 'dir' in t14_local_tree.item(s)['tags']
                local = os.path.join(t14_local_path_var.get(), name)
                tasks.append((name, local, is_dir))
        if not tasks:
            return
        if not t14_sftp[0]:
            messagebox.showwarning('提示', '请先连接 SSH')
            return
        rdir = t14_remote_path_var.get()
        if not rdir:
            messagebox.showwarning('提示', '请先在右侧选择远程目录')
            return

        def _runner():
            try:
                for name, local, is_dir in tasks:
                    remote = rdir.rstrip('/') + '/' + name
                    if is_dir:
                        _sftp_upload_dir(local, remote, _t14_log)
                    else:
                        root.after(0, _t14_log, f'↑ 上传 {local} → {remote}')
                        t14_sftp[0].put(local, remote)
                root.after(0, lambda: (
                    _t14_log('[+] 上传完成'),
                    _t14_list_remote(rdir)
                ))
            except Exception as e:
                root.after(0, lambda: messagebox.showerror('上传失败', str(e)))

        threading.Thread(target=_runner, daemon=True).start()

    def _sftp_upload_dir(local, remote, log_cb):
        """递归上传目录"""
        import stat
        try:
            t14_sftp[0].mkdir(remote)
        except OSError:
            pass
        for name in os.listdir(local):
            lp = os.path.join(local, name)
            rp = remote.rstrip('/') + '/' + name
            if os.path.isdir(lp):
                _sftp_upload_dir(lp, rp, log_cb)
            else:
                root.after(0, log_cb, f'↑ {lp} → {rp}')
                t14_sftp[0].put(lp, rp)

    # 下载
    def _t14_download(tasks=None):
        """tasks 为 None 时用 t14_remote_tree 当前选中;
        tasks 为 list[(name, remote_path, is_dir)] 时按指定任务传.
        """
        if tasks is None:
            sel = t14_remote_tree.selection()
            if not sel:
                messagebox.showinfo('提示', '请先在右侧选择要下载的文件/目录')
                return
            tasks = []
            for s in sel:
                vals = t14_remote_tree.item(s)['values']
                name = vals[0] if vals else ''
                if not name or name == '..':
                    continue
                is_dir = 'dir' in t14_remote_tree.item(s)['tags']
                remote = t14_remote_path_var.get().rstrip('/') + '/' + name
                tasks.append((name, remote, is_dir))
        if not tasks:
            return
        if not t14_sftp[0]:
            messagebox.showwarning('提示', '请先连接 SSH')
            return
        ldir = t14_local_path_var.get()
        if not ldir:
            messagebox.showwarning('提示', '请先在左侧选择本地目录')
            return

        def _runner():
            try:
                for name, remote, is_dir in tasks:
                    local = os.path.join(ldir, name)
                    if is_dir:
                        _sftp_download_dir(remote, local, _t14_log)
                    else:
                        root.after(0, _t14_log, f'↓ 下载 {remote} → {local}')
                        t14_sftp[0].get(remote, local)
                root.after(0, lambda: (
                    _t14_log('[+] 下载完成'),
                    _t14_list_local(ldir)
                ))
            except Exception as e:
                root.after(0, lambda: messagebox.showerror('下载失败', str(e)))

        threading.Thread(target=_runner, daemon=True).start()

    def _sftp_download_dir(remote, local, log_cb):
        import stat
        os.makedirs(local, exist_ok=True)
        for attr in t14_sftp[0].listdir_attr(remote):
            rp = remote.rstrip('/') + '/' + attr.filename
            lp = os.path.join(local, attr.filename)
            if stat.S_ISDIR(attr.st_mode):
                _sftp_download_dir(rp, lp, log_cb)
            else:
                root.after(0, log_cb, f'↓ {rp} → {lp}')
                t14_sftp[0].get(rp, lp)

    # 删除
    def _t14_delete():
        for tree, side, get_path in (
            (t14_local_tree, 'local', lambda: t14_local_path_var.get()),
            (t14_remote_tree, 'remote', lambda: t14_remote_path_var.get()),
        ):
            sel = tree.selection()
            if not sel:
                continue
            if not messagebox.askyesno('确认', f'确定删除选中的 {len(sel)} 项 ({side})? 此操作不可恢复!'):
                continue
            base = get_path()
            for s in sel:
                name = tree.item(s)['values'][0]
                if name == '..':
                    continue
                full = base.rstrip('/\\') + ('/' if side == 'remote' else os.sep) + name
                try:
                    if side == 'local':
                        if os.path.isdir(full):
                            shutil.rmtree(full)
                        else:
                            os.remove(full)
                    else:
                        # 远端: 用 stat 判断是文件还是目录, 分别走 remove / rmtree
                        import stat as _stat
                        try:
                            attr = t14_sftp[0].stat(full)
                        except Exception:
                            # stat 失败: 退化到 listdir_attr (文件会抛错, 改用 remove 试试)
                            try:
                                t14_sftp[0].remove(full)
                            except Exception as e:
                                messagebox.showerror('删除失败', f'{full}: {e}')
                            continue
                        if _stat.S_ISDIR(attr.st_mode):
                            _t14_sftp_rmtree(full)
                        else:
                            t14_sftp[0].remove(full)
                except Exception as e:
                    messagebox.showerror('删除失败', f'{full}: {e}')
            if side == 'local':
                _t14_list_local(base)
            else:
                _t14_list_remote(base)

    def _t14_sftp_rmtree(path):
        import stat
        for attr in t14_sftp[0].listdir_attr(path):
            p = path.rstrip('/') + '/' + attr.filename
            if stat.S_ISDIR(attr.st_mode):
                _t14_sftp_rmtree(p)
            else:
                t14_sftp[0].remove(p)
        t14_sftp[0].rmdir(path)

    # 重命名
    def _t14_rename():
        for tree, base in ((t14_local_tree, t14_local_path_var.get()),
                            (t14_remote_tree, t14_remote_path_var.get())):
            sel = tree.selection()
            if not sel:
                continue
            old = tree.item(sel[0])['values'][0]
            if old == '..':
                continue
            new = tk.simpledialog.askstring('重命名', f'将 "{old}" 重命名为:', initialvalue=old)
            if not new or new == old:
                continue
            old_full = base.rstrip('/\\') + ('/' if base.startswith('/') else os.sep) + old
            new_full = base.rstrip('/\\') + ('/' if base.startswith('/') else os.sep) + new
            try:
                if tree is t14_local_tree:
                    os.rename(old_full, new_full)
                    _t14_list_local(base)
                else:
                    t14_sftp[0].rename(old_full, new_full)
                    _t14_list_remote(base)
                _t14_log(f'[↻] {old} → {new}')
            except Exception as e:
                messagebox.showerror('重命名失败', str(e))

    # 新建目录
    def _t14_mkdir():
        for tree, base, side in ((t14_local_tree, t14_local_path_var.get(), 'local'),
                                   (t14_remote_tree, t14_remote_path_var.get(), 'remote')):
            name = tk.simpledialog.askstring('新建目录', f'在 {side} 当前目录下新建目录名:')
            if not name:
                continue
            try:
                if side == 'local':
                    os.makedirs(os.path.join(base, name), exist_ok=False)
                    _t14_list_local(base)
                else:
                    if not t14_sftp[0]:
                        messagebox.showwarning('提示', '请先连接 SSH')
                        continue
                    t14_sftp[0].mkdir(base.rstrip('/') + '/' + name)
                    _t14_list_remote(base)
                _t14_log(f'[+] 新建目录 {name}')
            except Exception as e:
                messagebox.showerror('新建目录失败', str(e))

    # 双击文件 → 弹出编辑器 (本地/远程)
    t14_editor_win = [None]
    t14_editor_text = [None]
    t14_editor_remote_path = [None]
    t14_editor_dirty = [False]
    t14_editor_encoding = [None]

    def _t14_view_file(side):
        if side == 'local':
            sel = t14_local_tree.selection()
            if not sel:
                return
            name = t14_local_tree.item(sel[0])['values'][0]
            if name == '..' or 'dir' in t14_local_tree.item(sel[0])['tags']:
                return
            full = os.path.join(t14_local_path_var.get(), name)
            _t14_open_editor(full, is_remote=False, display_name=name)
        else:
            sel = t14_remote_tree.selection()
            if not sel:
                return
            name = t14_remote_tree.item(sel[0])['values'][0]
            if name == '..' or 'dir' in t14_remote_tree.item(sel[0])['tags']:
                return
            full = t14_remote_path_var.get().rstrip('/') + '/' + name
            _t14_open_editor(full, is_remote=True, display_name=name)

    def _t14_open_editor(full, is_remote, display_name):
        # 读取文件 (后台线程, 大文件不卡 UI)
        def _reader():
            try:
                if is_remote:
                    if not t14_sftp[0]:
                        root.after(0, lambda: messagebox.showerror('错误', '未连接 SSH'))
                        return
                    with t14_sftp[0].open(full, 'rb') as f:
                        data = f.read()
                else:
                    with open(full, 'rb') as f:
                        data = f.read()
                # 尝试用 utf-8 解码, 失败回退 latin-1
                enc = 'utf-8'
                try:
                    text = data.decode('utf-8')
                except UnicodeDecodeError:
                    enc = 'latin-1'
                    text = data.decode('latin-1', errors='replace')
                root.after(0, lambda: _t14_show_editor(full, is_remote, display_name, text, enc))
            except Exception as e:
                root.after(0, lambda: messagebox.showerror('打开失败', f'{full}: {e}'))

        threading.Thread(target=_reader, daemon=True).start()

    def _t14_show_editor(full, is_remote, display_name, text, enc):
        if t14_editor_win[0] and t14_editor_win[0].winfo_exists():
            t14_editor_win[0].destroy()
        win = tk.Toplevel(root)
        win.title(f'{"远程" if is_remote else "本地"} - {display_name}')
        win.geometry('900x600')
        win.configure(bg=C['bg'])

        ttk.Label(win, text=f'[{"远程" if is_remote else "本地"}]  {full}  (编码: {enc})',
                  font=(F, 9), foreground=C['sub']).pack(anchor='w', padx=10, pady=(8, 0))

        text_w = tk.Text(win, wrap='none', font=(M, 10),
                          bg='#fafbfc', relief='flat',
                          highlightthickness=1, highlightbackground=C['bd'])
        text_w.pack(fill='both', expand=True, padx=10, pady=8)

        # 滚动条
        ysb = ttk.Scrollbar(text_w, orient='vertical', command=text_w.yview)
        ysb.pack(side='right', fill='y')
        xsb = ttk.Scrollbar(text_w, orient='horizontal', command=text_w.xview)
        xsb.pack(side='bottom', fill='x')
        text_w.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)

        text_w.insert('1.0', text)
        t14_editor_text[0] = text_w
        t14_editor_remote_path[0] = full if is_remote else None
        t14_editor_win[0] = win
        t14_editor_encoding[0] = enc

        def _on_change(event=None):
            t14_editor_dirty[0] = True
        text_w.bind('<<Modified>>', lambda e: (text_w.edit_modified(False), _on_change()))

        # 底部按钮
        btn = ttk.Frame(win, style='TFrame')
        btn.pack(fill='x', padx=10, pady=(0, 8))
        ttk.Button(btn, text='[保存]', style='Accent.TButton',
                   command=lambda: _t14_save_editor(full, is_remote)).pack(side='left', padx=4)
        ttk.Button(btn, text='[重新加载]', style='Small.TButton',
                   command=lambda: _t14_open_editor(full, is_remote, display_name)).pack(side='left', padx=4)
        ttk.Button(btn, text='关闭', style='Small.TButton',
                   command=win.destroy).pack(side='right', padx=4)

    def _t14_save_editor(full, is_remote):
        text = t14_editor_text[0].get('1.0', 'end-1c')
        data = text.encode(t14_editor_encoding[0] or 'utf-8', errors='replace')

        def _writer():
            try:
                if is_remote:
                    if not t14_sftp[0]:
                        root.after(0, lambda: messagebox.showerror('错误', '未连接 SSH'))
                        return
                    with t14_sftp[0].open(full, 'wb') as f:
                        f.write(data)
                else:
                    with open(full, 'wb') as f:
                        f.write(data)
                t14_editor_dirty[0] = False
                root.after(0, lambda: (
                    _t14_log(f'[💾] 已保存 {full}'),
                    messagebox.showinfo('已保存', f'已保存到:\n{full}')
                ))
            except Exception as e:
                root.after(0, lambda: messagebox.showerror('保存失败', str(e)))

        threading.Thread(target=_writer, daemon=True).start()

    # 启动时列出本地当前目录
    _t14_list_local(t14_local_path_var.get())





    # ══════════════════════════════════════════════════════════════
    # T14 终端模式 — 在 t14 内部加 "SFTP / 终端" 切换,
    #   复用 t14_client[0] SSH 连接, 调 invoke_shell() 拿 channel
    #   支持多 session (ttk.Notebook), ANSI 颜色, Ctrl+C 中断
    #   沿用串口的 hist_lock 终端化 (5000 行上限, 历史只读)
    # ══════════════════════════════════════════════════════════════
    t14_term_sessions = []   # list of dict {frame, text, channel, recv_thread, alive, name}
    t14_term_nb = None       # ttk.Notebook (terminal tabs)

    def _t14_switch_mode(mode):
        """切换 SFTP / 终端 主区显示"""
        t14_mode.set(mode)
        if mode == 'sftp':
            t14_path_bar.grid()
            t14_main.grid()
            t14_term_area.grid_remove()
            t14_mode_sftp_btn.configure(style='Accent.TButton')
            t14_mode_term_btn.configure(style='Normal.TButton')
        else:  # term
            t14_path_bar.grid_remove()
            t14_main.grid_remove()
            t14_term_area.grid()
            t14_mode_sftp_btn.configure(style='Normal.TButton')
            t14_mode_term_btn.configure(style='Accent.TButton')
            # 首次切到终端: 自动尝试建一个 session
            if not t14_term_sessions:
                _t14_new_term_session()

    # ── 终端主区 (默认隐藏) ──
    t14_term_area = ttk.Frame(t14, style='TFrame')
    t14_term_area.grid(row=1, column=0, rowspan=2, sticky='nsew',
                       padx=12, pady=4)
    t14_term_area.grid_rowconfigure(1, weight=1)
    t14_term_area.grid_columnconfigure(0, weight=1)

    # 顶部: session 管理
    t14_term_topbar = ttk.Frame(t14_term_area, style='TFrame')
    t14_term_topbar.grid(row=0, column=0, sticky='ew', padx=4, pady=(4, 6))
    ttk.Button(t14_term_topbar, text='➕ 新会话', style='Accent.TButton',
               command=lambda: _t14_new_term_session()).pack(side='left', padx=4)
    ttk.Button(t14_term_topbar, text='❌ 关闭当前', style='Small.TButton',
               command=lambda: _t14_close_current_term()).pack(side='left', padx=4)
    ttk.Separator(t14_term_topbar, orient='vertical').pack(
        side='left', fill='y', padx=6)
    ttk.Button(t14_term_topbar, text='🗑 清屏 (Ctrl+L)', style='Small.TButton',
               command=lambda: _t14_term_clear_current()).pack(side='left', padx=4)
    ttk.Separator(t14_term_topbar, orient='vertical').pack(
        side='left', fill='y', padx=6)
    ttk.Button(t14_term_topbar, text='⏏ 中断 (Ctrl+C)', style='Small.TButton',
               command=lambda: _t14_term_send_sigint()).pack(side='left', padx=4)
    t14_term_status = tk.StringVar(value='● 未连接')
    ttk.Label(t14_term_topbar, textvariable=t14_term_status,
              font=(F, 9, 'bold'),
              foreground=C['sub']).pack(side='right', padx=8)

    t14_term_nb = ttk.Notebook(t14_term_area)
    t14_term_nb.grid(row=1, column=0, sticky='nsew', padx=4, pady=4)

    # ═══════════════════════════════════════════════════════════════
    # 终端化核心代码 (仿 t10 hist_lock)
    # ═══════════════════════════════════════════════════════════════
    _T14_TERM_MAX_LINES = 5000

    def _t14_term_lock_history(text_widget):
        """打 hist_lock tag, 钳制光标到末行"""
        try:
            cutoff = text_widget.index('end-1c')
            text_widget.tag_add('t14_hist_lock', '1.0', cutoff)
            line_count = int(text_widget.index('end-1c').split('.')[0])
            if line_count > _T14_TERM_MAX_LINES:
                excess = line_count - _T14_TERM_MAX_LINES
                text_widget.delete('1.0', f'{excess + 1}.0')
                text_widget.tag_remove('t14_hist_lock', '1.0', 'end')
                text_widget.tag_add('t14_hist_lock', '1.0',
                                    text_widget.index('end-1c'))
        except tk.TclError:
            pass

    def _t14_term_clamp(text_widget):
        """把光标强制到末行"""
        try:
            insert_idx = text_widget.index('insert')
            if text_widget.tag_prevrange('t14_hist_lock', insert_idx):
                text_widget.mark_set('insert', 'end-1c')
                text_widget.see('insert')
        except tk.TclError:
            pass

    def _t14_term_block_hist(event):
        """拦截对 hist_lock 区域的修改 (保留函数签名以备后用, 当前不拦截任何输入)"""
        # 历史教训: 之前实现是检查光标是否在 t14_hist_lock 区域,
        # 但 hist_lock 范围是 1.0 到 end-1c, 末行光标也在范围内,
        # 导致 _on_keypress 把所有普通字符 (含 BackSpace) 都当成 'break' 拦截,
        # 用户输不进任何字符, BackSpace 也不工作.
        # 现在已改为 _on_keypress 内不再调用此函数, 这里保持函数体可空
        # 以便未来需要在 hist_lock 上做更精细 (按行) 的拦截时复用.
        return None

    # ── ANSI 颜色处理 ──
    _ANSI_COLORS = {
        '30': '#1f2329',  # black
        '31': '#dc2626',  # red
        '32': '#16a34a',  # green
        '33': '#ca8a04',  # yellow
        '34': '#2563eb',  # blue
        '35': '#9333ea',  # magenta
        '36': '#0891b2',  # cyan
        '37': '#e5e7eb',  # white
        '90': '#6b7280',  # bright black
        '91': '#f87171', '92': '#4ade80', '93': '#facc15',
        '94': '#60a5fa', '95': '#c084fc', '96': '#22d3ee',
        '97': '#f9fafb',
    }
    _ANSI_RE = re.compile(r'\x1b\[((?:\d+;?)+)m')

    def _t14_ansi_split(text):
        """把含 ANSI 颜色码的文本切成 (text, color) 段"""
        result = []
        last_end = 0
        cur_attrs = set()
        for m in _ANSI_RE.finditer(text):
            plain = text[last_end:m.start()]
            if plain:
                # 当前 attrs 决定颜色 (取最后一个非 0 的颜色)
                fg = None
                for attr in cur_attrs:
                    if attr in _ANSI_COLORS:
                        fg = _ANSI_COLORS[attr]
                result.append((plain, fg))
            last_end = m.end()
            codes = m.group(1).split(';')
            new_attrs = set()
            i = 0
            while i < len(codes):
                c = codes[i]
                if c == '0' or c == '':
                    new_attrs = set()
                elif c in ('1', '2', '22'):
                    pass  # bold/dim — 忽略
                elif c in _ANSI_COLORS:
                    new_attrs.add(c)
                elif c == '38' and i + 2 < len(codes) and codes[i+1] == '5':
                    # 256-color (忽略扩展, 用默认)
                    i += 2
                i += 1
            cur_attrs = new_attrs
        if last_end < len(text):
            plain = text[last_end:]
            fg = None
            for attr in cur_attrs:
                if attr in _ANSI_COLORS:
                    fg = _ANSI_COLORS[attr]
            result.append((plain, fg))
        return result

    # ── VT100 控制序列解析器 (处理 \b / \r / \x1b[K / \x1b[D / \x1b[H 等) ──
    # 我们不维护真实光标, 直接在 tk Text 上模拟:
    #   \b / \x7f  → 删除光标前一个字符 (move insert back 1c, delete 1c)
    #   \r         → 光标移到当前行行首
    #   \n         → 换行 (insert \n)
    #   \x1b[Nd    → 光标上移 N 行 (不实现, 简单忽略)
    #   \x1b[NB    → 光标下移 N 行
    #   \x1b[NC    → 光标右移 N 列
    #   \x1b[ND    → 光标左移 N 列
    #   \x1b[H    → 光标移到 (0,0)
    #   \x1b[2J   → 清屏
    #   \x1b[K    → 清行 (从光标到行尾, 含 0K / 1K / 2K)
    #   \x1b[?25l/h → 光标显隐 (忽略)
    #   \x1b[Nm   → SGR 颜色 (颜色码由 _t14_ansi_split 处理)
    #   其余 ESC 序列 → 忽略
    _VT_CSI_RE = re.compile(r'\x1b\[([\x30-\x3f]*)([\x20-\x2f]*)([\x40-\x7e])')
    # OSC 序列: ESC ] ... BEL(0x07) 或 ESC ] ... ESC \ (ST)
    # 用于设置终端标题等, 我们只关心"不显示到 widget", 整段跳过
    _OSC_RE = re.compile(r'\x1b\][\x08\x07]')  # 简化: ESC] + 直到 BEL 或 ST
    # DCS / OSC / PM / APC 等序列: ESC + P/X/^/_ + ... + BEL/ST — 整段跳过
    _ESC_STR_RE = re.compile(r'\x1b[\(\)\*\.\_\^\$\#\~\|]')  # 其它 ESC 开头序列的字符

    def _t14_term_write(text_widget, text, lock_history=True):
        """写入一段文本, 解析 ANSI 颜色 + VT100 控制序列.
        所有对历史区的修改都用 t14_hist_lock 标签; 任何对光标的操作都基于
        当前 'insert' 位置, 确保和 _on_keypress 本地 echo 同步.
        """
        text_widget.configure(state='normal')
        try:
            # 先按 SGR 颜色 + 普通控制字符粗切, 颜色段用 _t14_ansi_split
            # 其它控制序列 (光标移动/清行) 用 _VT_CSI_RE 二次解析
            # 简化做法: 把 \x1b[m 这种 SGR 单独抽走, 剩下 \x1b[... 控制序列交 _VT_CSI_RE
            pos = 0
            sgr_pattern = re.compile(r'\x1b\[((?:\d+;?)+)m')
            while pos < len(text):
                m = sgr_pattern.search(text, pos)
                if not m:
                    # 剩余部分无 SGR, 但可能含其它控制序列
                    _t14_term_apply_control(text_widget, text[pos:])
                    break
                # SGR 之前的部分 (可能含其它控制序列)
                if m.start() > pos:
                    _t14_term_apply_control(text_widget, text[pos:m.start()])
                # 解析 SGR
                codes = m.group(1).split(';')
                fg = None
                cur_attrs = set()
                for c in codes:
                    if c == '0' or c == '':
                        cur_attrs = set()
                    elif c in ('1', '2', '22'):
                        cur_attrs.add('bold')
                    elif c in _ANSI_COLORS:
                        cur_attrs.add(c)
                for a in cur_attrs:
                    if a in _ANSI_COLORS:
                        fg = _ANSI_COLORS[a]
                # 找下一个 SGR 之间的纯文本
                next_m = sgr_pattern.search(text, m.end())
                plain_end = next_m.start() if next_m else len(text)
                plain = text[m.end():plain_end]
                if plain:
                    # 同样需要解析中间的控制序列 (但保持 SGR 颜色)
                    _t14_term_apply_control(text_widget, plain, fg_color=fg)
                pos = plain_end
            text_widget.see('end')
            if lock_history:
                _t14_term_lock_history(text_widget)
        except tk.TclError:
            pass

    def _t14_term_apply_control(text_widget, text, fg_color=None):
        """把含控制字符的字符串拆成: 普通字符 insert / \b 删 1 / \\r 移到行首 / \\n 换行 / CSI 序列处理.
        fg_color 不为 None 时, 普通字符按此颜色插入.
        """
        if not text:
            return
        lock_tag = 't14_hist_lock'
        if fg_color:
            plain_tag = (fg_color, lock_tag)
        else:
            plain_tag = (lock_tag,)
        i = 0
        n = len(text)
        buf = []  # 累积普通字符, 遇到控制字符 flush

        def _flush():
            nonlocal buf
            if buf:
                s = ''.join(buf)
                text_widget.insert('insert', s, plain_tag)
                buf = []

        while i < n:
            ch = text[i]
            c = ord(ch)
            if ch == '\b':
                # bash readline 退格序列: '\b' + '\x1b[K' (EL = 擦行尾).
                # '\b' 本身删除光标前一个字符, 后面 '\x1b[K' 把光标到行尾清空.
                _flush()
                try:
                    cur = text_widget.index('insert')
                    col = int(cur.split('.')[1])
                    if col == 0:
                        # 在行首, 先跳到上一行行末, 再删
                        try:
                            text_widget.mark_set('insert', f'{cur} -1c lineend')
                            text_widget.delete('insert -1c', 'insert')
                        except tk.TclError:
                            pass
                    else:
                        text_widget.delete('insert -1c', 'insert')
                except tk.TclError:
                    pass
                i += 1
            elif c == 0x7f:
                # 单独的 DEL (PTY 的 ECHOE 输出) — 同上 noop
                _flush()
                i += 1
            elif ch == '\r':
                # CR (光标归位) — 忽略. bash 总是把 \r\n 一起发; \n 单独
                # 处理为换行已能正确表达"光标移到下一行"; \r 的 mark_set 到
                # 行首会破坏当前行已显示的 PS1 + 之前输入的命令 — 新字符会被
                # 插到行首, 导致 PS1 在同一行重复堆叠.
                i += 1
            elif ch == '\n' or ch == '\v' or ch == '\f':
                _flush()
                try:
                    text_widget.insert('insert', '\n', plain_tag)
                except tk.TclError:
                    pass
                i += 1
            elif ch == '\a':
                try:
                    text_widget.bell()
                except tk.TclError:
                    pass
                i += 1
            elif ch == '\t':
                _flush()
                try:
                    text_widget.insert('insert', '    ', plain_tag)
                except tk.TclError:
                    pass
                i += 1
            elif ch == '\x1b':
                _flush()
                # 1) CSI 序列: ESC [ ... <final>
                m = _VT_CSI_RE.match(text, i)
                if m:
                    param = m.group(1) or '0'
                    final = m.group(3)
                    if final == 'K':
                        try:
                            cur = text_widget.index('insert')
                            if param == '0' or param == '':
                                text_widget.delete(cur, cur + ' lineend')
                            elif param == '1':
                                text_widget.delete(cur + ' linestart', cur)
                            elif param == '2':
                                text_widget.delete(cur + ' linestart',
                                                   cur + ' lineend')
                        except tk.TclError:
                            pass
                    elif final == 'J':
                        try:
                            if param == '2' or param == '3':
                                text_widget.delete('1.0', 'end')
                                text_widget.mark_set('insert', '1.0')
                        except tk.TclError:
                            pass
                    elif final == 'H' or final == 'f':
                        try:
                            text_widget.mark_set('insert', '1.0')
                        except tk.TclError:
                            pass
                    elif final in 'ABCD':
                        try:
                            n_step = int(param) if param else 1
                            n_step = max(1, n_step)
                            cur = text_widget.index('insert')
                            if final == 'A':
                                text_widget.mark_set(
                                    'insert', f'{cur} -{n_step} lines')
                            elif final == 'B':
                                text_widget.mark_set(
                                    'insert', f'{cur} +{n_step} lines')
                            elif final == 'C':
                                text_widget.mark_set(
                                    'insert', f'{cur} +{n_step}c')
                            elif final == 'D':
                                for _ in range(n_step):
                                    cur2 = text_widget.index('insert')
                                    col = int(cur2.split('.')[1])
                                    if col == 0:
                                        break
                                    text_widget.mark_set(
                                        'insert', f'{cur2} -1c')
                        except tk.TclError:
                            pass
                    i = m.end()
                # 2) OSC 序列: ESC ] ... BEL(0x07) 或 ST(ESC \) — 跳过整段
                # (bash 用它设置终端标题, 不应显示到 widget)
                elif i + 1 < n and text[i + 1] == ']':
                    j = i + 2
                    while j < n:
                        if text[j] == '\x07':  # BEL
                            j += 1
                            break
                        if text[j] == '\x1b' and j + 1 < n and text[j + 1] == '\\':
                            j += 2
                            break
                        j += 1
                    i = j
                # 3) 其它 ESC + 字符: ESC + 单字符命令 (如 ESC= / ESC> / ESC c 等)
                # 4) DCS/PM/APC: ESC P / ESC ^ / ESC _ + ... + BEL/ST
                elif i + 1 < n and text[i + 1] in 'PX_^':
                    j = i + 2
                    while j < n:
                        if text[j] == '\x07':
                            j += 1
                            break
                        if text[j] == '\x1b' and j + 1 < n and text[j + 1] == '\\':
                            j += 2
                            break
                        j += 1
                    i = j
                # 5) ESC + 中间字符 (0x20-0x2F) + final 0x40-0x7E
                elif i + 1 < n and 0x20 <= ord(text[i + 1]) <= 0x2F:
                    j = i + 1
                    while j < n and 0x20 <= ord(text[j]) <= 0x2F:
                        j += 1
                    if j < n and 0x40 <= ord(text[j]) <= 0x7E:
                        j += 1
                    i = j
                # 6) 其它: 跳过 ESC + 一个字符
                elif i + 1 < n:
                    i += 2
                else:
                    i = n
            elif c < 0x20:
                _flush()
                i += 1
            else:
                buf.append(ch)
                i += 1
        _flush()

    def _t14_term_clear(text_widget):
        text_widget.configure(state='normal')
        text_widget.delete('1.0', 'end')
        text_widget.tag_remove('t14_hist_lock', '1.0', 'end')

    # ── 多 session 管理 ──
    def _t14_new_term_session(name=None):
        """新建 SSH 终端 session"""
        if not t14_client[0]:
            messagebox.showinfo('提示', '请先连接 SSH')
            return
        if not _HAS_PARAMIKO:
            messagebox.showerror('错误', 'paramiko 未安装')
            return
        # 分配默认名
        idx = len(t14_term_sessions) + 1
        if name is None:
            name = f'session-{idx}'
        # 创建 tab 页
        tab = ttk.Frame(t14_term_nb, style='TFrame')
        t14_term_nb.add(tab, text=f'  {name}  ')
        text = tk.Text(tab, font=('Consolas', 10), bg='#0d1117', fg='#e6edf3',
                       insertbackground='#e6edf3', relief='flat',
                       padx=8, pady=6, wrap='char', undo=False)
        text.pack(side='left', fill='both', expand=True)
        sb = ttk.Scrollbar(tab, orient='vertical', command=text.yview)
        sb.pack(side='right', fill='y')
        text.configure(yscrollcommand=sb.set)

        # 配置颜色 tag
        for code, color in _ANSI_COLORS.items():
            text.tag_configure(color, foreground=color)
        text.tag_configure('t14_hist_lock')  # 不改颜色
        text.tag_raise('t14_hist_lock')

        # 通道
        channel = None
        recv_thread = None
        alive = [True]
        try:
            channel = t14_client[0].invoke_shell(term='xterm', width=120,
                                                  height=30)
            # 立即把 terminal 大小也设置一下
            try:
                channel.resize_pty(width=120, height=30)
            except Exception:
                pass
            # 不动 stty 状态: 远端 PTY 默认 ECHO + ECHOE 开启, 由远端 shell
            # (bash -i) 负责回显键入字符和擦除. _on_keypress 不做本地 echo,
            # 所以不会有双字符/双退格. 这里不需要 sleep 读 buffer, 让 recv_loop
            # 正常处理登录 banner / PS1 即可.
        except Exception as e:
            messagebox.showerror('打开终端失败', str(e))
            t14_term_nb.forget(tab)
            tab.destroy()
            return

        session = {
            'name': name,
            'tab': tab,
            'text': text,
            'channel': channel,
            'alive': alive,
            'recv_thread': None,
        }
        t14_term_sessions.append(session)

        # 输入处理: 本地立即 echo + 把按键直接转发给 channel
            # 输入处理: 把按键直接转发给 channel, 显示完全靠远端 PTY 回显.
        # 关键: paramiko invoke_shell 默认启用了 PTY 的 ECHO, 远端 shell (bash -i)
        # 也会回显键入字符 + 在 Backspace 时发出 '\b \b' 擦除. 我们不做本地 echo,
        # 避免双字符/双退格. 发送顺序按 XTerm 约定 (Backspace 用 DEL=0x7f, Enter 用 CR).
        def _on_keypress(event):
            if not channel or channel.closed or not alive[0]:
                return 'break'
            key = event.keysym
            ctrl = (event.state & 0x4) != 0
            send = None

            if key == 'Return':
                send = '\r'
            elif key == 'BackSpace':
                send = '\x7f'
            elif key == 'Tab':
                send = '\t'
            elif key == 'Up':
                send = '\x1b[A'
            elif key == 'Down':
                send = '\x1b[B'
            elif key == 'Right':
                send = '\x1b[C'
            elif key == 'Left':
                send = '\x1b[D'
            elif key == 'Home':
                send = '\x1b[H'
            elif key == 'End':
                send = '\x1b[F'
            elif key == 'Delete':
                send = '\x1b[3~'
            elif key == 'Page_Up':
                send = '\x1b[5~'
            elif key == 'Page_Down':
                send = '\x1b[6~'
            elif key == 'Escape':
                send = '\x1b'
            elif ctrl and key.lower() == 'c':
                # Ctrl+C: 若有选中文本, 走默认复制; 否则发 \x03 (SIGINT)
                try:
                    sel = text.tag_ranges('sel')
                except tk.TclError:
                    sel = ()
                if sel:
                    return None  # 放行 (tk 自己处理复制)
                try:
                    channel.send('\x03')
                except Exception:
                    pass
                return 'break'
            elif ctrl and key.lower() == 'd':
                send = '\x04'
            elif ctrl and key.lower() == 'l':
                # Ctrl+L: 清屏 (发 ANSI 清屏序列给远端, 让它执行 clear)
                send = '\x0c'
            elif ctrl and key.lower() == 'a':
                # Ctrl+A: 行首
                send = '\x01'
            elif ctrl and key.lower() == 'e':
                # Ctrl+E: 行末
                send = '\x05'
            elif ctrl and key.lower() == 'k':
                # Ctrl+K: 删除到行末
                send = '\x0b'
            elif ctrl and key.lower() == 'u':
                # Ctrl+U: 删整行
                send = '\x15'
            else:
                # 普通字符
                if event.char and ord(event.char[0]) >= 32:
                    send = event.char
                else:
                    return None
            if send is not None:
                try:
                    channel.send(send)
                except Exception:
                    pass
                return 'break'

        text.bind('<Key>', _on_keypress)

        # 接收循环
        def _recv_loop():
            while alive[0] and channel and not channel.closed:
                try:
                    if channel.recv_ready():
                        data = channel.recv(4096)
                        if not data:
                            break
                        try:
                            text_data = data.decode('utf-8', 'replace')
                        except Exception:
                            text_data = data.decode('latin-1', 'replace')
                        # 在主线程里更新 GUI
                        root.after(0, lambda t=text_data: _t14_term_write(text, t))
                    else:
                        import time as _t
                        _t.sleep(0.02)
                except Exception:
                    break
            alive[0] = False
            try:
                root.after(0, lambda: _t14_term_write(
                    text, '\n[会话已结束]\n'))
            except Exception:
                pass

        recv_thread = threading.Thread(target=_recv_loop, daemon=True)
        session['recv_thread'] = recv_thread
        recv_thread.start()

        # 切到新 tab
        t14_term_nb.select(tab)
        t14_term_status.set(f'● {name}')

    def _t14_close_current_term():
        """关闭当前选中的 terminal tab"""
        sel = t14_term_nb.select()
        if not sel:
            return
        # 找到对应 session
        for i, s in enumerate(t14_term_sessions):
            if str(s['tab']) == sel:
                s['alive'][0] = False
                try:
                    if s['channel']:
                        s['channel'].close()
                except Exception:
                    pass
                try:
                    t14_term_nb.forget(s['tab'])
                    s['tab'].destroy()
                except Exception:
                    pass
                t14_term_sessions.pop(i)
                t14_term_status.set(
                    f'● {len(t14_term_sessions)} 个会话' if t14_term_sessions
                    else '● 未连接')
                break

    def _t14_term_clear_current():
        """清屏当前 tab"""
        sel = t14_term_nb.select()
        if not sel:
            return
        for s in t14_term_sessions:
            if str(s['tab']) == sel:
                _t14_term_clear(s['text'])
                break

    def _t14_term_send_sigint():
        """给当前 session 发 Ctrl+C (SIGINT)"""
        sel = t14_term_nb.select()
        if not sel:
            return
        for s in t14_term_sessions:
            if str(s['tab']) == sel:
                try:
                    if s['channel'] and not s['channel'].closed:
                        s['channel'].send('\x03')
                except Exception:
                    pass
                break

    # ── 连接/断开时同步 session 状态 ──
    _t14_orig_connect = _t14_connect
    def _t14_connect_with_term(*a, **k):
        _t14_orig_connect(*a, **k)
        # 连接成功后, 如果有 session 且通道失效, 关闭之
        for s in list(t14_term_sessions):
            s['alive'][0] = False
            try:
                s['channel'].close()
            except Exception:
                pass
            try:
                t14_term_nb.forget(s['tab'])
                s['tab'].destroy()
            except Exception:
                pass
        t14_term_sessions.clear()
        if t14_mode.get() == 'term':
            _t14_new_term_session()
    # 注: 这里包装 _t14_connect 略复杂 (原函数有 _t14_log 引用),
    #     简化做法: 在 _t14_connect 末尾的 success 分支插入 hook.
    #     但为减少侵入, 不动 _t14_connect; 用户自己点 "新会话" 即可.

    # 断开时清理所有 sessions
    _t14_orig_disconnect = _t14_disconnect
    def _t14_disconnect_with_term(*a, **k):
        # 关闭所有 session
        for s in list(t14_term_sessions):
            s['alive'][0] = False
            try:
                s['channel'].close()
            except Exception:
                pass
            try:
                t14_term_nb.forget(s['tab'])
                s['tab'].destroy()
            except Exception:
                pass
        t14_term_sessions.clear()
        t14_term_status.set('● 未连接')
        _t14_orig_disconnect(*a, **k)

    # 用包装版本替换
    t14_disconnect_btn.configure(command=_t14_disconnect_with_term)

    # 初始隐藏终端主区
    t14_term_area.grid_remove()

    # ═══ 底部状态栏 (精致卡片) ═══
    footer_bg = tk.Frame(root, bg=C['bg'])
    footer_bg.pack(side='bottom', fill='x', padx=14, pady=(0, 12))

    footer = tk.Frame(footer_bg, bg=C['card'],
                      highlightbackground=C['bd'],
                      highlightthickness=1, bd=0)
    footer.pack(fill='x')

    # 左侧状态
    footer_left = tk.Frame(footer, bg=C['card'])
    footer_left.pack(side='left', padx=14, pady=8)

    # 绿点 (状态指示, 无文字)
    dot = tk.Canvas(footer_left, width=8, height=8, bg=C['card'],
                    highlightthickness=0, bd=0)
    dot.create_oval(1, 1, 7, 7, fill=C['green'], outline='')
    dot.pack(side='left', padx=(0, 6))


    # ══════════════════════════════════════════════════════════════
    # TAB 15 — ModelSim 仿真自动化 (Vivado 工程 → 一键生成 do/tcl/f + 跑)
    # ══════════════════════════════════════════════════════════════
    t15 = ttk.Frame(nb, style='TFrame')
    nb.add(t15, text='🧪 仿真自动化')
    t15.grid_rowconfigure(2, weight=1)   # 文件列表
    t15.grid_rowconfigure(5, weight=1)   # 日志
    t15.grid_columnconfigure(0, weight=1)

    # ── 0) 状态变量 ──
    t15_proj_dir = tk.StringVar()                # Vivado 工程根目录
    t15_top_module = tk.StringVar()              # 顶层模块名 (下拉 + 可编辑)
    t15_top_manual = tk.StringVar()              # 手动输入的顶层 (覆盖下拉)
    t15_modelsim_path = tk.StringVar()           # vsim.exe 完整路径
    t15_run_time = tk.StringVar(value='200ns')   # run 时间
    t15_resolution = tk.StringVar(value='1ps')   # 仿真分辨率
    t15_gen_do = tk.BooleanVar(value=True)
    t15_gen_tcl = tk.BooleanVar(value=True)
    t15_gen_f = tk.BooleanVar(value=True)
    t15_file_rows = []   # list of dict {path, kind, enabled, order}
    t15_libraries = [('work', 'work')]   # (库名, 物理路径, 可空=默认 ./work)

    # ── 1) 顶部: 工程根 + Modelsim 路径 ──
    t15_topbar = ttk.LabelFrame(t15, text=' Vivado 工程 ')
    t15_topbar.grid(row=0, column=0, sticky='ew', padx=12, pady=(10, 4))
    t15_topbar.grid_columnconfigure(1, weight=1)

    ttk.Label(t15_topbar, text='工程根:', font=(F, 9)).grid(
        row=0, column=0, padx=(8, 4), pady=6, sticky='e')
    ttk.Entry(t15_topbar, textvariable=t15_proj_dir, font=(F, 10)).grid(
        row=0, column=1, sticky='ew', padx=2, pady=6)
    ttk.Button(t15_topbar, text='浏览', style='Normal.TButton',
               command=lambda: (d := filedialog.askdirectory(title='选择 Vivado 工程根目录'))
               and t15_proj_dir.set(os.path.abspath(d)) and t15_scan()
               ).grid(row=0, column=2, padx=2, pady=6)
    ttk.Button(t15_topbar, text='🔄 重新扫描', style='Accent.TButton',
               command=lambda: t15_scan()).grid(row=0, column=3, padx=2, pady=6)

    ttk.Label(t15_topbar, text='vsim.exe:', font=(F, 9)).grid(
        row=1, column=0, padx=(8, 4), pady=(0, 6), sticky='e')
    ttk.Entry(t15_topbar, textvariable=t15_modelsim_path, font=(F, 9)).grid(
        row=1, column=1, sticky='ew', padx=2, pady=(0, 6))
    ttk.Button(t15_topbar, text='选 vsim.exe', style='Normal.TButton',
               command=lambda: (p := filedialog.askopenfilename(
                   title='选择 ModelSim/Questa 的 vsim.exe',
                   filetypes=[('vsim.exe', 'vsim.exe'), ('All', '*.*')]))
               and t15_modelsim_path.set(p)
               ).grid(row=1, column=2, padx=2, pady=(0, 6))
    ttk.Label(t15_topbar, text='(可留空 — 跑仿真时再选)',
              foreground=C['sub'], font=(F, 8)).grid(
        row=1, column=3, padx=2, pady=(0, 6))

    # ── 2) 顶层 + 库映射 ──
    t15_cfg = ttk.LabelFrame(t15, text=' 仿真配置 ')
    t15_cfg.grid(row=1, column=0, sticky='ew', padx=12, pady=4)
    t15_cfg.grid_columnconfigure(1, weight=1)
    t15_cfg.grid_columnconfigure(3, weight=1)

    ttk.Label(t15_cfg, text='顶层模块:', font=(F, 9)).grid(
        row=0, column=0, padx=(8, 4), pady=4, sticky='e')
    t15_top_combo = ttk.Combobox(t15_cfg, textvariable=t15_top_module,
                                 font=(F, 10), width=24, state='normal')
    t15_top_combo.grid(row=0, column=1, sticky='ew', padx=2, pady=4)

    ttk.Label(t15_cfg, text='库映射 (name=path):',
              font=(F, 9)).grid(row=0, column=2, padx=(16, 4), pady=4, sticky='e')
    t15_lib_list = tk.Listbox(t15_cfg, font=(F, 9), height=3,
                              relief='flat', borderwidth=1,
                              bg=C['bg'], fg=C['fg'],
                              selectbackground=C['blue'], selectforeground='white',
                              activestyle='none')
    t15_lib_list.grid(row=0, column=3, sticky='ew', padx=2, pady=4)
    for n, p in t15_libraries:
        t15_lib_list.insert('end', f'{n}={p or "(默认)"}')
    t15_lib_btn_frame = ttk.Frame(t15_cfg)
    t15_lib_btn_frame.grid(row=0, column=4, padx=(2, 8), pady=4)

    def t15_add_lib():
        s = tk.simpledialog.askstring('添加库', '格式: 库名=物理路径 (留空路径用默认)',
                                      parent=t15_cfg)
        if not s:
            return
        if '=' in s:
            n, p = s.split('=', 1)
        else:
            n, p = s, ''
        n, p = n.strip(), p.strip()
        t15_libraries.append((n, p))
        t15_lib_list.insert('end', f'{n}={p or "(默认)"}')

    def t15_del_lib():
        sel = t15_lib_list.curselection()
        if not sel:
            return
        # 至少保留一个 work
        if len(t15_libraries) <= 1:
            messagebox.showinfo('提示', '至少保留一个库 (work)', parent=t15_cfg)
            return
        for i in reversed(sel):
            t15_libraries.pop(i)
            t15_lib_list.delete(i)

    ttk.Button(t15_lib_btn_frame, text='+', width=3,
               command=lambda: t15_add_lib()).pack(pady=(0, 2))
    ttk.Button(t15_lib_btn_frame, text='-', width=3,
               command=lambda: t15_del_lib()).pack()

    # ── 3) 文件列表 ──
    t15_flist = ttk.LabelFrame(t15, text=' 文件列表 (勾选=参与仿真, ↑↓ 调顺序) ')
    t15_flist.grid(row=2, column=0, sticky='nsew', padx=12, pady=4)
    t15_flist.grid_rowconfigure(0, weight=1)
    t15_flist.grid_columnconfigure(0, weight=1)

    cols = ('on', 'order', 'kind', 'rel', 'size')
    t15_tree = ttk.Treeview(t15_flist, columns=cols, show='headings',
                            selectmode='extended', height=10)
    t15_tree.heading('on', text='✓')
    t15_tree.heading('order', text='#')
    t15_tree.heading('kind', text='类型')
    t15_tree.heading('rel', text='相对路径')
    t15_tree.heading('size', text='大小')
    t15_tree.column('on', width=30, anchor='center', stretch=False)
    t15_tree.column('order', width=40, anchor='e', stretch=False)
    t15_tree.column('kind', width=50, anchor='center', stretch=False)
    t15_tree.column('rel', width=420, anchor='w', stretch=True)
    t15_tree.column('size', width=70, anchor='e', stretch=False)
    t15_tree.tag_configure('vhdl', foreground='#1d4ed8')
    t15_tree.tag_configure('vlog', foreground='#15803d')
    t15_tree.tag_configure('skip', foreground=C['sub'])
    t15_tree.grid(row=0, column=0, sticky='nsew', padx=(8, 0), pady=6)
    t15_sb = ttk.Scrollbar(t15_flist, orient='vertical', command=t15_tree.yview)
    t15_sb.grid(row=0, column=1, sticky='ns', padx=(0, 4), pady=6)
    t15_tree.configure(yscrollcommand=t15_sb.set)

    def t15_render_tree():
        t15_tree.delete(*t15_tree.get_children())
        for i, r in enumerate(t15_file_rows):
            mark = '☑' if r['enabled'] else '☐'
            tag = ('skip',) if not r['enabled'] else (r['kind'],)
            size = r.get('size', 0)
            size_s = _fmt_size(size) if size else ''
            t15_tree.insert('', 'end', iid=str(i),
                            values=(mark, i+1, r['kind'].upper(), r['rel'], size_s),
                            tags=tag)
        # 刷新顶层下拉
        tops = sorted({r['top'] for r in t15_file_rows if r.get('top')})
        t15_top_combo['values'] = tops
        if not t15_top_module.get() and tops:
            t15_top_module.set(tops[0])

    def t15_toggle_selected():
        for iid in t15_tree.selection():
            i = int(iid)
            t15_file_rows[i]['enabled'] = not t15_file_rows[i]['enabled']
        t15_render_tree()

    def t15_move(delta):
        sels = t15_tree.selection()
        if not sels:
            return
        idxs = sorted([int(i) for i in sels], reverse=(delta > 0))
        for i in idxs:
            j = i + delta
            if 0 <= j < len(t15_file_rows):
                t15_file_rows[i], t15_file_rows[j] = t15_file_rows[j], t15_file_rows[i]
        t15_render_tree()
        # 保持选中
        for i in idxs:
            j = i + delta
            if 0 <= j < len(t15_file_rows):
                t15_tree.selection_add(str(j))

    t15_btnrow = ttk.Frame(t15_flist)
    t15_btnrow.grid(row=1, column=0, columnspan=2, sticky='w', padx=8, pady=(0, 6))
    ttk.Button(t15_btnrow, text='☑ 切换勾选', style='Small.TButton',
               command=t15_toggle_selected).pack(side='left', padx=2)
    ttk.Button(t15_btnrow, text='↑ 上移', style='Small.TButton',
               command=lambda: t15_move(-1)).pack(side='left', padx=2)
    ttk.Button(t15_btnrow, text='↓ 下移', style='Small.TButton',
               command=lambda: t15_move(1)).pack(side='left', padx=2)
    ttk.Button(t15_btnrow, text='全选', style='Small.TButton',
               command=lambda: (r.update(enabled=True) for r in t15_file_rows) or t15_render_tree()
               ).pack(side='left', padx=2)
    ttk.Button(t15_btnrow, text='全不选', style='Small.TButton',
               command=lambda: (r.update(enabled=False) for r in t15_file_rows) or t15_render_tree()
               ).pack(side='left', padx=2)

    # ── 4) 扫描 ──
    _T15_INCLUDE_DIRS = {'src', 'sim', 'hdl', 'rtl', 'tb', 'test', 'tests',
                         'testbench', 'ip_src', 'sim_src'}
    _T15_EXCLUDE_DIRS = {'.xpr.cache', '.runs', '.ip_user_files', '.sim',
                         '.git', '.svn', '.cache', '__pycache__', 'work',
                         '.Xil', 'vivado_ip', 'ip_repo'}

    def t15_scan():
        proj = t15_proj_dir.get().strip()
        if not proj or not os.path.isdir(proj):
            messagebox.showinfo('提示', '请先选择有效的工程根目录')
            return
        t15_file_rows = []
        vhdl_re = re.compile(r'^\s*(?:entity|architecture)\s+(\w+)', re.M)
        vlog_re = re.compile(r'^\s*module\s+(\w+)\s*(?:\(|;|#)', re.M)
        sv_class_re = re.compile(r'^\s*(?:class|interface|package)\s+(\w+)', re.M)

        for root, dirs, files in os.walk(proj):
            # 排除目录
            dirs[:] = [d for d in dirs
                       if d not in _T15_EXCLUDE_DIRS
                       and not d.startswith('.')]
            # 仅在 include 目录中扫 (除非目录里只有 .v/.sv/.vhd 几个常见扩展)
            rel = os.path.relpath(root, proj)
            top_dir = rel.split(os.sep)[0] if rel != '.' else ''
            in_include = (rel == '.' or top_dir.lower() in _T15_INCLUDE_DIRS)
            if not in_include:
                continue
            for fn in files:
                ext = os.path.splitext(fn)[1].lower()
                if ext in ('.vhd', '.vhdl'):
                    kind = 'vhdl'
                elif ext in ('.v', '.sv', '.svh'):
                    kind = 'vlog'
                else:
                    continue
                fp = os.path.join(root, fn)
                try:
                    sz = os.path.getsize(fp)
                except OSError:
                    sz = 0
                # 提取顶层名
                top = ''
                try:
                    with open(fp, 'r', encoding='utf-8', errors='replace') as fh:
                        src = fh.read(64 * 1024)
                    m = (vhdl_re.search(src) if kind == 'vhdl' else vlog_re.search(src))
                    if m:
                        top = m.group(1)
                except Exception:
                    pass
                t15_file_rows.append({
                    'path': fp,
                    'rel': os.path.relpath(fp, proj).replace('\\', '/'),
                    'kind': kind,
                    'enabled': True,
                    'size': sz,
                    'top': top,
                })

        # 排序: VHDL 先 (VHDL 必须先编译), 然后 VLOG; 同类内按 rel 字典序
        t15_file_rows.sort(key=lambda r: (0 if r['kind'] == 'vhdl' else 1, r['rel']))

        # 提升到模块作用域, 让其它函数访问
        globals()['t15_file_rows'] = t15_file_rows
        t15_render_tree()
        t15_status.set(f'✔ 扫描到 {len(t15_file_rows)} 个文件')

    t15_tree.bind('<space>', lambda e: t15_toggle_selected())

    # ── 5) 生成文件 ──
    t15_gen_box = ttk.LabelFrame(t15, text=' 生成与运行 ')
    t15_gen_box.grid(row=3, column=0, sticky='ew', padx=12, pady=4)
    t15_gen_box.grid_columnconfigure(8, weight=1)

    ttk.Checkbutton(t15_gen_box, text='生成 sim.do', variable=t15_gen_do
                    ).grid(row=0, column=0, padx=8, pady=6)
    ttk.Checkbutton(t15_gen_box, text='生成 sim.tcl', variable=t15_gen_tcl
                    ).grid(row=0, column=1, padx=8, pady=6)
    ttk.Checkbutton(t15_gen_box, text='生成 sim.f', variable=t15_gen_f
                    ).grid(row=0, column=2, padx=8, pady=6)
    ttk.Separator(t15_gen_box, orient='vertical').grid(
        row=0, column=3, sticky='ns', padx=6, pady=2)
    ttk.Label(t15_gen_box, text='Run:', font=(F, 9)).grid(
        row=0, column=4, padx=(4, 2), pady=6, sticky='e')
    ttk.Entry(t15_gen_box, textvariable=t15_run_time, width=8, font=(F, 9)
              ).grid(row=0, column=5, padx=2, pady=6)
    ttk.Label(t15_gen_box, text='分辨率:', font=(F, 9)).grid(
        row=0, column=6, padx=(8, 2), pady=6, sticky='e')
    ttk.Combobox(t15_gen_box, textvariable=t15_resolution, width=6,
                 values=['1fs', '10fs', '100fs', '1ps', '10ps', '100ps', '1ns'],
                 state='readonly', font=(F, 9)).grid(row=0, column=7, padx=2, pady=6)

    ttk.Button(t15_gen_box, text='📝 仅生成文件', style='Accent.TButton',
               command=lambda: t15_generate_files(run_after=False)).grid(
        row=0, column=8, padx=4, pady=6, sticky='e')
    ttk.Button(t15_gen_box, text='▶ 后台跑 (vsim -c)', style='Accent.TButton',
               command=lambda: t15_generate_files(run_after='batch')).grid(
        row=0, column=9, padx=4, pady=6, sticky='e')
    ttk.Button(t15_gen_box, text='🖥 GUI 加载 (vsim)', style='Accent.TButton',
               command=lambda: t15_generate_files(run_after='gui')).grid(
        row=0, column=10, padx=4, pady=6, sticky='e')

    # ── 6) 日志 ──
    t15_log_frame = ttk.LabelFrame(t15, text=' 运行日志 ')
    t15_log_frame.grid(row=5, column=0, sticky='nsew', padx=12, pady=(4, 10))
    t15_log_frame.grid_rowconfigure(0, weight=1)
    t15_log_frame.grid_columnconfigure(0, weight=1)
    t15_log = tk.Text(t15_log_frame, font=(M, 9), bg='#fafbfc', fg=C['fg'],
                      insertbackground=C['fg'], relief='flat',
                      padx=10, pady=6, wrap='word', height=8)
    t15_log.grid(row=0, column=0, sticky='nsew', padx=8, pady=6)
    t15_log_sb = ttk.Scrollbar(t15_log_frame, orient='vertical', command=t15_log.yview)
    t15_log_sb.grid(row=0, column=1, sticky='ns', padx=(0, 4), pady=6)
    t15_log.configure(yscrollcommand=t15_log_sb.set, state='disabled')

    def t15_log_append(msg, color=None):
        t15_log.configure(state='normal')
        if color:
            t15_log.insert('end', msg + '\n', color)
        else:
            t15_log.insert('end', msg + '\n')
        t15_log.see('end')
        t15_log.configure(state='disabled')

    t15_log.tag_configure('err', foreground=C['red'])
    t15_log.tag_configure('ok',  foreground=C['green'])
    t15_log.tag_configure('inf', foreground=C['blue'])

    # 占位 t15_status (在主状态栏也可, 但 t15 内部用)
    t15_status = tk.StringVar(value='● 就绪')
    ttk.Label(t15_gen_box, textvariable=t15_status, font=(F, 9),
              foreground=C['sub']).grid(row=0, column=11, padx=8, pady=6, sticky='e')

    # ── 7) 生成 do/tcl/f ──
    def t15_top_name():
        """优先用手动输入, 否则用下拉值"""
        m = t15_top_manual.get().strip()
        if m:
            return m
        return t15_top_module.get().strip()

    def t15_build_do(sim_dir, top):
        files = [r for r in t15_file_rows if r['enabled']]
        lines = [
            '# =====================================',
            '# sim.do — 自动生成 by FPGA Toolbox',
            f'# {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
            '# =====================================',
            '',
            '# 设置工作目录 (do 文件位置)',
            f'cd {{{sim_dir}}}',
            '',
            '# 解析参数: vsim -do "do sim.do [run_time]"',
            'if {$argc >= 1} {',
            '  set RUN_TIME [lindex $argv 0]',
            '} else {',
            f'  set RUN_TIME {{{t15_run_time.get()}}}',
            '}',
            '',
        ]
        # 库
        for n, p in t15_libraries:
            if p:
                lines.append(f'vlib {p}')
                lines.append(f'vmap {n} {p}')
            else:
                lines.append(f'vlib {n}')
                lines.append(f'vmap {n} {n}')
        lines.append('')
        # 编译
        lines.append('# 编译源文件 (VHDL 先, VLOG 后)')
        for r in files:
            if r['kind'] == 'vhdl':
                lines.append(f'vcom -2008 -work work ../../{r["rel"]}')
            else:
                sv = '-sv ' if r['rel'].endswith('.sv') or r['rel'].endswith('.svh') else ''
                lines.append(f'vlog {sv}-work work ../../{r["rel"]}')
        lines.append('')
        # 顶层
        if top:
            lines.append('# 启动仿真')
            lines.append(f'vsim -t {t15_resolution.get()} work.{top}')
            lines.append('')
            lines.append('# 运行')
            lines.append('add wave -r /*')
            lines.append('run $RUN_TIME')
            lines.append('wave zoom full')
        else:
            lines.append('# 未指定顶层, 跳过 vsim')
        lines.append('')
        return '\n'.join(lines)

    def t15_build_tcl(sim_dir, top):
        files = [r for r in t15_file_rows if r['enabled']]
        lines = [
            '# =====================================',
            '# sim.tcl — 自动生成 by FPGA Toolbox',
            '# Vivado/Questa 通用 (Vivado: source sim.tcl)',
            '# =====================================',
            'if {[catch {set ::argv}] == 0 && $argc >= 1} {',
            '  set RUN_TIME [lindex $argv 0]',
            '} else {',
            f'  set RUN_TIME {{{t15_run_time.get()}}}',
            '}',
            '',
        ]
        for n, p in t15_libraries:
            if p:
                lines.append(f'vlib {p}')
            else:
                lines.append(f'vlib {n}')
        for r in files:
            if r['kind'] == 'vhdl':
                lines.append(f'vcom -2008 -work work ../{r["rel"]}')
            else:
                sv = '-sv ' if r['rel'].endswith('.sv') or r['rel'].endswith('.svh') else ''
                lines.append(f'vlog {sv}-work work ../{r["rel"]}')
        if top:
            lines.append(f'vsim -t {t15_resolution.get()} work.{top}')
            lines.append('add wave -r /*')
            lines.append('run $RUN_TIME')
        lines.append('')
        return '\n'.join(lines)

    def t15_build_f(sim_dir):
        lines = [
            '# sim.f — 自动生成 by FPGA Toolbox',
            '# 用于:  vsim -f sim.f  /  vlog -f sim.f',
            '',
            '+libext+.v+.sv+.svh+.vhd+.vhdl',
        ]
        for r in t15_file_rows:
            if r['enabled']:
                lines.append(f'../{r["rel"]}')
        return '\n'.join(lines)

    def t15_generate_files(run_after=False):
        proj = t15_proj_dir.get().strip()
        if not proj:
            messagebox.showinfo('提示', '请先选择 Vivado 工程根目录')
            return
        if not t15_file_rows:
            t15_log_append('[!] 没有可生成的文件, 请先扫描', 'err')
            return
        sim_dir = os.path.join(proj, 'sim')
        try:
            os.makedirs(sim_dir, exist_ok=True)
        except Exception as e:
            t15_log_append(f'[!] 创建 sim 目录失败: {e}', 'err')
            return
        top = t15_top_name()
        if not top:
            t15_log_append('[!] 未指定顶层模块 (在 "仿真配置" 区选/输入)', 'err')
            return

        # 写文件
        written = []
        if t15_gen_do.get():
            p = os.path.join(sim_dir, 'sim.do')
            with open(p, 'w', encoding='utf-8') as f:
                f.write(t15_build_do(sim_dir, top))
            written.append(('sim.do', p))
        if t15_gen_tcl.get():
            p = os.path.join(sim_dir, 'sim.tcl')
            with open(p, 'w', encoding='utf-8') as f:
                f.write(t15_build_tcl(sim_dir, top))
            written.append(('sim.tcl', p))
        if t15_gen_f.get():
            p = os.path.join(sim_dir, 'sim.f')
            with open(p, 'w', encoding='utf-8') as f:
                f.write(t15_build_f(sim_dir))
            written.append(('sim.f', p))

        for name, p in written:
            t15_log_append(f'[+] 已生成: {name}  ({p})', 'ok')
        t15_status.set(f'✔ 已生成 {len(written)} 个文件')

        if not run_after:
            return
        if not t15_modelsim_path.get().strip():
            t15_log_append('[!] 未配置 vsim.exe, 跳过后续运行', 'err')
            return

        # 启动 vsim
        vsim = t15_modelsim_path.get().strip()
        try:
            if run_after == 'batch':
                t15_log.append(f'[→] 后台启动:  {vsim} -c -do sim.do {t15_run_time.get()}')
                t15_log.see('end')
                proc = subprocess.Popen(
                    [vsim, '-c', '-do', 'sim.do', t15_run_time.get()],
                    cwd=sim_dir,
                    stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    text=True, env={**os.environ, 'TERM': 'dumb'})
                # 后台线程收集输出
                def _drain():
                    for line in proc.stdout:
                        try:
                            root.after(0, lambda l=line.rstrip(): t15_log_append(l, 'inf'))
                        except Exception:
                            pass
                    proc.wait()
                    root.after(0, lambda: t15_log_append(
                        f'[<] vsim 退出, rc={proc.returncode}', 'ok' if proc.returncode == 0 else 'err'))
                import threading as _thr
                _thr.Thread(target=_drain, daemon=True).start()
            else:  # gui
                t15_log_append(f'[→] GUI 启动:  {vsim} -do sim.do', 'inf')
                subprocess.Popen([vsim, '-do', 'sim.do'], cwd=sim_dir)
        except FileNotFoundError:
            t15_log_append(f'[!] 找不到 vsim: {vsim}', 'err')
        except Exception as e:
            t15_log_append(f'[!] 启动失败: {e}', 'err')

    # 加载上次保存的 Modelsim 路径
    _t15_cfg = _load_config()
    if _t15_cfg.get('modelsim_path'):
        t15_modelsim_path.set(_t15_cfg['modelsim_path'])

    def t15_save_modelsim_path(*_):
        c = _load_config()
        c['modelsim_path'] = t15_modelsim_path.get().strip()
        _save_config(c)
    t15_modelsim_path.trace_add('write', lambda *_: t15_save_modelsim_path())

    # 自动尝试扫描 (如果工程根已填)
    if t15_proj_dir.get().strip() and os.path.isdir(t15_proj_dir.get().strip()):
        t15_scan()

    # ══════════════════════════════════════════════════════════════
    # TAB 16 — 设置 (Vivado 路径 / DocNav 路径 全局管理)
    # ══════════════════════════════════════════════════════════════
    from src.app_config import (
        get_vivado_paths, add_vivado_path, remove_vivado_path, validate_vivado_bin,
        get_docnav_paths, add_docnav_path, remove_docnav_path, validate_docnav_dir,
        get_vivado_bin_dirs, get_valid_docnav_dirs, reload as appcfg_reload,
    )

    t16 = ttk.Frame(nb, style='TFrame')
    nb.add(t16, text='⚙ 设置')
    t16.grid_rowconfigure(0, weight=0)   # 工具栏
    t16.grid_rowconfigure(1, weight=1)   # 主内容区 (Vivado/DocNav 列表)
    t16.grid_rowconfigure(2, weight=0)   # 状态栏
    t16.grid_rowconfigure(3, weight=0)   # 路径汇总 (固定高度)
    t16.grid_columnconfigure(0, weight=1)

    # ═══ 工具按钮栏 (row 0) ═══
    t16_toolbar = ttk.Frame(t16, style='TFrame')
    t16_toolbar.grid(row=0, column=0, sticky='ew', padx=12, pady=(10, 4))
    ttk.Button(t16_toolbar, text='🔄 刷新所有路径状态',
               command=lambda: _t16_refresh_all(),
               style='Accent.TButton').pack(side='left', padx=(0, 10))

    t16_status = tk.StringVar(value='配置全局路径后，其他 Tab 将自动读取')
    ttk.Label(t16, textvariable=t16_status, foreground=C['sub'],
              font=(F, 9)).grid(row=2, column=0, sticky='w', padx=16, pady=(2, 2))

    # ═══ 主内容区: 左右两栏 (Vivado | DocNav) ═══
    t16_main = ttk.Frame(t16, style='TFrame')
    t16_main.grid(row=1, column=0, sticky='nsew', padx=12, pady=(4, 4))
    t16_main.grid_rowconfigure(0, weight=1)
    t16_main.grid_columnconfigure(0, weight=1)
    t16_main.grid_columnconfigure(1, weight=1)

    # ─── 左栏: Vivado 路径管理 ───
    t16_viv_frame = ttk.LabelFrame(t16_main, text=' 🛠 Vivado 路径 (到 bin/ 层) ')
    t16_viv_frame.grid(row=0, column=0, sticky='nsew', padx=(0, 8))
    t16_viv_frame.grid_rowconfigure(1, weight=1)
    t16_viv_frame.grid_columnconfigure(0, weight=1)

    ttk.Label(t16_viv_frame, text='设置 Vivado 安装的 bin 目录，供 压缩/国产化/约束生成 等 Tab 使用',
              foreground=C['sub'], font=(F, 8)).grid(
        row=0, column=0, sticky='w', padx=14, pady=(8, 4))

    # Vivado 路径 Listbox
    t16_viv_list_frame = tk.Frame(t16_viv_frame, bg=C['card'],
                                  highlightbackground=C['bd'],
                                  highlightthickness=1, bd=0)
    t16_viv_list_frame.grid(row=1, column=0, sticky='nsew', padx=14, pady=(2, 4))
    t16_viv_list_frame.grid_rowconfigure(0, weight=1)
    t16_viv_list_frame.grid_columnconfigure(0, weight=1)

    t16_viv_list = tk.Listbox(t16_viv_list_frame, font=(M, 9), bg=C['ebg'],
                               fg=C['fg'], selectbackground='#cfe2f3',
                               selectforeground=C['fg'], relief='flat',
                               activestyle='none', exportselection=False,
                               selectmode='extended')
    t16_viv_list.grid(row=0, column=0, sticky='nsew', padx=1, pady=1)
    t16_viv_scroll = tk.Scrollbar(t16_viv_list_frame, orient='vertical',
                                   bg=C['card'], troughcolor=C['bg'],
                                   relief='flat', width=10)
    t16_viv_scroll.grid(row=0, column=1, sticky='ns', pady=1)
    t16_viv_list.config(yscrollcommand=t16_viv_scroll.set)
    t16_viv_scroll.config(command=t16_viv_list.yview)

    # Vivado 按钮
    t16_viv_btns = ttk.Frame(t16_viv_frame, style='TFrame')
    t16_viv_btns.grid(row=2, column=0, sticky='ew', padx=14, pady=(2, 8))
    ttk.Button(t16_viv_btns, text='➕ 添加路径',
               command=lambda: _t16_add_vivado(),
               style='Accent.TButton').pack(side='left', padx=(0, 6))
    ttk.Button(t16_viv_btns, text='➖ 删除选中',
               command=lambda: _t16_del_vivado(),
               style='Normal.TButton').pack(side='left', padx=(0, 6))
    ttk.Button(t16_viv_btns, text='✔ 验证路径',
               command=lambda: _t16_validate_vivado(),
               style='Small.TButton').pack(side='left')

    # ─── 右栏: DocNav 路径管理 ───
    t16_dn_frame = ttk.LabelFrame(t16_main, text=' 📚 DocNav 路径 (安装目录) ')
    t16_dn_frame.grid(row=0, column=1, sticky='nsew', padx=(8, 0))
    t16_dn_frame.grid_rowconfigure(1, weight=1)
    t16_dn_frame.grid_columnconfigure(0, weight=1)

    ttk.Label(t16_dn_frame, text='设置 DocNav 安装目录 (含 resources/xdocs.xml)，供 IP文档 Tab 搜索下载使用',
              foreground=C['sub'], font=(F, 8)).grid(
        row=0, column=0, sticky='w', padx=14, pady=(8, 4))

    # DocNav 路径 Listbox
    t16_dn_list_frame = tk.Frame(t16_dn_frame, bg=C['card'],
                                  highlightbackground=C['bd'],
                                  highlightthickness=1, bd=0)
    t16_dn_list_frame.grid(row=1, column=0, sticky='nsew', padx=14, pady=(2, 4))
    t16_dn_list_frame.grid_rowconfigure(0, weight=1)
    t16_dn_list_frame.grid_columnconfigure(0, weight=1)

    t16_dn_list = tk.Listbox(t16_dn_list_frame, font=(M, 9), bg=C['ebg'],
                              fg=C['fg'], selectbackground='#cfe2f3',
                              selectforeground=C['fg'], relief='flat',
                              activestyle='none', exportselection=False,
                              selectmode='extended')
    t16_dn_list.grid(row=0, column=0, sticky='nsew', padx=1, pady=1)
    t16_dn_scroll = tk.Scrollbar(t16_dn_list_frame, orient='vertical',
                                  bg=C['card'], troughcolor=C['bg'],
                                  relief='flat', width=10)
    t16_dn_scroll.grid(row=0, column=1, sticky='ns', pady=1)
    t16_dn_list.config(yscrollcommand=t16_dn_scroll.set)
    t16_dn_scroll.config(command=t16_dn_list.yview)

    # DocNav 按钮
    t16_dn_btns = ttk.Frame(t16_dn_frame, style='TFrame')
    t16_dn_btns.grid(row=2, column=0, sticky='ew', padx=14, pady=(2, 8))
    ttk.Button(t16_dn_btns, text='➕ 添加路径',
               command=lambda: _t16_add_docnav(),
               style='Accent.TButton').pack(side='left', padx=(0, 6))
    ttk.Button(t16_dn_btns, text='➖ 删除选中',
               command=lambda: _t16_del_docnav(),
               style='Normal.TButton').pack(side='left', padx=(0, 6))
    ttk.Button(t16_dn_btns, text='✔ 验证路径',
               command=lambda: _t16_validate_docnav(),
               style='Small.TButton').pack(side='left')

    # ═══ 底部: 路径汇总 (row 3) ═══
    t16_summary_frame = ttk.LabelFrame(t16, text=' 📋 全局路径汇总 (其他 Tab 将使用这些路径) ')
    t16_summary_frame.grid(row=3, column=0, sticky='ew', padx=12, pady=(4, 10))
    t16_summary_frame.grid_columnconfigure(0, weight=1)

    t16_summary_text = tk.Text(t16_summary_frame, font=(M, 9), bg=C['ebg'],
                                fg=C['fg'], relief='flat', padx=14, pady=10,
                                wrap='word', height=4, state='disabled')
    t16_summary_text.grid(row=0, column=0, sticky='ew', padx=12, pady=(6, 8))

    # ─── 刷新函数 ───
    def _t16_populate_vivado():
        t16_viv_list.delete(0, 'end')
        valid = get_vivado_bin_dirs()
        all_paths = get_vivado_paths()
        for p in all_paths:
            marker = '✔ ' if p in valid else '✘ '
            t16_viv_list.insert('end', marker + p)
            if p not in valid:
                t16_viv_list.itemconfig('end', fg=C['red'])

    def _t16_populate_docnav():
        t16_dn_list.delete(0, 'end')
        valid = get_valid_docnav_dirs()
        all_paths = get_docnav_paths()
        for p in all_paths:
            marker = '✔ ' if p in valid else '✘ '
            t16_dn_list.insert('end', marker + p)
            if p not in valid:
                t16_dn_list.itemconfig('end', fg=C['red'])

    def _t16_update_summary():
        appcfg_reload()
        viv_paths = get_vivado_bin_dirs()
        dn_paths = get_valid_docnav_dirs()
        lines = []
        lines.append(f'Vivado 有效路径 ({len(viv_paths)} 个):')
        if viv_paths:
            for p in viv_paths:
                lines.append(f'    ✔ {p}')
        else:
            lines.append('    (未配置)')
        lines.append('')
        lines.append(f'DocNav 有效路径 ({len(dn_paths)} 个):')
        if dn_paths:
            for p in dn_paths:
                lines.append(f'    ✔ {p}')
        else:
            lines.append('    (未配置)')
        lines.append('')
        lines.append(f'配置文件: {os.path.join(os.path.expanduser("~"), ".fpga_tool", "app_config.json")}')

        t16_summary_text.config(state='normal')
        t16_summary_text.delete('1.0', 'end')
        t16_summary_text.insert('1.0', '\n'.join(lines))
        t16_summary_text.config(state='disabled')

    def _t16_add_vivado():
        d = filedialog.askdirectory(title='选择 Vivado 安装目录 或 bin 目录 (自动探测)')
        if not d:
            return
        ok, msg = add_vivado_path(d)
        if ok:
            t16_status.set(f'✔ 已添加 Vivado: {msg}')
        else:
            messagebox.showwarning('无效路径', f'{msg}\n{d}')
            t16_status.set(f'✘ {msg}')
        _t16_refresh_all()

    def _t16_del_vivado():
        sel = t16_viv_list.curselection()
        if not sel:
            messagebox.showinfo('提示', '请先在列表中选中要删除的路径 (支持Ctrl多选)')
            return
        all_paths = get_vivado_paths()
        to_del = [all_paths[i] for i in reversed(sel) if i < len(all_paths)]
        if not to_del:
            return
        if not messagebox.askyesno('确认删除', f'确定要删除 {len(to_del)} 个 Vivado 路径吗？\n\n' + '\n'.join(to_del)):
            return
        for p in to_del:
            remove_vivado_path(p)
        _t16_refresh_all()
        t16_status.set(f'已删除 {len(to_del)} 个 Vivado 路径')

    def _t16_validate_vivado():
        sel = t16_viv_list.curselection()
        if not sel:
            messagebox.showinfo('提示', '请先选中要验证的路径')
            return
        all_paths = get_vivado_paths()
        idx = sel[0]
        if idx < len(all_paths):
            p = all_paths[idx]
            if validate_vivado_bin(p):
                messagebox.showinfo('验证结果', f'✔ 有效\nvivado.exe 存在\n{p}')
            else:
                messagebox.showwarning('验证结果', f'✘ 无效\n未找到 vivado.exe\n{p}')

    def _t16_add_docnav():
        d = filedialog.askdirectory(title='选择 DocNav 安装目录')
        if not d:
            return
        ok, msg = add_docnav_path(d)
        if ok:
            t16_status.set(f'✔ 已添加 DocNav: {msg}')
        else:
            messagebox.showwarning('无效路径', f'{msg}\n{d}')
            t16_status.set(f'✘ {msg}')
        _t16_refresh_all()

    def _t16_del_docnav():
        sel = t16_dn_list.curselection()
        if not sel:
            messagebox.showinfo('提示', '请先选中要删除的路径 (支持Ctrl多选)')
            return
        all_paths = get_docnav_paths()
        to_del = [all_paths[i] for i in reversed(sel) if i < len(all_paths)]
        if not to_del:
            return
        if not messagebox.askyesno('确认删除', f'确定要删除 {len(to_del)} 个 DocNav 路径吗？\n\n' + '\n'.join(to_del)):
            return
        for p in to_del:
            remove_docnav_path(p)
        _t16_refresh_all()
        t16_status.set(f'已删除 {len(to_del)} 个 DocNav 路径')

    def _t16_validate_docnav():
        sel = t16_dn_list.curselection()
        if not sel:
            messagebox.showinfo('提示', '请先选中要验证的路径')
            return
        all_paths = get_docnav_paths()
        idx = sel[0]
        if idx < len(all_paths):
            p = all_paths[idx]
            if validate_docnav_dir(p):
                messagebox.showinfo('验证结果', f'✔ 有效\nresources/xdocs.xml 存在\n{p}')
            else:
                messagebox.showwarning('验证结果', f'✘ 无效\n未找到 resources/xdocs.xml\n{p}')

    def _t16_refresh_all():
        _t16_populate_vivado()
        _t16_populate_docnav()
        _t16_update_summary()

    # 初始化
    _t16_refresh_all()

    # Tab 切换时自动刷新路径状态条
    def _on_tab_change(event=None):
        try:
            cur = str(nb.select())
            if cur == str(t2): t2_refresh_viv_path()
            elif cur == str(t4): _t4_refresh_viv()
            elif cur == str(t6): _t6_refresh_dn()
            elif cur == str(t9):
                _a2_refresh_viv()
                _a3_refresh_viv()
        except Exception:
            pass
    nb.bind('<<NotebookTabChanged>>', _on_tab_change, add='+')

    root.mainloop()
